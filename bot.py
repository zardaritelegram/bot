

import os
from threading import Thread
from http.server import HTTPServer, BaseHTTPRequestHandler

# ╔══════════════════════════════════════════════════════════════╗
# ║                    IMPORTS & CONFIG                          ║
# ╚══════════════════════════════════════════════════════════════╝
import httpx
import sqlite3
import asyncio
import io
import ast
import re
import time as time_module
import csv
import multiprocessing
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes, MessageHandler, filters

TOKEN = os.environ.get("TOKEN", "")
# ADMIN_ID می‌تواند یک عدد تنها یا چند عدد جدا‌شده با کاما باشد (چند مالک اصلی)
# مثال: ADMIN_ID=111111,222222
_admin_id_raw = os.environ.get("ADMIN_ID", "0")
PRIMARY_ADMIN_IDS = set()
for _part in _admin_id_raw.split(","):
    _part = _part.strip()
    if _part.isdigit() or (_part.startswith("-") and _part[1:].isdigit()):
        PRIMARY_ADMIN_IDS.add(int(_part))
if not PRIMARY_ADMIN_IDS:
    PRIMARY_ADMIN_IDS.add(0)
ADMIN_ID = next(iter(PRIMARY_ADMIN_IDS))  # اولین مالک اصلی -- برای سازگاری با کد قدیمی که یک ADMIN_ID تکی انتظار دارد
GOLD_API_KEY = "goldapi-d23da414dfdcbbe06a2e2ce8d28a095c-io"
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")


# ╔══════════════════════════════════════════════════════════════╗
# ║                       DATABASE                               ║
# ║  توابع دیتابیس - دست نزن                                    ║
# ╚══════════════════════════════════════════════════════════════╝
def init_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        user_id INTEGER PRIMARY KEY, username TEXT, first_name TEXT,
        status TEXT DEFAULT 'pending', approved_at TEXT,
        expires_at TEXT, daily_msg INTEGER DEFAULT 1,
        referral_code TEXT, referred_by INTEGER, referred_count INTEGER DEFAULT 0,
        join_method TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS watchlist (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, symbol TEXT, market TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS alarms (
        id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER,
        symbol TEXT, market TEXT, target REAL, direction TEXT,
        active INTEGER DEFAULT 1, created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS user_state (
        user_id INTEGER PRIMARY KEY, state TEXT, data TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS referrals (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        referrer_id INTEGER, referred_id INTEGER UNIQUE, created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS failed_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, module TEXT, query_text TEXT, created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS capital_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, fulfilled INTEGER DEFAULT 0, created_at TEXT)''')  # ← حذف ماژول ربات مدیریت سرمایه = پاک کن این خط
    c.execute('''CREATE TABLE IF NOT EXISTS msg_chats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, status TEXT DEFAULT 'pending', admin_id INTEGER, created_at TEXT)''')  # ← حذف فیچر پیام به ادمین = پاک کن این ۲ جدول
    c.execute('''CREATE TABLE IF NOT EXISTS msg_chat_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chat_id INTEGER, user_id INTEGER, admin_id INTEGER,
        direction TEXT, content_summary TEXT, created_at TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS bot_settings (
        key TEXT PRIMARY KEY, value TEXT)''')  # ← حذف اگه هیچ فیچری ازش استفاده نکرد = پاک کن این خط
    c.execute('''CREATE TABLE IF NOT EXISTS trial_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, created_at TEXT)''')  # ← حذف فیچر نسخه تست ربات مدیریت سرمایه = پاک کن این خط
    c.execute('''CREATE TABLE IF NOT EXISTS license_chats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER, status TEXT DEFAULT 'pending', created_at TEXT)''')  # ← حذف فیچر چت لایسنس = پاک کن این خط
    c.execute('''CREATE TABLE IF NOT EXISTS admin_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        label TEXT, file_type TEXT, file_id TEXT, caption TEXT, created_at TEXT)''')  # ← حذف کتابخانه فایل ادمین = پاک کن این خط
    c.execute('''CREATE TABLE IF NOT EXISTS admin_texts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        label TEXT, content TEXT, created_at TEXT)''')  # ← حذف کتابخانه متن ادمین = پاک کن این خط
    c.execute('''CREATE TABLE IF NOT EXISTS admins (
        user_id INTEGER PRIMARY KEY, added_by INTEGER, added_at TEXT)''')  # ← حذف چندادمینی = پاک کن این خط
    c.execute('''CREATE TABLE IF NOT EXISTS social_links (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        label TEXT, url TEXT, created_at TEXT)''')  # ← حذف لینک‌های تبلیغاتی = پاک کن این خط
    for col_def in [
        "ALTER TABLE users ADD COLUMN daily_msg INTEGER DEFAULT 1",
        "ALTER TABLE users ADD COLUMN referral_code TEXT",
        "ALTER TABLE users ADD COLUMN referred_by INTEGER",
        "ALTER TABLE users ADD COLUMN referred_count INTEGER DEFAULT 0",
        "ALTER TABLE users ADD COLUMN join_method TEXT",
        "ALTER TABLE users ADD COLUMN email TEXT",
        "ALTER TABLE users ADD COLUMN phone TEXT",
        "ALTER TABLE users ADD COLUMN email_alerts_enabled INTEGER DEFAULT 0",
        "ALTER TABLE users ADD COLUMN birthdate TEXT",      # ← حذف فیلدهای پروفایل تکمیلی = این ۴ خط را پاک کن
        "ALTER TABLE users ADD COLUMN city TEXT",
        "ALTER TABLE users ADD COLUMN country TEXT",
        "ALTER TABLE users ADD COLUMN occupation TEXT",
        "ALTER TABLE license_chats ADD COLUMN admin_id INTEGER",  # ← حذف چندادمینی = پاک کن این خط
    ]:
        try:
            c.execute(col_def)
        except:
            pass
    conn.commit()
    conn.close()

def get_user(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row

# ─── چندادمینی: PRIMARY_ADMIN_IDS (متغیر محیطی ADMIN_ID، می‌تواند چند ──
# ─── نفر باشد) همیشه «مالک اصلی» است و از طریق پنل قابل حذف نیست؛ ──
# ─── ادمین‌های اضافه از جدول admins میان ────────────────────────
def is_admin(user_id):
    """آیا این کاربر ادمین است؟ (هر مالک اصلی از ENV یا هر کسی که در جدول admins باشد)"""
    if user_id in PRIMARY_ADMIN_IDS:
        return True
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT 1 FROM admins WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row is not None

def is_primary_admin(user_id):
    """فقط مالک(های) اصلی (PRIMARY_ADMIN_IDS از متغیر محیطی) -- برای عملیات حساس مثل افزودن/حذف ادمین دیگر یا دسترسی به اطلاعات کاربران"""
    return user_id in PRIMARY_ADMIN_IDS

def add_admin(user_id, added_by):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO admins (user_id, added_by, added_at) VALUES (?, ?, ?) "
              "ON CONFLICT(user_id) DO NOTHING",
              (user_id, added_by, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def remove_admin(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM admins WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()

def get_all_extra_admins():
    """لیست ادمین‌های اضافه‌شده (بدون هیچ‌کدام از مالکین اصلی) -- برای نمایش در پنل"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT user_id, added_at FROM admins ORDER BY added_at")
    rows = c.fetchall()
    conn.close()
    return rows

# ─── لینک‌های تبلیغاتی/شبکه‌های اجتماعی (فقط مالکین اصلی مدیریت می‌کنند) ──
def add_social_link(label, url):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO social_links (label, url, created_at) VALUES (?, ?, ?)",
              (label, url, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    link_id = c.lastrowid
    conn.commit()
    conn.close()
    return link_id

def get_social_links():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, label, url FROM social_links ORDER BY id")
    rows = c.fetchall()
    conn.close()
    return rows

def delete_social_link(link_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM social_links WHERE id=?", (link_id,))
    conn.commit()
    conn.close()

def get_all_admin_ids():
    """لیست همه‌ی آیدی‌های ادمین (همه‌ی مالکین اصلی + همه‌ی ادمین‌های اضافه) -- برای اطلاع‌رسانی به همه"""
    ids = set(PRIMARY_ADMIN_IDS)
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT user_id FROM admins")
    for row in c.fetchall():
        ids.add(row[0])
    conn.close()
    return list(ids)

async def notify_admins(bot, text, reply_markup=None):
    """
    ارسال یک پیام متنی به همه‌ی ادمین‌ها (مالک اصلی + همه‌ی ادمین‌های
    اضافه‌شده از پنل). اگه پیام یه دکمه‌ی عمل (مثل «آماده‌ام ارسال کنم»)
    داشته باشه، همون دکمه برای همه‌ی ادمین‌ها فرستاده می‌شه؛ هر کدوم
    اول بزنه، همون درخواست رو به عهده می‌گیره (چون state با آیدی خودِ
    همون ادمین ذخیره می‌شه، نه یه مقدار ثابت). ورودی bot می‌تواند
    context.bot یا update.get_bot() باشد -- هر دو یک شیء Bot هستند.
    """
    for aid in get_all_admin_ids():
        try:
            await bot.send_message(aid, text, reply_markup=reply_markup)
        except Exception as e:
            print(f"خطای اطلاع‌رسانی به ادمین {aid}: {e}")

def log_failed_request(user_id, module, query_text):
    """ثبت درخواستی که هیچ نتیجه‌ای برایش پیدا نشد (برای بررسی بعدی توسط ادمین)"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO failed_requests (user_id, module, query_text, created_at) VALUES (?, ?, ?, ?)",
              (user_id, module, query_text, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def save_user(user_id, username, first_name):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT OR IGNORE INTO users (user_id, username, first_name, status, daily_msg) VALUES (?, ?, ?, 'pending', 1)",
              (user_id, username, first_name))
    conn.commit()
    conn.close()

def generate_referral_code(user_id):
    """کد رفرال یکتا برای هر کاربر بر اساس user_id (پایدار و قابل بازتولید)"""
    return f"R{user_id}"

def ensure_referral_code(user_id):
    """اگر کاربر کد رفرال نداشته باشد، یکی برایش می‌سازد و ذخیره می‌کند"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT referral_code FROM users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    if row and row[0]:
        conn.close()
        return row[0]
    code = generate_referral_code(user_id)
    c.execute("UPDATE users SET referral_code=? WHERE user_id=?", (code, user_id))
    conn.commit()
    conn.close()
    return code

def get_user_by_referral_code(code):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT user_id FROM users WHERE referral_code=?", (code,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else None

def record_referral(referrer_id, referred_id):
    """
    ثبت یک دعوت موفق. اگر این کاربر (referred_id) قبلاً یک‌بار توسط هرکسی
    دعوت شده باشد (یعنی از قبل در جدول referrals با هر referrer_id ثبت
    شده)، دوباره ثبت نمی‌شود -- طبق درخواست، هر فرد فقط یک‌بار می‌تواند
    به‌عنوان «دعوت‌شده» برای افزایش شمارنده‌ی معرف محسوب شود.
    خروجی: True اگر ثبت جدید و معتبر بود، False اگر تکراری/نامعتبر بود.
    """
    if referrer_id == referred_id:
        return False
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id FROM referrals WHERE referred_id=?", (referred_id,))
    if c.fetchone():
        conn.close()
        return False
    c.execute("INSERT INTO referrals (referrer_id, referred_id, created_at) VALUES (?, ?, ?)",
              (referrer_id, referred_id, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    c.execute("UPDATE users SET referred_count = referred_count + 1 WHERE user_id=?", (referrer_id,))
    conn.commit()
    conn.close()
    return True

def get_referred_count(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT referred_count FROM users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row and row[0] else 0

def approve_user_by_referral(user_id):
    """تأیید کاربر از طریق مسیر دعوت ۳ نفر (نه ادمین)"""
    now = get_iran_now()
    expires = now + timedelta(days=365)
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE users SET status='active', approved_at=?, expires_at=?, daily_msg=1, join_method='referral' WHERE user_id=?",
              (now.strftime("%Y-%m-%d %H:%M"), expires.strftime("%Y-%m-%d %H:%M"), user_id))
    conn.commit()
    conn.close()
    set_state(user_id, "first_active_start_pending")  # ← حذف تور خوش‌آمد = پاک کن این خط

def approve_user(user_id):
    now = get_iran_now()
    expires = now + timedelta(days=365)
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE users SET status='active', approved_at=?, expires_at=?, daily_msg=1, join_method='admin' WHERE user_id=?",
              (now.strftime("%Y-%m-%d %H:%M"), expires.strftime("%Y-%m-%d %H:%M"), user_id))
    conn.commit()
    conn.close()
    set_state(user_id, "first_active_start_pending")  # ← حذف تور خوش‌آمد = پاک کن این خط

def reject_user(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE users SET status='rejected' WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()

def set_state(user_id, state, data=""):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO user_state (user_id, state, data) VALUES (?, ?, ?)", (user_id, state, data))
    conn.commit()
    conn.close()

def get_state(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT state, data FROM user_state WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row if row else (None, None)

def clear_state(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM user_state WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()

def get_watchlist(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, symbol, market FROM watchlist WHERE user_id=?", (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def add_watchlist(user_id, symbol, market):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id FROM watchlist WHERE user_id=? AND symbol=?", (user_id, symbol))
    if c.fetchone():
        conn.close()
        return False
    c.execute("INSERT INTO watchlist (user_id, symbol, market) VALUES (?, ?, ?)", (user_id, symbol, market))
    conn.commit()
    conn.close()
    return True

def remove_watchlist(wid):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM watchlist WHERE id=?", (wid,))
    conn.commit()
    conn.close()

def get_alarms(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, symbol, market, target, direction FROM alarms WHERE user_id=? AND active=1", (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def add_alarm(user_id, symbol, market, target, direction):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO alarms (user_id, symbol, market, target, direction, active, created_at) VALUES (?, ?, ?, ?, ?, 1, ?)",
              (user_id, symbol, market, target, direction, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def deactivate_alarm(alarm_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE alarms SET active=0 WHERE id=?", (alarm_id,))
    conn.commit()
    conn.close()

def remove_alarm(alarm_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM alarms WHERE id=?", (alarm_id,))
    conn.commit()
    conn.close()

def get_all_active_alarms():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, user_id, symbol, market, target, direction FROM alarms WHERE active=1")
    rows = c.fetchall()
    conn.close()
    return rows


# ╔══════════════════════════════════════════════════════════════╗
# ║                  MODULE: FINANCIAL                           ║
# ║                                                              ║
# ║  برای حذف کامل این ماژول:                                   ║
# ║  ۱. این بلوک رو پاک کن (از اینجا تا خط ═══ بعدی)          ║
# ║  ۲. در MAIN MENU خط financial رو کامنت کن                  ║
# ║  ۳. در ROUTER خط financial رو کامنت کن                     ║
# ╚══════════════════════════════════════════════════════════════╝

IRAN_SYMBOLS = {
    "USD": "price_dollar_rl", "EUR": "price_eur", "GBP": "price_gbp",
    "GOLD18": "geram18", "COIN": "sekee", "HALFCOIN": "nim",
    "QTRCOIN": "rob", "SILVER": "silver_999", "MITHGAL": "mesghal",
}
CRYPTO_SYMBOLS = {
    "BTC": "bitcoin", "ETH": "ethereum", "USDT": "tether",
    "BNB": "binancecoin", "TRX": "tron", "SOL": "solana",
    "ADA": "cardano", "XRP": "ripple", "DOGE": "dogecoin",
    "DOT": "polkadot", "MATIC": "matic-network", "LTC": "litecoin",
    "AVAX": "avalanche-2", "LINK": "chainlink", "UNI": "uniswap",
}
GOLDAPI_SYMBOLS = {
    "GOLD": ("XAU", "USD"), "XAUUSD": ("XAU", "USD"),
    "SILVER": ("XAG", "USD"), "XAGUSD": ("XAG", "USD"),
}
FOREX_YAHOO = {
    "EURUSD": "EURUSD%3DX", "GBPUSD": "GBPUSD%3DX",
    "USDCAD": "USDCAD%3DX", "USDJPY": "USDJPY%3DX",
    "AUDUSD": "AUDUSD%3DX", "USDCHF": "USDCHF%3DX",
    "NZDUSD": "NZDUSD%3DX", "EURGBP": "EURGBP%3DX",
    "EURJPY": "EURJPY%3DX", "GBPJPY": "GBPJPY%3DX",
    "OIL": "BZ%3DF", "BRENT": "BZ%3DF", "WTI": "CL%3DF",
}

def detect_market(symbol):
    s = symbol.upper()
    if s in IRAN_SYMBOLS: return "iran"
    if s in CRYPTO_SYMBOLS: return "crypto"
    if s in GOLDAPI_SYMBOLS: return "goldapi"
    return "forex"

def fmt_price(price, unit):
    if price is None: return "---"
    if unit == "IRR": return f"{price:,}"
    if price < 1000: return f"{price:.5f}"
    return f"{price:,.2f}"

async def get_goldapi_price(metal="XAU", currency="USD"):
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(f"https://www.goldapi.io/api/{metal}/{currency}",
                headers={"x-access-token": GOLD_API_KEY})
            data = r.json()
            price = data.get("price")
            change = data.get("chp", 0)
            return price, change, "🟢" if change >= 0 else "🔴"
    except Exception as e:
        print(f"GoldAPI error: {e}")
        return None, 0, "⚪"

async def get_yahoo_price(symbol):
    try:
        async with httpx.AsyncClient(timeout=10,
            headers={"User-Agent": "Mozilla/5.0"}) as client:
            r = await client.get(f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1m&range=1d")
            meta = r.json()["chart"]["result"][0]["meta"]
            price = meta.get("regularMarketPrice", 0)
            prev = meta.get("chartPreviousClose", meta.get("previousClose", price))
            change = ((price - prev) / prev * 100) if prev else 0
            return price, change, "🟢" if change >= 0 else "🔴"
    except:
        return None, 0, "⚪"

async def get_price_by_symbol(symbol, market):
    symbol = symbol.upper()
    try:
        async with httpx.AsyncClient(timeout=10,
            headers={"User-Agent": "Mozilla/5.0"}) as client:
            if market == "iran":
                key = IRAN_SYMBOLS.get(symbol)
                if not key: return None, "نامشخص"
                r = await client.get("https://call4.tgju.org/ajax.json")
                val = r.json().get("current", {}).get(key, {}).get("p", "").replace(",", "")
                return int(float(val)), "IRR"
            elif market == "crypto":
                cg_id = CRYPTO_SYMBOLS.get(symbol)
                if not cg_id: return None, "نامشخص"
                r = await client.get("https://api.coingecko.com/api/v3/simple/price",
                    params={"ids": cg_id, "vs_currencies": "usd"})
                return r.json().get(cg_id, {}).get("usd"), "USD"
            elif market == "goldapi":
                metal, currency = GOLDAPI_SYMBOLS.get(symbol, ("XAU", "USD"))
                r = await client.get(f"https://www.goldapi.io/api/{metal}/{currency}",
                    headers={"x-access-token": GOLD_API_KEY})
                return r.json().get("price"), "USD"
            else:
                yahoo_sym = FOREX_YAHOO.get(symbol, symbol + "%3DX")
                r = await client.get(f"https://query1.finance.yahoo.com/v8/finance/chart/{yahoo_sym}?interval=1m&range=1d")
                return r.json()["chart"]["result"][0]["meta"].get("regularMarketPrice"), "USD"
    except:
        return None, "نامشخص"

async def validate_symbol(symbol):
    market = detect_market(symbol.upper())
    price, unit = await get_price_by_symbol(symbol, market)
    if price and price > 0:
        return True, market, price, unit
    return False, None, None, None

async def get_iran_prices():
    try:
        async with httpx.AsyncClient(timeout=10,
            headers={"User-Agent": "Mozilla/5.0"}) as client:
            current = (await client.get("https://call4.tgju.org/ajax.json")).json().get("current", {})
            def gp(key):
                try: return int(float(current[key]["p"].replace(",", "")))
                except: return None
            return {"dollar": gp("price_dollar_rl"), "gold18": gp("geram18"),
                    "gold_coin": gp("sekee"), "half_coin": gp("nim"),
                    "quarter": gp("rob"), "gram_gold": gp("mesghal"), "silver": gp("silver_999")}
    except Exception as e:
        print(f"Iran prices error: {e}")
        return None

async def get_crypto_prices():
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            data = (await client.get("https://api.coingecko.com/api/v3/simple/price",
                params={"ids": "bitcoin,ethereum,tether,binancecoin,tron",
                        "vs_currencies": "usd", "include_24hr_change": "true"})).json()
            def fmt(cid):
                c = data.get(cid, {})
                p, ch = c.get("usd", 0), c.get("usd_24h_change", 0)
                return p, ch, "🟢" if ch >= 0 else "🔴"
            return {"bitcoin": fmt("bitcoin"), "ethereum": fmt("ethereum"),
                    "tether": fmt("tether"), "bnb": fmt("binancecoin"), "tron": fmt("tron")}
    except Exception as e:
        print(f"Crypto error: {e}")
        return None

async def get_forex_prices():
    try:
        gp, gc, ga = await get_goldapi_price("XAU", "USD")
        sp, sc, sa = await get_goldapi_price("XAG", "USD")
        return {"gold": (gp, gc, ga), "silver": (sp, sc, sa),
                "oil": await get_yahoo_price("BZ%3DF"),
                "eur": await get_yahoo_price("EURUSD%3DX"),
                "gbp": await get_yahoo_price("GBPUSD%3DX")}
    except Exception as e:
        print(f"Forex error: {e}")
        return None

async def check_alarms(app):
    for alarm_id, user_id, symbol, market, target, direction in get_all_active_alarms():
        try:
            price, unit = await get_price_by_symbol(symbol, market)
            if price is None: continue
            if (direction == "above" and price >= target) or (direction == "below" and price <= target):
                arrow = "📈" if direction == "above" else "📉"
                alert_text = (
                    f"🔔 آلارم فعال شد!\n\n{arrow} {symbol}\n"
                    f"قیمت هدف: {fmt_price(target, unit)}\n"
                    f"قیمت فعلی: {fmt_price(price, unit)} {unit}\n"
                    f"🕐 {get_iran_now().strftime('%H:%M:%S')}"
                )
                await app.bot.send_message(user_id, alert_text)
                # اگر کاربر ایمیل ثبت کرده و هشدار ایمیلی را فعال کرده باشد، ایمیل هم ارسال شود
                db_user = get_user(user_id)
                email = db_user[11] if db_user and len(db_user) > 11 else None
                email_enabled = db_user[13] if db_user and len(db_user) > 13 else 0
                if email and email_enabled:
                    await send_email(email, f"آلارم قیمتی: {symbol}", alert_text)
                deactivate_alarm(alarm_id)
        except Exception as e:
            print(f"Alarm error {alarm_id}: {e}")

async def alarm_loop(app):
    while True:
        await check_alarms(app)
        await asyncio.sleep(60)

def save_capital_request(user_id):
    """ثبت درخواست کاربر برای ربات مدیریت سرمایه، برای پیگیری توسط ادمین"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO capital_requests (user_id, created_at) VALUES (?, ?)",
              (user_id, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    request_id = c.lastrowid
    conn.commit()
    conn.close()
    return request_id

def get_capital_request(request_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT user_id, fulfilled FROM capital_requests WHERE id=?", (request_id,))
    row = c.fetchone()
    conn.close()
    return row

def mark_capital_request_fulfilled(request_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE capital_requests SET fulfilled=1 WHERE id=?", (request_id,))
    conn.commit()
    conn.close()

# ─── تنظیمات کلی کلید-مقدار (برای فایل نسخه‌ی تست و مشابه) ──────
def set_bot_setting(key, value):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO bot_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
              (key, value))
    conn.commit()
    conn.close()

def get_bot_setting(key):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT value FROM bot_settings WHERE key=?", (key,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else None

# ─── درخواست نسخه‌ی تست ربات مدیریت سرمایه ──────────────────────
def save_trial_request(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO trial_requests (user_id, created_at) VALUES (?, ?)",
              (user_id, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

# ─── چت زنده‌ی لایسنس (بین ادمین و کاربر، دوطرفه) ────────────────
def create_license_chat(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    # admin_id از همون اول روی مالک اصلی تنظیم می‌شود -- یعنی ابتدا همیشه
    # اول به دست ادمین اصلی می‌رسد؛ اگر دستی منتقل شود یا ۱۵ دقیقه بی‌پاسخ
    # بماند (escalation خودکار)، این مقدار عوض می‌شود.
    c.execute("INSERT INTO license_chats (user_id, status, admin_id, created_at) VALUES (?, 'pending', ?, ?)",
              (user_id, ADMIN_ID, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    chat_id = c.lastrowid
    conn.commit()
    conn.close()
    return chat_id

def get_license_chat(chat_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT user_id, status, admin_id FROM license_chats WHERE id=?", (chat_id,))
    row = c.fetchone()
    conn.close()
    return row

def update_license_chat_status(chat_id, status):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE license_chats SET status=? WHERE id=?", (status, chat_id))
    conn.commit()
    conn.close()

def assign_license_chat_admin(chat_id, admin_id):
    """ثبت این‌که کدام ادمین مسئول این مکالمه‌ی لایسنس است -- برای مسیریابی درست پیام‌های کاربر"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE license_chats SET admin_id=? WHERE id=?", (admin_id, chat_id))
    conn.commit()
    conn.close()

def get_overdue_pending_license_chats(minutes=15):
    """
    درخواست‌های لایسنسی که هنوز 'pending' هستند (کسی شروعش نکرده)، هنوز
    دست ادمین اصلی مانده (منتقل نشده)، و بیش از {minutes} دقیقه از
    ساختشان گذشته -- برای واگذاری خودکار توسط license_escalation_loop.
    """
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    cutoff = (get_iran_now() - timedelta(minutes=minutes)).strftime("%Y-%m-%d %H:%M")
    c.execute("SELECT id, user_id FROM license_chats WHERE status='pending' AND admin_id=? AND created_at<=?",
              (ADMIN_ID, cutoff))
    rows = c.fetchall()
    conn.close()
    return rows

def get_next_round_robin_admin(setting_key="license_round_robin_index"):
    """
    آیدی ادمین بعدی برای واگذاری خودکار -- به‌ترتیب و مساوی بین
    ادمین‌های اضافه‌شده (نه مالک اصلی) می‌چرخد. اگر هیچ ادمین اضافه‌ای
    وجود نداشته باشد، None برمی‌گرداند (یعنی جایی برای واگذاری نیست).
    setting_key جدا برای هر صف (مثلاً چت لایسنس و پیام به ادمین) تا
    نوبت‌دهی این دو مستقل از هم بماند.
    """
    extra_admins = get_all_extra_admins()
    if not extra_admins:
        return None
    extra_ids = [a[0] for a in extra_admins]
    idx_str = get_bot_setting(setting_key)
    idx = int(idx_str) if idx_str and idx_str.isdigit() else 0
    idx = idx % len(extra_ids)
    next_admin_id = extra_ids[idx]
    set_bot_setting(setting_key, str((idx + 1) % len(extra_ids)))
    return next_admin_id

# ─── کتابخانه‌ی فایل‌های آماده‌ی ادمین (حداکثر ۱۰ عدد) ────────────
def add_admin_file(label, file_type, file_id, caption):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO admin_files (label, file_type, file_id, caption, created_at) VALUES (?, ?, ?, ?, ?)",
              (label, file_type, file_id, caption, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    file_db_id = c.lastrowid
    conn.commit()
    conn.close()
    return file_db_id

def get_admin_files():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, label, file_type, file_id, caption FROM admin_files ORDER BY id")
    rows = c.fetchall()
    conn.close()
    return rows

def get_admin_file(file_db_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT label, file_type, file_id, caption FROM admin_files WHERE id=?", (file_db_id,))
    row = c.fetchone()
    conn.close()
    return row

def update_admin_file(file_db_id, label, file_type, file_id, caption):
    """جایگزینی کامل اسم و محتوای یک فایل کتابخانه، بدون نیاز به حذف و افزودن دوباره"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE admin_files SET label=?, file_type=?, file_id=?, caption=? WHERE id=?",
              (label, file_type, file_id, caption, file_db_id))
    conn.commit()
    conn.close()

def delete_admin_file(file_db_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM admin_files WHERE id=?", (file_db_id,))
    conn.commit()
    conn.close()

def count_admin_files():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM admin_files")
    n = c.fetchone()[0]
    conn.close()
    return n

# ─── کتابخانه‌ی متن‌های آماده‌ی ادمین (نامحدود، قابل ویرایش) ───────
def add_admin_text(label, content):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO admin_texts (label, content, created_at) VALUES (?, ?, ?)",
              (label, content, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    text_id = c.lastrowid
    conn.commit()
    conn.close()
    return text_id

def get_admin_texts():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, label, content FROM admin_texts ORDER BY id")
    rows = c.fetchall()
    conn.close()
    return rows

def get_admin_text(text_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT label, content FROM admin_texts WHERE id=?", (text_id,))
    row = c.fetchone()
    conn.close()
    return row

def update_admin_text(text_id, label, content):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE admin_texts SET label=?, content=? WHERE id=?", (label, content, text_id))
    conn.commit()
    conn.close()

def delete_admin_text(text_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM admin_texts WHERE id=?", (text_id,))
    conn.commit()
    conn.close()

def detect_message_content(msg):
    """
    تشخیص نوع و شناسه‌ی فایل یک پیام تلگرام (برای ذخیره‌سازی طولانی‌مدت،
    نه ارسال مستقیم). خروجی: (file_type, file_id) یا (None, None) اگر
    نوع پشتیبانی نشود. برای متن، file_id همیشه None است (چون متن خودش
    در caption/content ذخیره می‌شود، نه با file_id).
    """
    if msg.document:
        return "document", msg.document.file_id
    if msg.photo:
        return "photo", msg.photo[-1].file_id
    if msg.audio:
        return "audio", msg.audio.file_id
    if msg.video:
        return "video", msg.video.file_id
    if msg.voice:
        return "voice", msg.voice.file_id
    if msg.text:
        return "text", None
    return None, None

# ─── ربات مدیریت سرمایه: نسخه‌ی تست و لایسنس پولی ────────────────
# قیمت واقعی و روش پرداخت از پنل مدیریت → کتابخانه‌ی فایل و متن →
# «💰 متن قیمت‌گذاری لایسنس» تنظیم و ذخیره می‌شود (چون ممکن است
# تغییر کند و نباید داخل کد ثابت باشد). متن زیر فقط یک fallback است
# که فقط تا وقتی ادمین هنوز چیزی تنظیم نکرده نمایش داده می‌شود.
CAPITAL_LICENSE_PRICING_TEXT_FALLBACK = (
    "💳 لایسنس ربات مدیریت سرمایه\n\n"
    "قیمت‌ها و روش پرداخت هنوز از پنل مدیریت تنظیم نشده است.\n\n"
    "برای اطلاع از قیمت و شرایط، «✅ درخواست دارم» را بزنید تا مستقیم "
    "با ادمین در ارتباط باشید."
)

def capital_bot_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🎁 دریافت نسخه تست ۷ روزه", callback_data="cap_trial"),
         InlineKeyboardButton("💳 نسخه پولی (لایسنس‌دار)", callback_data="cap_paid")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")],
    ])

async def send_content_by_type(context, chat_id, file_type, file_id, caption, default_caption=None):
    """
    ارسال محتوا بر اساس نوعش (که قبلاً ذخیره یا از یک پیام تلگرام تشخیص
    داده شده) به یک چت مشخص. برای فایل تست ذخیره‌شده و هر فیچر مشابه
    دیگری قابل استفاده است. خروجی: True در صورت موفقیت، False در خطا.
    """
    try:
        final_caption = caption or default_caption
        if file_type == "document":
            await context.bot.send_document(chat_id=chat_id, document=file_id, caption=final_caption)
        elif file_type == "photo":
            await context.bot.send_photo(chat_id=chat_id, photo=file_id, caption=final_caption)
        elif file_type == "audio":
            await context.bot.send_audio(chat_id=chat_id, audio=file_id, caption=final_caption)
        elif file_type == "video":
            await context.bot.send_video(chat_id=chat_id, video=file_id, caption=final_caption)
        elif file_type == "voice":
            await context.bot.send_voice(chat_id=chat_id, voice=file_id, caption=final_caption)
        elif file_type == "text":
            await context.bot.send_message(chat_id=chat_id, text=final_caption or "")
        else:
            return False
        return True
    except Exception as e:
        print(f"خطای ارسال محتوا با send_content_by_type: {e}")
        return False

async def relay_free_message(msg, context, target_chat_id, prefix=None):
    """
    تشخیص نوع یک پیام تلگرام واقعی (متن/عکس/فایل/صدا/ویدیو -- حتی اگر
    خود پیام از یک کانال یا چت دیگر فوروارد شده باشد، چون msg.photo/
    msg.document/... مستقل از فوروارد بودن همیشه در دسترس‌اند) و ارسال
    مستقیم آن به یک چت مقصد، بدون افشای منبع اصلی. برای ربات مدیریت
    سرمایه، پاسخ به پیام کاربر، و چت زنده‌ی لایسنس مشترک است.
    خروجی: True در صورت موفقیت، False اگر نوع پیام پشتیبانی نشود یا خطا بدهد.
    """
    try:
        caption = msg.caption or None
        if prefix and caption:
            caption = f"{prefix}{caption}"
        elif prefix and not caption:
            caption = prefix.strip()

        if msg.document:
            await context.bot.send_document(chat_id=target_chat_id, document=msg.document.file_id, caption=caption)
        elif msg.photo:
            await context.bot.send_photo(chat_id=target_chat_id, photo=msg.photo[-1].file_id, caption=caption)
        elif msg.audio:
            await context.bot.send_audio(chat_id=target_chat_id, audio=msg.audio.file_id, caption=caption)
        elif msg.video:
            await context.bot.send_video(chat_id=target_chat_id, video=msg.video.file_id, caption=caption)
        elif msg.voice:
            await context.bot.send_voice(chat_id=target_chat_id, voice=msg.voice.file_id, caption=caption)
        elif msg.text:
            text = f"{prefix}{msg.text}" if prefix else msg.text
            await context.bot.send_message(chat_id=target_chat_id, text=text)
        else:
            return False
        return True
    except Exception as e:
        print(f"خطای relay_free_message: {e}")
        return False

def financial_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💵 بازار ایران", callback_data="fin_currency"),
         InlineKeyboardButton("₿ بازار کریپتو", callback_data="fin_crypto")],
        [InlineKeyboardButton("📈 بازار فارکس", callback_data="fin_forex"),
         InlineKeyboardButton("👁 واچ لیست شخصی", callback_data="fin_watchlist")],
        [InlineKeyboardButton("🔔 آلارم قیمت", callback_data="fin_alarm"),
         InlineKeyboardButton("📐 میزان خرید", callback_data="ps_size")],  # ← حذف = پاک کن این خط
        [InlineKeyboardButton("📊 بک‌تست", callback_data="fin_backtest"),  # ← حذف = پاک کن این خط
         InlineKeyboardButton("🧮 بک‌تست اندیکاتور", callback_data="ind_backtest")],  # ← حذف ماژول اندیکاتوری = پاک کن این خط
        [InlineKeyboardButton("📶 اندیکاتور شخصی", callback_data="ind_custom"),  # ← حذف ماژول اندیکاتور شخصی = پاک کن این خط
         InlineKeyboardButton("🧠 تحلیل هوش مصنوعی", callback_data="menu_ai_market")],  # ← حذف ماژول تحلیل بازار = پاک کن این خط
        [InlineKeyboardButton("📰 اخبار مهم امروز/هفته", callback_data="fin_news")],  # ← حذف ماژول اخبار = پاک کن این خط
        [InlineKeyboardButton("🤖 ربات مدیریت سرمایه", callback_data="fin_capital_bot")],  # ← حذف ماژول ربات مدیریت سرمایه = پاک کن این خط
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_main")],
    ])

def watchlist_menu(user_id):
    items = get_watchlist(user_id)
    keyboard = []
    
    # دکمه‌های حذف - 3تایی در هر سطر
    row = []
    for wid, symbol, market in items:
        row.append(InlineKeyboardButton(f"❌ {symbol}", callback_data=f"wl_del_{wid}"))
        if len(row) == 3:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    
    keyboard.append([InlineKeyboardButton("➕ افزودن ارز", callback_data="wl_add")])
    keyboard.append([InlineKeyboardButton("📊 مشاهده قیمت‌ها", callback_data="wl_view")])
    keyboard.append([InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")])
    return InlineKeyboardMarkup(keyboard)

def alarm_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ آلارم جدید", callback_data="alarm_new"),
         InlineKeyboardButton("📋 آلارم‌های فعال", callback_data="alarm_list")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")],
    ])

async def handle_financial(query, user_id, context):
    data = query.data

    if data == "menu_financial":
        await query.edit_message_text("💰 بازارهای مالی:", reply_markup=financial_menu())

    elif data == "back_financial":
        clear_state(user_id)
        await query.edit_message_text("💰 بازارهای مالی:", reply_markup=financial_menu())

    elif data == "fin_capital_bot":
        await query.edit_message_text(
            "🤖 ربات مدیریت سرمایه\n\nلطفاً یکی از گزینه‌های زیر را انتخاب کنید:",
            reply_markup=capital_bot_menu()
        )

    elif data == "cap_trial":
        requester = query.from_user
        save_trial_request(user_id)
        trial_setting = get_bot_setting("trial_file")

        if trial_setting:
            import json
            info = json.loads(trial_setting)
            sent = await send_content_by_type(
                context, user_id, info["file_type"], info.get("file_id"), info.get("caption"),
                default_caption="🎁 نسخه‌ی تست ۷ روزه‌ی ربات مدیریت سرمایه"
            )
            if sent:
                await query.edit_message_text(
                    "✅ نسخه‌ی تست برای شما ارسال شد؛ لطفاً پیام‌های چت را بررسی کنید.",
                    reply_markup=capital_bot_menu()
                )
                await notify_admins(
                    context.bot,
                    f"ℹ️ نسخه‌ی تست به‌صورت خودکار برای {requester.first_name} "
                    f"(@{requester.username or 'ندارد'}, {user_id}) ارسال شد."
                )
                return
            # اگر ارسال خودکار خطا داد (مثلاً file_id دیگر معتبر نیست)، به مسیر دستی زیر می‌رویم

        # فایل تست هنوز تنظیم نشده (یا ارسال خودکارش خطا داد) -- به ادمین اطلاع بده
        request_id = save_capital_request(user_id)
        await notify_admins(
            context.bot,
            f"🎁 درخواست نسخه‌ی تست ۷ روزه (فایل تست خودکار تنظیم نشده)\n\n"
            f"👤 از طرف: {requester.first_name} (@{requester.username or 'ندارد'})\n"
            f"🆔 آیدی: {user_id}\n\n"
            f"می‌توانید از پنل مدیریت، فایل تست را یک‌بار تنظیم کنید تا از این پس "
            f"خودکار ارسال شود، یا همین یک‌بار با دکمه‌ی زیر دستی برایش بفرستید.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📤 آماده‌ام ارسال کنم", callback_data=f"admin_send_capital_{request_id}")]
            ])
        )
        await query.edit_message_text(
            "✅ درخواست شما ثبت شد؛ ادمین به‌زودی نسخه‌ی تست را برایتان ارسال می‌کند.",
            reply_markup=capital_bot_menu()
        )

    elif data == "cap_paid":
        pricing_text = get_bot_setting("pricing_text") or CAPITAL_LICENSE_PRICING_TEXT_FALLBACK
        await query.edit_message_text(
            pricing_text,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ درخواست دارم", callback_data="cap_paid_confirm")],
                [InlineKeyboardButton("❌ انصراف", callback_data="cap_paid_reject")],
            ])
        )

    elif data == "cap_paid_reject":
        await query.edit_message_text(
            "🤖 ربات مدیریت سرمایه\n\nلطفاً یکی از گزینه‌های زیر را انتخاب کنید:",
            reply_markup=capital_bot_menu()
        )

    elif data == "cap_paid_confirm":
        requester = query.from_user
        chat_id = create_license_chat(user_id)
        for primary_id in PRIMARY_ADMIN_IDS:
            try:
                await context.bot.send_message(
                    primary_id,
                    license_request_text(requester.first_name, requester.username, user_id),
                    reply_markup=license_request_primary_menu(chat_id)
                )
            except Exception as e:
                print(f"خطای اطلاع‌رسانی درخواست لایسنس به مالک اصلی {primary_id}: {e}")
        await query.edit_message_text(
            "✅ درخواست شما برای ادمین ارسال شد؛ به‌زودی مکالمه شروع می‌شود.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")]])
        )

    elif data == "fin_currency":
        await query.edit_message_text("⏳ در حال دریافت قیمت‌ها...")
        prices = await get_iran_prices()
        if prices:
            def f(v): return f"{v:,}" if v else "---"
            text = ("💵 Iran Market\n`━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
                f"`{'USD':<10} {f(prices['dollar']):>15} IRR`\n"
                f"`{'Gold 18':<10} {f(prices['gold18']):>15} IRR`\n"
                f"`{'Coin':<10} {f(prices['gold_coin']):>15} IRR`\n"
                f"`{'Half Coin':<10} {f(prices['half_coin']):>15} IRR`\n"
                f"`{'Qtr Coin':<10} {f(prices['quarter']):>15} IRR`\n"
                f"`{'Mithgal':<10} {f(prices['gram_gold']):>15} IRR`\n"
                f"`{'Silver':<10} {f(prices['silver']):>15} IRR`\n"
                f"`━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━`\n🕐 {get_iran_now().strftime('%H:%M:%S')}")
        else:
            text = "❌ خطا در دریافت قیمت‌ها."
        await query.edit_message_text(text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")]]))

    elif data == "fin_crypto":
        await query.edit_message_text("⏳ در حال دریافت قیمت‌ها...")
        prices = await get_crypto_prices()
        if prices:
            def row(n, d):
                p, ch, ar = d
                return f"`{n:<5} ${p:>10,.2f}  {ar} {ch:>+6.2f}%`"
            text = ("₿ Crypto Market\n`━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
                + row("BTC", prices["bitcoin"]) + "\n"
                + row("ETH", prices["ethereum"]) + "\n"
                + row("USDT", prices["tether"]) + "\n"
                + row("BNB", prices["bnb"]) + "\n"
                + row("TRX", prices["tron"]) + "\n"
                f"`━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━`\n🕐 {get_iran_now().strftime('%H:%M:%S')}")
        else:
            text = "❌ خطا در دریافت قیمت‌ها."
        await query.edit_message_text(text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")]]))

    elif data == "fin_forex":
        await query.edit_message_text("⏳ در حال دریافت قیمت‌ها...")
        prices = await get_forex_prices()
        if prices:
            def row(n, d):
                p, ch, ar = d
                if p is None: return f"`{n:<8} {'---':>12}`"
                return f"`{n:<8} {fmt_price(p, 'USD'):>12}  {ar} {ch:>+6.2f}%`"
            text = ("📈 Forex Market\n`━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
                + row("GOLD", prices["gold"]) + "\n"
                + row("SILVER", prices["silver"]) + "\n"
                + row("OIL", prices["oil"]) + "\n"
                + row("EUR/USD", prices["eur"]) + "\n"
                + row("GBP/USD", prices["gbp"]) + "\n"
                f"`━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━`\n🕐 {get_iran_now().strftime('%H:%M:%S')}")
        else:
            text = "❌ خطا در دریافت قیمت‌ها."
        await query.edit_message_text(text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")]]))

    elif data == "fin_watchlist":
        await query.edit_message_text("👁 واچ لیست شخصی:", reply_markup=watchlist_menu(user_id))

    elif data == "wl_add":
        set_state(user_id, "wl_add")
        await query.edit_message_text(
            "➕ نماد ارز را تایپ کنید:\n\nمثال:\n`BTC` `GOLD` `USDCAD` `EURUSD`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="fin_watchlist")]]))

    elif data == "wl_view":
        items = get_watchlist(user_id)
        if not items:
            await query.answer("واچ لیست خالی است!", show_alert=True)
            return
        await query.edit_message_text("⏳ در حال دریافت قیمت‌ها...")
        lines = []
        for _, symbol, market in items:
            price, unit = await get_price_by_symbol(symbol, market)
            lines.append(f"`{symbol:<10} {fmt_price(price, unit):>15} {unit}`")
        text = ("👁 واچ لیست من\n`━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
            + "\n".join(lines) + f"\n`━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━`\n🕐 {get_iran_now().strftime('%H:%M:%S')}")
        await query.edit_message_text(text, parse_mode="Markdown", reply_markup=watchlist_menu(user_id))

    elif data.startswith("wl_del_"):
        remove_watchlist(int(data.replace("wl_del_", "")))
        await query.edit_message_text("✅ حذف شد.", reply_markup=watchlist_menu(user_id))

    elif data == "fin_alarm":
        await query.edit_message_text("🔔 آلارم قیمت:", reply_markup=alarm_menu())

    elif data == "alarm_new":
        set_state(user_id, "alarm_symbol")
        await query.edit_message_text(
            "🔔 نماد ارز را وارد کنید:\n\nمثال:\n`BTC` `GOLD` `USDCAD`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="fin_alarm")]]))

    elif data == "alarm_list":
        alarms = get_alarms(user_id)
        if not alarms:
            await query.answer("هیچ آلارم فعالی ندارید!", show_alert=True)
            return
        text = "📋 آلارم‌های فعال:\n\n"
        keyboard = []
        for aid, symbol, market, target, direction in alarms:
            unit = "IRR" if market == "iran" else "USD"
            arrow = "📈" if direction == "above" else "📉"
            text += f"{arrow} {symbol} ← {fmt_price(target, unit)} {unit}\n"
            keyboard.append([InlineKeyboardButton(f"🗑 حذف {symbol}", callback_data=f"alarm_del_{aid}")])
        keyboard.append([InlineKeyboardButton("🔙 برگشت", callback_data="fin_alarm")])
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))

    elif data.startswith("alarm_del_"):
        remove_alarm(int(data.replace("alarm_del_", "")))
        await query.edit_message_text("✅ آلارم حذف شد.", reply_markup=alarm_menu())

    elif data in ["alarm_dir_above", "alarm_dir_below"]:
        state, sdata = get_state(user_id)
        if state == "alarm_direction" and sdata:
            parts = sdata.split("|")
            symbol, market, target, unit = parts[0], parts[1], float(parts[2]), parts[3]
            direction = "above" if data == "alarm_dir_above" else "below"
            add_alarm(user_id, symbol, market, target, direction)
            clear_state(user_id)
            arrow = "📈" if direction == "above" else "📉"
            await query.edit_message_text(
                f"✅ آلارم ثبت شد!\n\n{arrow} {symbol}\nقیمت هدف: {fmt_price(target, unit)} {unit}",
                reply_markup=alarm_menu())

async def handle_financial_message(user_id, text, update):
    state, data = get_state(user_id)

    if state == "wl_add":
        symbol = text.upper()
        await update.message.reply_text("⏳ در حال بررسی...")
        valid, market, price, unit = await validate_symbol(symbol)
        if valid:
            if add_watchlist(user_id, symbol, market):
                await update.message.reply_text(
                    f"✅ {symbol} اضافه شد!\n💰 قیمت: {fmt_price(price, unit)} {unit}",
                    reply_markup=watchlist_menu(user_id))
            else:
                await update.message.reply_text(f"⚠️ {symbol} قبلاً هست.", reply_markup=watchlist_menu(user_id))
        else:
            await update.message.reply_text(f"❌ '{symbol}' پیدا نشد.", reply_markup=watchlist_menu(user_id))
        clear_state(user_id)
        return True

    elif state == "alarm_symbol":
        symbol = text.upper()
        await update.message.reply_text("⏳ در حال بررسی...")
        valid, market, price, unit = await validate_symbol(symbol)
        if valid:
            set_state(user_id, "alarm_price", f"{symbol}|{market}|{price}|{unit}")
            await update.message.reply_text(
                f"✅ {symbol} پیدا شد!\n💰 قیمت: {fmt_price(price, unit)} {unit}\n\nقیمت هدف را وارد کنید:")
        else:
            await update.message.reply_text(f"❌ '{symbol}' پیدا نشد. دوباره امتحان کنید:")
        return True

    elif state == "alarm_price":
        try:
            target = float(text.replace(",", ""))
            parts = data.split("|")
            symbol, market, unit = parts[0], parts[1], parts[3]
            set_state(user_id, "alarm_direction", f"{symbol}|{market}|{target}|{unit}")
            await update.message.reply_text("جهت آلارم:",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(f"📈 بالاتر از {fmt_price(target, unit)}", callback_data="alarm_dir_above")],
                    [InlineKeyboardButton(f"📉 پایین‌تر از {fmt_price(target, unit)}", callback_data="alarm_dir_below")],
                ]))
        except:
            await update.message.reply_text("❌ عدد وارد کنید:")
        return True

    return False

# ═══════════════ پایان MODULE: FINANCIAL ═══════════════════════




# ╔══════════════════════════════════════════════════════════════╗
# ║                  MODULE: REMINDER                            ║
# ║  برای حذف: این بلوک رو پاک کن                               ║
# ║  + خط rem_/menu_reminder رو از ROUTER اصلی پاک کن            ║
# ║  + state های rem_ رو از message_handler پاک کن              ║
# ║  + خط reminder_loop رو از post_init پاک کن                  ║
# ╚══════════════════════════════════════════════════════════════╝

def get_iran_now():
    """
    تاریخ و ساعت کامل فعلی به وقت ایران (datetime object، naive -- بدون
    اطلاعات timezone). عمداً naive نگه داشته شده چون تمام تاریخ‌های ذخیره‌شده
    در دیتابیس هم naive هستند (با strptime خوانده می‌شوند)؛ اگر این تابع
    aware (با timezone.utc) برمی‌گرداند، هرگونه مقایسه‌ی مستقیم با تاریخ‌های
    دیتابیس (مثل چک انقضای اشتراک در تابع start) با خطای قطعی
    "can't compare offset-naive and offset-aware datetimes" کرش می‌کرد --
    این دقیقاً همان باگی بود که باعث بی‌پاسخ ماندن /start بعد از تأیید
    ادمین می‌شد.
    """
    from datetime import timezone
    return datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3, minutes=30)

def gregorian_to_jalali(gy, gm, gd):
    """
    تبدیل تاریخ میلادی به شمسی (جلالی) با یک الگوریتم مستقل و استاندارد --
    بدون نیاز به هیچ کتابخانه‌ی خارجی (مثل jdatetime)، پس هرگز به مشکل نصب
    پکیج روی سرور Railway نمی‌خورد. این الگوریتم با چند تاریخ شناخته‌شده
    (از جمله نوروز ۱۴۰۳ و ۱۴۰۴) تست و تأیید شده است.
    خروجی: (jalali_year, jalali_month, jalali_day)
    """
    g_days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    j_days_in_month = [31, 31, 31, 31, 31, 31, 30, 30, 30, 30, 30, 29]

    gy2 = gy - 1600
    gm2 = gm - 1
    gd2 = gd - 1

    g_day_no = 365 * gy2 + (gy2 + 3) // 4 - (gy2 + 99) // 100 + (gy2 + 399) // 400
    for i in range(gm2):
        g_day_no += g_days_in_month[i]
    if gm2 > 1 and ((gy % 4 == 0 and gy % 100 != 0) or (gy % 400 == 0)):
        g_day_no += 1
    g_day_no += gd2

    j_day_no = g_day_no - 79
    j_np = j_day_no // 12053
    j_day_no %= 12053

    jy = 979 + 33 * j_np + 4 * (j_day_no // 1461)
    j_day_no %= 1461

    if j_day_no >= 366:
        jy += (j_day_no - 1) // 365
        j_day_no = (j_day_no - 1) % 365

    for i in range(11):
        if j_day_no < j_days_in_month[i]:
            jm = i + 1
            jd = j_day_no + 1
            break
        j_day_no -= j_days_in_month[i]
    else:
        jm = 12
        jd = j_day_no + 1

    return jy, jm, jd

JALALI_MONTH_NAMES_FA = [
    "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
    "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
]

_PERSIAN_DIGITS_MAP = str.maketrans("0123456789", "۰۱۲۳۴۵۶۷۸۹")

def to_persian_digits(value):
    """
    تبدیل ارقام لاتین به فارسی در یک رشته یا عدد -- برای یکدست کردن
    نمایش تاریخ/ساعت/شمارش به کاربر. فقط برای محتوای فارسیِ خواندنی
    استفاده می‌شود (تاریخ شمسی، ساعت)؛ عمداً برای اعداد مالی/قیمت
    (دلار، لات، پیپ) و آیدی‌های عددی (که استاندارد جهانی لاتین دارند)
    استفاده نمی‌شود، چون تبدیل آن‌ها غیرمعمول و کمتر خوانا می‌شود.
    """
    return str(value).translate(_PERSIAN_DIGITS_MAP)

def format_date_fa(dt_str_or_obj, with_time=True):
    """
    تابع مرکزی فرمت‌دهی تاریخ/ساعت برای نمایش به کاربر -- طبق تصمیم شما،
    همیشه هم میلادی و هم شمسی با هم نشان داده می‌شود. ورودی می‌تواند یک
    رشته‌ی ذخیره‌شده در دیتابیس (فرمت "%Y-%m-%d %H:%M" یا "%Y-%m-%d") یا
    مستقیم یک شیء datetime باشد. ذخیره‌سازی داخلی در دیتابیس همچنان
    میلادی باقی می‌ماند -- این تابع فقط برای نمایش نهایی به کاربر است.
    بخش شمسی و ساعت با ارقام فارسی نمایش داده می‌شود (طبق قرارداد رایج
    فارسی‌نویسی)؛ بخش میلادی داخل پرانتز عمداً با ارقام لاتین می‌ماند
    (چون یک قالب مرجع/فنی است، مثل تاریخ‌های ISO).
    خروجی نمونه: "۱۸ تیر ۱۴۰۵ (2026-07-09) ساعت ۲۳:۲۸"
    """
    if isinstance(dt_str_or_obj, str):
        try:
            if " " in dt_str_or_obj:
                dt = datetime.strptime(dt_str_or_obj, "%Y-%m-%d %H:%M")
            else:
                dt = datetime.strptime(dt_str_or_obj, "%Y-%m-%d")
        except ValueError:
            return dt_str_or_obj  # اگر فرمت ناشناخته بود، همان رشته‌ی اصلی برگردانده شود
    else:
        dt = dt_str_or_obj

    jy, jm, jd = gregorian_to_jalali(dt.year, dt.month, dt.day)
    jalali_str = to_persian_digits(f"{jd} {JALALI_MONTH_NAMES_FA[jm-1]} {jy}")
    gregorian_str = dt.strftime("%Y-%m-%d")

    if with_time and (dt.hour or dt.minute):
        time_str = to_persian_digits(dt.strftime("%H:%M"))
        return f"{jalali_str} ({gregorian_str}) ساعت {time_str}"
    return f"{jalali_str} ({gregorian_str})"

# ─── دیتابیس ───────────────────────────────────────────────────
def init_reminder_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS tasks (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id      INTEGER,
        title        TEXT,
        due_date     TEXT,
        due_time     TEXT,
        repeat_type  TEXT DEFAULT 'once',
        alert_before INTEGER DEFAULT 0,
        alert_sent   INTEGER DEFAULT 0,
        due_notified INTEGER DEFAULT 0,
        done         INTEGER DEFAULT 0,
        created_at   TEXT
    )''')
    try:
        c.execute("ALTER TABLE tasks ADD COLUMN due_notified INTEGER DEFAULT 0")
    except:
        pass
    conn.commit()
    conn.close()

def add_task(user_id, title, due_date, due_time, repeat_type, alert_before_minutes):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''INSERT INTO tasks (user_id, title, due_date, due_time, repeat_type, alert_before, created_at)
                 VALUES (?, ?, ?, ?, ?, ?, ?)''',
              (user_id, title, due_date, due_time, repeat_type, alert_before_minutes,
               get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_user_tasks(user_id, only_pending=True):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    if only_pending:
        c.execute('''SELECT id, title, due_date, due_time, repeat_type, alert_before, done
                     FROM tasks WHERE user_id=? AND done=0 ORDER BY due_date, due_time''', (user_id,))
    else:
        c.execute('''SELECT id, title, due_date, due_time, repeat_type, alert_before, done
                     FROM tasks WHERE user_id=? AND done=1 ORDER BY due_date DESC, due_time DESC''', (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

KIDS_HOMEWORK_PREFIX = "📚 تکلیف: "

def get_kids_homework_tasks(user_id):
    """تکالیف مدرسه از همان جدول tasks مشترک، فیلترشده با پیشوند مخصوص تکلیف"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT title, due_date, due_time, done FROM tasks
                 WHERE user_id=? AND title LIKE ? ORDER BY due_date, due_time''',
              (user_id, f"{KIDS_HOMEWORK_PREFIX}%"))
    rows = c.fetchall()
    conn.close()
    return [(title.replace(KIDS_HOMEWORK_PREFIX, ""), due_date, due_time, done) for title, due_date, due_time, done in rows]

def get_tasks_for_date(user_id, date_str):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT id, title, due_time FROM tasks
                 WHERE user_id=? AND due_date=? AND done=0 ORDER BY due_time''', (user_id, date_str))
    rows = c.fetchall()
    conn.close()
    return rows

def mark_task_done(task_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE tasks SET done=1 WHERE id=?", (task_id,))
    conn.commit()
    conn.close()

def delete_task(task_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM tasks WHERE id=?", (task_id,))
    conn.commit()
    conn.close()

def mark_alert_sent(task_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE tasks SET alert_sent=1 WHERE id=?", (task_id,))
    conn.commit()
    conn.close()

def mark_due_notified(task_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE tasks SET due_notified=1 WHERE id=?", (task_id,))
    conn.commit()
    conn.close()

def advance_recurring_task(task_id, repeat_type, due_date):
    """تسک تکرارشونده را به دور بعدی منتقل می‌کند (تاریخ جدید + ریست وضعیت هشدار)"""
    current = datetime.strptime(due_date, "%Y-%m-%d")
    if repeat_type == "daily":
        next_date = current + timedelta(days=1)
    elif repeat_type == "weekly":
        next_date = current + timedelta(days=7)
    else:
        return
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE tasks SET due_date=?, alert_sent=0, due_notified=0, done=0 WHERE id=?",
              (next_date.strftime("%Y-%m-%d"), task_id))
    conn.commit()
    conn.close()

def get_all_pending_tasks_raw():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT id, user_id, title, due_date, due_time, repeat_type, alert_before, alert_sent, due_notified
                 FROM tasks WHERE done=0''')
    rows = c.fetchall()
    conn.close()
    return rows

def get_morning_summary_users():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT DISTINCT user_id FROM tasks WHERE done=0")
    rows = [r[0] for r in c.fetchall()]
    conn.close()
    return rows

# ─── منوها ────────────────────────────────────────────────────
def reminder_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ تسک جدید", callback_data="rem_new"),
         InlineKeyboardButton("📋 لیست تسک‌ها", callback_data="rem_list")],
        [InlineKeyboardButton("✅ انجام شده‌ها", callback_data="rem_done_list")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_main")],
    ])

def repeat_type_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔂 یک‌بار", callback_data="rem_rep_once"),
         InlineKeyboardButton("🔁 هر روز", callback_data="rem_rep_daily")],
        [InlineKeyboardButton("📅 هر هفته", callback_data="rem_rep_weekly")],
    ])

def alert_yes_no_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ بله", callback_data="rem_alert_yes"),
         InlineKeyboardButton("❌ خیر", callback_data="rem_alert_no")],
    ])

def alert_unit_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⏱ دقیقه", callback_data="rem_unit_minute"),
         InlineKeyboardButton("🕐 ساعت", callback_data="rem_unit_hour")],
    ])

REPEAT_LABELS = {"once": "یک‌بار", "daily": "هر روز", "weekly": "هر هفته"}

# ─── هندلر دکمه‌ها ────────────────────────────────────────────
async def handle_reminder(query, user_id):
    data = query.data

    if data == "menu_reminder":
        await query.edit_message_text("⏰ یادآور و تسک:", reply_markup=reminder_menu())

    elif data == "rem_new":
        set_state(user_id, "rem_title")
        await query.edit_message_text(
            "➕ تسک جدید\n\nعنوان تسک را وارد کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="menu_reminder")]])
        )

    elif data == "rem_list":
        await handle_reminder_refresh_list(query, user_id)

    elif data == "rem_done_list":
        tasks = get_user_tasks(user_id, only_pending=False)
        if not tasks:
            await query.edit_message_text("📭 هیچ تسک انجام‌شده‌ای ندارید.", reply_markup=reminder_menu())
            return
        text = "✅ تسک‌های انجام‌شده (۱۰ مورد آخر):\n\n"
        for tid, title, due_date, due_time, repeat_type, alert_before, done in tasks[:10]:
            text += f"✓ {title} — {format_date_fa(due_date, with_time=False)}\n"
        await query.edit_message_text(
            text, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="menu_reminder")]])
        )

    elif data.startswith("rem_done_"):
        tid = int(data.replace("rem_done_", ""))
        mark_task_done(tid)
        await query.answer("✅ تسک انجام‌شده ثبت شد.", show_alert=True)
        await handle_reminder_refresh_list(query, user_id)

    elif data.startswith("rem_del_"):
        tid = int(data.replace("rem_del_", ""))
        delete_task(tid)
        await query.answer("🗑 تسک حذف شد.", show_alert=True)
        await handle_reminder_refresh_list(query, user_id)

    elif data.startswith("rem_rep_"):
        repeat_type = data.replace("rem_rep_", "")
        state, sdata = get_state(user_id)
        if state == "rem_repeat":
            set_state(user_id, "rem_alert_choice", f"{sdata}|{repeat_type}")
            await query.edit_message_text(
                "🔔 می‌خواهید قبل از موعد، یادآوری هشدار برایتان ارسال شود؟",
                reply_markup=alert_yes_no_menu()
            )

    elif data in ["rem_alert_yes", "rem_alert_no"]:
        state, sdata = get_state(user_id)
        if state == "rem_alert_choice":
            if data == "rem_alert_no":
                await finalize_task_creation(query, user_id, sdata, alert_before_minutes=0)
            else:
                set_state(user_id, "rem_alert_unit", sdata)
                await query.edit_message_text(
                    "⏳ واحد زمانی هشدار را انتخاب کنید:",
                    reply_markup=alert_unit_menu()
                )

    elif data in ["rem_unit_minute", "rem_unit_hour"]:
        state, sdata = get_state(user_id)
        if state == "rem_alert_unit":
            unit = "minute" if data == "rem_unit_minute" else "hour"
            set_state(user_id, "rem_alert_value", f"{sdata}|{unit}")
            unit_fa = "دقیقه" if unit == "minute" else "ساعت"
            await query.edit_message_text(f"چند {unit_fa} قبل از موعد هشدار بدهم؟ (عدد وارد کنید)")

async def handle_reminder_refresh_list(query, user_id):
    tasks = get_user_tasks(user_id, only_pending=True)
    if not tasks:
        await query.edit_message_text("📭 هیچ تسک فعالی ندارید.", reply_markup=reminder_menu())
        return
    text = "📋 تسک‌های فعال:\n\n"
    keyboard = []
    for tid, title, due_date, due_time, repeat_type, alert_before, done in tasks:
        rep_label = REPEAT_LABELS.get(repeat_type, "یک‌بار")
        text += f"🔸 {title}\n📅 {format_date_fa(due_date, with_time=False)} ⏰ {due_time} ({rep_label})\n\n"
        keyboard.append([
            InlineKeyboardButton(f"✅ {title[:15]}", callback_data=f"rem_done_{tid}"),
            InlineKeyboardButton("🗑", callback_data=f"rem_del_{tid}")
        ])
    keyboard.append([InlineKeyboardButton("🔙 برگشت", callback_data="menu_reminder")])
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))

async def finalize_task_creation(query, user_id, sdata, alert_before_minutes):
    parts = sdata.split("|")
    title, due_date, due_time, repeat_type = parts[0], parts[1], parts[2], parts[3]
    add_task(user_id, title, due_date, due_time, repeat_type, alert_before_minutes)
    clear_state(user_id)
    rep_label = REPEAT_LABELS.get(repeat_type, "یک‌بار")
    alert_line = f"\n🔔 هشدار: {alert_before_minutes} دقیقه قبل" if alert_before_minutes > 0 else ""
    await query.edit_message_text(
        f"✅ تسک ثبت شد!\n\n"
        f"🔸 {title}\n📅 {format_date_fa(due_date, with_time=False)} ⏰ {due_time}\n🔁 تکرار: {rep_label}{alert_line}",
        reply_markup=reminder_menu()
    )

async def handle_reminder_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    if state == "rem_title":
        set_state(user_id, "rem_date", text)
        await update.message.reply_text("📅 تاریخ انجام را وارد کنید (فرمت: YYYY-MM-DD):\nمثال: `2026-06-28`", parse_mode="Markdown")
        return True

    elif state == "rem_date":
        try:
            datetime.strptime(text, "%Y-%m-%d")
            set_state(user_id, "rem_time", f"{data}|{text}")
            await update.message.reply_text("⏰ ساعت انجام را وارد کنید (فرمت 24 ساعته HH:MM):\nمثال: `14:30`", parse_mode="Markdown")
        except:
            await update.message.reply_text("❌ فرمت تاریخ اشتباه است.\nمثال: `2026-06-28`", parse_mode="Markdown")
        return True

    elif state == "rem_time":
        try:
            datetime.strptime(text, "%H:%M")
            set_state(user_id, "rem_repeat", f"{data}|{text}")
            await update.message.reply_text("🔁 نوع تکرار تسک را انتخاب کنید:", reply_markup=repeat_type_menu())
        except:
            await update.message.reply_text("❌ فرمت ساعت اشتباه است.\nمثال: `14:30`", parse_mode="Markdown")
        return True

    elif state == "rem_alert_value":
        try:
            value = int(text)
            if value <= 0:
                await update.message.reply_text("❌ عدد باید بزرگ‌تر از صفر باشد:")
                return True
            parts = data.split("|")
            title, due_date, due_time, repeat_type, unit = parts[0], parts[1], parts[2], parts[3], parts[4]
            alert_minutes = value if unit == "minute" else value * 60

            add_task(user_id, title, due_date, due_time, repeat_type, alert_minutes)
            clear_state(user_id)
            rep_label = REPEAT_LABELS.get(repeat_type, "یک‌بار")
            unit_fa = "دقیقه" if unit == "minute" else "ساعت"
            await update.message.reply_text(
                f"✅ تسک ثبت شد!\n\n"
                f"🔸 {title}\n📅 {format_date_fa(due_date, with_time=False)} ⏰ {due_time}\n🔁 تکرار: {rep_label}\n"
                f"🔔 هشدار: {value} {unit_fa} قبل از موعد",
                reply_markup=reminder_menu()
            )
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    return False

# ─── حلقه پس‌زمینه: هشدار، یادآوری زمان، خلاصه صبحگاهی ────────
async def check_reminders(app):
    now_iran = get_iran_now()
    today_str = now_iran.strftime("%Y-%m-%d")
    now_time_str = now_iran.strftime("%H:%M")
    now_dt = datetime.strptime(f"{today_str} {now_time_str}", "%Y-%m-%d %H:%M")

    tasks = get_all_pending_tasks_raw()
    for tid, user_id, title, due_date, due_time, repeat_type, alert_before, alert_sent, due_notified in tasks:
        try:
            due_dt = datetime.strptime(f"{due_date} {due_time}", "%Y-%m-%d %H:%M")

            # هشدار پیش از موعد (بازه‌ای، نه تساوی دقیق - مقاوم در برابر تأخیر حلقه)
            if alert_before > 0 and not alert_sent:
                alert_dt = due_dt - timedelta(minutes=alert_before)
                if alert_dt <= now_dt < due_dt:
                    await app.bot.send_message(
                        user_id,
                        f"🔔 یادآوری زودهنگام!\n\n🔸 {title}\n⏰ موعد: {due_time} امروز\n"
                        f"({alert_before} دقیقه دیگر فرا می‌رسد)"
                    )
                    mark_alert_sent(tid)

            # رسیدن به خود موعد تسک - بازه‌ای (از لحظه موعد تا ۵ دقیقه بعد) تا اگر
            # حلقه دقیقاً همان دقیقه چک نکرد، پیام از قلم نیفتد؛ با فلگ due_notified
            # از ارسال تکراری جلوگیری می‌شود
            if not due_notified and due_dt <= now_dt < due_dt + timedelta(minutes=5):
                await app.bot.send_message(user_id, f"⏰ وقتشه!\n\n🔸 {title}")
                mark_due_notified(tid)
                if repeat_type in ["daily", "weekly"]:
                    advance_recurring_task(tid, repeat_type, due_date)
                else:
                    mark_task_done(tid)
        except Exception as e:
            print(f"خطای بررسی تسک {tid}: {e}")

async def send_morning_summary(app):
    today_str = get_iran_now().strftime("%Y-%m-%d")
    for user_id in get_morning_summary_users():
        tasks_today = get_tasks_for_date(user_id, today_str)
        if not tasks_today:
            continue
        text = "🌅 خلاصه صبحگاهی — کارهای امروز:\n\n"
        for tid, title, due_time in tasks_today:
            text += f"🔸 {due_time} — {title}\n"
        try:
            await app.bot.send_message(user_id, text)
        except Exception as e:
            print(f"خطای ارسال خلاصه صبحگاهی به {user_id}: {e}")

async def reminder_loop(app):
    last_summary_date = None
    while True:
        try:
            now_iran = get_iran_now()
            await check_reminders(app)

            # خلاصه صبحگاهی دقیقاً ساعت ۸:۰۰ تا ۸:۰۱ به وقت ایران، فقط یک‌بار در روز
            if now_iran.strftime("%H:%M") == "08:00":
                today_str = now_iran.strftime("%Y-%m-%d")
                if last_summary_date != today_str:
                    await send_morning_summary(app)
                    last_summary_date = today_str
        except Exception as e:
            print(f"خطای حلقه یادآور: {e}")
        await asyncio.sleep(60)

# ═══════════════ پایان MODULE: REMINDER ════════════════════════



# ╔══════════════════════════════════════════════════════════════╗
# ║                  MODULE: AI CHAT                             ║
# ║  چت با چند منبع هوش مصنوعی (Groq اصلی، Gemini در صورت        ║
# ║  داشتن کلید به‌عنوان پشتیبان)                                 ║
# ║  برای حذف: این بلوک رو پاک کن                                ║
# ║  + خط menu_ai رو از main_menu پاک کن                        ║
# ║  + خط ai_ رو از ROUTER پاک کن                                ║
# ║  + state های ai_ رو از message_handler پاک کن               ║
# ╚══════════════════════════════════════════════════════════════╝

GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")  # اختیاری - وقتی گرفتی همینجا اضافه می‌شود

GROQ_MODEL = "openai/gpt-oss-120b"
GEMINI_MODEL = "gemini-2.0-flash"

# ─── دیتابیس این ماژول ──────────────────────────────────────
def init_ai_db():
    pass  # این ماژول در حال حاضر جدول اختصاصی نیاز ندارد

# ─── فراخوانی Groq (منبع اصلی چت) ───────────────────────────
async def call_groq_chat(messages):
    if not GROQ_API_KEY:
        return None
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {GROQ_API_KEY}"},
                json={"model": GROQ_MODEL, "messages": messages, "temperature": 0.7}
            )
            data = r.json()
            if "choices" in data:
                return data["choices"][0]["message"]["content"]
            print(f"Groq error response: {data}")
            return None
    except Exception as e:
        print(f"خطای Groq: {e}")
        return None

# ─── فراخوانی Gemini (منبع پشتیبان، فعال فقط در صورت داشتن کلید) ─
async def call_gemini_chat(messages):
    if not GEMINI_API_KEY:
        return None
    try:
        # تبدیل فرمت پیام‌های OpenAI-style به فرمت Gemini
        contents = []
        for m in messages:
            if m["role"] == "system":
                continue
            role = "user" if m["role"] == "user" else "model"
            contents.append({"role": role, "parts": [{"text": m["content"]}]})

        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent",
                params={"key": GEMINI_API_KEY},
                json={"contents": contents}
            )
            data = r.json()
            candidates = data.get("candidates", [])
            if candidates:
                return candidates[0]["content"]["parts"][0]["text"]
            print(f"Gemini error response: {data}")
            return None
    except Exception as e:
        print(f"خطای Gemini: {e}")
        return None

async def get_ai_response(messages):
    """تلاش با Groq، در صورت خطا تلاش با Gemini (اگر کلیدش موجود باشد)"""
    response = await call_groq_chat(messages)
    if response:
        return response
    response = await call_gemini_chat(messages)
    if response:
        return response
    return None

# ─── منوها ────────────────────────────────────────────────────
def ai_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💬 شروع چت", callback_data="ai_start")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="menu_tools")],
    ])

# ─── هندلر دکمه‌ها ────────────────────────────────────────────
async def handle_ai(query, user_id):
    data = query.data
    if data == "menu_ai":
        await query.edit_message_text(
            "🤖 دستیار هوش مصنوعی\n\nهر سوالی داری بپرس، بدون محدودیت تعداد.",
            reply_markup=ai_menu()
        )
    elif data == "ai_start":
        set_state(user_id, "ai_chatting")
        await query.edit_message_text(
            "💬 چت با دستیار هوش مصنوعی\n\nسوال خود را بنویسید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 پایان چت", callback_data="menu_ai")]])
        )

async def handle_ai_message(user_id, text, update):
    state, _ = get_state(user_id)
    if state != "ai_chatting":
        return False

    await update.message.reply_text("⏳ در حال فکر کردن...")
    messages = [
        {"role": "system", "content": "تو یک دستیار هوشمند و مفید هستی که به زبان فارسی پاسخ می‌دهی."},
        {"role": "user", "content": text}
    ]
    response = await get_ai_response(messages)

    if response:
        await update.message.reply_text(
            response,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 پایان چت", callback_data="menu_ai")]])
        )
    else:
        await update.message.reply_text(
            "❌ متأسفانه دستیار هوش مصنوعی موقتاً در دسترس نیست. کمی بعد دوباره امتحان کنید.",
            reply_markup=ai_menu()
        )
    return True

# ═══════════════ پایان MODULE: AI CHAT ═════════════════════════


# ╔══════════════════════════════════════════════════════════════╗
# ║                  MODULE: KIDS                                ║
# ║  قصه‌گویی هوشمند متناسب با سن و جنسیت کودک (فقط متن؛ خروجی    ║
# ║  PDF طبق تصمیم کاربر حذف شد)                                 ║
# ║  از همان منبع AI ماژول قبلی (Groq/Gemini) استفاده می‌کند      ║
# ║  برای حذف: این بلوک رو پاک کن                                ║
# ║  + خط menu_kids رو از main_menu پاک کن                      ║
# ║  + خط kids_ رو از ROUTER پاک کن                              ║
# ║  + state های kids_ رو از message_handler پاک کن             ║
# ╚══════════════════════════════════════════════════════════════╝

def kids_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📖 قصه جدید", callback_data="kids_story"),
         InlineKeyboardButton("✏️ تکالیف و مطالعه", callback_data="kids_homework")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_main")],
    ])

def kids_homework_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 ثبت تکلیف جدید (با یادآور)", callback_data="kids_hw_add"),
         InlineKeyboardButton("📋 لیست تکالیف ثبت‌شده", callback_data="kids_hw_list")],
        [InlineKeyboardButton("🤖 کمک برای حل تمرین", callback_data="kids_hw_help")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="menu_kids")],
    ])

def kids_gender_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👧 دختر", callback_data="kids_gender_girl"),
         InlineKeyboardButton("👦 پسر", callback_data="kids_gender_boy")],
    ])

def kids_story_source_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✨ قصه جدید بساز", callback_data="kids_src_new")],
        [InlineKeyboardButton("📚 از قصه‌ها و حکایت‌های قدیمی پیدا کن", callback_data="kids_src_old")],
        [InlineKeyboardButton("🔍 قصه‌ای با اسم مشخص می‌خواهم", callback_data="kids_src_named")],
    ])

def build_story_prompt(age, gender, topic):
    gender_fa = "دختر" if gender == "girl" else "پسر"
    return (
        f"یک داستان کودکانه به زبان فارسی برای یک {gender_fa} {age} ساله بنویس.\n\n"
        f"موضوع داستان: {topic}\n\n"
        f"قوانین مهم:\n"
        f"- داستان باید کاملاً متناسب با سن {age} سال باشد (واژگان ساده و قابل‌فهم برای این سن)\n"
        f"- داستان باید آموزنده باشد و یک نتیجه‌ی اخلاقی ساده و مثبت در پایان داشته باشد\n"
        f"- شخصیت اصلی باید {gender_fa} باشد تا کودک بهتر با آن ارتباط بگیرد\n"
        f"- طول داستان متناسب با سن انتخاب شود: برای سنین کوچک‌تر (۲ تا ۵ سال) کوتاه‌تر "
        f"(در حد ۲ تا ۳ صفحه، حدود ۵۰۰ تا ۸۰۰ کلمه) و برای سنین بزرگ‌تر (۶ تا ۱۴ سال) "
        f"بلندتر و با جزئیات بیشتر (تا ۷ یا ۸ صفحه، حدود ۱۸۰۰ تا ۲۵۰۰ کلمه)؛ در هر حالت "
        f"طول نهایی باید بین ۲ صفحه و ۸ صفحه بماند، نه کمتر و نه بیشتر\n"
        f"- فقط متن داستان را بنویس، بدون مقدمه یا توضیح اضافه\n"
        f"- در پایان، یک خط با عنوان «🌟 پیام داستان:» و نتیجه‌ی اخلاقی کوتاه اضافه کن"
    )

# ─── جستجوی قصه/حکایت قدیمی (منابع فارسی معتبر، مستقل از AI) ────
KIDS_OLD_STORY_SLUG_PREFIXES = ["قصه بچگانه با موضوع", "قصه کودکانه با موضوع", "قصه با موضوع", "داستان کودکانه با موضوع"]
KIDS_OLD_STORY_BASE_PATHS = ["fun/story", "fun/intresting-subjects", "fun/sms"]

KIDS_OLD_STORY_FALLBACK_URLS = [
    "https://roozaneh.net/fun/story/%D9%82%D8%B5%D9%87-%D9%87%D8%A7%DB%8C-%DA%A9%D9%88%D8%AF%DA%A9%D8%A7%D9%86%D9%87/",
    "https://roozaneh.net/fun/intresting-subjects/%D9%82%D8%B5%D9%87-%D8%B4%DB%8C%D8%B1%DB%8C%D9%86-%DA%A9%D9%88%D8%AF%DA%A9%D8%A7%D9%86%D9%87/",
    "https://roozaneh.net/fun/story/%D9%82%D8%B5%D9%87-%D8%B4%D8%A8-%D8%A8%D8%B1%D8%A7%DB%8C-%DA%A9%D9%88%D8%AF%DA%A9-%D8%AF%D8%A8%D8%B3%D8%AA%D8%A7%D9%86%DB%8C/",
]

async def kids_try_direct_story_urls(topic):
    """امتحان مستقیم چند الگوی URL رایج برای قصه با موضوع مشخص (بدون نیاز به موتور جستجوی خارجی)"""
    if not topic:
        return None
    import urllib.parse
    for base_path in KIDS_OLD_STORY_BASE_PATHS:
        for prefix in KIDS_OLD_STORY_SLUG_PREFIXES:
            slug = f"{prefix}-{topic}"
            encoded = urllib.parse.quote(slug)
            url = f"https://roozaneh.net/{base_path}/{encoded}/"
            try:
                async with httpx.AsyncClient(timeout=10, headers={"User-Agent": "Mozilla/5.0"}, follow_redirects=True) as client:
                    r = await client.get(url)
                    if r.status_code == 200:
                        return url
            except Exception:
                continue
    return None

async def kids_search_wordpress_internal(topic):
    """جستجوی داخلی خود سایت روزانه برای قصه با موضوع مشخص (مستقل از Bing/DuckDuckGo)"""
    query_text = f"قصه کودکانه {topic}".strip()
    try:
        async with httpx.AsyncClient(timeout=15, headers={"User-Agent": "Mozilla/5.0"}, follow_redirects=True) as client:
            r = await client.get("https://roozaneh.net/", params={"s": query_text})
            if r.status_code != 200:
                return None
            links = re.findall(r'<h2[^>]*entry-title[^>]*>\s*<a[^>]+href="(https?://roozaneh\.net/[^"]+)"', r.text)
            if not links:
                links = re.findall(r'<a[^>]+rel="bookmark"[^>]+href="(https?://roozaneh\.net/[^"]+)"', r.text)
            return links[0] if links else None
    except Exception as e:
        print(f"خطای جستجوی داخلی روزانه برای قصه: {e}")
        return None

async def kids_fetch_story_paragraphs(url, min_length=25, max_items=40):
    """دریافت صفحه و استخراج پاراگراف‌های محتوایی اصلی داستان (نه منو/تبلیغات/معرفی سایت)"""
    import html as html_escape_module
    try:
        async with httpx.AsyncClient(timeout=15, headers={"User-Agent": "Mozilla/5.0"}, follow_redirects=True) as client:
            r = await client.get(url)
            if r.status_code != 200:
                return []
            html_text = r.text
            content_match = re.search(r'entry-content[^>]*>(.*?)(?:related-posts|comments|footer)', html_text, re.DOTALL)
            search_area = content_match.group(1) if content_match else html_text
            paragraphs = re.findall(r'<p[^>]*>(.*?)</p>', search_area, re.DOTALL)
            cleaned = []
            skip_markers = [
                "مطلب مشابه", "مطلب پیشنهادی", "اشتراک گذاری",
                "روزانه »", "در این بخش", "در این مطلب",
                "گردآوری کرده", "آماده کرده ایم", "آماده کرده‌ایم",
                "با ما باشید", "با ما همراه", "قرار داده‌ایم", "قرار داده ایم",
                "را در روزانه", "امیدواریم", "اهمیت زیادی در رشد",
            ]
            for p in paragraphs:
                text = re.sub(r'<[^>]+>', '', p)
                text = html_escape_module.unescape(text)
                text = re.sub(r'\s+', ' ', text).strip()
                if any(skip in text for skip in skip_markers):
                    continue
                if len(text) >= min_length:
                    cleaned.append(text)
                if len(cleaned) >= max_items:
                    break
            return cleaned
    except Exception as e:
        print(f"خطای دریافت/پارس صفحه قصه قدیمی: {e}")
        return []

async def kids_extract_single_story(raw_text, topic):
    """
    از هوش مصنوعی می‌خواهد از میان متن خام (که ممکن است شامل چند قصه‌ی
    کوتاه جدا از هم چسبیده به هم باشد) فقط یک قصه‌ی کامل و مستقل را جدا
    کند، آن را تمیز و روان بازنویسی کند (بدون تغییر داستان اصلی)، و در
    پایان بگوید که آیا نویسنده/منبع/کتاب این قصه را می‌شناسد یا نه.
    خروجی: dict با کلیدهای story, moral, source (source ممکن است خالی باشد)
    """
    prompt = (
        f"متن زیر ممکن است شامل چند قصه‌ی کوتاه کودکانه‌ی جدا از هم باشد که "
        f"به‌اشتباه به هم چسبیده‌اند:\n\n«{raw_text[:3000]}»\n\n"
        f"موضوع درخواستی: {topic or 'عمومی'}\n\n"
        f"لطفاً این کارها را انجام بده:\n"
        f"۱. فقط یک قصه‌ی کامل و مستقل (با شروع، میانه و پایان مشخص) را از "
        f"میان این متن پیدا کن و جدا کن -- بدون تغییر محتوای اصلی داستان، "
        f"فقط آن را تمیز و روان بازنویسی کن.\n"
        f"۲. اگر این قصه پیام یا نتیجه‌ی اخلاقی مشخصی دارد، آن را جدا بنویس.\n"
        f"۳. اگر می‌دانی این قصه از کدام نویسنده، کتاب، یا مجموعه‌ی شناخته‌شده‌ی "
        f"قصه‌های کودکانه (مثل کلیله و دمنه، مثنوی، حکایت‌های عامیانه ایرانی، "
        f"داستان‌های اپیک تنتین و... ) گرفته شده، نام آن را بنویس؛ در غیر این "
        f"صورت بنویس «نامشخص».\n\n"
        f"دقیقاً و فقط به این فرمت جواب بده:\n"
        f"قصه: <متن کامل قصه>\n"
        f"پیام: <پیام اخلاقی، یا خالی بگذار>\n"
        f"منبع: <نام نویسنده/کتاب یا نامشخص>"
    )
    result = await get_ai_response([{"role": "user", "content": prompt}])
    if not result:
        return None

    story_match = re.search(r'قصه\s*:?\s*(.+?)(?=\nپیام\s*:|\Z)', result, re.DOTALL)
    moral_match = re.search(r'پیام\s*:?\s*(.+?)(?=\nمنبع\s*:|\Z)', result, re.DOTALL)
    source_match = re.search(r'منبع\s*:?\s*(.+)', result, re.DOTALL)

    if not story_match:
        return None

    story = story_match.group(1).strip()
    moral = moral_match.group(1).strip() if moral_match else ""
    source = source_match.group(1).strip() if source_match else "نامشخص"
    return {"story": story, "moral": moral, "source": source}

async def kids_find_old_story(topic):
    """
    جستجوی قصه/حکایت قدیمی واقعی برای موضوع داده‌شده:
    ۰. امتحان چند الگوی URL مستقیم رایج
    ۰.۵ جستجوی داخلی سایت روزانه
    ۱. اگر موضوع خاصی نبود یا پیدا نشد، یکی از صفحات معتبر عمومی قصه کودکانه
    ۲. از میان محتوای خام (که ممکن است چند قصه‌ی چسبیده باشد)، یک قصه‌ی
       کامل با کمک هوش مصنوعی جدا و بازنویسی می‌شود
    خروجی: (dict شامل story/moral/source, url منبع) یا (None, None)
    """
    url = await kids_try_direct_story_urls(topic) if topic else None
    if not url and topic:
        url = await kids_search_wordpress_internal(topic)
    if not url:
        import random
        url = random.choice(KIDS_OLD_STORY_FALLBACK_URLS)

    paragraphs = await kids_fetch_story_paragraphs(url)
    if not paragraphs:
        return None, None

    raw_text = "\n\n".join(paragraphs[:20])
    story_data = await kids_extract_single_story(raw_text, topic)
    if not story_data:
        return None, None

    return story_data, url

async def kids_find_story_by_name(story_name):
    """
    جستجوی یک قصه/حکایت با اسم دقیق درخواستی کاربر (نه موضوع کلی):
    ۰. امتحان چند الگوی URL مستقیم با خود اسم قصه
    ۰.۵ جستجوی داخلی سایت روزانه با اسم قصه
    ۱. اگر در منابع فارسی معتبر پیدا نشد، از هوش مصنوعی خواسته می‌شود
       اگر این قصه/حکایت را می‌شناسد (از دانش خودش، نه ساختگی)، آن را
       با وفاداری کامل به داستان اصلی بازگو کند
    خروجی: (dict شامل story/moral/source, نوع منبع) یا (None, None)
    """
    import urllib.parse
    name_clean = story_name.strip()

    # مرحله ۰: چند الگوی URL مستقیم رایج با خود اسم قصه
    slug_prefixes = ["قصه", "داستان", "حکایت", "قصه بچگانه با موضوع", "قصه کودکانه با موضوع"]
    base_paths = ["fun/story", "fun/intresting-subjects", "fun/sms"]
    url = None
    for base_path in base_paths:
        for prefix in slug_prefixes:
            slug = f"{prefix}-{name_clean}"
            encoded = urllib.parse.quote(slug)
            candidate = f"https://roozaneh.net/{base_path}/{encoded}/"
            try:
                async with httpx.AsyncClient(timeout=10, headers={"User-Agent": "Mozilla/5.0"}, follow_redirects=True) as client:
                    r = await client.get(candidate)
                    if r.status_code == 200:
                        url = candidate
                        break
            except Exception:
                continue
        if url:
            break

    # مرحله ۰.۵: جستجوی داخلی سایت روزانه با اسم دقیق قصه
    if not url:
        try:
            async with httpx.AsyncClient(timeout=15, headers={"User-Agent": "Mozilla/5.0"}, follow_redirects=True) as client:
                r = await client.get("https://roozaneh.net/", params={"s": f"قصه {name_clean}"})
                if r.status_code == 200:
                    links = re.findall(r'<h2[^>]*entry-title[^>]*>\s*<a[^>]+href="(https?://roozaneh\.net/[^"]+)"', r.text)
                    if not links:
                        links = re.findall(r'<a[^>]+rel="bookmark"[^>]+href="(https?://roozaneh\.net/[^"]+)"', r.text)
                    if links:
                        url = links[0]
        except Exception as e:
            print(f"خطای جستجوی داخلی برای قصه با اسم: {e}")

    if url:
        paragraphs = await kids_fetch_story_paragraphs(url)
        if paragraphs:
            raw_text = "\n\n".join(paragraphs[:20])
            story_data = await kids_extract_single_story(raw_text, name_clean)
            if story_data:
                return story_data, "منبع فارسی معتبر"

    # مرحله ۱: هیچ صفحه‌ای پیدا نشد؛ از دانش خود هوش مصنوعی کمک بگیر
    prompt = (
        f"آیا داستان/حکایت/قصه‌ای کودکانه به اسم «{name_clean}» را می‌شناسی؟\n"
        f"اگر واقعاً این داستان را می‌شناسی (نه اینکه آن را بسازی)، آن را کامل "
        f"و با وفاداری به داستان اصلی به فارسی روان بازگو کن.\n\n"
        f"دقیقاً و فقط به این فرمت جواب بده:\n"
        f"قصه: <متن کامل قصه>\n"
        f"پیام: <پیام اخلاقی، یا خالی بگذار>\n"
        f"منبع: <نام نویسنده/کتاب یا نامشخص>\n\n"
        f"اگر داستانی با این اسم دقیق را نمی‌شناسی، فقط بنویس: هیچ"
    )
    result = await get_ai_response([{"role": "user", "content": prompt}])
    if not result or "هیچ" in result.strip()[:10]:
        return None, None

    story_match = re.search(r'قصه\s*:?\s*(.+?)(?=\nپیام\s*:|\Z)', result, re.DOTALL)
    moral_match = re.search(r'پیام\s*:?\s*(.+?)(?=\nمنبع\s*:|\Z)', result, re.DOTALL)
    source_match = re.search(r'منبع\s*:?\s*(.+)', result, re.DOTALL)
    if not story_match:
        return None, None

    story_data = {
        "story": story_match.group(1).strip(),
        "moral": moral_match.group(1).strip() if moral_match else "",
        "source": source_match.group(1).strip() if source_match else "نامشخص",
    }
    return story_data, "دانش هوش مصنوعی"

async def kids_send_story(update_or_query, story_text, title, is_message):
    """ارسال قصه (تکه‌تکه در صورت طولانی بودن)"""
    full_text = f"📖 {title}\n\n{story_text}"
    TELEGRAM_LIMIT = 4000
    if is_message:
        send_func = update_or_query.message.reply_text
    else:
        send_func = update_or_query.message.reply_text  # پیام جدید (نه ادیت) چون معمولاً طولانی است

    if len(full_text) <= TELEGRAM_LIMIT:
        await send_func(full_text, reply_markup=kids_menu())
    else:
        chunks = [full_text[i:i + TELEGRAM_LIMIT] for i in range(0, len(full_text), TELEGRAM_LIMIT)]
        for i, chunk in enumerate(chunks):
            is_last = (i == len(chunks) - 1)
            await send_func(chunk, reply_markup=kids_menu() if is_last else None)

# ─── هندلر دکمه‌ها ────────────────────────────────────────────
async def handle_kids(query, user_id):
    data = query.data

    if data == "menu_kids":
        await query.edit_message_text("👶 بخش فرزندان:", reply_markup=kids_menu())

    elif data == "kids_story":
        set_state(user_id, "kids_age")
        await query.edit_message_text(
            "📖 ساخت/جستجوی قصه\n\nسن کودک را وارد کنید (عدد):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="menu_kids")]])
        )

    elif data in ["kids_gender_girl", "kids_gender_boy"]:
        gender = "girl" if data == "kids_gender_girl" else "boy"
        state, sdata = get_state(user_id)
        if state == "kids_gender":
            set_state(user_id, "kids_topic", f"{sdata}|{gender}")
            await query.edit_message_text(
                "✏️ موضوع داستان را وارد کنید:\n\n"
                "مثال: دوستی، شجاعت، صداقت، حیوانات، فضا، احترام به بزرگ‌ترها\n\n"
                "یا اگر موضوع خاصی مدنظر نیست، فقط بنویسید: عمومی"
            )

    elif data in ["kids_src_new", "kids_src_old", "kids_src_named"]:
        state, sdata = get_state(user_id)
        if state != "kids_source_choice":
            return
        parts = sdata.split("|")
        age, gender, topic = parts[0], parts[1], parts[2]

        if data == "kids_src_new":
            clear_state(user_id)
            await query.edit_message_text("⏳ در حال نوشتن قصه... (ممکن است کمی طول بکشد)")
            prompt = build_story_prompt(age, gender, topic)
            messages = [
                {"role": "system", "content": "تو یک نویسنده‌ی خلاق داستان‌های کودکانه به زبان فارسی هستی."},
                {"role": "user", "content": prompt}
            ]
            story = await get_ai_response(messages)
            if not story:
                await query.message.reply_text(
                    "❌ متأسفانه ساخت قصه موقتاً ممکن نیست. کمی بعد دوباره امتحان کنید.",
                    reply_markup=kids_menu()
                )
                return
            title = f"قصه‌ای درباره {topic}"
            await kids_send_story(query, story, title, is_message=False)

        elif data == "kids_src_old":
            clear_state(user_id)
            await query.edit_message_text("⏳ در حال جستجوی قصه/حکایت قدیمی مرتبط...")
            topic_for_search = "" if topic == "عمومی" else topic
            story_data, source_url = await kids_find_old_story(topic_for_search)
            if not story_data:
                await query.message.reply_text(
                    f"❌ برای موضوع «{topic}» قصه/حکایت قدیمی مناسبی پیدا نشد.\n\n"
                    f"می‌توانید موضوع دیگری امتحان کنید یا از گزینه «قصه جدید بساز» استفاده کنید.",
                    reply_markup=kids_menu()
                )
                return
            title = f"قصه/حکایت قدیمی درباره {topic}" if topic_for_search else "یک قصه قدیمی برای شما"
            full_story = story_data["story"]
            if story_data.get("moral"):
                full_story += f"\n\n🌟 پیام داستان: {story_data['moral']}"
            source = story_data.get("source", "نامشخص")
            full_story += f"\n\n✍️ نویسنده/منبع: {source}"
            await kids_send_story(query, full_story, title, is_message=False)

        else:  # kids_src_named
            set_state(user_id, "kids_story_name", sdata)
            await query.edit_message_text(
                "🔍 اسم قصه یا حکایتی که می‌خواهید را بنویسید:\n\n"
                "مثال: لاک‌پشت و خرگوش، کلاغ و روباه، شنگول و منگول"
            )

    elif data == "kids_homework":
        await query.edit_message_text("✏️ تکالیف و مطالعه:", reply_markup=kids_homework_menu())

    elif data == "kids_hw_add":
        set_state(user_id, "kids_hw_subject")
        await query.edit_message_text(
            "📝 اسم درس یا عنوان تکلیف را بنویسید:\n\nمثال: ریاضی صفحه ۴۵",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="kids_homework")]])
        )

    elif data == "kids_hw_list":
        tasks = get_kids_homework_tasks(user_id)
        if not tasks:
            await query.edit_message_text("📭 هیچ تکلیفی ثبت نشده.", reply_markup=kids_homework_menu())
            return
        lines = []
        for title, due_date, due_time, done in tasks:
            status = "✅" if done else "⏳"
            lines.append(f"{status} {title} -- مهلت: {format_date_fa(due_date, with_time=False)} ساعت {due_time}")
        await query.edit_message_text("📋 تکالیف ثبت‌شده:\n\n" + "\n".join(lines), reply_markup=kids_homework_menu())

    elif data == "kids_hw_help":
        set_state(user_id, "kids_hw_question")
        await query.edit_message_text(
            "🤖 سؤال یا تمرین را بنویسید تا کمک کنم (مثلاً یک مسئله ریاضی ساده):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="kids_homework")]])
        )

async def handle_kids_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    if state == "kids_age":
        try:
            age = int(text)
            if age < 2 or age > 14:
                await update.message.reply_text("❌ لطفاً سنی بین ۲ تا ۱۴ سال وارد کنید:")
                return True
            set_state(user_id, "kids_gender", str(age))
            await update.message.reply_text("جنسیت کودک را انتخاب کنید:", reply_markup=kids_gender_menu())
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    elif state == "kids_topic":
        topic = text
        parts = data.split("|")
        age, gender = parts[0], parts[1]
        set_state(user_id, "kids_source_choice", f"{age}|{gender}|{topic}")
        await update.message.reply_text(
            "می‌خواهید یک قصه‌ی تازه بسازیم یا از قصه‌ها و حکایت‌های قدیمی برایتان پیدا کنیم؟",
            reply_markup=kids_story_source_menu()
        )
        return True

    elif state == "kids_story_name":
        story_name = text
        clear_state(user_id)
        await update.message.reply_text(f"⏳ در حال جستجوی قصه‌ی «{story_name}»...")
        story_data, source_type = await kids_find_story_by_name(story_name)
        if not story_data:
            log_failed_request(user_id, "قصه با اسم", story_name)
            await update.message.reply_text(
                f"❌ قصه/حکایتی با اسم «{story_name}» پیدا نشد.\n\n"
                f"لطفاً اسم دقیق‌تری امتحان کنید یا از گزینه‌های دیگر استفاده کنید.",
                reply_markup=kids_menu()
            )
            return True
        title = story_name
        full_story = story_data["story"]
        if story_data.get("moral"):
            full_story += f"\n\n🌟 پیام داستان: {story_data['moral']}"
        source = story_data.get("source", "نامشخص")
        full_story += f"\n\n✍️ نویسنده/منبع: {source}"
        await kids_send_story(update, full_story, title, is_message=True)
        return True

    elif state == "kids_hw_subject":
        subject = text
        set_state(user_id, "kids_hw_date", subject)
        await update.message.reply_text("مهلت انجام تکلیف را وارد کنید (فرمت: YYYY-MM-DD)\nمثال: 2026-06-25")
        return True

    elif state == "kids_hw_date":
        try:
            datetime.strptime(text, "%Y-%m-%d")
            set_state(user_id, "kids_hw_time", f"{data}|{text}")
            await update.message.reply_text("ساعت مهلت را وارد کنید (فرمت ۲۴ ساعته، مثلاً 18:00):")
        except:
            await update.message.reply_text("❌ فرمت تاریخ اشتباه است.\nمثال: 2026-06-25")
        return True

    elif state == "kids_hw_time":
        try:
            datetime.strptime(text, "%H:%M")
            subject, due_date = data.split("|")
            add_task(user_id, f"{KIDS_HOMEWORK_PREFIX}{subject}", due_date, text, "once", 60)
            clear_state(user_id)
            await update.message.reply_text(
                f"✅ تکلیف «{subject}» با مهلت {format_date_fa(due_date, with_time=False)} ساعت {text} ثبت شد.\n"
                f"۶۰ دقیقه قبل از موعد یادآوری می‌فرستم.",
                reply_markup=kids_homework_menu()
            )
        except:
            await update.message.reply_text("❌ فرمت ساعت اشتباه است.\nمثال: 18:00")
        return True

    elif state == "kids_hw_question":
        question = text
        clear_state(user_id)
        await update.message.reply_text("⏳ در حال بررسی سؤال...")
        prompt = (
            f"یک دانش‌آموز این سؤال یا تمرین مدرسه را پرسیده:\n«{question}»\n\n"
            f"به فارسی و ساده (متناسب با سطح مدرسه)، مرحله‌به‌مرحله راه‌حل را توضیح بده، "
            f"طوری که دانش‌آموز خودش یاد بگیرد، نه فقط جواب نهایی را بگویی."
        )
        result = await get_ai_response([{"role": "user", "content": prompt}])
        if not result:
            await update.message.reply_text(
                "❌ در حال حاضر امکان کمک نیست. کمی بعد دوباره امتحان کنید.",
                reply_markup=kids_homework_menu()
            )
            return True
        await update.message.reply_text(f"🤖 {result}", reply_markup=kids_homework_menu())
        return True

    return False

# ═══════════════ پایان MODULE: KIDS ════════════════════════════





# ╔══════════════════════════════════════════════════════════════╗
# ║                  MODULE: FUN (سرگرمی)                        ║
# ║  این ماژول شامل دو زیربخش مستقل است:                         ║
# ║    - سلامت و ورزش (منتقل‌شده از منوی اصلی)                  ║
# ║    - موزیک (جستجو در کانال + منابع رایگان قانونی)           ║
# ║  کتاب‌خوانی طبق تصمیم کاربر حذف شد (۱۹ تیر ۱۴۰۵).            ║
# ║  برای حذف کامل: این بلوک رو پاک کن                          ║
# ║  + خط menu_fun رو از main_menu پاک کن                       ║
# ║  + خط fun_/health_/music_ رو از ROUTER پاک کن               ║
# ║  + state های fun_/music_ رو از message_handler پاک کن       ║
# ╚══════════════════════════════════════════════════════════════╝

MUSIC_CHANNEL_USERNAME = "LiteMusics"  # بدون @ و بدون t.me/

# ─── دیتابیس این ماژول ──────────────────────────────────────
def init_fun_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS music_requests (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id     INTEGER,
        query_text  TEXT,
        fulfilled   INTEGER DEFAULT 0,
        created_at  TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS family_members (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id     INTEGER,
        name        TEXT,
        created_at  TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS health_weight_log (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id     INTEGER,
        member_id   INTEGER,
        weight_kg   REAL,
        logged_at   TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS health_profile (
        member_id   INTEGER PRIMARY KEY,
        height_cm   REAL,
        age         INTEGER,
        gender      TEXT,
        goal        TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS health_water_log (
        member_id   INTEGER,
        log_date    TEXT,
        glasses     INTEGER DEFAULT 0,
        PRIMARY KEY (member_id, log_date)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS health_sleep_log (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        member_id   INTEGER,
        sleep_time  TEXT,
        wake_time   TEXT,
        quality     TEXT,
        logged_at   TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS health_exercise_plan (
        member_id   INTEGER PRIMARY KEY,
        goal        TEXT,
        plan_text   TEXT,
        created_at  TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS music_index (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        artist      TEXT,
        title       TEXT,
        file_id     TEXT,
        message_id  INTEGER,
        added_at    TEXT
    )''')
    conn.commit()
    conn.close()

def save_music_request(user_id, query_text):
    """ثبت درخواست موزیک کاربر برای پیگیری بعدی توسط ادمین"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO music_requests (user_id, query_text, created_at) VALUES (?, ?, ?)",
              (user_id, query_text, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    request_id = c.lastrowid
    conn.commit()
    conn.close()
    return request_id

def get_music_request(request_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT user_id, query_text, fulfilled FROM music_requests WHERE id=?", (request_id,))
    row = c.fetchone()
    conn.close()
    return row

def mark_music_request_fulfilled(request_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE music_requests SET fulfilled=1 WHERE id=?", (request_id,))
    conn.commit()
    conn.close()

def index_music_track(artist, title, file_id, message_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    # از تکرار جلوگیری شود (همان message_id دوباره ایندکس نشود)
    c.execute("SELECT id FROM music_index WHERE message_id=?", (message_id,))
    if c.fetchone():
        conn.close()
        return
    c.execute("INSERT INTO music_index (artist, title, file_id, message_id, added_at) VALUES (?, ?, ?, ?, ?)",
              (artist, title, file_id, message_id, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def search_music_index(query_text):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    like = f"%{query_text}%"
    c.execute("SELECT artist, title, file_id FROM music_index WHERE artist LIKE ? OR title LIKE ? ORDER BY id DESC LIMIT 5",
              (like, like))
    rows = c.fetchall()
    conn.close()
    return rows

# ─── پارس کپشن پست کانال برای استخراج خواننده/اسم آهنگ ──────
def parse_music_caption(caption):
    """
    فرمت پست‌های کانال:
    🎤  #نام_خواننده
    🎼    نام آهنگ
    [بقیه متن...]
    """
    if not caption:
        return None, None
    artist, title = None, None
    for line in caption.split("\n"):
        line = line.strip()
        if line.startswith("🎤"):
            artist = line.replace("🎤", "").replace("#", "").strip()
        elif line.startswith("🎼"):
            title = line.replace("🎼", "").strip()
    return artist, title

async def music_refine_query_with_ai(query_text):
    """
    از هوش مصنوعی می‌خواهد عبارت جستجوی آهنگ را دقیق‌تر کند (مثلاً حدس زدن
    نام کامل خواننده یا اصلاح غلط املایی رایج) تا نتیجه‌ی جستجوی وب دقیق‌تر
    شود. اگر AI جواب نداد، همان عبارت اصلی استفاده می‌شود.
    """
    prompt = (
        f"کاربری این عبارت را برای جستجوی یک آهنگ فارسی نوشته: «{query_text}». "
        f"اگر این نام خواننده یا آهنگ خاصی را به ذهنت می‌آورد، نام کامل و درست "
        f"خواننده و/یا آهنگ را (برای جستجوی دقیق‌تر در گوگل) در یک خط کوتاه بنویس. "
        f"اگر مطمئن نیستی یا نمی‌شناسی، فقط همان عبارت اصلی را برگردان."
    )
    result = await get_ai_response([{"role": "user", "content": prompt}])
    if not result:
        return None
    return result.strip().split("\n")[0][:100]

async def music_test_link_reachable(url):
    """
    تست واقعی و فنی (نه ادعا یا حدس) اینکه آیا یک لینک از سمت سرور واقعاً
    قابل‌دسترسی است -- با یک درخواست HEAD/GET واقعی و بررسی کد وضعیت.
    توجه صادقانه: این فقط تضمین می‌کند که سرور ما به آن لینک دسترسی دارد؛
    اگر خود کاربر (مثلاً به‌خاطر تحریم یا فیلترینگ در کشورش) به آن سایت
    دسترسی نداشته باشد، این تست نمی‌تواند آن را از قبل تشخیص دهد، چون
    وضعیت شبکه‌ی کاربر با وضعیت شبکه‌ی سرور متفاوت است.
    """
    try:
        async with httpx.AsyncClient(timeout=8, headers={"User-Agent": "Mozilla/5.0"}, follow_redirects=True) as client:
            r = await client.get(url)
            return r.status_code == 200
    except Exception as e:
        print(f"خطای تست دسترسی لینک {url}: {e}")
        return False

async def music_build_search_links(query_text):
    """
    ساخت لینک‌های جستجو برای سه منبع. برای Radio Javan و گوگل، قبل از
    ارسال یک تست واقعی HTTP انجام می‌شود (چون این سایت‌ها robots.txt
    محدودکننده‌ای برای این نوع دسترسی ندارند). اما برای rozmusic.com،
    که در robots.txt صراحتاً اعلام کرده دسترسی خودکار/ربات‌ها را نمی‌خواهد،
    هیچ درخواست HTTP از سمت سرور زده نمی‌شود -- فقط یک لینک مستقیم به
    خود صفحه‌ی جستجوی این سایت (با پارامتر استاندارد وردپرس ?s=، که
    rozmusic روی آن ساخته شده) ساخته می‌شود. با باز کردن این لینک،
    عبارت جستجو از قبل در کادر جستجوی خود سایت نوشته شده است و کاربر
    فقط دکمه‌ی جستجو را می‌زند -- این یک لینک ساده است، نه اسکرپینگ،
    و هیچ محتوایی از سایت توسط سرور خوانده نمی‌شود.
    """
    import urllib.parse
    encoded_query = urllib.parse.quote(query_text)

    verified_links = []

    # Radio Javan و گوگل: تست واقعی HTTP قبل از ارسال (این سایت‌ها
    # محدودیت robots.txt برای این نوع دسترسی ندارند)
    testable_candidates = [
        ("Radio Javan", f"https://www.radiojavan.com/search?query={encoded_query}&type=mp3"),
        ("گوگل", f"https://www.google.com/search?q={urllib.parse.quote(query_text + ' آهنگ دانلود')}"),
    ]
    for label, url in testable_candidates:
        is_reachable = await music_test_link_reachable(url)
        if is_reachable:
            verified_links.append((label, url))
        else:
            print(f"⚠️ لینک {label} در تست واقعی سرور در دسترس نبود؛ فرستاده نمی‌شود.")

    # rozmusic.com: بدون تست HTTP (به احترام robots.txt این سایت)،
    # فقط لینک مستقیم صفحه‌ی جستجو با عبارت از پیش پرشده ساخته می‌شود
    rozmusic_link = f"https://rozmusic.com/?s={encoded_query}"
    verified_links.append(("رز موزیک (سایت ایرانی -- عبارت جستجو از قبل نوشته شده، فقط دکمه‌ی جستجوی خود سایت را بزنید)", rozmusic_link))

    return verified_links

# ─── منوها ────────────────────────────────────────────────────
def fun_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🏋️ سلامت و ورزش", callback_data="menu_health"),
         InlineKeyboardButton("🎵 موزیک", callback_data="menu_music")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_main")],
    ])

def family_members_menu(user_id):
    """منوی لیست اعضای خانواده + دکمه‌ی افزودن عضو جدید"""
    members = get_family_members(user_id)
    buttons = []
    for member_id, name in members:
        buttons.append([InlineKeyboardButton(f"👤 {name}", callback_data=f"health_member_{member_id}")])
    if len(members) < 5:
        buttons.append([InlineKeyboardButton("➕ افزودن عضو جدید", callback_data="health_add_member")])
    if members:
        buttons.append([InlineKeyboardButton("🗑 حذف یک عضو", callback_data="health_delete_member_menu")])
    buttons.append([InlineKeyboardButton("🔙 برگشت", callback_data="back_fun")])
    return InlineKeyboardMarkup(buttons)

def health_menu(member_id):
    """منوی سلامت مخصوص یک عضو خاص (member_id در callback_data هر دکمه گنجانده شده)"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⚖️ ثبت وزن", callback_data=f"health_weight_{member_id}"),
         InlineKeyboardButton("📊 نمودار وزن", callback_data=f"health_chart_{member_id}")],
        [InlineKeyboardButton("🏃 برنامه ورزشی", callback_data=f"health_exercise_{member_id}"),
         InlineKeyboardButton("🧮 محاسبه BMI و کالری", callback_data=f"health_bmi_{member_id}")],
        [InlineKeyboardButton("💧 یادآور آب", callback_data=f"health_water_{member_id}"),
         InlineKeyboardButton("😴 ردیاب خواب", callback_data=f"health_sleep_{member_id}")],
        [InlineKeyboardButton("🔙 برگشت به لیست اعضا", callback_data="menu_health")],
    ])

def music_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔍 جستجوی آهنگ", callback_data="music_search")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_fun")],
    ])

def add_family_member(user_id, name):
    """افزودن یک عضو جدید خانواده (حداکثر ۵ نفر برای هر user_id)"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM family_members WHERE user_id=?", (user_id,))
    count = c.fetchone()[0]
    if count >= 5:
        conn.close()
        return None
    c.execute("INSERT INTO family_members (user_id, name, created_at) VALUES (?, ?, ?)",
              (user_id, name, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    member_id = c.lastrowid
    conn.commit()
    conn.close()
    return member_id

def get_family_members(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, name FROM family_members WHERE user_id=? ORDER BY id", (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_family_member_name(member_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT name FROM family_members WHERE id=?", (member_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else None

def delete_family_member(member_id):
    """حذف یک عضو خانواده و تمام سوابق سلامت او"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("DELETE FROM family_members WHERE id=?", (member_id,))
    c.execute("DELETE FROM health_weight_log WHERE member_id=?", (member_id,))
    c.execute("DELETE FROM health_profile WHERE member_id=?", (member_id,))
    c.execute("DELETE FROM health_water_log WHERE member_id=?", (member_id,))
    c.execute("DELETE FROM health_sleep_log WHERE member_id=?", (member_id,))
    c.execute("DELETE FROM health_exercise_plan WHERE member_id=?", (member_id,))
    conn.commit()
    conn.close()

def log_weight(user_id, member_id, weight_kg):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO health_weight_log (user_id, member_id, weight_kg, logged_at) VALUES (?, ?, ?, ?)",
              (user_id, member_id, weight_kg, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_weight_history(member_id, limit=10):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT weight_kg, logged_at FROM health_weight_log WHERE member_id=? ORDER BY id DESC LIMIT ?", (member_id, limit))
    rows = c.fetchall()
    conn.close()
    return rows

def save_health_profile(member_id, height_cm=None, age=None, gender=None, goal=None):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT member_id FROM health_profile WHERE member_id=?", (member_id,))
    exists = c.fetchone()
    if exists:
        updates, params = [], []
        if height_cm is not None:
            updates.append("height_cm=?"); params.append(height_cm)
        if age is not None:
            updates.append("age=?"); params.append(age)
        if gender is not None:
            updates.append("gender=?"); params.append(gender)
        if goal is not None:
            updates.append("goal=?"); params.append(goal)
        if updates:
            params.append(member_id)
            c.execute(f"UPDATE health_profile SET {', '.join(updates)} WHERE member_id=?", params)
    else:
        c.execute("INSERT INTO health_profile (member_id, height_cm, age, gender, goal) VALUES (?, ?, ?, ?, ?)",
                  (member_id, height_cm, age, gender, goal))
    conn.commit()
    conn.close()

def get_health_profile(member_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT height_cm, age, gender, goal FROM health_profile WHERE member_id=?", (member_id,))
    row = c.fetchone()
    conn.close()
    return row

def calculate_bmi(weight_kg, height_cm):
    height_m = height_cm / 100
    bmi = weight_kg / (height_m ** 2)
    if bmi < 18.5:
        category = "کمبود وزن"
    elif bmi < 25:
        category = "طبیعی"
    elif bmi < 30:
        category = "اضافه وزن"
    else:
        category = "چاقی"
    return round(bmi, 1), category

def log_water(member_id):
    """یک لیوان آب به رکورد امروز اضافه می‌کند"""
    today = get_iran_now().strftime("%Y-%m-%d")
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''INSERT INTO health_water_log (member_id, log_date, glasses) VALUES (?, ?, 1)
                 ON CONFLICT(member_id, log_date) DO UPDATE SET glasses = glasses + 1''', (member_id, today))
    conn.commit()
    c.execute("SELECT glasses FROM health_water_log WHERE member_id=? AND log_date=?", (member_id, today))
    count = c.fetchone()[0]
    conn.close()
    return count

def get_today_water(member_id):
    today = get_iran_now().strftime("%Y-%m-%d")
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT glasses FROM health_water_log WHERE member_id=? AND log_date=?", (member_id, today))
    row = c.fetchone()
    conn.close()
    return row[0] if row else 0

def log_sleep(member_id, sleep_time, wake_time, quality):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO health_sleep_log (member_id, sleep_time, wake_time, quality, logged_at) VALUES (?, ?, ?, ?, ?)",
              (member_id, sleep_time, wake_time, quality, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_recent_sleep(member_id, limit=7):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT sleep_time, wake_time, quality, logged_at FROM health_sleep_log WHERE member_id=? ORDER BY id DESC LIMIT ?", (member_id, limit))
    rows = c.fetchall()
    conn.close()
    return rows

def save_exercise_plan(member_id, goal, plan_text):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''INSERT INTO health_exercise_plan (member_id, goal, plan_text, created_at) VALUES (?, ?, ?, ?)
                 ON CONFLICT(member_id) DO UPDATE SET goal=excluded.goal, plan_text=excluded.plan_text, created_at=excluded.created_at''',
              (member_id, goal, plan_text, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

# ─── هندلر دکمه‌ها ────────────────────────────────────────────
async def handle_fun(query, user_id):
    data = query.data

    if data == "menu_fun":
        await query.edit_message_text("🎮 سرگرمی:", reply_markup=fun_menu())

    elif data == "back_fun":
        clear_state(user_id)
        try:
            await query.edit_message_text("🎮 سرگرمی:", reply_markup=fun_menu())
        except Exception:
            await query.message.reply_text("🎮 سرگرمی:", reply_markup=fun_menu())

    # ─── سلامت و ورزش ───
    elif data == "menu_health":
        await query.edit_message_text("👨‍👩‍👧‍👦 پروفایل‌های سلامت خانواده:\n\nیک عضو را انتخاب کنید یا عضو جدید اضافه کنید.", reply_markup=family_members_menu(user_id))

    elif data == "health_add_member":
        set_state(user_id, "health_awaiting_member_name")
        await query.edit_message_text(
            "➕ اسم عضو جدید را وارد کنید (مثلاً «پدر»، «مادر»، «سارا»):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="menu_health")]])
        )

    elif data == "health_delete_member_menu":
        members = get_family_members(user_id)
        buttons = [[InlineKeyboardButton(f"🗑 {name}", callback_data=f"health_delete_confirm_{member_id}")] for member_id, name in members]
        buttons.append([InlineKeyboardButton("🔙 برگشت", callback_data="menu_health")])
        await query.edit_message_text("🗑 کدام عضو حذف شود؟ (همه‌ی سوابق او هم پاک می‌شود)", reply_markup=InlineKeyboardMarkup(buttons))

    elif data.startswith("health_delete_confirm_"):
        member_id = int(data.replace("health_delete_confirm_", ""))
        member_name = get_family_member_name(member_id)
        delete_family_member(member_id)
        await query.edit_message_text(f"✅ «{member_name}» و تمام سوابق او حذف شد.", reply_markup=family_members_menu(user_id))

    elif data.startswith("health_member_"):
        member_id = int(data.replace("health_member_", ""))
        member_name = get_family_member_name(member_id)
        if not member_name:
            await query.answer("این عضو دیگر وجود ندارد.", show_alert=True)
            return
        await query.edit_message_text(f"🏋️ سلامت و ورزش -- {member_name}", reply_markup=health_menu(member_id))

    elif data.startswith("health_weight_") and not data.startswith("health_weight_add_"):
        member_id = int(data.replace("health_weight_", ""))
        member_name = get_family_member_name(member_id)
        set_state(user_id, "health_weight_input", str(member_id))
        await query.edit_message_text(
            f"⚖️ وزن فعلی «{member_name}» را به کیلوگرم وارد کنید (مثلاً 72.5):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data=f"health_member_{member_id}")]])
        )

    elif data.startswith("health_chart_"):
        member_id = int(data.replace("health_chart_", ""))
        member_name = get_family_member_name(member_id)
        history = get_weight_history(member_id, limit=10)
        if not history:
            await query.edit_message_text(f"📭 هنوز وزنی برای «{member_name}» ثبت نشده.", reply_markup=health_menu(member_id))
            return
        lines = [f"{w} کیلوگرم -- {format_date_fa(t)}" for w, t in reversed(history)]
        trend = ""
        if len(history) >= 2:
            diff = history[0][0] - history[-1][0]
            trend = f"\n\n{'📉 کاهش' if diff < 0 else '📈 افزایش' if diff > 0 else '➡️ ثابت'} {abs(diff):.1f} کیلوگرم نسبت به {len(history)} ثبت اخیر"
        await query.edit_message_text(
            f"📊 تاریخچه وزن -- {member_name}:\n\n" + "\n".join(lines) + trend,
            reply_markup=health_menu(member_id)
        )

    elif data.startswith("health_exercise_"):
        member_id = int(data.replace("health_exercise_", ""))
        set_state(user_id, "health_exercise_goal", str(member_id))
        await query.edit_message_text(
            "🏃 هدف ورزشی را انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("💪 عضله‌سازی", callback_data=f"health_goal_muscle_{member_id}")],
                [InlineKeyboardButton("🔥 لاغری و چربی‌سوزی", callback_data=f"health_goal_fatloss_{member_id}")],
                [InlineKeyboardButton("🏃‍♂️ فیتنس عمومی", callback_data=f"health_goal_fitness_{member_id}")],
                [InlineKeyboardButton("🔙 برگشت", callback_data=f"health_member_{member_id}")],
            ])
        )

    elif data.startswith("health_goal_"):
        parts = data.replace("health_goal_", "").rsplit("_", 1)
        goal_key, member_id = parts[0], int(parts[1])
        member_name = get_family_member_name(member_id)
        goal_map = {"muscle": "عضله‌سازی", "fatloss": "لاغری و چربی‌سوزی", "fitness": "فیتنس عمومی"}
        goal_fa = goal_map.get(goal_key, goal_key)
        save_health_profile(member_id, goal=goal_fa)
        await query.edit_message_text(f"⏳ در حال آماده‌سازی برنامه‌ی ورزشی برای «{goal_fa}»...")
        prompt = (
            f"یک برنامه‌ی ورزشی ساده و عملی برای هدف «{goal_fa}» بنویس. "
            f"برنامه باید هفتگی (۳ تا ۴ روز در هفته)، با حرکات ساده و بدون نیاز به تجهیزات "
            f"خاص باشد. برای هر روز، حرکات + تعداد ست/تکرار را مشخص کن. کوتاه و عملی باش."
        )
        result = await get_ai_response([{"role": "user", "content": prompt}])
        if not result:
            await query.message.reply_text("❌ در حال حاضر امکان ساخت برنامه نیست.", reply_markup=health_menu(member_id))
            return
        save_exercise_plan(member_id, goal_fa, result)
        await query.message.reply_text(f"🏃 برنامه‌ی ورزشی -- {member_name} ({goal_fa})\n\n{result}", reply_markup=health_menu(member_id))

    elif data.startswith("health_bmi_"):
        member_id = int(data.replace("health_bmi_", ""))
        member_name = get_family_member_name(member_id)
        profile = get_health_profile(member_id)
        if profile and profile[0]:
            history = get_weight_history(member_id, limit=1)
            if history:
                bmi, category = calculate_bmi(history[0][0], profile[0])
                await query.edit_message_text(
                    f"🧮 BMI «{member_name}»: {bmi} ({category})\n\nبر اساس آخرین وزن ثبت‌شده و قد {profile[0]:.0f} سانتی‌متر.",
                    reply_markup=health_menu(member_id)
                )
                return
        set_state(user_id, "health_bmi_height", str(member_id))
        await query.edit_message_text(
            f"🧮 برای محاسبه BMI «{member_name}»، قد را به سانتی‌متر وارد کنید (مثلاً 175):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data=f"health_member_{member_id}")]])
        )

    elif data.startswith("health_water_add_"):
        member_id = int(data.replace("health_water_add_", ""))
        member_name = get_family_member_name(member_id)
        count = log_water(member_id)
        await query.edit_message_text(
            f"💧 آب امروز «{member_name}»: {count} لیوان\n\n✅ ثبت شد!",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("➕ یک لیوان دیگر", callback_data=f"health_water_add_{member_id}")],
                [InlineKeyboardButton("🔙 برگشت", callback_data=f"health_member_{member_id}")],
            ])
        )

    elif data.startswith("health_water_"):
        member_id = int(data.replace("health_water_", ""))
        member_name = get_family_member_name(member_id)
        count = get_today_water(member_id)
        await query.edit_message_text(
            f"💧 آب امروز «{member_name}»: {count} لیوان\n\nهر بار که آب می‌نوشد، دکمه زیر را بزنید:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("➕ یک لیوان نوشید", callback_data=f"health_water_add_{member_id}")],
                [InlineKeyboardButton("🔙 برگشت", callback_data=f"health_member_{member_id}")],
            ])
        )

    elif data.startswith("health_sleep_add_"):
        member_id = int(data.replace("health_sleep_add_", ""))
        set_state(user_id, "health_sleep_time", str(member_id))
        await query.edit_message_text("😴 ساعت خوابیدن دیشب را وارد کنید (فرمت ۲۴ ساعته، مثلاً 23:30):")

    elif data.startswith("health_sleep_"):
        member_id = int(data.replace("health_sleep_", ""))
        member_name = get_family_member_name(member_id)
        recent = get_recent_sleep(member_id, limit=5)
        lines = [f"😴 {s} تا {w} -- کیفیت: {q} ({format_date_fa(t)})" for s, w, q, t in recent] if recent else []
        text = f"😴 ردیاب خواب -- {member_name}\n\n" + ("\n".join(lines) if lines else "هنوز خوابی ثبت نشده.")
        await query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("➕ ثبت خواب دیشب", callback_data=f"health_sleep_add_{member_id}")],
                [InlineKeyboardButton("🔙 برگشت", callback_data=f"health_member_{member_id}")],
            ])
        )

    elif data.startswith("health_quality_"):
        state, sdata = get_state(user_id)
        if state != "health_sleep_quality":
            return
        quality = data.replace("health_quality_", "")
        member_id_str, sleep_time, wake_time = sdata.split("|")
        member_id = int(member_id_str)
        member_name = get_family_member_name(member_id)
        log_sleep(member_id, sleep_time, wake_time, quality)
        clear_state(user_id)
        await query.edit_message_text(
            f"✅ خواب دیشب «{member_name}» ثبت شد.\n\n😴 خواب: {sleep_time} | بیداری: {wake_time} | کیفیت: {quality}",
            reply_markup=health_menu(member_id)
        )

    elif data == "menu_music":
        await query.edit_message_text(
            f"🎵 موزیک\n\n"
            f"نام آهنگ یا خواننده را تایپ کنید تا جستجو کنم.\n\n"
            f"اول در کانال موزیک جستجو می‌کنم، اگر نبود سراغ "
            f"منابع رایگان و قانونی می‌روم.",
            reply_markup=music_menu()
        )

    elif data == "music_search":
        set_state(user_id, "music_query")
        await query.edit_message_text(
            "🔍 نام آهنگ یا خواننده را وارد کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="menu_music")]])
        )

async def handle_fun_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    if state == "health_awaiting_member_name":
        member_name = text.strip()[:30]
        member_id = add_family_member(user_id, member_name)
        clear_state(user_id)
        if member_id is None:
            await update.message.reply_text("❌ حداکثر ۵ عضو مجاز است.", reply_markup=family_members_menu(user_id))
            return True
        await update.message.reply_text(f"✅ «{member_name}» اضافه شد.", reply_markup=family_members_menu(user_id))
        return True

    if state == "health_weight_input":
        member_id = int(data)
        member_name = get_family_member_name(member_id)
        try:
            weight = float(text.replace(",", "."))
            if not (20 <= weight <= 300):
                await update.message.reply_text("❌ لطفاً وزنی منطقی به کیلوگرم وارد کنید (بین ۲۰ تا ۳۰۰):")
                return True
            log_weight(user_id, member_id, weight)
            clear_state(user_id)
            await update.message.reply_text(f"✅ وزن {weight} کیلوگرم برای «{member_name}» ثبت شد.", reply_markup=health_menu(member_id))
        except ValueError:
            await update.message.reply_text("❌ لطفاً فقط عدد وارد کنید (مثلاً 72.5):")
        return True

    elif state == "health_bmi_height":
        member_id = int(data)
        member_name = get_family_member_name(member_id)
        try:
            height = float(text.replace(",", "."))
            if not (100 <= height <= 250):
                await update.message.reply_text("❌ لطفاً قدی منطقی به سانتی‌متر وارد کنید (بین ۱۰۰ تا ۲۵۰):")
                return True
            save_health_profile(member_id, height_cm=height)
            clear_state(user_id)
            history = get_weight_history(member_id, limit=1)
            if not history:
                await update.message.reply_text(
                    f"✅ قد ثبت شد.\n\n⚖️ برای محاسبه‌ی BMI «{member_name}»، ابتدا وزن را هم ثبت کنید.",
                    reply_markup=health_menu(member_id)
                )
                return True
            bmi, category = calculate_bmi(history[0][0], height)
            await update.message.reply_text(
                f"🧮 BMI «{member_name}»: {bmi} ({category})\n\nبر اساس قد {height:.0f} سانتی‌متر و آخرین وزن ثبت‌شده.",
                reply_markup=health_menu(member_id)
            )
        except ValueError:
            await update.message.reply_text("❌ لطفاً فقط عدد وارد کنید (مثلاً 175):")
        return True

    elif state == "health_sleep_time":
        member_id = int(data)
        if not re.match(r'^\d{1,2}:\d{2}$', text):
            await update.message.reply_text("❌ فرمت اشتباه است. مثال: 23:30")
            return True
        set_state(user_id, "health_wake_time", f"{member_id}|{text}")
        await update.message.reply_text("😴 ساعت بیدار شدن امروز صبح را وارد کنید (فرمت ۲۴ ساعته، مثلاً 07:00):")
        return True

    elif state == "health_wake_time":
        if not re.match(r'^\d{1,2}:\d{2}$', text):
            await update.message.reply_text("❌ فرمت اشتباه است. مثال: 07:00")
            return True
        member_id_str, sleep_time = data.split("|")
        set_state(user_id, "health_sleep_quality", f"{member_id_str}|{sleep_time}|{text}")
        await update.message.reply_text(
            "😴 کیفیت خواب دیشب را انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("😃 عالی", callback_data="health_quality_عالی")],
                [InlineKeyboardButton("🙂 خوب", callback_data="health_quality_خوب")],
                [InlineKeyboardButton("😐 متوسط", callback_data="health_quality_متوسط")],
                [InlineKeyboardButton("😔 بد", callback_data="health_quality_بد")],
            ])
        )
        return True

    elif state == "music_query":
        query_text = text
        clear_state(user_id)

        # ثبت درخواست و اطلاع‌رسانی به ادمین -- تا ادمین بتواند در صورت
        # پیدا کردن آهنگ در کانال، مستقیم از طریق ربات برای همین کاربر بفرستد
        requester = update.effective_user
        request_id = save_music_request(user_id, query_text)
        await notify_admins(
            update.get_bot(),
            f"🎵 درخواست آهنگ جدید\n\n"
            f"🔍 جستجو: {query_text}\n"
            f"👤 از طرف: {requester.first_name} (@{requester.username or 'ندارد'})\n"
            f"🆔 آیدی: {user_id}\n\n"
            f"اگر این آهنگ را در کانال @LiteMusics پیدا کردید، روی دکمه‌ی زیر بزنید "
            f"و فایل صوتی را برای من (در همین چت) بفرستید تا خودکار برای کاربر ارسال شود.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📤 آماده‌ام فایل را بفرستم", callback_data=f"admin_send_music_{request_id}")]
            ])
        )

        # ۱) اول از همه، پیشنهاد جستجوی مستقیم در کانال موزیک تلگرام
        # (اکثر آهنگ‌های فارسی معمولاً همان‌جا پیدا می‌شوند)
        await update.message.reply_text(
            f"🎵 برای «{query_text}»، اول این کانال موزیک تلگرام را امتحان کنید "
            f"(از داخل خود تلگرام، دکمه‌ی جستجو 🔍 را در کانال بزنید و اسم آهنگ را وارد کنید):\n"
            f"https://t.me/LiteMusics\n\n"
            f"⏳ در همین حال، در حال جستجوی منابع دیگر هم هستم..."
        )

        # ۲) جستجو در ایندکس داخلی (در صورت وجود آهنگ‌های قبلاً ایندکس‌شده)
        results = search_music_index(query_text)
        if results:
            artist, title, file_id = results[0]
            await update.message.reply_audio(
                audio=file_id,
                caption=f"🎤 {artist}\n🎼 {title}"
            )
            await update.message.reply_text("نتیجه پیدا شد ✅", reply_markup=music_menu())
            return True

        # ۳) پیدا نشد در ایندکس داخلی -- با کمک هوش مصنوعی عبارت جستجو دقیق‌تر
        # می‌شود (مثلاً حدس اسم کامل خواننده/آلبوم)، سپس هر لینک جستجو قبل از
        # ارسال با یک درخواست HTTP واقعی تست می‌شود؛ فقط لینک‌هایی که واقعاً
        # از سمت سرور در دسترس بودند فرستاده می‌شوند (نه فقط ساخته می‌شوند)
        await update.message.reply_text("⏳ در حال تست واقعی لینک‌ها قبل از ارسال...")
        refined_query = await music_refine_query_with_ai(query_text)
        verified_links = await music_build_search_links(refined_query or query_text)

        if not verified_links:
            await update.message.reply_text(
                f"❌ متأسفانه در حال حاضر هیچ منبع جستجویی در دسترس نبود. "
                f"لطفاً کانال @LiteMusics را امتحان کنید.",
                reply_markup=music_menu()
            )
            return True

        lines = [f"✅ برای «{query_text}»، این لینک‌ها را امتحان کنید:\n"]
        for i, (label, url) in enumerate(verified_links, 1):
            lines.append(f"{i}. {label}:\n{url}")
        lines.append(
            "\n⚠️ توجه: در دسترس بودن Radio Javan و گوگل از سمت سرور تست شده؛ اگر خود "
            "شما (به‌خاطر تحریم یا فیلترینگ در کشورتان) به یکی از این سایت‌ها دسترسی "
            "نداشته باشید، ممکن است باز نشود. همچنین از نظر قانونی بودن دانلود در کشور "
            "خودتان مطمئن شوید."
        )
        await update.message.reply_text("\n".join(lines), reply_markup=music_menu(), disable_web_page_preview=True)
        return True

    return False

# ─── ایندکس کردن پست‌های جدید کانال موزیک (خودکار) ──────────
async def handle_channel_post(update: Update, context: ContextTypes.DEFAULT_TYPE):
    post = update.channel_post
    if not post:
        return
    chat_username = post.chat.username
    if not chat_username or chat_username.lower() != MUSIC_CHANNEL_USERNAME.lower():
        return
    if not post.audio:
        return
    caption = post.caption or ""
    artist, title = parse_music_caption(caption)
    if not artist:
        artist = post.audio.performer or "نامشخص"
    if not title:
        title = post.audio.title or "نامشخص"
    index_music_track(artist, title, post.audio.file_id, post.message_id)
    print(f"🎵 آهنگ جدید ایندکس شد: {artist} - {title}")

# ═══════════════ پایان MODULE: FUN ══════════════════════════════












# ╔══════════════════════════════════════════════════════════════╗
# ║        MODULE: TOOLS (ابزارهای کاربردی)                       ║
# ║  منوی مرکزی برای دستیار هوش مصنوعی و فیلترشکن.                ║
# ║  OCR و ترجمه طبق تصمیم کاربر حذف شدند (۱۹ تیر ۱۴۰۵).         ║
# ╚══════════════════════════════════════════════════════════════╝

def tools_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🤖 دستیار هوش مصنوعی", callback_data="menu_ai"),
         InlineKeyboardButton("🛡 فیلترشکن", callback_data="menu_antifilter")],  # ← حذف ماژول فیلترشکن = پاک کن این خط
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_main")],
    ])

async def handle_tools(query, user_id):
    data = query.data

    if data == "menu_tools":
        await query.edit_message_text("🛠 ابزارهای کاربردی:", reply_markup=tools_menu())

# ═══════════════ پایان MODULE: TOOLS ══════════════════════════════















# ╔══════════════════════════════════════════════════════════════╗
# ║                  MODULE: SETTINGS                            ║
# ╚══════════════════════════════════════════════════════════════╝

def save_user_profile_field(user_id, field, value):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute(f"UPDATE users SET {field}=? WHERE user_id=?", (value, user_id))
    conn.commit()
    conn.close()

def toggle_email_alerts(user_id, enabled):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE users SET email_alerts_enabled=? WHERE user_id=?", (1 if enabled else 0, user_id))
    conn.commit()
    conn.close()

# ─── پیام محدود کاربر به ادمین (حداکثر ۳ چت جدید در ۲۴ ساعت) -────
# ─── دقیقاً با همون الگوی چت لایسنس: اول مالکین اصلی، امکان انتقال ──
# ─── دستی، و واگذاری خودکار بعد از ۱۵ دقیقه بین ادمین‌های اضافه ────
def get_user_new_chat_count_24h(user_id):
    """تعداد چت‌های جدیدی که این کاربر در ۲۴ ساعت گذشته با ادمین شروع کرده (نه هر پیام داخل یک چت فعال)"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    cutoff = (get_iran_now() - timedelta(hours=24)).strftime("%Y-%m-%d %H:%M")
    c.execute("SELECT COUNT(*) FROM msg_chats WHERE user_id=? AND created_at >= ?", (user_id, cutoff))
    count = c.fetchone()[0]
    conn.close()
    return count

def create_msg_chat(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO msg_chats (user_id, status, admin_id, created_at) VALUES (?, 'pending', ?, ?)",
              (user_id, ADMIN_ID, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    chat_id = c.lastrowid
    conn.commit()
    conn.close()
    return chat_id

def get_msg_chat(chat_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT user_id, status, admin_id FROM msg_chats WHERE id=?", (chat_id,))
    row = c.fetchone()
    conn.close()
    return row

def update_msg_chat_status(chat_id, status):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE msg_chats SET status=? WHERE id=?", (status, chat_id))
    conn.commit()
    conn.close()

def assign_msg_chat_admin(chat_id, admin_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("UPDATE msg_chats SET admin_id=? WHERE id=?", (admin_id, chat_id))
    conn.commit()
    conn.close()

def get_overdue_pending_msg_chats(minutes=15):
    """چت‌های پشتیبانی که هنوز 'pending' مانده، دست مالک اصلی است (منتقل نشده)، و بیش از {minutes} دقیقه گذشته"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    cutoff = (get_iran_now() - timedelta(minutes=minutes)).strftime("%Y-%m-%d %H:%M")
    c.execute("SELECT id, user_id FROM msg_chats WHERE status='pending' AND admin_id=? AND created_at<=?",
              (ADMIN_ID, cutoff))
    rows = c.fetchall()
    conn.close()
    return rows

def log_msg_chat_message(chat_id, user_id, admin_id, direction, content_summary):
    """ثبت هر پیام رد‌و‌بدل‌شده (هر دو جهت) برای گزارش اکسل بعدی؛ فقط ۱۰ روز اخیر نگه داشته می‌شود"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''INSERT INTO msg_chat_logs (chat_id, user_id, admin_id, direction, content_summary, created_at)
                 VALUES (?, ?, ?, ?, ?, ?)''',
              (chat_id, user_id, admin_id, direction, content_summary, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def purge_old_msg_chat_logs(days=10):
    """حذف خودکار لاگ گفتگوهای پشتیبانی قدیمی‌تر از {days} روز -- طبق درخواست کاربر"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    cutoff = (get_iran_now() - timedelta(days=days)).strftime("%Y-%m-%d %H:%M")
    c.execute("DELETE FROM msg_chat_logs WHERE created_at < ?", (cutoff,))
    conn.commit()
    conn.close()

def get_msg_chat_logs():
    """همه‌ی ردیف‌های لاگ باقی‌مانده (همیشه حداکثر ۱۰ روز اخیر، چون قدیمی‌تر خودکار پاک می‌شود)"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT created_at, user_id, admin_id, direction, content_summary FROM msg_chat_logs ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def get_msg_chat_original_text(chat_id):
    """متن اولین پیامی که کاربر برای شروع این گفتگو فرستاده -- برای نمایش دوباره هنگام انتقال/انصراف انتقال"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT content_summary FROM msg_chat_logs WHERE chat_id=? AND direction='user_to_admin'
                 ORDER BY id ASC LIMIT 1''', (chat_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else "(متن اصلی در دسترس نیست)"

def settings_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👤 پروفایل من", callback_data="set_profile"),
         InlineKeyboardButton("✏️ تکمیل پروفایل", callback_data="set_complete_profile")],
        [InlineKeyboardButton("📅 تاریخ انقضا", callback_data="set_expiry"),
         InlineKeyboardButton("📩 پیام به ادمین", callback_data="set_msg_admin")],  # ← حذف فیچر پیام به ادمین = پاک کن این خط
        [InlineKeyboardButton("📢 صفحات و شبکه‌های ما", callback_data="menu_social_links")],  # ← حذف لینک‌های تبلیغاتی = پاک کن این خط
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_main")],
    ])

def profile_complete_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📧 ثبت/ویرایش ایمیل", callback_data="set_email"),
         InlineKeyboardButton("📱 ثبت/ویرایش شماره تلفن", callback_data="set_phone")],
        [InlineKeyboardButton("🎂 ثبت/ویرایش تاریخ تولد", callback_data="set_birthdate"),  # ← حذف فیلدهای پروفایل تکمیلی = این ۳ خط را پاک کن
         InlineKeyboardButton("🏙 ثبت/ویرایش شهر", callback_data="set_city")],
        [InlineKeyboardButton("🌍 ثبت/ویرایش کشور", callback_data="set_country"),
         InlineKeyboardButton("💼 ثبت/ویرایش شغل", callback_data="set_occupation")],
        [InlineKeyboardButton("🔔 فعال/غیرفعال کردن هشدار ایمیلی", callback_data="set_toggle_email_alerts")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="menu_settings")],
    ])

async def handle_settings(query, user_id, context):
    data = query.data
    if data == "menu_settings":
        await query.edit_message_text("⚙️ تنظیمات:", reply_markup=settings_menu())

    elif data == "menu_social_links":  # ← حذف لینک‌های تبلیغاتی = پاک کن این بلوک
        links = get_social_links()
        if not links:
            await query.edit_message_text(
                "📢 هنوز لینکی ثبت نشده.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="menu_settings")]])
            )
            return
        rows = [[InlineKeyboardButton(label, url=url)] for _id, label, url in links]
        rows.append([InlineKeyboardButton("🔙 برگشت", callback_data="menu_settings")])
        await query.edit_message_text("📢 صفحات و شبکه‌های ما:", reply_markup=InlineKeyboardMarkup(rows))

    elif data == "set_profile":
        db_user = get_user(user_id)
        expires = db_user[5] if db_user else "---"
        email = db_user[11] if db_user and len(db_user) > 11 and db_user[11] else "ثبت نشده"
        phone = db_user[12] if db_user and len(db_user) > 12 and db_user[12] else "ثبت نشده"
        email_alerts = "فعال ✅" if db_user and len(db_user) > 13 and db_user[13] else "غیرفعال ❌"
        birthdate = db_user[14] if db_user and len(db_user) > 14 and db_user[14] else "ثبت نشده"
        city = db_user[15] if db_user and len(db_user) > 15 and db_user[15] else "ثبت نشده"
        country = db_user[16] if db_user and len(db_user) > 16 and db_user[16] else "ثبت نشده"
        occupation = db_user[17] if db_user and len(db_user) > 17 and db_user[17] else "ثبت نشده"
        await query.edit_message_text(
            f"👤 پروفایل من\n\n🆔 آیدی: {user_id}\n📅 انقضا: {expires}\n"
            f"📧 ایمیل: {email}\n📱 تلفن: {phone}\n🔔 هشدار ایمیلی: {email_alerts}\n"
            f"🎂 تاریخ تولد: {birthdate}\n🏙 شهر: {city}\n🌍 کشور: {country}\n💼 شغل: {occupation}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="menu_settings")]]))

    elif data == "set_expiry":
        db_user = get_user(user_id)
        expires = db_user[5] if db_user else "---"
        await query.answer(f"📅 انقضا: {expires}", show_alert=True)

    elif data == "set_complete_profile":
        await query.edit_message_text("✏️ تکمیل پروفایل:", reply_markup=profile_complete_menu())

    elif data == "set_birthdate":  # ← حذف فیلدهای پروفایل تکمیلی = این ۴ بلوک elif را پاک کن
        set_state(user_id, "set_awaiting_birthdate")
        await query.edit_message_text(
            "🎂 تاریخ تولد خود را وارد کنید (مثلاً 1370/05/12 یا هر فرمتی که راحت‌اید):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="set_complete_profile")]])
        )

    elif data == "set_city":
        set_state(user_id, "set_awaiting_city")
        await query.edit_message_text(
            "🏙 نام شهر محل سکونت خود را وارد کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="set_complete_profile")]])
        )

    elif data == "set_country":
        set_state(user_id, "set_awaiting_country")
        await query.edit_message_text(
            "🌍 نام کشور محل سکونت خود را وارد کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="set_complete_profile")]])
        )

    elif data == "set_occupation":
        set_state(user_id, "set_awaiting_occupation")
        await query.edit_message_text(
            "💼 شغل خود را وارد کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="set_complete_profile")]])
        )

    elif data == "set_msg_admin":  # ← حذف فیچر پیام به ادمین = پاک کن این بلوک elif
        msg_count = get_user_new_chat_count_24h(user_id)
        if msg_count >= 3:
            await query.answer(
                "⛔ شما در ۲۴ ساعت گذشته به حداکثر تعداد پیام (۳ عدد) رسیده‌اید. لطفاً بعداً دوباره امتحان کنید.",
                show_alert=True
            )
            return
        set_state(user_id, "set_awaiting_admin_msg")
        await query.edit_message_text(
            f"📩 پیام خود را برای ادمین بنویسید (شما {3 - msg_count} پیام دیگر در ۲۴ ساعت آینده می‌توانید بفرستید):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="menu_settings")]])
        )

    elif data == "set_email":
        set_state(user_id, "set_awaiting_email")
        await query.edit_message_text(
            "📧 آدرس ایمیل خود را وارد کنید:\n\n"
            "این ایمیل فقط برای دریافت هشدارهای قیمتی و تحلیل روزانه (در صورت تأیید شما) استفاده می‌شود.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="set_complete_profile")]])
        )

    elif data == "set_phone":
        set_state(user_id, "set_awaiting_phone")
        await query.edit_message_text(
            "📱 شماره تلفن خود را وارد کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="set_complete_profile")]])
        )

    elif data == "set_toggle_email_alerts":
        db_user = get_user(user_id)
        email = db_user[11] if db_user and len(db_user) > 11 else None
        if not email:
            await query.answer("❌ ابتدا باید ایمیل خود را ثبت کنید.", show_alert=True)
            return
        current = db_user[13] if len(db_user) > 13 and db_user[13] else 0
        toggle_email_alerts(user_id, not current)
        new_state = "فعال ✅" if not current else "غیرفعال ❌"
        await query.edit_message_text(f"🔔 هشدار ایمیلی اکنون: {new_state}", reply_markup=profile_complete_menu())

    elif data == "set_test_email":
        db_user = get_user(user_id)
        email = db_user[11] if db_user and len(db_user) > 11 else None
        if not email:
            await query.answer("❌ ابتدا باید ایمیل خود را ثبت کنید.", show_alert=True)
            return
        await query.edit_message_text("⏳ در حال ارسال ایمیل تستی...")
        success = await send_email(
            email, "تست ایمیل ربات Zardari_ai",
            "این یک ایمیل تستی است. اگر این را دریافت کرده‌اید، ارسال ایمیل از ربات به شما به‌درستی کار می‌کند."
        )
        if success:
            await query.message.reply_text(
                f"✅ ایمیل تستی با موفقیت به «{email}» ارسال شد.\n\n"
                f"اگر آن را در اینباکس نمی‌بینید، پوشه‌ی اسپم/Junk را هم بررسی کنید.",
                reply_markup=profile_complete_menu()
            )
        else:
            await query.message.reply_text(
                "❌ ارسال ایمیل ناموفق بود. این معمولاً یعنی RESEND_API_KEY "
                "روی سرور Railway درست تنظیم نشده یا سرویس Resend درخواست را رد کرده. "
                "لطفاً به ادمین اطلاع دهید تا بررسی کند.",
                reply_markup=profile_complete_menu()
            )

    else:
        await query.answer("🔧 به زودی اضافه می‌شود!", show_alert=True)

async def handle_settings_message(user_id, text, update, context):
    state, data = get_state(user_id)

    if state == "set_awaiting_email":
        email = text.strip()
        if "@" not in email or "." not in email.split("@")[-1]:
            await update.message.reply_text("❌ این آدرس ایمیل معتبر به نظر نمی‌رسد. دوباره وارد کنید:")
            return True
        save_user_profile_field(user_id, "email", email)
        toggle_email_alerts(user_id, True)  # به‌طور پیش‌فرض بلافاصله فعال می‌شود تا کاربر مجبور به گام اضافه نباشد
        clear_state(user_id)
        await update.message.reply_text(
            f"✅ ایمیل «{email}» ثبت شد.\n\n"
            f"🔔 هشدار ایمیلی برای این آدرس به‌طور خودکار فعال شد؛ اگر نمی‌خواهید "
            f"ایمیل دریافت کنید، از همین منو می‌توانید غیرفعالش کنید.",
            reply_markup=profile_complete_menu()
        )
        return True

    elif state == "set_awaiting_phone":
        phone = text.strip()
        save_user_profile_field(user_id, "phone", phone)
        clear_state(user_id)
        await update.message.reply_text(f"✅ شماره تلفن «{phone}» ثبت شد.", reply_markup=profile_complete_menu())
        return True

    elif state == "set_awaiting_birthdate":  # ← حذف فیلدهای پروفایل تکمیلی = این ۴ بلوک elif را پاک کن
        birthdate = text.strip()
        save_user_profile_field(user_id, "birthdate", birthdate)
        clear_state(user_id)
        await update.message.reply_text(f"✅ تاریخ تولد «{birthdate}» ثبت شد.", reply_markup=profile_complete_menu())
        return True

    elif state == "set_awaiting_city":
        city = text.strip()
        save_user_profile_field(user_id, "city", city)
        clear_state(user_id)
        await update.message.reply_text(f"✅ شهر «{city}» ثبت شد.", reply_markup=profile_complete_menu())
        return True

    elif state == "set_awaiting_country":
        country = text.strip()
        save_user_profile_field(user_id, "country", country)
        clear_state(user_id)
        await update.message.reply_text(f"✅ کشور «{country}» ثبت شد.", reply_markup=profile_complete_menu())
        return True

    elif state == "set_awaiting_occupation":
        occupation = text.strip()
        save_user_profile_field(user_id, "occupation", occupation)
        clear_state(user_id)
        await update.message.reply_text(f"✅ شغل «{occupation}» ثبت شد.", reply_markup=profile_complete_menu())
        return True

    elif state == "set_awaiting_admin_msg":  # ← حذف فیچر پیام به ادمین = پاک کن این بلوک elif
        msg_count = get_user_new_chat_count_24h(user_id)
        if msg_count >= 3:
            clear_state(user_id)
            await update.message.reply_text(
                "⛔ شما در ۲۴ ساعت گذشته به حداکثر تعداد پیام (۳ عدد) رسیده‌اید. لطفاً بعداً دوباره امتحان کنید.",
                reply_markup=settings_menu()
            )
            return True
        message_text = text.strip()
        clear_state(user_id)
        chat_id = create_msg_chat(user_id)
        log_msg_chat_message(chat_id, user_id, None, "user_to_admin", message_text)
        requester = update.effective_user
        for primary_id in PRIMARY_ADMIN_IDS:
            try:
                await context.bot.send_message(
                    primary_id,
                    msg_request_text(requester.first_name, requester.username, user_id, message_text),
                    reply_markup=msg_request_primary_menu(chat_id)
                )
            except Exception as e:
                print(f"خطای اطلاع‌رسانی پیام کاربر به مالک اصلی {primary_id}: {e}")
        await update.message.reply_text(
            "✅ پیام شما ارسال شد؛ ادمین به‌زودی پاسخ می‌دهد.",
            reply_markup=settings_menu()
        )
        return True

    return False

# ═══════════════ پایان MODULE: SETTINGS ════════════════════════



# ╔══════════════════════════════════════════════════════════════╗
# ║              MODULE: ANTIFILTER (فیلترشکن و پروکسی)          ║
# ║  پروکسی‌های MTProto از یک منبع رایگان که هر ۱۲ ساعت خودکار    ║
# ║  تست و به‌روزرسانی می‌شود دریافت می‌شود (نه تست دستی توسط      ║
# ║  ربات -- این منبع خودش این کار را انجام می‌دهد). همچنین لینک   ║
# ║  دانلود رسمی و مستقیم Psiphon (اندروید+آیفون) و Soren VPN     ║
# ║  (فقط اندروید -- نسخه iOS رسمی برای این اپ پیدا نشد) ارائه    ║
# ║  می‌شود.                                                        ║
# ║  برای حذف: این بلوک رو پاک کن                                 ║
# ║  + خط menu_antifilter رو از main_menu پاک کن                  ║
# ║  + خط af_ رو از ROUTER اصلی پاک کن                             ║
# ╚══════════════════════════════════════════════════════════════╝

# منابع پروکسی‌های MTProto که هر ۱۲ ساعت به‌صورت خودکار تست و
# به‌روزرسانی می‌شوند (چند منبع، تا اگر یکی در دسترس نبود بقیه امتحان شوند)
ANTIFILTER_PROXY_SOURCE_URLS = [
    "https://raw.githubusercontent.com/SoliSpirit/mtproto/master/all_proxies.txt",
    "https://raw.githubusercontent.com/Grim1313/mtproto-for-telegram/master/all_proxies.txt",
    "https://raw.githubusercontent.com/telegram-proxy/telegram-proxy/main/all_proxies.txt",
]

# لینک‌های رسمی دانلود اپلیکیشن‌های ضدفیلتر
PSIPHON_ANDROID_APK_URL = "https://github.com/Psiphon-Inc/psiphon-downloads/raw/master/Downloads/PsiphonAndroid.apk"
PSIPHON_IOS_APPSTORE_URL = "https://apps.apple.com/us/app/psiphon-vpn-freedom-online/id1276263909"
SOREN_VPN_ANDROID_URL = "https://play.google.com/store/apps/details?id=com.provpn.soren"
JAMJAM_VPN_ANDROID_URL = "https://play.google.com/store/apps/details?id=pepe.vpn.free"

async def antifilter_fetch_proxy_list():
    """
    دریافت لیست پروکسی‌های MTProto از منابعی که خودشان هر ۱۲ ساعت پروکسی‌ها
    را تست می‌کنند. چند منبع به‌ترتیب امتحان می‌شوند تا اگر یکی در دسترس
    نبود یا خالی بود، منبع بعدی جایگزین شود.
    خروجی: لیست رشته‌های tg://proxy?server=...&port=...&secret=...
    """
    for source_url in ANTIFILTER_PROXY_SOURCE_URLS:
        try:
            async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
                r = await client.get(source_url)
                if r.status_code != 200:
                    continue
                lines = [
                    line.strip() for line in r.text.splitlines()
                    if line.strip().startswith("tg://proxy") or line.strip().startswith("https://t.me/proxy")
                ]
                if lines:
                    return lines
        except Exception as e:
            print(f"خطای دریافت لیست پروکسی از {source_url}: {e}")
            continue
    return []

def antifilter_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🌐 دریافت ۵ پروکسی تلگرام", callback_data="af_proxies")],
        [InlineKeyboardButton("📱 دانلود Psiphon", callback_data="af_psiphon"),
         InlineKeyboardButton("📱 دانلود Soren VPN", callback_data="af_soren")],
        [InlineKeyboardButton("📱 دانلود JamJam VPN", callback_data="af_jamjam")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="menu_tools")],
    ])

async def handle_antifilter(query, user_id):
    data = query.data

    if data == "menu_antifilter":
        await query.edit_message_text(
            "🛡 فیلترشکن و پروکسی\n\n"
            "پروکسی‌های تلگرام از منبعی دریافت می‌شوند که هر ۱۲ ساعت "
            "به‌صورت خودکار تست می‌شوند (نه لیست ثابت و قدیمی).",
            reply_markup=antifilter_menu()
        )

    elif data == "af_proxies":
        await query.edit_message_text("⏳ در حال دریافت جدیدترین پروکسی‌های تست‌شده...")
        proxy_list = await antifilter_fetch_proxy_list()
        if not proxy_list:
            await query.message.reply_text(
                "❌ در حال حاضر دریافت لیست پروکسی ممکن نیست. کمی بعد دوباره امتحان کنید.",
                reply_markup=antifilter_menu()
            )
            return
        top5 = proxy_list[:5]
        lines = [f"{i+1}. {p}" for i, p in enumerate(top5)]
        text = (
            "🌐 ۵ پروکسی تازه‌ی تلگرام (تست‌شده در ۱۲ ساعت اخیر):\n\n"
            + "\n\n".join(lines) +
            "\n\n💡 روی هرکدام که در تلگرام باز کنید، به‌صورت خودکار تنظیم می‌شود.\n"
            "⚠️ پروکسی‌های رایگان معمولاً چند روز بیشتر دوام ندارند؛ اگر یکی کار نکرد، بقیه را امتحان کنید."
        )
        await query.message.reply_text(text, reply_markup=antifilter_menu(), disable_web_page_preview=True)

    elif data == "af_psiphon":
        await query.edit_message_text("⏳ در حال آماده‌سازی فایل نصب Psiphon...")
        try:
            async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
                r = await client.get(PSIPHON_ANDROID_APK_URL)
                if r.status_code == 200 and len(r.content) > 100000:
                    await query.message.reply_document(
                        document=io.BytesIO(r.content),
                        filename="Psiphon-Android.apk",
                        caption=(
                            "📱 Psiphon برای اندروید (فایل نصب مستقیم، از مخزن رسمی Psiphon)\n\n"
                            f"📱 برای آیفون: {PSIPHON_IOS_APPSTORE_URL}"
                        )
                    )
                else:
                    raise Exception("دریافت فایل APK ناموفق بود")
        except Exception as e:
            print(f"خطای دانلود Psiphon APK: {e}")
            await query.message.reply_text(
                f"⚠️ دریافت مستقیم فایل نصب اندروید ممکن نشد.\n\n"
                f"لینک مستقیم دانلود اندروید:\n{PSIPHON_ANDROID_APK_URL}\n\n"
                f"📱 برای آیفون (اپ استور رسمی):\n{PSIPHON_IOS_APPSTORE_URL}",
                reply_markup=antifilter_menu(),
                disable_web_page_preview=True
            )
            return
        await query.message.reply_text("✅ فایل نصب Psiphon ارسال شد.", reply_markup=antifilter_menu())

    elif data == "af_soren":
        await query.edit_message_text(
            f"📱 Soren VPN\n\n"
            f"اندروید (گوگل پلی، رسمی):\n{SOREN_VPN_ANDROID_URL}\n\n"
            f"⚠️ توضیح صادقانه: برخلاف Psiphon که یک مخزن رسمی گیت‌هاب برای "
            f"فایل نصب مستقیم دارد، Soren VPN هیچ منبع رسمی مستقیمی برای "
            f"فایل APK ندارد و فقط از طریق Google Play توزیع می‌شود. فایل‌های "
            f"APK این اپ که در سایت‌های واسطه (غیررسمی) پیدا می‌شود، اصالت "
            f"تضمین‌شده‌ای ندارند، پس ترجیح دادیم آن‌ها را نفرستیم و فقط لینک "
            f"رسمی Google Play را بدهیم.\n\n"
            f"⚠️ همچنین: نسخه‌ی رسمی Soren VPN برای آیفون هیچ‌جا پیدا نشد؛ "
            f"در حال حاضر این اپلیکیشن فقط برای اندروید در دسترس است.",
            reply_markup=antifilter_menu(),
            disable_web_page_preview=True
        )

    elif data == "af_jamjam":
        await query.edit_message_text(
            f"📱 JamJam VPN\n\n"
            f"اندروید (گوگل پلی، رسمی):\n{JAMJAM_VPN_ANDROID_URL}\n\n"
            f"⚠️ توضیح صادقانه: JamJam VPN هم مثل Soren VPN هیچ منبع رسمی "
            f"مستقیمی برای فایل APK ندارد و فقط از طریق Google Play توزیع "
            f"می‌شود، پس فقط لینک رسمی داده شد، نه فایل از منبع نامطمئن.\n\n"
            f"⚠️ همچنین: هیچ نسخه‌ی رسمی iOS برای JamJam VPN پیدا نشد "
            f"(یک اپ دیگر به اسم مشابه «Jamjams» در اپ‌استور وجود دارد که "
            f"کاملاً برنامه‌ی متفاوتی است، نه همین JamJam VPN؛ برای جلوگیری "
            f"از گمراهی، لینک آن داده نشد). در حال حاضر JamJam VPN فقط "
            f"برای اندروید در دسترس است.",
            reply_markup=antifilter_menu(),
            disable_web_page_preview=True
        )

# ═══════════════ پایان MODULE: ANTIFILTER ════════════════════════



# ╔══════════════════════════════════════════════════════════════╗
# ║              MODULE: ADMIN PANEL (پنل مدیریت)                ║
# ║  فقط برای ادمین‌ها قابل دسترسی است (مالک اصلی + هر کسی که از    ║
# ║  لیست اندیکاتورهای شخصی ارسالی، آمار کلی، مسدود/فعال کردن     ║
# ║  سریع، و لاگ درخواست‌های ناموفق.                               ║
# ║  برای حذف: این بلوک رو پاک کن                                 ║
# ║  + خط admin_panel رو از main_menu پاک کن                      ║
# ║  + خط adm_ رو از ROUTER اصلی پاک کن                            ║
# ║  + state های adm_ رو از message_handler پاک کن                 ║
# ╚══════════════════════════════════════════════════════════════╝

def admin_get_all_users(limit=30):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT user_id, first_name, username, status, approved_at, join_method, email, phone,
                        birthdate, city, country, occupation
                 FROM users ORDER BY user_id DESC LIMIT ?''', (limit,))
    rows = c.fetchall()
    conn.close()
    return rows

def admin_generate_users_excel():
    """
    ساخت فایل اکسل کامل از همه‌ی کاربران (بدون محدودیت تعداد) با تمام
    فیلدهای پروفایل -- برای دانلود توسط ادمین. از openpyxl (که در
    requirements.txt موجود است) استفاده می‌کند و در حافظه (io.BytesIO)
    ساخته می‌شود، بدون نوشتن فایل موقت روی دیسک.
    """
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT user_id, first_name, username, status, approved_at, expires_at,
                        join_method, email, phone, birthdate, city, country, occupation
                 FROM users ORDER BY user_id DESC''')
    rows = c.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "کاربران"
    headers = ["آیدی", "نام", "یوزرنیم", "وضعیت", "تاریخ تأیید", "تاریخ انقضا", "روش عضویت",
               "ایمیل", "تلفن", "تاریخ تولد", "شهر", "کشور", "شغل"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for col_cells in ws.columns:
        length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_msg_chat_logs_excel():
    """
    ساخت فایل اکسل از لاگ گفتگوهای «پیام به ادمین» -- همیشه فقط شامل
    ۱۰ روز اخیر است (چون قدیمی‌تر خودکار توسط msg_chat_escalation_loop
    پاک می‌شود). ستون‌ها مشخص می‌کنند کدام ادمین به کدام کاربر پاسخ داده.
    """
    rows = get_msg_chat_logs()
    wb = Workbook()
    ws = wb.active
    ws.title = "گفتگوهای پشتیبانی (۱۰ روز اخیر)"
    headers = ["زمان", "آیدی کاربر", "آیدی ادمین", "جهت", "متن/نوع پیام"]
    direction_fa = {"user_to_admin": "کاربر → ادمین", "admin_to_user": "ادمین → کاربر"}
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for created_at, user_id, admin_id, direction, content_summary in rows:
        ws.append([created_at, user_id, admin_id or "---", direction_fa.get(direction, direction), content_summary])
    for col_cells in ws.columns:
        length = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def admin_get_all_active_user_ids():
    """دریافت آیدی همه‌ی کاربران فعال (بدون محدودیت تعداد) برای ارسال پیام همگانی"""
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT user_id FROM users WHERE status='active'")
    rows = c.fetchall()
    conn.close()
    return [r[0] for r in rows]

def parse_broadcast_target_ids(data):
    """
    تبدیل مقدار ذخیره‌شده در state (که می‌تواند "all" یا رشته‌ای از
    آیدی‌های جدا‌شده با کاما باشد) به لیست آیدی‌های عددی مقصد. مشترک
    بین مسیر متنی و مسیر انتخاب از کتابخانه در ارسال پیام همگانی.
    """
    if data == "all":
        return admin_get_all_active_user_ids()
    return [int(x.strip()) for x in data.split(",") if x.strip().isdigit()]

def admin_get_stats():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM users")
    total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM users WHERE status='active'")
    active = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM users WHERE status='pending'")
    pending = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM users WHERE status='rejected'")
    rejected = c.fetchone()[0]
    today = get_iran_now().strftime("%Y-%m-%d")
    c.execute("SELECT COUNT(*) FROM users WHERE approved_at LIKE ?", (f"{today}%",))
    approved_today = c.fetchone()[0]
    conn.close()
    return {"total": total, "active": active, "pending": pending, "rejected": rejected, "approved_today": approved_today}

def admin_get_all_custom_indicators(limit=20):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT ci.id, u.first_name, u.user_id, ci.name, ci.source, ci.created_at
                 FROM custom_indicators ci LEFT JOIN users u ON ci.user_id = u.user_id
                 ORDER BY ci.id DESC LIMIT ?''', (limit,))
    rows = c.fetchall()
    conn.close()
    return rows

def admin_get_failed_requests(limit=20):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''SELECT fr.module, fr.query_text, fr.created_at, u.first_name
                 FROM failed_requests fr LEFT JOIN users u ON fr.user_id = u.user_id
                 ORDER BY fr.id DESC LIMIT ?''', (limit,))
    rows = c.fetchall()
    conn.close()
    return rows

JOIN_METHOD_LABELS = {"admin": "تأیید ادمین", "referral": "دعوت", "paid": "پرداخت", None: "نامشخص"}
STATUS_LABELS = {"active": "✅ فعال", "pending": "⏳ در انتظار", "rejected": "❌ رد شده"}

def admin_panel_menu(user_id):
    """
    منوی پنل مدیریت بر اساس سطح دسترسی متفاوت است: مالکین اصلی همه‌ی
    بخش‌ها را می‌بینند؛ ادمین‌های اضافه‌شده به اطلاعات کاربران (لیست/
    اکسل/مدیریت سریع)، پیام همگانی، و مدیریت ادمین‌ها دسترسی ندارند.
    """
    rows = [[InlineKeyboardButton("📊 آمار کلی", callback_data="adm_stats")]]
    if is_primary_admin(user_id):
        rows[0].append(InlineKeyboardButton("👥 لیست کاربران", callback_data="adm_users"))
        rows.append([InlineKeyboardButton("📥 خروجی اکسل کاربران", callback_data="adm_export_users"),  # ← حذف = پاک کن این خط
                     InlineKeyboardButton("📝 لاگ درخواست‌های ناموفق", callback_data="adm_failed")])
        rows.append([InlineKeyboardButton("📢 ارسال پیام به کاربران", callback_data="adm_broadcast_menu")])
    else:
        rows.append([InlineKeyboardButton("📝 لاگ درخواست‌های ناموفق", callback_data="adm_failed")])
    rows.append([InlineKeyboardButton("📶 لیست اندیکاتورهای شخصی ارسالی", callback_data="adm_indicators")])
    rows.append([InlineKeyboardButton("📁 کتابخانه‌ی فایل و متن", callback_data="adm_library")])  # ← حذف کتابخانه = پاک کن این خط
    if is_primary_admin(user_id):
        rows.append([InlineKeyboardButton("🔎 مدیریت سریع یک کاربر (با آیدی)", callback_data="adm_manage_user")])
        rows.append([InlineKeyboardButton("👥 مدیریت ادمین‌ها", callback_data="adm_manage_admins")])  # ← حذف چندادمینی = پاک کن این خط
        rows.append([InlineKeyboardButton("🔗 لینک‌های تبلیغاتی", callback_data="adm_social_links")])  # ← حذف لینک‌های تبلیغاتی = پاک کن این خط
        rows.append([InlineKeyboardButton("📥 اکسل گفتگوهای پشتیبانی (۱۰ روز)", callback_data="adm_export_msg_logs")])  # ← حذف فیچر پیام به ادمین = پاک کن این خط
    rows.append([InlineKeyboardButton("🔙 برگشت", callback_data="back_main")])
    return InlineKeyboardMarkup(rows)

def admin_manage_admins_menu():
    extra_admins = get_all_extra_admins()
    rows = [[InlineKeyboardButton(f"🗑 {uid} (اضافه‌شده {added_at})", callback_data=f"adm_remove_admin_{uid}")]
            for uid, added_at in extra_admins]
    rows.append([InlineKeyboardButton("➕ افزودن ادمین جدید", callback_data="adm_add_admin")])
    rows.append([InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")])
    return InlineKeyboardMarkup(rows)

def admin_library_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🎁 فایل نسخه‌ی تست", callback_data="adm_set_trial_file"),  # ← حذف کتابخانه = پاک کن این خط
         InlineKeyboardButton("💳 فایل نسخه‌ی پولی", callback_data="adm_set_paid_file")],
        [InlineKeyboardButton("📚 فایل‌های آماده (حداکثر ۱۰)", callback_data="adm_files_list"),
         InlineKeyboardButton("💬 متن‌های آماده", callback_data="adm_texts_list")],
        [InlineKeyboardButton("💰 متن قیمت‌گذاری لایسنس", callback_data="adm_set_pricing_text")],  # ← حذف = پاک کن این خط
        [InlineKeyboardButton("❓ متن راهنما", callback_data="adm_set_help_text")],  # ← حذف راهنما = پاک کن این خط
        [InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")],
    ])

def admin_files_list_menu():
    files = get_admin_files()
    rows = [[InlineKeyboardButton(f"📄 {label}", callback_data=f"adm_file_view_{fid}")]
            for fid, label, *_rest in files]
    if len(files) < 10:
        rows.append([InlineKeyboardButton("➕ افزودن فایل جدید", callback_data="adm_files_add")])
    rows.append([InlineKeyboardButton("🔙 برگشت", callback_data="adm_library")])
    return InlineKeyboardMarkup(rows), len(files)

def admin_texts_list_menu():
    texts = get_admin_texts()
    rows = [[InlineKeyboardButton(f"💬 {label}", callback_data=f"adm_text_view_{tid}")]
            for tid, label, _content in texts]
    rows.append([InlineKeyboardButton("➕ افزودن متن جدید", callback_data="adm_texts_add")])
    rows.append([InlineKeyboardButton("🔙 برگشت", callback_data="adm_library")])
    return InlineKeyboardMarkup(rows)

def license_chat_action_menu(chat_id):
    """
    دکمه‌های سریع در حین چت زنده‌ی لایسنس -- علاوه بر پایان مکالمه،
    امکان ارسال فوری فایل نسخه‌ی پولی، انتخاب از کتابخانه‌ی فایل‌های
    آماده، یا انتخاب از متن‌های آماده، بدون خروج از مکالمه.
    """
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💳 ارسال فایل نسخه‌ی پولی", callback_data=f"lic_send_paid_{chat_id}")],
        [InlineKeyboardButton("📚 ارسال فایل آماده", callback_data=f"lic_pick_file_{chat_id}"),
         InlineKeyboardButton("💬 ارسال متن آماده", callback_data=f"lic_pick_text_{chat_id}")],
        [InlineKeyboardButton("🔚 پایان مکالمه", callback_data=f"lic_end_{chat_id}")],
    ])

def license_request_primary_menu(chat_id):
    """پیامی که فقط مالک اصلی می‌بیند -- هم می‌تواند خودش شروع کند، هم منتقل کند"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ شروع مکالمه", callback_data=f"lic_start_{chat_id}")],
        [InlineKeyboardButton("↪️ انتقال به ادمین دیگر", callback_data=f"lic_transfer_pick_{chat_id}")],
    ])

def license_request_assigned_menu(chat_id):
    """پیامی که به ادمینی که چت به او منتقل/واگذار شده می‌رسد -- فقط شروع، بدون انتقال بیشتر"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ شروع مکالمه", callback_data=f"lic_start_{chat_id}")],
    ])

def license_request_text(requester_first_name, requester_username, user_id):
    return (
        f"💳 درخواست نسخه‌ی پولی (لایسنس) ربات مدیریت سرمایه\n\n"
        f"👤 از طرف: {requester_first_name} (@{requester_username or 'ندارد'})\n"
        f"🆔 آیدی: {user_id}\n\n"
        f"برای شروع مکالمه‌ی دوطرفه (متن/عکس/فایل/صدا/ویدیو، شامل فوروارد از "
        f"کانال‌های دیگر) با این کاربر، دکمه‌ی زیر را بزنید."
    )

def msg_chat_action_menu(chat_id):
    """در حین چت فعال «پیام به ادمین» -- فقط پایان مکالمه (برخلاف چت لایسنس، دکمه‌ی سریع فایل/متن ندارد)"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔚 پایان مکالمه", callback_data=f"msgc_end_{chat_id}")],
    ])

def msg_request_primary_menu(chat_id):
    """پیامی که فقط مالکین اصلی می‌بینند -- هم می‌توانند خودشان پاسخ بدهند، هم منتقل کنند"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("↩️ پاسخ بده", callback_data=f"msgc_start_{chat_id}")],
        [InlineKeyboardButton("↪️ انتقال به ادمین دیگر", callback_data=f"msgc_transfer_pick_{chat_id}")],
    ])

def msg_request_assigned_menu(chat_id):
    """پیامی که به ادمینی که این گفتگو به او منتقل/واگذار شده می‌رسد -- فقط پاسخ، بدون انتقال بیشتر"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("↩️ پاسخ بده", callback_data=f"msgc_start_{chat_id}")],
    ])

def msg_request_text(requester_first_name, requester_username, user_id, message_text):
    return (
        f"📩 پیام جدید از کاربر\n\n"
        f"👤 از طرف: {requester_first_name} (@{requester_username or 'ندارد'})\n"
        f"🆔 آیدی: {user_id}\n\n"
        f"💬 متن پیام:\n{message_text}"
    )

async def handle_admin_panel(query, user_id, context):
    if not is_admin(user_id):
        await query.answer("⛔ این بخش فقط برای ادمین است.", show_alert=True)
        return

    data = query.data

    # ─── دفاع در عمق: علاوه بر مخفی بودن این دکمه‌ها در منو برای ادمین‌های
    # اضافه، اگر کسی مستقیم callback_data را بزند هم همینجا رد می‌شود ───
    _primary_only_prefixes = ("adm_users", "adm_export_users", "adm_manage_user", "adm_block_",
                               "adm_activate_", "adm_broadcast_", "adm_social_link", "adm_export_msg_logs")
    if data.startswith(_primary_only_prefixes) and not is_primary_admin(user_id):
        await query.answer("⛔ این بخش فقط برای مالک اصلی ربات است.", show_alert=True)
        return

    if data == "admin_panel":
        await query.edit_message_text("🛠 پنل مدیریت:", reply_markup=admin_panel_menu(user_id))

    elif data == "adm_broadcast_menu":
        await query.edit_message_text(
            "📢 ارسال پیام به کاربران\n\nمقصد پیام را انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📣 ارسال به همه‌ی کاربران فعال", callback_data="adm_broadcast_all")],
                [InlineKeyboardButton("📋 ارسال به لیست خاص (چند آیدی)", callback_data="adm_broadcast_list")],
                [InlineKeyboardButton("👤 ارسال به یک کاربر خاص", callback_data="adm_broadcast_one")],
                [InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")],
            ])
        )

    elif data == "adm_broadcast_all":
        set_state(user_id, "adm_awaiting_broadcast_content", "all")
        await query.edit_message_text(
            "📣 محتوایی که می‌خواهید به همه‌ی کاربران فعال ارسال شود را همین‌جا بفرستید "
            "(متن، عکس، فایل، صدا، ویدیو -- حتی فوروارد از یک کانال دیگر)، یا از "
            "کتابخانه انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📚 انتخاب فایل آماده", callback_data="adm_broadcast_pick_file"),
                 InlineKeyboardButton("💬 انتخاب متن آماده", callback_data="adm_broadcast_pick_text")],
                [InlineKeyboardButton("🔙 برگشت", callback_data="adm_broadcast_menu")],
            ])
        )

    elif data == "adm_broadcast_list":
        set_state(user_id, "adm_awaiting_broadcast_ids")
        await query.edit_message_text(
            "📋 آیدی‌های عددی کاربران مقصد را با کاما (,) جدا از هم بنویسید.\n\nمثال: 123456789, 987654321",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="adm_broadcast_menu")]])
        )

    elif data == "adm_broadcast_one":
        set_state(user_id, "adm_awaiting_broadcast_single_id")
        await query.edit_message_text(
            "👤 آیدی عددی کاربر مقصد را وارد کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="adm_broadcast_menu")]])
        )

    elif data == "adm_broadcast_pick_file":  # ← حذف = پاک کن این ۴ بلوک elif
        state, target_data = get_state(user_id)
        if state != "adm_awaiting_broadcast_content" or not target_data:
            await query.answer("اول باید مقصد پیام را انتخاب کنید.", show_alert=True)
            return
        files = get_admin_files()
        if not files:
            await query.answer("هنوز فایلی در کتابخانه ثبت نشده.", show_alert=True)
            return
        rows = [[InlineKeyboardButton(f"📄 {label}", callback_data=f"adm_broadcast_use_file_{fid}")]
                for fid, label, *_rest in files]
        rows.append([InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")])
        await query.edit_message_text("📚 کدام فایل ارسال شود؟", reply_markup=InlineKeyboardMarkup(rows))

    elif data.startswith("adm_broadcast_use_file_"):
        state, target_data = get_state(user_id)
        if state != "adm_awaiting_broadcast_content" or not target_data:
            await query.answer("این عملیات منقضی شده؛ دوباره از منوی ارسال پیام شروع کنید.", show_alert=True)
            return
        file_db_id = int(data.replace("adm_broadcast_use_file_", ""))
        row = get_admin_file(file_db_id)
        if not row:
            await query.answer("این فایل پیدا نشد.", show_alert=True)
            return
        target_ids = parse_broadcast_target_ids(target_data)
        label, file_type, file_id, caption = row
        import json
        set_state(user_id, "adm_broadcast_ready",
                   json.dumps({"target_ids": target_ids, "file_type": file_type, "file_id": file_id, "caption": caption}))
        await send_content_by_type(context, user_id, file_type, file_id, caption, default_caption=label)
        await query.message.reply_text(
            f"👆 پیش‌نمایش بالا. این محتوا به {len(target_ids)} کاربر ارسال خواهد شد. تأیید می‌کنید؟",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ بله، ارسال کن", callback_data="adm_broadcast_confirm_send")],
                [InlineKeyboardButton("❌ لغو", callback_data="admin_panel")],
            ])
        )

    elif data == "adm_broadcast_pick_text":
        state, target_data = get_state(user_id)
        if state != "adm_awaiting_broadcast_content" or not target_data:
            await query.answer("اول باید مقصد پیام را انتخاب کنید.", show_alert=True)
            return
        texts = get_admin_texts()
        if not texts:
            await query.answer("هنوز متنی در کتابخانه ثبت نشده.", show_alert=True)
            return
        rows = [[InlineKeyboardButton(f"💬 {label}", callback_data=f"adm_broadcast_use_text_{tid}")]
                for tid, label, _content in texts]
        rows.append([InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")])
        await query.edit_message_text("💬 کدام متن ارسال شود؟", reply_markup=InlineKeyboardMarkup(rows))

    elif data.startswith("adm_broadcast_use_text_"):
        state, target_data = get_state(user_id)
        if state != "adm_awaiting_broadcast_content" or not target_data:
            await query.answer("این عملیات منقضی شده؛ دوباره از منوی ارسال پیام شروع کنید.", show_alert=True)
            return
        text_id = int(data.replace("adm_broadcast_use_text_", ""))
        row = get_admin_text(text_id)
        if not row:
            await query.answer("این متن پیدا نشد.", show_alert=True)
            return
        target_ids = parse_broadcast_target_ids(target_data)
        label, content = row
        import json
        set_state(user_id, "adm_broadcast_ready",
                   json.dumps({"target_ids": target_ids, "file_type": "text", "file_id": None, "caption": content}))
        await query.edit_message_text(
            f"👆 پیش‌نمایش:\n\n{content}\n\nاین متن به {len(target_ids)} کاربر ارسال خواهد شد. تأیید می‌کنید؟",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ بله، ارسال کن", callback_data="adm_broadcast_confirm_send")],
                [InlineKeyboardButton("❌ لغو", callback_data="admin_panel")],
            ])
        )

    elif data == "adm_broadcast_confirm_send":
        state, sdata = get_state(user_id)
        if state != "adm_broadcast_ready":
            return
        import json
        payload = json.loads(sdata)
        target_ids = payload["target_ids"]
        file_type = payload.get("file_type", "text")
        file_id = payload.get("file_id")
        caption = payload.get("caption")
        if caption is None and "message_text" in payload:  # سازگاری با فرمت قدیمی، اگر جایی باقی مانده باشد
            caption = payload["message_text"]
        await query.edit_message_text(f"⏳ در حال ارسال به {len(target_ids)} کاربر...")
        success_count, fail_count = 0, 0
        for target_id in target_ids:
            sent = await send_content_by_type(context, target_id, file_type, file_id, caption)
            if sent:
                success_count += 1
            else:
                fail_count += 1
            await asyncio.sleep(0.05)  # فاصله‌ی کوچک برای احترام به محدودیت سرعت تلگرام
        clear_state(user_id)
        await query.message.reply_text(
            f"✅ ارسال تمام شد.\n\n📤 موفق: {success_count} | ❌ ناموفق: {fail_count}",
            reply_markup=admin_panel_menu(user_id)
        )

    elif data == "adm_stats":
        stats = admin_get_stats()
        text = (
            f"📊 آمار کلی ربات\n\n"
            f"👥 کل کاربران ثبت‌شده: {stats['total']}\n"
            f"✅ فعال: {stats['active']}\n"
            f"⏳ در انتظار تأیید: {stats['pending']}\n"
            f"❌ رد شده: {stats['rejected']}\n"
            f"🆕 تأییدشده امروز: {stats['approved_today']}"
        )
        await query.edit_message_text(text, reply_markup=admin_panel_menu(user_id))

    elif data == "adm_users":
        users = admin_get_all_users()
        if not users:
            await query.edit_message_text("📭 هنوز کاربری ثبت نشده.", reply_markup=admin_panel_menu(user_id))
            return
        lines = []
        for uid, first_name, username, status, approved_at, join_method, email, phone, birthdate, city, country, occupation in users:
            status_label = STATUS_LABELS.get(status, status)
            method_label = JOIN_METHOD_LABELS.get(join_method, join_method or "نامشخص")
            contact_parts = []
            if email:
                contact_parts.append(f"📧 {email}")
            if phone:
                contact_parts.append(f"📱 {phone}")
            contact_line = f"\n   {' | '.join(contact_parts)}" if contact_parts else ""
            profile_parts = []
            if birthdate:
                profile_parts.append(f"🎂 {birthdate}")
            if city:
                profile_parts.append(f"🏙 {city}")
            if country:
                profile_parts.append(f"🌍 {country}")
            if occupation:
                profile_parts.append(f"💼 {occupation}")
            profile_line = f"\n   {' | '.join(profile_parts)}" if profile_parts else ""
            lines.append(
                f"{status_label} {first_name} (@{username or '---'})\n"
                f"   🆔 {uid} | روش: {method_label} | تاریخ: {format_date_fa(approved_at) if approved_at else '---'}{contact_line}{profile_line}"
            )
        text = "👥 آخرین ۳۰ کاربر:\n\n" + "\n\n".join(lines)
        # تلگرام محدودیت طول پیام دارد؛ در صورت طولانی بودن تکه‌تکه ارسال شود
        if len(text) > 4000:
            chunks = [text[i:i+4000] for i in range(0, len(text), 4000)]
            await query.edit_message_text(chunks[0])
            for chunk in chunks[1:]:
                await query.message.reply_text(chunk)
            await query.message.reply_text("👆 لیست کاربران", reply_markup=admin_panel_menu(user_id))
        else:
            await query.edit_message_text(text, reply_markup=admin_panel_menu(user_id))

    elif data == "adm_export_users":  # ← حذف = پاک کن این خط
        await query.answer("⏳ در حال ساخت فایل اکسل...")
        buffer = admin_generate_users_excel()
        await context.bot.send_document(
            user_id,
            document=buffer,
            filename=f"users_{get_iran_now().strftime('%Y-%m-%d_%H-%M')}.xlsx",
            caption="📊 خروجی اکسل کامل لیست کاربران (همه‌ی فیلدهای پروفایل)"
        )
        await query.message.reply_text("👆 فایل اکسل ارسال شد.", reply_markup=admin_panel_menu(user_id))

    elif data == "adm_export_msg_logs":  # ← حذف فیچر پیام به ادمین = پاک کن این بلوک
        await query.answer("⏳ در حال ساخت فایل اکسل...")
        rows = get_msg_chat_logs()
        if not rows:
            await query.edit_message_text("📭 در ۱۰ روز اخیر هیچ گفتگویی ثبت نشده.", reply_markup=admin_panel_menu(user_id))
            return
        buffer = generate_msg_chat_logs_excel()
        await context.bot.send_document(
            user_id,
            document=buffer,
            filename=f"support_chats_{get_iran_now().strftime('%Y-%m-%d_%H-%M')}.xlsx",
            caption="📊 گفتگوهای «پیام به ادمین» -- ۱۰ روز اخیر (قدیمی‌تر خودکار پاک می‌شود)"
        )
        await query.message.reply_text("👆 فایل اکسل ارسال شد.", reply_markup=admin_panel_menu(user_id))

    elif data == "adm_indicators":
        indicators = admin_get_all_custom_indicators()
        if not indicators:
            await query.edit_message_text("📭 هنوز هیچ اندیکاتور شخصی ارسال نشده.", reply_markup=admin_panel_menu(user_id))
            return
        keyboard = []
        for ind_id, first_name, uid, name, source, created_at in indicators:
            label = f"📶 {name} - {first_name or uid} ({format_date_fa(created_at, with_time=False)})"
            keyboard.append([InlineKeyboardButton(label[:60], callback_data=f"adm_ind_dl_{ind_id}")])
        keyboard.append([InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")])
        await query.edit_message_text("📶 اندیکاتورهای شخصی ارسالی (برای دانلود کد روی هرکدام بزنید):", reply_markup=InlineKeyboardMarkup(keyboard))

    elif data.startswith("adm_ind_dl_"):
        ind_id = int(data.replace("adm_ind_dl_", ""))
        name, code = get_custom_indicator_by_id(ind_id)
        if not code:
            await query.answer("این اندیکاتور پیدا نشد.", show_alert=True)
            return
        await query.message.reply_document(
            document=io.BytesIO(code.encode("utf-8")),
            filename=f"{name}.py",
            caption=f"📶 کد اندیکاتور «{name}»"
        )

    elif data == "adm_failed":
        failed = admin_get_failed_requests()
        if not failed:
            await query.edit_message_text("📭 هیچ درخواست ناموفقی ثبت نشده.", reply_markup=admin_panel_menu(user_id))
            return
        lines = [f"🔸 [{module}] «{query_text}» -- {first_name or 'نامشخص'} -- {format_date_fa(created_at)}"
                 for module, query_text, created_at, first_name in failed]
        text = "📝 آخرین درخواست‌های ناموفق:\n\n" + "\n".join(lines)
        await query.edit_message_text(text[:4000], reply_markup=admin_panel_menu(user_id))

    elif data == "adm_library":  # ← حذف کتابخانه = پاک کن این بلوک
        await query.edit_message_text("📁 کتابخانه‌ی فایل و متن:", reply_markup=admin_library_menu())

    elif data == "adm_set_pricing_text":  # ← حذف = پاک کن این بلوک
        set_state(user_id, "adm_awaiting_pricing_text")
        current = get_bot_setting("pricing_text")
        preview = f"\n\nمتن فعلی:\n\n{current}" if current else "\n\n(هنوز متنی تنظیم نشده؛ فعلاً یک متن پیش‌فرض عمومی نمایش داده می‌شود.)"
        full_text = (
            f"💰 متن جدید قیمت‌گذاری لایسنس را بنویسید (شامل قیمت‌های ۳/۶/۱۲ ماهه و روش پرداخت). "
            f"همین متن دقیقاً به کاربر نمایش داده می‌شود.{preview}"
        )[:4000]
        await query.edit_message_text(
            full_text,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="adm_library")]])
        )

    elif data == "adm_set_help_text":  # ← حذف راهنما = پاک کن این بلوک
        set_state(user_id, "adm_awaiting_help_text")
        current = get_bot_setting("help_text")
        preview = f"\n\nمتن فعلی:\n\n{current}" if current else "\n\n(هنوز متنی تنظیم نشده؛ فعلاً متن پیش‌فرض عمومی نمایش داده می‌شود.)"
        full_text = f"❓ متن جدید راهنما را بنویسید. همین متن دقیقاً به کاربر نمایش داده می‌شود.{preview}"[:4000]
        await query.edit_message_text(
            full_text,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="adm_library")]])
        )

    elif data == "adm_set_trial_file":
        set_state(user_id, "adm_awaiting_trial_file_content")
        current = get_bot_setting("trial_file")
        status_line = "\n\n(الان یک فایل تنظیم‌شده دارید؛ با ارسال محتوای جدید جایگزین می‌شود.)" if current else ""
        await query.edit_message_text(
            "🎁 متن، عکس، فایل، صدا یا ویدیوی نسخه‌ی تست را همین‌جا بفرستید. "
            "این محتوا ذخیره می‌شود و از این پس هر کاربری که «دریافت نسخه تست» را بزند، "
            f"خودکار همین را دریافت می‌کند.{status_line}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="adm_library")]])
        )

    elif data == "adm_set_paid_file":
        set_state(user_id, "adm_awaiting_paid_file_content")
        current = get_bot_setting("paid_file")
        status_line = "\n\n(الان یک فایل تنظیم‌شده دارید؛ با ارسال محتوای جدید جایگزین می‌شود.)" if current else ""
        await query.edit_message_text(
            "💳 متن، عکس، فایل، صدا یا ویدیوی نسخه‌ی پولی را همین‌جا بفرستید. این محتوا "
            "ذخیره می‌شود و در حین چت زنده‌ی لایسنس، با یک دکمه قابل ارسال فوری برای "
            f"کاربر خواهد بود.{status_line}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="adm_library")]])
        )

    elif data == "adm_files_list":
        menu, count = admin_files_list_menu()
        await query.edit_message_text(f"📚 فایل‌های آماده ({count}/10):", reply_markup=menu)

    elif data == "adm_files_add":
        if count_admin_files() >= 10:
            await query.answer("⛔ در حال حاضر ۱۰ فایل ثبت شده (حداکثر مجاز)؛ اول یکی را حذف کنید.", show_alert=True)
            return
        set_state(user_id, "adm_awaiting_new_file_label")
        await query.edit_message_text(
            "یک اسم کوتاه برای این فایل بنویسید (برای شناسایی سریع بعداً، مثلاً «بروشور فارسی»):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 انصراف", callback_data="adm_files_list")]])
        )

    elif data.startswith("adm_file_view_"):
        file_db_id = int(data.replace("adm_file_view_", ""))
        row = get_admin_file(file_db_id)
        if not row:
            await query.answer("این فایل پیدا نشد.", show_alert=True)
            return
        label, file_type, file_id, caption = row
        await query.edit_message_text(
            f"📄 {label}\nنوع: {file_type}",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("👁 پیش‌نمایش (ارسال برای خودم)", callback_data=f"adm_file_preview_{file_db_id}")],
                [InlineKeyboardButton("✏️ ویرایش", callback_data=f"adm_file_edit_{file_db_id}"),  # ← حذف ویرایش کتابخانه = پاک کن این خط
                 InlineKeyboardButton("🗑 حذف", callback_data=f"adm_file_del_{file_db_id}")],
                [InlineKeyboardButton("🔙 برگشت", callback_data="adm_files_list")],
            ])
        )

    elif data.startswith("adm_file_preview_"):
        file_db_id = int(data.replace("adm_file_preview_", ""))
        row = get_admin_file(file_db_id)
        if not row:
            await query.answer("این فایل پیدا نشد.", show_alert=True)
            return
        label, file_type, file_id, caption = row
        await send_content_by_type(context, user_id, file_type, file_id, caption, default_caption=label)
        await query.answer("✅ برای شما ارسال شد.")

    elif data.startswith("adm_file_edit_"):  # ← حذف ویرایش کتابخانه = پاک کن این بلوک
        file_db_id = int(data.replace("adm_file_edit_", ""))
        row = get_admin_file(file_db_id)
        if not row:
            await query.answer("این فایل پیدا نشد.", show_alert=True)
            return
        label, file_type, file_id, caption = row
        set_state(user_id, "adm_awaiting_edit_file_label", str(file_db_id))
        await query.edit_message_text(
            f"✏️ ویرایش «{label}»\n\nاسم جدید را بنویسید (یا همان اسم قبلی را دوباره بفرستید اگر نمی‌خواهید عوض شود):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 انصراف", callback_data=f"adm_file_view_{file_db_id}")]])
        )

    elif data.startswith("adm_file_del_"):
        file_db_id = int(data.replace("adm_file_del_", ""))
        delete_admin_file(file_db_id)
        menu, count = admin_files_list_menu()
        await query.edit_message_text(f"✅ حذف شد.\n\n📚 فایل‌های آماده ({count}/10):", reply_markup=menu)

    elif data == "adm_texts_list":
        texts = get_admin_texts()
        if not texts:
            await query.edit_message_text(
                "💬 هنوز متنی ثبت نشده.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("➕ افزودن متن جدید", callback_data="adm_texts_add")],
                    [InlineKeyboardButton("🔙 برگشت", callback_data="adm_library")],
                ])
            )
            return
        await query.edit_message_text("💬 متن‌های آماده:", reply_markup=admin_texts_list_menu())

    elif data == "adm_texts_add":
        set_state(user_id, "adm_awaiting_new_text_label")
        await query.edit_message_text(
            "یک اسم کوتاه برای این متن بنویسید (مثلاً «شماره حساب تتر» یا «سوال: نحوه واریز»):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 انصراف", callback_data="adm_texts_list")]])
        )

    elif data.startswith("adm_text_view_"):
        text_id = int(data.replace("adm_text_view_", ""))
        row = get_admin_text(text_id)
        if not row:
            await query.answer("این متن پیدا نشد.", show_alert=True)
            return
        label, content = row
        await query.edit_message_text(
            f"💬 {label}\n\n{content}",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✏️ ویرایش", callback_data=f"adm_text_edit_{text_id}")],
                [InlineKeyboardButton("🗑 حذف", callback_data=f"adm_text_del_{text_id}")],
                [InlineKeyboardButton("🔙 برگشت", callback_data="adm_texts_list")],
            ])
        )

    elif data.startswith("adm_text_edit_"):
        text_id = int(data.replace("adm_text_edit_", ""))
        row = get_admin_text(text_id)
        if not row:
            await query.answer("این متن پیدا نشد.", show_alert=True)
            return
        set_state(user_id, "adm_awaiting_edit_text_content", str(text_id))
        await query.edit_message_text("متن جدید را بنویسید (جایگزین متن قبلی می‌شود؛ اسم همان قبلی باقی می‌ماند):")

    elif data.startswith("adm_text_del_"):
        text_id = int(data.replace("adm_text_del_", ""))
        delete_admin_text(text_id)
        texts = get_admin_texts()
        if not texts:
            await query.edit_message_text(
                "✅ حذف شد.\n\n💬 هنوز متنی ثبت نشده.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("➕ افزودن متن جدید", callback_data="adm_texts_add")],
                    [InlineKeyboardButton("🔙 برگشت", callback_data="adm_library")],
                ])
            )
            return
        await query.edit_message_text("✅ حذف شد.\n\n💬 متن‌های آماده:", reply_markup=admin_texts_list_menu())

    elif data == "adm_manage_user":
        set_state(user_id, "adm_awaiting_user_id")
        await query.edit_message_text(
            "🔎 آیدی عددی کاربر مورد نظر را وارد کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")]])
        )

    elif data == "adm_manage_admins":  # ← حذف چندادمینی = پاک کن این ۴ بلوک elif
        if not is_primary_admin(user_id):
            await query.answer("⛔ فقط مالک اصلی ربات می‌تواند ادمین‌ها را مدیریت کند.", show_alert=True)
            return
        extra_admins = get_all_extra_admins()
        count_line = f"تعداد ادمین‌های اضافه‌شده: {len(extra_admins)}"
        await query.edit_message_text(
            f"👥 مدیریت ادمین‌ها\n\n{count_line}\n\n"
            f"ادمین اصلی (مالک ربات) همیشه ثابت است و از اینجا قابل حذف نیست.",
            reply_markup=admin_manage_admins_menu()
        )

    elif data == "adm_add_admin":
        if not is_primary_admin(user_id):
            await query.answer("⛔ فقط مالک اصلی ربات می‌تواند ادمین‌ها را مدیریت کند.", show_alert=True)
            return
        set_state(user_id, "adm_awaiting_new_admin_id")
        await query.edit_message_text(
            "آیدی عددی تلگرام فردی که می‌خواهید ادمین شود را وارد کنید.\n\n"
            "(کاربر باید حداقل یک‌بار /start ربات را زده باشد تا بتوانیم بهش پیام بدهیم؛ "
            "برای گرفتن آیدی عددی هرکسی می‌توانید از ربات‌هایی مثل @userinfobot کمک بگیرید.)",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 انصراف", callback_data="adm_manage_admins")]])
        )

    elif data.startswith("adm_remove_admin_"):
        if not is_primary_admin(user_id):
            await query.answer("⛔ فقط مالک اصلی ربات می‌تواند ادمین‌ها را مدیریت کند.", show_alert=True)
            return
        target_admin_id = int(data.replace("adm_remove_admin_", ""))
        remove_admin(target_admin_id)
        clear_state(target_admin_id)  # اگه وسط یه کاری بود، حالتش پاک بشه که گیر نکنه
        try:
            await context.bot.send_message(target_admin_id, "ℹ️ دسترسی ادمین شما به این ربات لغو شد.")
        except Exception as e:
            print(f"خطای اطلاع‌رسانی لغو دسترسی ادمین: {e}")
        await query.edit_message_text("✅ حذف شد.", reply_markup=admin_manage_admins_menu())

    elif data == "adm_social_links":  # ← حذف لینک‌های تبلیغاتی = پاک کن این ۳ بلوک elif
        links = get_social_links()
        rows = [[InlineKeyboardButton(f"🗑 {label}", callback_data=f"adm_social_link_del_{lid}")]
                for lid, label, url in links]
        rows.append([InlineKeyboardButton("➕ افزودن لینک جدید", callback_data="adm_social_link_add")])
        rows.append([InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")])
        await query.edit_message_text(
            f"🔗 لینک‌های تبلیغاتی (تعداد: {len(links)})\n\n"
            f"این لینک‌ها زیر ⚙️ تنظیمات → 📢 صفحات و شبکه‌های ما به همه‌ی کاربران نشان داده می‌شود.",
            reply_markup=InlineKeyboardMarkup(rows)
        )

    elif data == "adm_social_link_add":
        set_state(user_id, "adm_awaiting_social_link_label")
        await query.edit_message_text(
            "اسمی که کاربر می‌بیند را بنویسید (مثلاً «📸 اینستاگرام» یا «📢 کانال تلگرام»):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 انصراف", callback_data="adm_social_links")]])
        )

    elif data.startswith("adm_social_link_del_"):
        link_id = int(data.replace("adm_social_link_del_", ""))
        delete_social_link(link_id)
        links = get_social_links()
        rows = [[InlineKeyboardButton(f"🗑 {label}", callback_data=f"adm_social_link_del_{lid}")]
                for lid, label, url in links]
        rows.append([InlineKeyboardButton("➕ افزودن لینک جدید", callback_data="adm_social_link_add")])
        rows.append([InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")])
        await query.edit_message_text(f"✅ حذف شد.\n\n🔗 لینک‌های تبلیغاتی (تعداد: {len(links)})", reply_markup=InlineKeyboardMarkup(rows))

    elif data.startswith("adm_block_") or data.startswith("adm_activate_"):
        target_uid = int(data.split("_")[-1])
        is_activate = data.startswith("adm_activate_")
        if is_activate:
            approve_user(target_uid)
        else:
            reject_user(target_uid)
        status_label = "✅ فعال" if is_activate else "❌ رد شده"
        await query.edit_message_text(
            f"✅ وضعیت کاربر {target_uid} به «{status_label}» تغییر کرد.",
            reply_markup=admin_panel_menu(user_id)
        )
        try:
            if is_activate:
                await context.bot.send_message(target_uid, "🎉 دسترسی شما توسط ادمین فعال شد.\n\n/start بزنید.")
            else:
                await context.bot.send_message(target_uid, "❌ دسترسی شما توسط ادمین مسدود شد.")
        except Exception as e:
            print(f"خطای اطلاع‌رسانی تغییر وضعیت به کاربر: {e}")

async def handle_admin_panel_message(user_id, text, update):
    if not is_admin(user_id):
        return False
    state, data = get_state(user_id)

    if state == "adm_awaiting_broadcast_content":
        # data شامل "all" یا رشته‌ی آیدی‌های جدا‌شده با کاما است (بسته به مبدأ)
        message_text = text
        target_ids = parse_broadcast_target_ids(data)
        if not target_ids:
            await update.message.reply_text("❌ هیچ مقصدی برای ارسال پیدا نشد.", reply_markup=admin_panel_menu(user_id))
            clear_state(user_id)
            return True
        import json
        set_state(user_id, "adm_broadcast_ready",
                  json.dumps({"target_ids": target_ids, "file_type": "text", "file_id": None, "caption": message_text}))
        await update.message.reply_text(
            f"📢 پیش‌نمایش پیام:\n\n{message_text}\n\n"
            f"این پیام به {len(target_ids)} کاربر ارسال خواهد شد. تأیید می‌کنید؟",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ بله، ارسال کن", callback_data="adm_broadcast_confirm_send")],
                [InlineKeyboardButton("❌ لغو", callback_data="admin_panel")],
            ])
        )
        return True

    elif state == "adm_awaiting_broadcast_ids":
        ids_text = text.strip()
        try:
            target_ids = [int(x.strip()) for x in ids_text.split(",") if x.strip()]
        except ValueError:
            await update.message.reply_text("❌ لطفاً فقط آیدی‌های عددی جدا‌شده با کاما وارد کنید. مثال: 123456789, 987654321")
            return True
        if not target_ids:
            await update.message.reply_text("❌ هیچ آیدی معتبری پیدا نشد. دوباره امتحان کنید:")
            return True
        set_state(user_id, "adm_awaiting_broadcast_content", ",".join(str(i) for i in target_ids))
        await update.message.reply_text(
            f"✅ {len(target_ids)} آیدی ثبت شد.\n\n"
            f"حالا محتوا را بفرستید (متن، عکس، فایل، صدا، ویدیو -- حتی فوروارد از یک "
            f"کانال دیگر)، یا از کتابخانه انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📚 انتخاب فایل آماده", callback_data="adm_broadcast_pick_file"),
                 InlineKeyboardButton("💬 انتخاب متن آماده", callback_data="adm_broadcast_pick_text")],
            ])
        )
        return True

    elif state == "adm_awaiting_broadcast_single_id":
        try:
            target_id = int(text.strip())
        except ValueError:
            await update.message.reply_text("❌ لطفاً فقط آیدی عددی وارد کنید:")
            return True
        set_state(user_id, "adm_awaiting_broadcast_content", str(target_id))
        await update.message.reply_text(
            "✅ آیدی ثبت شد.\n\n"
            "حالا محتوا را بفرستید (متن، عکس، فایل، صدا، ویدیو -- حتی فوروارد از یک "
            "کانال دیگر)، یا از کتابخانه انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📚 انتخاب فایل آماده", callback_data="adm_broadcast_pick_file"),
                 InlineKeyboardButton("💬 انتخاب متن آماده", callback_data="adm_broadcast_pick_text")],
            ])
        )
        return True

    elif state == "adm_awaiting_pricing_text":  # ← حذف = پاک کن این بلوک
        set_bot_setting("pricing_text", text.strip())
        clear_state(user_id)
        await update.message.reply_text("✅ متن قیمت‌گذاری ذخیره شد.", reply_markup=admin_library_menu())
        return True

    elif state == "adm_awaiting_help_text":  # ← حذف راهنما = پاک کن این بلوک
        set_bot_setting("help_text", text.strip())
        clear_state(user_id)
        await update.message.reply_text("✅ متن راهنما ذخیره شد.", reply_markup=admin_library_menu())
        return True

    elif state == "adm_awaiting_new_admin_id":  # ← حذف چندادمینی = پاک کن این بلوک
        if not is_primary_admin(user_id):
            clear_state(user_id)
            return True
        try:
            new_admin_id = int(text.strip())
        except ValueError:
            await update.message.reply_text("❌ لطفاً فقط آیدی عددی وارد کنید:")
            return True
        if is_admin(new_admin_id):
            clear_state(user_id)
            await update.message.reply_text("این فرد از قبل ادمین است.", reply_markup=admin_manage_admins_menu())
            return True
        add_admin(new_admin_id, added_by=user_id)
        clear_state(user_id)
        try:
            await update.get_bot().send_message(
                new_admin_id,
                "🎉 شما به‌عنوان ادمین این ربات تعیین شدید؛ حالا از منوی اصلی «🛠 پنل مدیریت» را می‌بینید."
            )
        except Exception as e:
            print(f"خطای اطلاع‌رسانی به ادمین جدید: {e}")
        await update.message.reply_text(
            f"✅ کاربر {new_admin_id} به‌عنوان ادمین اضافه شد.",
            reply_markup=admin_manage_admins_menu()
        )
        return True

    elif state == "adm_awaiting_social_link_label":  # ← حذف لینک‌های تبلیغاتی = پاک کن این ۲ بلوک elif
        if not is_primary_admin(user_id):
            clear_state(user_id)
            return True
        label = text.strip()[:60]
        set_state(user_id, "adm_awaiting_social_link_url", label)
        await update.message.reply_text(f"حالا آدرس (لینک) «{label}» را بفرستید (باید با http:// یا https:// شروع شود):")
        return True

    elif state == "adm_awaiting_social_link_url":
        if not is_primary_admin(user_id):
            clear_state(user_id)
            return True
        url = text.strip()
        if not (url.startswith("http://") or url.startswith("https://")):
            await update.message.reply_text("❌ لینک باید با http:// یا https:// شروع شود. دوباره بفرستید:")
            return True
        label = data
        add_social_link(label, url)
        clear_state(user_id)
        await update.message.reply_text(f"✅ لینک «{label}» اضافه شد.", reply_markup=admin_panel_menu(user_id))
        return True

    elif state == "adm_awaiting_new_file_label":  # ← حذف کتابخانه = پاک کن این ۴ بلوک elif
        label = text.strip()[:60]
        set_state(user_id, "adm_awaiting_new_file_content", label)
        await update.message.reply_text(f"حالا متن، عکس، فایل، صدا یا ویدیوی «{label}» را بفرستید:")
        return True

    elif state == "adm_awaiting_edit_file_label":  # ← حذف ویرایش کتابخانه = پاک کن این بلوک
        file_db_id = int(data)
        new_label = text.strip()[:60]
        set_state(user_id, "adm_awaiting_edit_file_content", f"{file_db_id}:{new_label}")
        await update.message.reply_text(f"حالا محتوای جدید «{new_label}» را بفرستید (متن، عکس، فایل، صدا یا ویدیو):")
        return True

    elif state == "adm_awaiting_new_text_label":
        label = text.strip()[:60]
        set_state(user_id, "adm_awaiting_new_text_content", label)
        await update.message.reply_text(f"حالا متن «{label}» را بنویسید:")
        return True

    elif state == "adm_awaiting_new_text_content":
        label = data
        content = text.strip()
        add_admin_text(label, content)
        clear_state(user_id)
        await update.message.reply_text(f"✅ متن «{label}» به کتابخانه اضافه شد.", reply_markup=admin_library_menu())
        return True

    elif state == "adm_awaiting_edit_text_content":
        text_id = int(data)
        row = get_admin_text(text_id)
        if not row:
            clear_state(user_id)
            await update.message.reply_text("❌ این متن دیگر وجود ندارد.", reply_markup=admin_library_menu())
            return True
        label, _old_content = row
        update_admin_text(text_id, label, text.strip())
        clear_state(user_id)
        await update.message.reply_text(f"✅ متن «{label}» به‌روزرسانی شد.", reply_markup=admin_library_menu())
        return True

    if state == "adm_awaiting_user_id":
        try:
            target_uid = int(text.strip())
        except:
            await update.message.reply_text("❌ لطفاً فقط آیدی عددی وارد کنید:")
            return True
        target = get_user(target_uid)
        clear_state(user_id)
        if not target:
            await update.message.reply_text("❌ کاربری با این آیدی پیدا نشد.", reply_markup=admin_panel_menu(user_id))
            return True
        status_label = STATUS_LABELS.get(target[3], target[3])
        method_label = JOIN_METHOD_LABELS.get(target[10] if len(target) > 10 else None, "نامشخص")
        email = target[11] if len(target) > 11 and target[11] else "ثبت نشده"
        phone = target[12] if len(target) > 12 and target[12] else "ثبت نشده"
        birthdate = target[14] if len(target) > 14 and target[14] else "ثبت نشده"
        city = target[15] if len(target) > 15 and target[15] else "ثبت نشده"
        country = target[16] if len(target) > 16 and target[16] else "ثبت نشده"
        occupation = target[17] if len(target) > 17 and target[17] else "ثبت نشده"
        text_out = (
            f"👤 {target[2]} (@{target[1] or '---'})\n"
            f"🆔 {target[0]}\n"
            f"وضعیت: {status_label}\n"
            f"روش عضویت: {method_label}\n"
            f"تاریخ تأیید: {format_date_fa(target[4]) if target[4] else '---'}\n"
            f"انقضا: {format_date_fa(target[5]) if target[5] else '---'}\n"
            f"📧 ایمیل: {email}\n"
            f"📱 تلفن: {phone}\n"
            f"🎂 تاریخ تولد: {birthdate}\n"
            f"🏙 شهر: {city}\n"
            f"🌍 کشور: {country}\n"
            f"💼 شغل: {occupation}"
        )
        await update.message.reply_text(
            text_out,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🚫 مسدود کردن", callback_data=f"adm_block_{target_uid}")],
                [InlineKeyboardButton("✅ فعال کردن", callback_data=f"adm_activate_{target_uid}")],
                [InlineKeyboardButton("🔙 برگشت", callback_data="admin_panel")],
            ])
        )
        return True

    return False

# ═══════════════ پایان MODULE: ADMIN PANEL ═══════════════════════



# ╔══════════════════════════════════════════════════════════════╗
# ║        MODULE: EMAIL (ارسال ایمیل با Resend API)             ║
# ║  توجه فنی مهم: قبلاً از Gmail SMTP استفاده می‌شد، اما Railway   ║
# ║  پورت‌های SMTP خروجی (۲۵، ۴۶۵، ۵۸۷) را روی پلن‌های رایگان/    ║
# ║  Hobby به‌طور کامل مسدود می‌کند (این یک محدودیت شناخته‌شده و   ║
# ║  عمدی خود Railway است، نه باگ در کد). Resend یک سرویس رایگان  ║
# ║  (۳۰۰۰ ایمیل در ماه) است که به‌جای SMTP، از HTTPS (پورت ۴۴۳،   ║
# ║  که هرگز توسط Railway مسدود نمی‌شود) استفاده می‌کند.           ║
# ╚══════════════════════════════════════════════════════════════╝

RESEND_FROM_ADDRESS = "Zardari AI <onboarding@resend.dev>"

async def send_email(to_address, subject, body):
    """
    ارسال ایمیل با Resend HTTPS API (بدون نیاز به SMTP که در Railway مسدود
    است). اگر RESEND_API_KEY تنظیم نشده باشد، به‌آرامی شکست می‌خورد (بدون
    کرش کردن ربات) و False برمی‌گرداند.
    """
    if not RESEND_API_KEY:
        print("⚠️ RESEND_API_KEY تنظیم نشده؛ ایمیل ارسال نشد.")
        return False

    html_body = body.replace("\n", "<br>")
    for attempt in range(2):
        try:
            async with httpx.AsyncClient(timeout=20) as client:
                r = await client.post(
                    "https://api.resend.com/emails",
                    headers={
                        "Authorization": f"Bearer {RESEND_API_KEY}",
                        "Content-Type": "application/json",
                    },
                    json={
                        "from": RESEND_FROM_ADDRESS,
                        "to": [to_address],
                        "subject": subject,
                        "html": f"<div dir='rtl' style='font-family:sans-serif'>{html_body}</div>",
                    }
                )
                if r.status_code in (200, 201):
                    return True
                print(f"خطای Resend (تلاش {attempt + 1}): وضعیت {r.status_code} | پاسخ: {r.text[:300]}")
        except Exception as e:
            print(f"خطای ارسال ایمیل با Resend (تلاش {attempt + 1}): {e}")
        if attempt == 0:
            await asyncio.sleep(3)
    return False

# ═══════════════ پایان MODULE: EMAIL ═════════════════════════════


# ╔══════════════════════════════════════════════════════════════╗
# ║        MODULE: MARKET AI ANALYSIS (تحلیل هوش مصنوعی بازار)   ║
# ║  دو نوع تحلیل: نمادهای ثابت (ایران/کریپتو/فارکس) یا واچ‌لیست  ║
# ║  شخصی کاربر. برای هر نماد، کندل ۴ساعته و روزانه (ساخته‌شده از ║
# ║  ترکیب کندل‌های ۱ساعته -- تنها تایم‌فریمی که BiQuote واقعاً    ║
# ║  پشتیبانی می‌کند) گرفته می‌شود و هوش مصنوعی دو محدوده (Buy/    ║
# ║  Sell) با ریسک و توضیح کوتاه پیشنهاد می‌دهد. تحلیل کاملاً       ║
# ║  فوری است (هروقت کاربر بخواهد، همان لحظه اجرا می‌شود، بدون     ║
# ║  زمان‌بندی روزانه). اگر کاربر یک پیشنهاد را تأیید کند، یک       ║
# ║  آلارم قیمتی (از سیستم آلارم موجود) ساخته می‌شود که با رسیدن   ║
# ║  قیمت، هم در ربات هم (در صورت تنظیم ایمیل و تأیید کاربر) با    ║
# ║  ایمیل به او اطلاع می‌دهد.                                    ║
# ║  برای حذف: این بلوک رو پاک کن                                 ║
# ║  + خط menu_ai_market رو از financial_menu پاک کن               ║
# ║  + خط aim_ رو از ROUTER اصلی پاک کن                            ║
# ║  + state های aim_ رو از message_handler پاک کن                 ║
# ╚══════════════════════════════════════════════════════════════╝

MARKET_AI_FIXED_SYMBOLS = {
    "iran": [("USD", "iran"), ("GOLD18", "iran")],
    "crypto": [("BTC", "crypto"), ("ETH", "crypto")],
    "forex": [("EURUSD", "forex"), ("GBPUSD", "forex"), ("XAUUSD", "forex"), ("XAGUSD", "forex")],
}

def aggregate_candles(hourly_candles, group_size):
    """ترکیب کندل‌های ۱ساعته پشت‌سرهم برای ساخت کندل‌های بزرگ‌تر (۴ساعته یا روزانه)"""
    aggregated = []
    for i in range(0, len(hourly_candles), group_size):
        group = hourly_candles[i:i + group_size]
        if not group:
            continue
        aggregated.append({
            "time_utc": group[0]["time_utc"],
            "open": group[0].get("open", group[0]["close"]),
            "high": max(c["high"] for c in group),
            "low": min(c["low"] for c in group),
            "close": group[-1]["close"],
        })
    return aggregated

async def get_market_ai_candles(symbol, market):
    """
    دریافت کندل ۴ساعته و روزانه برای یک نماد فارکس/کریپتو -- با گرفتن
    کندل‌های ۱ساعته (تا حداکثر ~۸۳ روز که BiQuote پشتیبانی می‌کند) و
    ترکیب دستی آن‌ها. توجه: این تابع فقط برای فارکس/کریپتو است، نه بازار
    ایران (که BiQuote اصلاً داده‌ی ریالی ندارد -- برای آن از تابع جداگانه
    get_market_ai_iran_snapshot استفاده می‌شود).
    خروجی: (candles_4h, candles_daily) یا (None, None) اگر داده در دسترس نبود
    """
    biquote_symbol_map = {"GOLD": "XAUUSD", "SILVER": "XAGUSD", "BTC": "BTCUSD", "ETH": "ETHUSD"}
    query_symbol = biquote_symbol_map.get(symbol.upper(), symbol.upper())

    to_dt_utc = datetime.now()
    from_dt_utc = to_dt_utc - timedelta(days=30)
    hourly_candles, error = await get_historical_range_candles_raw(query_symbol, from_dt_utc, to_dt_utc, "1h")
    if error or not hourly_candles:
        return None, None

    candles_4h = aggregate_candles(hourly_candles, 4)
    candles_daily = aggregate_candles(hourly_candles, 24)
    return candles_4h, candles_daily

async def ai_analyze_iran_symbol(symbol, current_price):
    """
    تحلیل ساده برای نمادهای بازار ایران (دلار/طلای ۱۸ عیار) که BiQuote
    داده‌ی کندل تاریخی ریالی ندارد. فقط بر اساس قیمت لحظه‌ای فعلی، هوش
    مصنوعی یک محدوده‌ی منطقی Buy/Sell (بر پایه‌ی نوسان معمول این بازار،
    نه کندل واقعی) پیشنهاد می‌دهد -- این کمتر دقیق از تحلیل کندل‌محور
    است و باید همین‌طور صادقانه به کاربر گفته شود.
    """
    prompt = (
        f"قیمت فعلی {symbol} در بازار ایران (به ریال): {current_price:,.0f}\n\n"
        f"چون داده‌ی کندل تاریخی دقیق برای این بازار در دسترس نیست، بر اساس "
        f"دانش عمومی خودت از نوسانات معمول این بازار در ایران، یک محدوده‌ی "
        f"قیمتی منطقی برای خرید (Buy) کمی پایین‌تر از قیمت فعلی و یک محدوده‌ی "
        f"فروش (Sell) کمی بالاتر پیشنهاد بده. سطح ریسک و یک توضیح خیلی کوتاه "
        f"(حداکثر ۲ جمله) هم بده. در توضیح تصریح کن که این تحلیل فقط بر پایه‌ی "
        f"قیمت لحظه‌ای است، نه کندل تاریخی.\n\n"
        f"دقیقاً و فقط به این فرمت جواب بده:\n"
        f"Buy: <قیمت>\n"
        f"Buy ریسک: <کم یا متوسط یا زیاد>\n"
        f"Buy توضیح: <توضیح کوتاه>\n"
        f"Sell: <قیمت>\n"
        f"Sell ریسک: <کم یا متوسط یا زیاد>\n"
        f"Sell توضیح: <توضیح کوتاه>"
    )
    result = await get_ai_response([{"role": "user", "content": prompt}])
    if not result:
        return None

    def extract(field):
        m = re.search(rf'{field}\s*:?\s*(.+)', result)
        return m.group(1).strip() if m else None

    try:
        buy_price = float(re.sub(r'[^\d.]', '', extract("Buy") or ""))
        sell_price = float(re.sub(r'[^\d.]', '', extract("Sell") or ""))
        return {
            "symbol": symbol,
            "buy_price": buy_price, "buy_risk": extract("Buy ریسک") or "نامشخص", "buy_note": extract("Buy توضیح") or "",
            "sell_price": sell_price, "sell_risk": extract("Sell ریسک") or "نامشخص", "sell_note": extract("Sell توضیح") or "",
        }
    except Exception:
        return None

async def ai_analyze_symbol(symbol, candles_4h, candles_daily):
    """
    از هوش مصنوعی می‌خواهد بر اساس کندل‌های ۴ساعته و روزانه، یک محدوده‌ی
    Buy و یک محدوده‌ی Sell پیشنهاد دهد، با سطح ریسک و توضیح کوتاه.
    خروجی: dict یا None اگر AI پاسخ قابل‌فهمی نداد.
    """
    def summarize(candles, label, max_points=30):
        recent = candles[-max_points:]
        lines = [f"{c['time_utc'].strftime('%Y-%m-%d %H:%M')}: O={c['open']:.5f} H={c['high']:.5f} L={c['low']:.5f} C={c['close']:.5f}" for c in recent]
        return f"کندل‌های {label} اخیر:\n" + "\n".join(lines)

    prompt = (
        f"تحلیل تکنیکال برای نماد {symbol} بر اساس داده‌های واقعی زیر:\n\n"
        f"{summarize(candles_4h, 'چهار ساعته')}\n\n"
        f"{summarize(candles_daily, 'روزانه')}\n\n"
        f"بر اساس این داده‌ها، یک محدوده‌ی قیمتی مناسب برای ورود خرید (Buy) و "
        f"یک محدوده‌ی قیمتی مناسب برای ورود فروش (Sell) پیشنهاد بده. سطح ریسک "
        f"هرکدام (کم/متوسط/زیاد) و یک توضیح خیلی کوتاه (حداکثر ۲ جمله) هم بده.\n\n"
        f"دقیقاً و فقط به این فرمت جواب بده:\n"
        f"Buy: <قیمت>\n"
        f"Buy ریسک: <کم یا متوسط یا زیاد>\n"
        f"Buy توضیح: <توضیح کوتاه>\n"
        f"Sell: <قیمت>\n"
        f"Sell ریسک: <کم یا متوسط یا زیاد>\n"
        f"Sell توضیح: <توضیح کوتاه>"
    )
    result = await get_ai_response([{"role": "user", "content": prompt}])
    if not result:
        return None

    def extract(field):
        m = re.search(rf'{field}\s*:?\s*(.+)', result)
        return m.group(1).strip() if m else None

    try:
        buy_price = float(re.sub(r'[^\d.]', '', extract("Buy") or ""))
        sell_price = float(re.sub(r'[^\d.]', '', extract("Sell") or ""))
        return {
            "symbol": symbol,
            "buy_price": buy_price, "buy_risk": extract("Buy ریسک") or "نامشخص", "buy_note": extract("Buy توضیح") or "",
            "sell_price": sell_price, "sell_risk": extract("Sell ریسک") or "نامشخص", "sell_note": extract("Sell توضیح") or "",
        }
    except Exception as e:
        print(f"خطای پارس پاسخ AI برای {symbol}: {e} | پاسخ خام: {result[:200] if result else None}")
        return None

def format_analysis_message(analysis):
    market = analysis.get("market", "forex")
    unit = "ریال" if market == "iran" else "دلار" if market == "crypto" else ""
    buy_str = f"{analysis['buy_price']:,.0f} {unit}".strip() if market == "iran" else f"{analysis['buy_price']}"
    sell_str = f"{analysis['sell_price']:,.0f} {unit}".strip() if market == "iran" else f"{analysis['sell_price']}"
    return (
        f"📊 {analysis['symbol']}\n\n"
        f"🟢 محدوده خرید پیشنهادی (Buy): {buy_str}\n"
        f"   ریسک: {analysis['buy_risk']} | {analysis['buy_note']}\n\n"
        f"🔴 محدوده فروش پیشنهادی (Sell): {sell_str}\n"
        f"   ریسک: {analysis['sell_risk']} | {analysis['sell_note']}\n\n"
        f"ℹ️ این دو، پیشنهاد ورودی هستند، نه معامله‌ی انجام‌شده. با تأیید هرکدام، "
        f"فقط یک آلارم قیمتی ساخته می‌شود که وقتی قیمت واقعاً به آن سطح برسد، به شما اطلاع می‌دهد."
    )

async def run_market_analysis_for_user(bot, user_id, analysis_type):
    """
    اجرای تحلیل بازار برای یک کاربر و ارسال نتیجه.
    نکته مهم: پارامتر bot باید همیشه و فقط یک شیء telegram.Bot واقعی باشد
    (مثلاً context.bot) -- هیچ تشخیص یا حدس‌زنی نوع دیگری اینجا انجام
    نمی‌شود، چون امتحان قبلی (تشخیص خودکار با hasattr) به‌طور نامعتبری گاهی
    یک شیء اشتباه (نه Bot واقعی) انتخاب می‌کرد و باعث خطای
    'User.send_message() got an unexpected keyword argument chat_id' می‌شد.
    """

    if analysis_type == "watchlist":
        symbols = [(s, m) for _, s, m in get_watchlist(user_id)]
    else:
        symbols = MARKET_AI_FIXED_SYMBOLS["iran"] + MARKET_AI_FIXED_SYMBOLS["crypto"] + MARKET_AI_FIXED_SYMBOLS["forex"]

    if not symbols:
        try:
            await bot.send_message(chat_id=user_id, text="❌ برای این نوع تحلیل هیچ نمادی وجود ندارد (واچ‌لیست شما خالی است؟).")
        except Exception as e:
            print(f"خطای اطلاع خالی بودن نمادها: {e}")
        return

    db_user = get_user(user_id)
    email = db_user[11] if db_user and len(db_user) > 11 else None
    email_enabled = db_user[13] if db_user and len(db_user) > 13 else 0

    all_analyses = []
    for i, (symbol, market) in enumerate(symbols[:10]):
        if i > 0:
            await asyncio.sleep(2)  # فاصله بین درخواست‌ها برای کاهش احتمال Rate Limit همزمان Groq
        try:
            if market == "iran":
                # بازار ایران: BiQuote داده‌ی کندل ریالی ندارد؛ فقط از قیمت لحظه‌ای استفاده می‌شود
                current_price, _ = await get_price_by_symbol(symbol, "iran")
                if current_price is None:
                    continue
                analysis = await ai_analyze_iran_symbol(symbol, current_price)
            else:
                candles_4h, candles_daily = await get_market_ai_candles(symbol, market)
                if not candles_4h or not candles_daily:
                    continue
                analysis = await ai_analyze_symbol(symbol, candles_4h, candles_daily)
            if analysis:
                analysis["market"] = market
                all_analyses.append(analysis)
        except Exception as e:
            print(f"خطای تحلیل نماد {symbol}: {e}")
            continue

    if not all_analyses:
        try:
            await bot.send_message(
                chat_id=user_id,
                text="❌ در حال حاضر تحلیل هیچ نمادی ممکن نشد (احتمالاً دریافت داده‌ی قیمتی یا پاسخ هوش مصنوعی ناموفق بود). کمی بعد دوباره امتحان کنید."
            )
        except Exception as e:
            print(f"خطای اطلاع عدم موفقیت تحلیل: {e}")
        return

    import json
    for analysis in all_analyses:
        try:
            text = format_analysis_message(analysis)
            keyboard = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ تأیید محدوده Buy (آلارم بساز)", callback_data=f"aim_conf_buy_{analysis['symbol']}_{analysis['buy_price']}_{analysis['market']}")],
                [InlineKeyboardButton("✅ تأیید محدوده Sell (آلارم بساز)", callback_data=f"aim_conf_sell_{analysis['symbol']}_{analysis['sell_price']}_{analysis['market']}")],
            ])
        except Exception as e:
            print(f"خطای ساخت پیام تحلیل برای {analysis.get('symbol')}: {type(e).__name__}: {e!r}")
            continue

        try:
            await bot.send_message(chat_id=user_id, text=f"🧠 تحلیل روزانه\n\n{text}", reply_markup=keyboard)
        except Exception as e:
            print(f"خطای ارسال تحلیل به کاربر {user_id}: {type(e).__name__}: {e!r} | طول متن: {len(text)}")

    if email and email_enabled:
        body_lines = [format_analysis_message(a) for a in all_analyses]
        await send_email(email, "تحلیل بازار", "\n\n---\n\n".join(body_lines))

def market_ai_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📌 تحلیل نمادهای ثابت", callback_data="aim_type_fixed"),
         InlineKeyboardButton("👁 تحلیل واچ‌لیست شخصی", callback_data="aim_type_watchlist")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")],
    ])

async def handle_market_ai(query, user_id, context):
    data = query.data

    if data == "menu_ai_market":
        await query.edit_message_text(
            "🧠 تحلیل هوش مصنوعی بازار\n\nیک گزینه را انتخاب کنید تا همین الان تحلیل شود:",
            reply_markup=market_ai_menu()
        )

    elif data in ["aim_type_fixed", "aim_type_watchlist"]:
        analysis_type = "fixed" if data == "aim_type_fixed" else "watchlist"
        await query.edit_message_text("⏳ در حال تحلیل همین الان...\n\nاین معمولاً ۱۵ تا ۴۵ ثانیه طول می‌کشد (چند نماد پشت‌سرهم بررسی می‌شوند). لطفاً صبر کنید، نتیجه پیام‌های جداگانه می‌آید.")
        await run_market_analysis_for_user(context.bot, user_id, analysis_type)

    elif data.startswith("aim_conf_"):
        parts = data.split("_")
        direction_word, symbol, price_str, market = parts[2], parts[3], parts[4], parts[5]
        target = float(price_str)
        alarm_direction = "below" if direction_word == "buy" else "above"
        add_alarm(user_id, symbol, market, target, alarm_direction)
        direction_fa = "رسیدن به یا پایین‌تر از (Buy)" if direction_word == "buy" else "رسیدن به یا بالاتر از (Sell)"
        await query.answer(
            f"✅ آلارم ساخته شد: وقتی {symbol} به {direction_fa} {price_str} برسد، در ربات و ایمیل (در صورت تنظیم) به شما خبر داده می‌شود.",
            show_alert=True
        )

async def handle_market_ai_message(user_id, text, update):
    """این ماژول دیگر state متنی ندارد (تحلیل کاملاً فوری و بدون زمان‌بندی است)"""
    return False

# ═══════════════ پایان MODULE: MARKET AI ANALYSIS ════════════════


# ╔══════════════════════════════════════════════════════════════╗
# ║        MODULE: FOREX NEWS (اخبار مهم فارکس/کریپتو)           ║
# ║  از تقویم اقتصادی رایگان ForexFactory (JSON عمومی) اخبار      ║
# ║  پرتأثیر هفته را می‌گیرد و با وقت ایران نمایش می‌دهد.          ║
# ╚══════════════════════════════════════════════════════════════╝

async def fetch_forex_factory_news():
    """
    دریافت اخبار اقتصادی هفته از فید JSON عمومی ForexFactory (رایگان، بدون کلید).
    خروجی: لیست دیکشنری با title/country/impact/date (UTC) یا [] اگر خطا داد.
    """
    try:
        async with httpx.AsyncClient(timeout=15, headers={"User-Agent": "Mozilla/5.0"}) as client:
            r = await client.get("https://nfs.faireconomy.media/ff_calendar_thisweek.json")
            if r.status_code != 200:
                return []
            data = r.json()
            events = []
            for item in data:
                if item.get("impact") == "High":
                    events.append({
                        "title": item.get("title", ""),
                        "country": item.get("country", ""),
                        "date_utc": item.get("date", ""),
                    })
            return events
    except Exception as e:
        print(f"خطای دریافت اخبار فارکس: {e}")
        return []

def forex_news_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔄 به‌روزرسانی", callback_data="fin_news")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")],
    ])

async def handle_forex_news(query, user_id):
    await query.edit_message_text("⏳ در حال دریافت اخبار مهم هفته...")
    events = await fetch_forex_factory_news()
    if not events:
        await query.message.reply_text(
            "❌ در حال حاضر دریافت اخبار ممکن نیست.",
            reply_markup=forex_news_menu()
        )
        return

    lines = []
    for e in events[:20]:
        try:
            event_dt_utc = datetime.fromisoformat(e["date_utc"].replace("Z", "+00:00"))
            event_dt_iran = event_dt_utc + IRAN_TZ_OFFSET
            time_str = event_dt_iran.strftime("%A %Y-%m-%d %H:%M")
        except Exception:
            time_str = e["date_utc"]
        lines.append(f"🔴 {e['title']} ({e['country']})\n   🕐 {time_str} (وقت ایران)")

    text = "📰 اخبار مهم این هفته (فارکس/کریپتو):\n\n" + "\n\n".join(lines)
    await query.message.reply_text(text[:4000], reply_markup=forex_news_menu())

# ═══════════════ پایان MODULE: FOREX NEWS ════════════════════════









# ╔══════════════════════════════════════════════════════════════╗
# ║                  MODULE: POSITION SIZE                       ║
# ║  برای حذف: این بلوک رو پاک کن                               ║
# ║  + خط ps_size رو از financial_menu پاک کن                   ║
# ║  + خط ps_ رو از ROUTER اصلی پاک کن                          ║
# ║  + state های ps_ رو از message_handler پاک کن               ║
# ╚══════════════════════════════════════════════════════════════╝

# ─── مشخصات قراردادی هر بازار برای محاسبه ارزش هر واحد حرکت قیمت ──
# pip_size: کوچک‌ترین واحد حرکت قیمت استاندارد آن نماد
# contract_size: اندازه ۱ لات استاندارد (تعداد واحد ارز پایه/کالا)
FX_CONTRACT_SPECS = {
    "EURUSD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": True},
    "GBPUSD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": True},
    "AUDUSD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": True},
    "NZDUSD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": True},
    "USDCAD": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": False},
    "USDJPY": {"pip_size": 0.01,   "contract_size": 100000, "quote_is_usd": False},
    "USDCHF": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": False},
    "EURGBP": {"pip_size": 0.0001, "contract_size": 100000, "quote_is_usd": False},
    "EURJPY": {"pip_size": 0.01,   "contract_size": 100000, "quote_is_usd": False},
    "GBPJPY": {"pip_size": 0.01,   "contract_size": 100000, "quote_is_usd": False},
}
METAL_CONTRACT_SPECS = {
    "GOLD": {"contract_size": 100},     # 1 لات طلا = 100 اونس
    "XAUUSD": {"contract_size": 100},
    "SILVER": {"contract_size": 5000},  # 1 لات نقره = 5000 اونس
    "XAGUSD": {"contract_size": 5000},
}

def ps_classify_symbol(symbol):
    """تشخیص نوع نماد برای انتخاب فرمول محاسبه: forex_pair / metal / crypto / iran / forex_other"""
    s = symbol.upper()
    if s in FX_CONTRACT_SPECS:
        return "forex_pair"
    if s in METAL_CONTRACT_SPECS:
        return "metal"
    if s in CRYPTO_SYMBOLS:
        return "crypto"
    if s in IRAN_SYMBOLS:
        return "iran"
    return "forex_other"

async def ps_calculate_position(symbol, direction, entry_price, sl_price, risk_usd):
    """
    محاسبه حجم پیشنهادی بر اساس مقدار ریسک دلاری.
    خروجی: dict شامل risk_distance, unit_value, lots/qty, explanation
    """
    symbol = symbol.upper()
    category = ps_classify_symbol(symbol)
    risk_distance = abs(entry_price - sl_price)

    if risk_distance == 0:
        return None

    if category == "forex_pair":
        spec = FX_CONTRACT_SPECS[symbol]
        contract_size = spec["contract_size"]
        if spec["quote_is_usd"]:
            # مثل EURUSD: ارزش حرکت قیمت برای 1 لات کامل = فاصله قیمت × اندازه قرارداد
            value_per_lot = risk_distance * contract_size
        else:
            # مثل USDJPY/USDCAD: ارز Quote دلار نیست، باید تقسیم بر نرخ لحظه‌ای جفت‌ارز شود
            # تا حرکت قیمت (که به واحد Quote است) به دلار تبدیل شود
            live_price, _ = await get_price_by_symbol(symbol, "forex")
            reference_price = live_price if live_price else entry_price
            value_per_lot = (risk_distance * contract_size) / reference_price
        lots = risk_usd / value_per_lot
        return {
            "category": "forex_pair", "risk_distance": risk_distance,
            "lots": lots, "value_per_lot": value_per_lot, "unit": "لات استاندارد"
        }

    elif category == "metal":
        spec = METAL_CONTRACT_SPECS[symbol]
        contract_size = spec["contract_size"]
        value_per_lot = risk_distance * contract_size
        lots = risk_usd / value_per_lot
        return {
            "category": "metal", "risk_distance": risk_distance,
            "lots": lots, "value_per_lot": value_per_lot, "unit": "لات استاندارد"
        }

    elif category == "crypto":
        qty = risk_usd / risk_distance
        return {
            "category": "crypto", "risk_distance": risk_distance,
            "lots": qty, "value_per_lot": risk_distance, "unit": f"واحد {symbol}"
        }

    elif category == "iran":
        qty = risk_usd / risk_distance
        return {
            "category": "iran", "risk_distance": risk_distance,
            "lots": qty, "value_per_lot": risk_distance, "unit": "واحد (ریال‌محور)"
        }

    else:  # forex_other - جفت ارز ناشناخته با pip استاندارد فرض می‌شود
        contract_size = 100000
        live_price, _ = await get_price_by_symbol(symbol, "forex")
        reference_price = live_price if live_price else entry_price
        value_per_lot = (risk_distance * contract_size) / reference_price if reference_price else 0
        lots = risk_usd / value_per_lot if value_per_lot else 0
        return {
            "category": "forex_other", "risk_distance": risk_distance,
            "lots": lots, "value_per_lot": value_per_lot, "unit": "لات استاندارد (تقریبی)"
        }

# ─── منوها ────────────────────────────────────────────────────
def position_size_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📐 محاسبه جدید", callback_data="ps_new")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")],
    ])

def ps_direction_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📈 Long (خرید)", callback_data="ps_dir_long"),
         InlineKeyboardButton("📉 Short (فروش)", callback_data="ps_dir_short")],
    ])

def ps_risk_type_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💵 مقدار دلاری ثابت", callback_data="ps_risk_fixed"),
         InlineKeyboardButton("📊 درصدی از بالانس", callback_data="ps_risk_percent")],
    ])

# ─── هندلر دکمه‌ها ────────────────────────────────────────────
async def handle_position_size(query, user_id):
    data = query.data

    if data == "ps_size":
        await query.edit_message_text(
            "📐 محاسبه میزان خرید (Position Sizing)\n\n"
            "با گرفتن نقطه ورود، حد ضرر و میزان ریسکی که می‌خواهید "
            "بپذیرید (دلاری یا درصدی از بالانس)، حجم مناسب معامله "
            "(لات استاندارد یا مقدار واحد) را محاسبه می‌کند.",
            reply_markup=position_size_menu()
        )

    elif data == "ps_new":
        set_state(user_id, "ps_symbol")
        await query.edit_message_text(
            "📐 محاسبه جدید\n\nنماد ارز را وارد کنید:\nمثال: `EURUSD` `GOLD` `BTC` `USD`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="ps_size")]])
        )

    elif data in ["ps_dir_long", "ps_dir_short"]:
        direction = "long" if data == "ps_dir_long" else "short"
        state, sdata = get_state(user_id)
        if state == "ps_direction":
            set_state(user_id, "ps_entry_price", f"{sdata}|{direction}")
            await query.edit_message_text(
                f"✅ جهت: {'Long 📈' if direction=='long' else 'Short 📉'}\n\nقیمت ورود را وارد کنید:"
            )

    elif data in ["ps_risk_fixed", "ps_risk_percent"]:
        state, sdata = get_state(user_id)
        if state == "ps_risk_type":
            if data == "ps_risk_fixed":
                set_state(user_id, "ps_risk_amount", f"{sdata}|fixed")
                await query.edit_message_text("💵 مقدار ریسک به دلار را وارد کنید:\nمثال: `10`", parse_mode="Markdown")
            else:
                set_state(user_id, "ps_balance", f"{sdata}|percent")
                await query.edit_message_text("💰 بالانس حساب را به دلار وارد کنید:\nمثال: `1000`", parse_mode="Markdown")

async def handle_position_size_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    if state == "ps_symbol":
        symbol = text.upper()
        set_state(user_id, "ps_direction", symbol)
        await update.message.reply_text("جهت معامله را انتخاب کنید:", reply_markup=ps_direction_menu())
        return True

    elif state == "ps_entry_price":
        try:
            entry_price = float(text.replace(",", ""))
            set_state(user_id, "ps_sl_price", f"{data}|{entry_price}")
            await update.message.reply_text("قیمت حد ضرر را وارد کنید:")
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    elif state == "ps_sl_price":
        try:
            sl_price = float(text.replace(",", ""))
            parts = data.split("|")
            symbol, direction, entry_price = parts[0], parts[1], float(parts[2])

            if (direction == "long" and sl_price >= entry_price) or (direction == "short" and sl_price <= entry_price):
                await update.message.reply_text("❌ حد ضرر با جهت معامله همخوانی ندارد.\nدوباره وارد کنید:")
                return True

            set_state(user_id, "ps_risk_type", f"{symbol}|{direction}|{entry_price}|{sl_price}")
            await update.message.reply_text("روش تعیین ریسک را انتخاب کنید:", reply_markup=ps_risk_type_menu())
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    elif state == "ps_risk_amount":
        try:
            risk_usd = float(text.replace(",", ""))
            if risk_usd <= 0:
                await update.message.reply_text("❌ مقدار ریسک باید بزرگ‌تر از صفر باشد:")
                return True
            parts = data.split("|")
            symbol, direction, entry_price, sl_price, risk_mode = parts
            await run_position_size_final(user_id, update, symbol, direction,
                                           float(entry_price), float(sl_price), risk_usd)
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    elif state == "ps_balance":
        try:
            balance = float(text.replace(",", ""))
            if balance <= 0:
                await update.message.reply_text("❌ بالانس باید بزرگ‌تر از صفر باشد:")
                return True
            set_state(user_id, "ps_percent", f"{data}|{balance}")
            await update.message.reply_text("درصد ریسک مجاز را وارد کنید:\nمثال: `2`")
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    elif state == "ps_percent":
        try:
            percent = float(text.replace(",", ""))
            if percent <= 0 or percent > 100:
                await update.message.reply_text("❌ درصد باید بین 0 و 100 باشد:")
                return True
            parts = data.split("|")
            symbol, direction, entry_price, sl_price, risk_mode, balance = parts
            risk_usd = float(balance) * (percent / 100)
            await run_position_size_final(user_id, update, symbol, direction,
                                           float(entry_price), float(sl_price), risk_usd,
                                           balance=float(balance), percent=percent)
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    return False

async def run_position_size_final(user_id, update, symbol, direction, entry_price, sl_price,
                                    risk_usd, balance=None, percent=None):
    await update.message.reply_text("⏳ در حال محاسبه...")

    result = await ps_calculate_position(symbol, direction, entry_price, sl_price, risk_usd)
    if not result:
        await update.message.reply_text("❌ خطا در محاسبه (فاصله ریسک صفر است).", reply_markup=position_size_menu())
        clear_state(user_id)
        return

    risk_line = f"💵 مقدار ریسک: ${risk_usd:,.2f}"
    if balance and percent:
        risk_line += f"\n💰 بالانس: ${balance:,.2f} | درصد ریسک: {percent}%"

    text = (
        f"📐 محاسبه میزان خرید — {symbol} ({'Long' if direction=='long' else 'Short'})\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
        f"ورود: {entry_price}\n"
        f"حد ضرر: {sl_price}\n"
        f"فاصله ریسک: {result['risk_distance']:.5f}\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
        f"{risk_line}\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
        f"📦 حجم پیشنهادی: {result['lots']:.4f} {result['unit']}\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`"
    )

    if result["category"] == "forex_other":
        text += "\n\n⚠️ این نماد در لیست دقیق ربات نبود؛ محاسبه با فرض pip استاندارد فارکس انجام شده و ممکن است تقریبی باشد."

    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=position_size_menu())
    clear_state(user_id)

# ═══════════════ پایان MODULE: POSITION SIZE ════════════════════


# ╔══════════════════════════════════════════════════════════════╗
# ║                  MODULE: BACKTEST                            ║
# ║  برای حذف: این بلوک رو پاک کن                               ║
# ║  + خط backtest رو از financial_menu پاک کن                  ║
# ║  + خط fin_backtest/bt_ رو از ROUTER اصلی پاک کن             ║
# ║  + state های bt_ رو از message_handler پاک کن               ║
# ╚══════════════════════════════════════════════════════════════╝

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

IRAN_TZ_OFFSET = timedelta(hours=3, minutes=30)
DEFAULT_R_LEVELS = [0.5, 1, 1.5, 2, 3, 4, 8, 10]

def get_strategy_levels(rules):
    """سطوح R این استراتژی را برمی‌گرداند (ذخیره‌شده در کلید __levels__، یا پیش‌فرض قدیمی)"""
    levels = rules.get("__levels__")
    if levels:
        return levels
    return DEFAULT_R_LEVELS

def describe_strategy(rules):
    """تولید خلاصه متنی خوانا از قوانین استراتژی برای نمایش به کاربر"""
    levels = get_strategy_levels(rules)
    lines = []
    for r in levels[:-1]:  # آخرین سطح همیشه خروج کامل است، جزو سوالات نیست
        key = str(r)
        rule = rules.get(key)
        if isinstance(rule, dict) and rule.get("type") == "r_trail":
            value = rule["value"]
            if value == 0:
                lines.append(f"🔸 R{r} → ریسک‌فری (SL = نقطه ورود)")
            else:
                lines.append(f"🔸 R{r} → SL به R{value} منتقل می‌شود")
        else:
            lines.append(f"🔸 R{r} → بدون تغییر (رد می‌شود)")
    lines.append(f"🔸 R{levels[-1]} → خروج کامل (همیشه ثابت)")
    return "\n".join(lines)

def to_iran_time(dt_naive_utc_like):
    """تبدیل datetime (که از Yahoo می‌آید، معمولاً UTC) به ساعت ایران برای نمایش"""
    return dt_naive_utc_like + IRAN_TZ_OFFSET

# ─── دیتابیس استراتژی ───────────────────────────────────────
def init_backtest_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS bt_strategies (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id    INTEGER,
        name       TEXT,
        rules_json TEXT,
        created_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS bt_results (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id     INTEGER,
        batch_id    TEXT,
        symbol      TEXT,
        direction   TEXT,
        entry_time  TEXT,
        entry_price REAL,
        sl_price    REAL,
        strategy    TEXT,
        max_r       REAL,
        result_r    REAL,
        exit_time   TEXT,
        exit_price  REAL,
        status      TEXT,
        created_at  TEXT
    )''')
    conn.commit()
    conn.close()

def save_strategy(user_id, name, rules):
    import json
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO bt_strategies (user_id, name, rules_json, created_at) VALUES (?, ?, ?, ?)",
              (user_id, name, json.dumps(rules), get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_strategies(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, name, rules_json FROM bt_strategies WHERE user_id=? ORDER BY id DESC", (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_strategy_by_id(sid):
    import json
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT name, rules_json FROM bt_strategies WHERE id=?", (sid,))
    row = c.fetchone()
    conn.close()
    if row:
        return row[0], json.loads(row[1])
    return None, None

def save_bt_result(user_id, batch_id, symbol, direction, entry_time, entry_price,
                    sl_price, strategy_name, max_r, result_r, exit_time, exit_price, status):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''INSERT INTO bt_results
        (user_id, batch_id, symbol, direction, entry_time, entry_price, sl_price,
         strategy, max_r, result_r, exit_time, exit_price, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (user_id, batch_id, symbol, direction, entry_time, entry_price, sl_price,
         strategy_name, max_r, result_r, exit_time, exit_price, status,
         get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_batch_results(user_id, batch_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT * FROM bt_results WHERE user_id=? AND batch_id=? ORDER BY id", (user_id, batch_id))
    rows = c.fetchall()
    conn.close()
    return rows

# ─── دریافت کندل‌های تاریخی (بدون تایم‌فریم روزانه) ─────────
BIQUOTE_INTERVAL_MAP = {"1m": "1m", "15m": "15m", "1h": "1h"}

async def get_historical_candles(symbol, entry_date, entry_hm, interval):
    """
    entry_date: YYYY-MM-DD | entry_hm: HH:MM (به وقت ایران) | interval: 1m/15m/1h
    منبع داده: BiQuote (https://biquote.io) -- داده‌ی مستقیم از MetaTrader 5
    رایگان، بدون نیاز به کلید API یا ثبت‌نام.
    """
    try:
        entry_dt_iran = datetime.strptime(f"{entry_date} {entry_hm}", "%Y-%m-%d %H:%M")
        entry_dt_utc = entry_dt_iran - IRAN_TZ_OFFSET
        days_ago = (datetime.now() - entry_dt_utc).days

        # BiQuote حداکثر ۲۰۰۰ کندل در هر درخواست می‌دهد؛ بر این اساس محدودیت روز تخمین زده می‌شود
        bars_per_day = {"1m": 1440, "15m": 96, "1h": 24}
        max_days_by_interval = {k: max(1, 2000 // v) for k, v in bars_per_day.items()}
        max_days = max_days_by_interval.get(interval, 30)
        if days_ago > max_days:
            return None, None, f"⚠️ تایم‌فریم انتخابی فقط تا حدود {max_days} روز گذشته را پشتیبانی می‌کند.\nتاریخ شما {days_ago} روز پیش است."

        biquote_interval = BIQUOTE_INTERVAL_MAP.get(interval, "15m")
        biquote_symbol = symbol.upper()

        # نگاشت نمادهای داخلی ربات به نمادهای BiQuote (طلا/نقره با همان نام ساده)
        symbol_map = {"GOLD": "XAUUSD", "SILVER": "XAGUSD"}
        biquote_symbol = symbol_map.get(biquote_symbol, biquote_symbol)

        # محدوده‌ی زمانی درخواست: از کمی قبل از لحظه‌ی ورود تا الان
        from_dt_utc = entry_dt_utc - timedelta(hours=1)
        from_iso = from_dt_utc.strftime("%Y-%m-%dT%H:%M:%SZ")

        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.get(
                f"https://biquote.io/api/{biquote_symbol}/ohlc",
                params={"interval": biquote_interval, "limit": 2000, "from": from_iso}
            )
            if r.status_code == 404:
                return None, None, f"❌ نماد «{symbol}» در منبع داده پیدا نشد."
            data = r.json()
            bars = data.get("bars", [])

            candles = []
            entry_ts = entry_dt_utc.timestamp()
            for bar in bars:
                bar_time = datetime.strptime(bar["openTime"], "%Y-%m-%dT%H:%M:%SZ")
                if bar_time.timestamp() >= entry_ts and not bar.get("isOpen", False):
                    high = bar.get("high")
                    low = bar.get("low")
                    close = bar.get("close")
                    if high is not None and low is not None:
                        candles.append({
                            "time_utc": bar_time,
                            "high": high, "low": low, "close": close
                        })
            # ترتیب صعودی زمان تضمین شود (BiQuote معمولاً جدیدترین را اول می‌دهد)
            candles.sort(key=lambda c: c["time_utc"])

            if not candles:
                return None, None, "❌ داده‌ای برای این بازه زمانی یافت نشد."
            return candles, entry_dt_utc, None
    except Exception as e:
        print(f"Historical data error (BiQuote): {e}")
        return None, None, "❌ خطا در دریافت داده‌های تاریخی."

# ─── موتور شبیه‌سازی با قوانین سفارشی کاربر ─────────────────
def run_custom_simulation(candles, entry_price, sl_price, direction, strategy_rules):
    """
    strategy_rules: dict مثل {"0.5": None, "1": {"type": "r_trail", "value": 0.2}, ...}
    و کلید "__levels__" که لیست سطوح R این استراتژی خاص را نگه می‌دارد.
    مقدار None یعنی SL تغییر نمی‌کند در آن سطح
    مقدار {"type": "r_trail", "value": X} یعنی SL روی سطح RX قرار می‌گیرد
        (مثلاً value=0 یعنی نقطه ورود، value=0.2 یعنی R0.2 از نقطه ورود)
    در آخرین سطح این استراتژی همیشه خروج کامل است
    """
    levels = get_strategy_levels(strategy_rules)
    final_level = levels[-1]

    R = abs(entry_price - sl_price)
    is_long = direction == "long"

    def level(r_mult):
        return entry_price + (r_mult * R) if is_long else entry_price - (r_mult * R)

    current_sl = sl_price
    max_r_reached = 0
    triggered_levels = set()

    for candle in candles:
        high, low = candle["high"], candle["low"]

        hit_sl = (low <= current_sl) if is_long else (high >= current_sl)
        if hit_sl:
            exit_r = (current_sl - entry_price) / R if is_long else (entry_price - current_sl) / R
            return {
                "exit_price": current_sl, "exit_time_utc": candle["time_utc"],
                "result_r": round(exit_r, 2), "max_r": max_r_reached, "status": "closed"
            }

        reached = high if is_long else low
        for r_mult in levels:
            if r_mult in triggered_levels:
                continue
            target_level = level(r_mult)
            condition = (reached >= target_level) if is_long else (reached <= target_level)
            if condition:
                triggered_levels.add(r_mult)
                max_r_reached = r_mult

                if r_mult == final_level:
                    return {
                        "exit_price": level(final_level), "exit_time_utc": candle["time_utc"],
                        "result_r": float(final_level), "max_r": final_level, "status": "closed"
                    }

                rule = strategy_rules.get(str(r_mult))
                if isinstance(rule, dict) and rule.get("type") == "r_trail":
                    current_sl = level(rule["value"])
                # rule None -> بدون تغییر

    last_close = candles[-1]["close"] if candles else entry_price
    open_r = (last_close - entry_price) / R if is_long else (entry_price - last_close) / R
    return {
        "exit_price": None, "exit_time_utc": None,
        "result_r": round(open_r, 2), "max_r": max_r_reached, "status": "open"
    }

# ─── ساخت فایل اکسل ──────────────────────────────────────────
def build_excel_report(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest Results"

    headers = ["نماد", "جهت", "تاریخ ورود (ایران)", "قیمت ورود", "SL اولیه",
               "استراتژی", "بالاترین R", "نتیجه (R)", "تاریخ خروج (ایران)",
               "قیمت خروج", "وضعیت"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        (_, _, _, symbol, direction, entry_time, entry_price, sl_price,
         strategy, max_r, result_r, exit_time, exit_price, status, _) = row
        ws.append([
            symbol, "Long" if direction == "long" else "Short",
            entry_time, entry_price, sl_price, strategy, max_r, result_r,
            exit_time or "---", exit_price or "---",
            "بسته شده" if status == "closed" else "باز"
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value is not None)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ─── منوها ────────────────────────────────────────────────────
def backtest_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ بک‌تست جدید", callback_data="bt_new"),
         InlineKeyboardButton("📁 آپلود فایل لیست معاملات", callback_data="bt_upload")],  # ← حذف = پاک کن این خط
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")],
    ])

def strategy_choice_menu(user_id):
    strategies = get_strategies(user_id)
    keyboard = []
    for sid, name, _ in strategies[:10]:
        keyboard.append([InlineKeyboardButton(f"📋 {name}", callback_data=f"bt_strat_use_{sid}")])
    keyboard.append([InlineKeyboardButton("🆕 ساخت استراتژی جدید", callback_data="bt_strat_create")])
    keyboard.append([InlineKeyboardButton("🔙 برگشت", callback_data="fin_backtest")])
    return InlineKeyboardMarkup(keyboard)

def r_level_question_menu(r_mult):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ بله، جابجا کن", callback_data=f"bt_r_yes_{r_mult}"),
         InlineKeyboardButton("❌ نه، رد شو", callback_data=f"bt_r_no_{r_mult}")],
    ])

def direction_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📈 Long (خرید)", callback_data="bt_dir_long"),
         InlineKeyboardButton("📉 Short (فروش)", callback_data="bt_dir_short")],
    ])

def timeframe_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("1 دقیقه (تا 7 روز پیش)", callback_data="bt_tf_1m")],
        [InlineKeyboardButton("15 دقیقه (تا 60 روز پیش)", callback_data="bt_tf_15m")],
        [InlineKeyboardButton("1 ساعت (تا 2 سال پیش)", callback_data="bt_tf_1h")],
    ])

# ─── هندلر دکمه‌ها ────────────────────────────────────────────
async def handle_backtest(query, user_id):
    data = query.data

    if data == "fin_backtest":
        await query.edit_message_text(
            "📊 بک‌تست مدیریت پوزیشن\n\n"
            "این ابزار با گرفتن نقطه ورود، حد ضرر و استراتژی شخصی شما، "
            "مسیر واقعی قیمت را شبیه‌سازی می‌کند.\n\n"
            "هنگام ساخت استراتژی، می‌توانید سطوح ریوارد (R) دلخواه خودتان "
            "را تعیین کنید یا از سطوح پیش‌فرض استفاده کنید:\n"
            f"{', '.join('R'+str(r) for r in DEFAULT_R_LEVELS)}\n"
            "(آخرین سطح همیشه خروج کامل است)",
            reply_markup=backtest_menu()
        )

    elif data == "bt_new":
        set_state(user_id, "bt_mode_choice")
        await query.edit_message_text(
            "📊 بک‌تست جدید\n\nنوع بک‌تست را انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔹 تکی (یک ارز)", callback_data="bt_mode_single")],
                [InlineKeyboardButton("🔸 گروهی (چند ارز)", callback_data="bt_mode_group")],
            ])
        )

    elif data in ["bt_mode_single", "bt_mode_group"]:
        mode = "single" if data == "bt_mode_single" else "group"
        set_state(user_id, "bt_strategy_choice", mode)
        strategies = get_strategies(user_id)
        if strategies:
            await query.edit_message_text(
                "📋 می‌خواهید از استراتژی قبلی استفاده کنید یا جدید بسازید؟",
                reply_markup=strategy_choice_menu(user_id)
            )
        else:
            set_state(user_id, "bt_strat_levels_choice", mode)
            await query.edit_message_text(
                "🆕 ساخت استراتژی جدید\n\n"
                "می‌خواهید چند سطح ریوارد (R) خودتان تعیین کنید یا از سطوح "
                f"پیش‌فرض ({', '.join('R'+str(r) for r in DEFAULT_R_LEVELS)}) استفاده کنید؟",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🎯 سطوح دلخواه خودم", callback_data="bt_levels_custom")],
                    [InlineKeyboardButton("⚡️ همان ۸ سطح پیش‌فرض", callback_data="bt_levels_default")],
                ])
            )

    elif data == "bt_strat_create":
        state, sdata = get_state(user_id)
        mode = sdata.split("|")[0] if sdata else "single"
        set_state(user_id, "bt_strat_levels_choice", mode)
        await query.edit_message_text(
            "🆕 ساخت استراتژی جدید\n\n"
            "می‌خواهید چند سطح ریوارد (R) خودتان تعیین کنید یا از سطوح "
            f"پیش‌فرض ({', '.join('R'+str(r) for r in DEFAULT_R_LEVELS)}) استفاده کنید؟",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🎯 سطوح دلخواه خودم", callback_data="bt_levels_custom")],
                [InlineKeyboardButton("⚡️ همان ۸ سطح پیش‌فرض", callback_data="bt_levels_default")],
            ])
        )

    elif data in ["bt_levels_custom", "bt_levels_default"]:
        state, sdata = get_state(user_id)
        mode = sdata if sdata else "single"
        if data == "bt_levels_default":
            import json
            rules = {"__levels__": DEFAULT_R_LEVELS}
            first_level = DEFAULT_R_LEVELS[0]
            set_state(user_id, f"bt_strat_r_{first_level}", f"{mode}|{json.dumps(rules)}")
            await query.edit_message_text(
                f"⚡️ سطوح پیش‌فرض انتخاب شد: {', '.join('R'+str(r) for r in DEFAULT_R_LEVELS)}\n\n"
                f"در سطح R{first_level}، حد ضرر را جابجا کنم؟",
                reply_markup=r_level_question_menu(str(first_level))
            )
        else:
            set_state(user_id, "bt_levels_count", mode)
            await query.edit_message_text(
                "🎯 چند سطح ریوارد نیاز داری؟\n\nیک عدد وارد کنید (مثلاً 5):"
            )

    elif data.startswith("bt_strat_use_"):
        sid = int(data.replace("bt_strat_use_", ""))
        name, rules = get_strategy_by_id(sid)
        state, sdata = get_state(user_id)
        mode = sdata.split("|")[0] if sdata else "single"
        if name:
            import json
            set_state(user_id, "bt_symbol", f"{mode}|{name}|{json.dumps(rules)}")
            await query.edit_message_text(
                f"✅ استراتژی «{name}» انتخاب شد.\n\n"
                f"📋 خلاصه قوانین این استراتژی:\n\n{describe_strategy(rules)}\n\n"
                f"نماد ارز را وارد کنید:\nمثال: `EURUSD` `GOLD` `USDCAD`",
                parse_mode="Markdown"
            )

    elif data.startswith("bt_r_yes_") or data.startswith("bt_r_no_"):
        is_yes = data.startswith("bt_r_yes_")
        r_mult = data.replace("bt_r_yes_", "").replace("bt_r_no_", "")
        state, sdata = get_state(user_id)

        if is_yes:
            set_state(user_id, f"bt_strat_slval_{r_mult}", sdata)
            await query.edit_message_text(
                f"🎯 در سطح R{r_mult}، حد ضرر را روی چه ریواردی بگذارم؟\n\n"
                f"یک عدد بین `0` و `{r_mult}` وارد کنید.\n"
                f"مثال: اگر بنویسید `0.2` یعنی حد ضرر روی R0.2 قرار می‌گیرد.\n\n"
                f"برای ریسک‌فری دقیق (نقطه ورود) بنویسید `0`",
                parse_mode="Markdown"
            )
        else:
            mode, rules_str = (sdata.split("|", 1) + [""])[:2]
            import json
            rules = json.loads(rules_str) if rules_str else {}
            rules[r_mult] = None
            levels = get_strategy_levels(rules)
            next_idx = levels.index(float(r_mult)) + 1
            await proceed_to_next_r_level(query, user_id, mode, rules, next_idx)

# عبور به سطح R بعدی یا پایان ساخت استراتژی
async def proceed_to_next_r_level(query_or_update, user_id, mode, rules, next_idx, is_message=False):
    import json
    levels = get_strategy_levels(rules)
    if next_idx < len(levels) - 1:  # آخرین سطح سوال ندارد (همیشه خروج کامل)
        next_r = levels[next_idx]
        set_state(user_id, f"bt_strat_r_{next_r}", f"{mode}|{json.dumps(rules)}")
        text = f"در سطح R{next_r}، حد ضرر را جابجا کنم؟"
        markup = r_level_question_menu(str(next_r))
        if is_message:
            await query_or_update.message.reply_text(text, reply_markup=markup)
        else:
            await query_or_update.edit_message_text(text, reply_markup=markup)
    else:
        set_state(user_id, "bt_strat_name", f"{mode}|{json.dumps(rules)}")
        text = "✅ تمام سطوح تنظیم شد!\n\nاسمی برای این استراتژی وارد کنید (مثال: استراتژی اصلی):"
        if is_message:
            await query_or_update.message.reply_text(text)
        else:
            await query_or_update.edit_message_text(text)

async def handle_backtest_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()

    # ─── دریافت تعداد سطوح ریوارد دلخواه ───
    if state == "bt_levels_count":
        try:
            count = int(text)
            if count < 2:
                await update.message.reply_text("❌ حداقل ۲ سطح نیاز است (یک سطح میانی + سطح خروج کامل).\nدوباره وارد کنید:")
                return True
            if count > 15:
                await update.message.reply_text("❌ حداکثر ۱۵ سطح مجاز است.\nدوباره وارد کنید:")
                return True
            mode = data
            set_state(user_id, "bt_levels_input", f"{mode}|{count}")
            await update.message.reply_text(
                f"🎯 حالا {count} سطح ریوارد را با کاما وارد کنید (به ترتیب صعودی):\n\n"
                f"مثال: `0.5, 1, 2, 3.5, 5`",
                parse_mode="Markdown"
            )
        except:
            await update.message.reply_text("❌ لطفاً یک عدد صحیح وارد کنید:")
        return True

    # ─── دریافت مقادیر سطوح ریوارد دلخواه ───
    elif state == "bt_levels_input":
        mode, count_str = data.split("|")
        count = int(count_str)
        try:
            parts = [p.strip() for p in text.split(",")]
            levels = [float(p) for p in parts]
        except:
            await update.message.reply_text(
                "❌ فرمت اشتباه است. اعداد را با کاما جدا کنید.\nمثال: `0.5, 1, 2, 3.5, 5`",
                parse_mode="Markdown"
            )
            return True

        if len(levels) != count:
            await update.message.reply_text(
                f"❌ شما گفتید {count} سطح می‌خواهید، ولی {len(levels)} عدد وارد کردید.\nدوباره وارد کنید:"
            )
            return True

        if any(l <= 0 for l in levels):
            await update.message.reply_text("❌ همه سطوح باید عدد مثبت باشند.\nدوباره وارد کنید:")
            return True

        if levels != sorted(levels):
            await update.message.reply_text("❌ سطوح باید به ترتیب صعودی وارد شوند.\nدوباره وارد کنید:")
            return True

        if len(set(levels)) != len(levels):
            await update.message.reply_text("❌ سطوح تکراری نباید وجود داشته باشد.\nدوباره وارد کنید:")
            return True

        import json
        rules = {"__levels__": levels}
        first_level = levels[0]
        set_state(user_id, f"bt_strat_r_{first_level}", f"{mode}|{json.dumps(rules)}")
        await update.message.reply_text(
            f"✅ سطوح ثبت شد: {', '.join('R'+str(l) for l in levels)}\n\n"
            f"در سطح R{first_level}، حد ضرر را جابجا کنم؟",
            reply_markup=r_level_question_menu(str(first_level))
        )
        return True

    # ─── دریافت مقدار SL برای یک سطح R خاص (به‌صورت عدد ریوارد، نه قیمت) ───
    elif state and state.startswith("bt_strat_slval_"):
        r_mult = state.replace("bt_strat_slval_", "")
        import json
        mode, rules_str = (data.split("|", 1) + [""])[:2]
        rules = json.loads(rules_str) if rules_str else {}

        try:
            r_value = float(text.replace(",", ""))
        except:
            await update.message.reply_text(
                f"❌ لطفاً یک عدد معتبر بین `0` و `{r_mult}` وارد کنید:",
                parse_mode="Markdown"
            )
            return True

        current_r = float(r_mult)
        if r_value < 0 or r_value >= current_r:
            await update.message.reply_text(
                f"❌ عدد باید بین `0` و `{r_mult}` باشد (کوچک‌تر از سطح فعلی).\n"
                f"عدد وارد‌شده ({r_value}) خارج از این بازه است.\nدوباره وارد کنید:",
                parse_mode="Markdown"
            )
            return True

        # ذخیره به‌صورت ضریب ریوارد نسبی (نه قیمت مطلق)
        # 0 یعنی دقیقاً نقطه ورود (ریسک‌فری کامل)
        rules[r_mult] = {"type": "r_trail", "value": r_value}

        levels = get_strategy_levels(rules)
        next_idx = levels.index(float(r_mult)) + 1
        await proceed_to_next_r_level(update, user_id, mode, rules, next_idx, is_message=True)
        return True

    # ─── ثبت نام استراتژی ───
    elif state == "bt_strat_name":
        import json
        mode, rules_str = (data.split("|", 1) + [""])[:2]
        rules = json.loads(rules_str) if rules_str else {}
        strategy_name = text
        save_strategy(user_id, strategy_name, rules)
        set_state(user_id, "bt_symbol", f"{mode}|{strategy_name}|{json.dumps(rules)}")
        await update.message.reply_text(
            f"✅ استراتژی «{strategy_name}» ذخیره شد.\n\n"
            f"📋 خلاصه قوانین این استراتژی:\n\n{describe_strategy(rules)}\n\n"
            f"نماد ارز را وارد کنید:\nمثال: `EURUSD` `GOLD` `USDCAD`",
            parse_mode="Markdown"
        )
        return True

    # ─── نماد ───
    elif state == "bt_symbol":
        symbol = text.upper()
        parts = data.split("|")
        mode, strategy_name, rules_str = parts[0], parts[1], parts[2]
        # اگر این ادامه‌ی یک batch گروهی است، batch_id قبلی را حفظ کن
        if len(parts) > 3 and parts[3].startswith("__BATCH__"):
            batch_id = parts[3].replace("__BATCH__", "")
            set_state(user_id, "bt_direction", f"{mode}|{strategy_name}|{rules_str}|{symbol}|__BATCHID__{batch_id}")
        else:
            set_state(user_id, "bt_direction", f"{mode}|{strategy_name}|{rules_str}|{symbol}")
        await update.message.reply_text("جهت معامله را انتخاب کنید:", reply_markup=direction_menu())
        return True

    # ─── تاریخ ورود ───
    elif state == "bt_entry_date":
        try:
            datetime.strptime(text, "%Y-%m-%d")
            set_state(user_id, "bt_entry_time", f"{data}|{text}")
            await update.message.reply_text(
                "⏰ ساعت ورود را به وقت ایران وارد کنید (فرمت 24 ساعته HH:MM)\nمثال: `14:35`",
                parse_mode="Markdown"
            )
        except:
            await update.message.reply_text("❌ فرمت تاریخ اشتباه است.\nمثال: `2026-03-10`", parse_mode="Markdown")
        return True

    # ─── ساعت ورود ───
    elif state == "bt_entry_time":
        try:
            datetime.strptime(text, "%H:%M")
            set_state(user_id, "bt_entry_price", f"{data}|{text}")
            await update.message.reply_text("قیمت ورود را وارد کنید:")
        except:
            await update.message.reply_text("❌ فرمت ساعت اشتباه است.\nمثال: `14:35` (24 ساعته)", parse_mode="Markdown")
        return True

    # ─── قیمت ورود ───
    elif state == "bt_entry_price":
        try:
            entry_price = float(text.replace(",", ""))
            set_state(user_id, "bt_sl_price", f"{data}|{entry_price}")
            await update.message.reply_text("قیمت حد ضرر اولیه را وارد کنید:")
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    # ─── حد ضرر اولیه ───
    elif state == "bt_sl_price":
        try:
            sl_price = float(text.replace(",", ""))
            parts = data.split("|")
            # اگر مسیر گروهی است، یک قطعه اضافه (__BATCHID__xxx) در میانه وجود دارد
            if len(parts) == 9 and parts[4].startswith("__BATCHID__"):
                mode, strategy_name, rules_str, symbol, batch_tag, direction, entry_date, entry_time, entry_price = parts
            else:
                mode, strategy_name, rules_str, symbol, direction, entry_date, entry_time, entry_price = parts
                batch_tag = None
            entry_price = float(entry_price)

            if (direction == "long" and sl_price >= entry_price) or (direction == "short" and sl_price <= entry_price):
                await update.message.reply_text("❌ حد ضرر با جهت معامله همخوانی ندارد.\nدوباره وارد کنید:")
                return True

            symbol_field = f"{symbol}{batch_tag}" if batch_tag else symbol
            set_state(user_id, "bt_timeframe",
                      f"{mode}|{strategy_name}|{rules_str}|{symbol_field}|{direction}|{entry_date}|{entry_time}|{entry_price}|{sl_price}")
            await update.message.reply_text("⏱ تایم‌فریم را انتخاب کنید:", reply_markup=timeframe_menu())
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    return False

# ─── دکمه‌های جهت و تایم‌فریم و ادامه/پایان گروهی ────────────
async def handle_backtest_buttons_extra(query, user_id, data):
    if data in ["bt_dir_long", "bt_dir_short"]:
        direction = "long" if data == "bt_dir_long" else "short"
        state, sdata = get_state(user_id)
        if state == "bt_direction":
            set_state(user_id, "bt_entry_date", f"{sdata}|{direction}")
            await query.edit_message_text(
                f"✅ جهت: {'Long 📈' if direction=='long' else 'Short 📉'}\n\n"
                f"تاریخ ورود را وارد کنید (فرمت: YYYY-MM-DD)\nمثال: `2026-03-10`",
                parse_mode="Markdown"
            )
        return True

    if data.startswith("bt_tf_"):
        timeframe = data.replace("bt_tf_", "")
        state, sdata = get_state(user_id)
        if state == "bt_timeframe":
            set_state(user_id, "bt_processing", f"{sdata}|{timeframe}")
            await query.edit_message_text("⏳ در حال دریافت داده‌های تاریخی و شبیه‌سازی...")
            await run_backtest_final(user_id, query)
        return True

    if data == "bt_add_another":
        state, sdata = get_state(user_id)
        parts = sdata.split("|")
        mode, strategy_name, rules_str, batch_id = parts[0], parts[1], parts[2], parts[3]
        set_state(user_id, "bt_symbol", f"{mode}|{strategy_name}|{rules_str}|__BATCH__{batch_id}")
        await query.edit_message_text(
            "➕ نماد ارز بعدی را وارد کنید:\nمثال: `EURUSD` `GOLD` `USDCAD`",
            parse_mode="Markdown"
        )
        return True

    if data == "bt_finish_group":
        state, sdata = get_state(user_id)
        parts = sdata.split("|")
        batch_id = parts[3]
        await send_batch_excel(query, user_id, batch_id)
        clear_state(user_id)
        return True

    return False

async def run_backtest_final(user_id, query):
    state, sdata = get_state(user_id)
    if state != "bt_processing" or not sdata:
        return

    parts = sdata.split("|")
    mode, strategy_name, rules_str, symbol, direction, entry_date, entry_time, entry_price, sl_price, timeframe = parts
    entry_price, sl_price = float(entry_price), float(sl_price)

    import json
    rules = json.loads(rules_str)

    existing_batch_id = None
    if "__BATCHID__" in symbol:
        symbol, batch_tag = symbol.split("__BATCHID__")
        existing_batch_id = batch_tag

    batch_id = existing_batch_id if existing_batch_id else f"{user_id}_{int(time_module.time())}"

    candles, entry_dt_utc, error = await get_historical_candles(symbol, entry_date, entry_time, timeframe)
    if error:
        await query.message.reply_text(error, reply_markup=backtest_menu())
        clear_state(user_id)
        return

    result = run_custom_simulation(candles, entry_price, sl_price, direction, rules)
    R = abs(entry_price - sl_price)

    entry_dt_iran = datetime.strptime(f"{entry_date} {entry_time}", "%Y-%m-%d %H:%M")
    exit_dt_iran = to_iran_time(result["exit_time_utc"]) if result["exit_time_utc"] else None

    save_bt_result(
        user_id, batch_id, symbol, direction,
        entry_dt_iran.strftime("%Y-%m-%d %H:%M"), entry_price, sl_price, strategy_name,
        result["max_r"], result["result_r"],
        exit_dt_iran.strftime("%Y-%m-%d %H:%M") if exit_dt_iran else None,
        result["exit_price"], result["status"]
    )

    status_text = "🟢 بسته شده" if result["status"] == "closed" else "🟡 هنوز باز (تا آخرین داده)"
    exit_line = (f"خروج: {result['exit_price']:.5f}\nتاریخ خروج (ایران): {exit_dt_iran.strftime('%Y-%m-%d %H:%M')}"
                 if result["exit_price"] else "هنوز خارج نشده")

    text = (
        f"📊 نتیجه بک‌تست — {symbol} ({'Long' if direction=='long' else 'Short'})\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
        f"استراتژی: {strategy_name}\n"
        f"ورود: {entry_price:.5f}\n"
        f"SL اولیه: {sl_price:.5f}  (R = {R:.5f})\n"
        f"تاریخ ورود (ایران): {format_date_fa(entry_dt_iran)}\n"
        f"تایم‌فریم: {timeframe}\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
        f"وضعیت: {status_text}\n"
        f"بالاترین R رسیده: R{result['max_r']}\n"
        f"{exit_line}\n"
        f"نتیجه: {'+' if result['result_r'] >= 0 else ''}{result['result_r']}R\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`"
    )

    if mode == "group":
        set_state(user_id, "bt_group_next", f"{mode}|{strategy_name}|{rules_str}|{batch_id}")
        await query.message.reply_text(
            text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("➕ ارز بعدی", callback_data="bt_add_another")],
                [InlineKeyboardButton("✅ پایان و دریافت اکسل", callback_data="bt_finish_group")],
            ])
        )
    else:
        await send_single_excel(query, user_id, batch_id, text)
        clear_state(user_id)

async def send_single_excel(query, user_id, batch_id, summary_text):
    rows = get_batch_results(user_id, batch_id)
    excel_buffer = build_excel_report(rows)
    await query.message.reply_text(summary_text, parse_mode="Markdown", reply_markup=backtest_menu())
    await query.message.reply_document(
        document=excel_buffer, filename=f"backtest_{batch_id}.xlsx",
        caption="📎 فایل اکسل نتیجه بک‌تست"
    )

async def send_batch_excel(query, user_id, batch_id):
    rows = get_batch_results(user_id, batch_id)
    excel_buffer = build_excel_report(rows)

    lines = ["📊 خلاصه نتایج بک‌تست گروهی:\n"]
    for row in rows:
        symbol, direction, max_r, result_r, status = row[3], row[4], row[9], row[10], row[13]
        emoji = "🟢" if status == "closed" else "🟡"
        lines.append(f"{emoji} {symbol} ({'Long' if direction=='long' else 'Short'}) → R{max_r} | نتیجه: {result_r:+.1f}R")

    await query.message.reply_text("\n".join(lines), reply_markup=backtest_menu())
    await query.message.reply_document(
        document=excel_buffer, filename=f"backtest_group_{batch_id}.xlsx",
        caption="📎 فایل اکسل نتایج بک‌تست گروهی"
    )

# ═══════════════ پایان MODULE: BACKTEST ════════════════════════



# ╔══════════════════════════════════════════════════════════════╗
# ║        MODULE: BACKTEST FILE UPLOAD (بک‌تست از فایل)         ║
# ║  کاربر فایل CSV یا Excel با لیست معاملات می‌فرستد (نماد،      ║
# ║  جهت، تاریخ، ساعت، قیمت ورود، حد ضرر، و اختیاری تایم‌فریم).   ║
# ║  همه ردیف‌ها با یک استراتژی مشترک بک‌تست می‌شوند و یک فایل    ║
# ║  اکسل با نتیجه‌ی کامل تمام معاملات برگردانده می‌شود.          ║
# ║  برای حذف: این بلوک رو پاک کن                                ║
# ║  + خط bt_upload رو از backtest_menu پاک کن                    ║
# ║  + خط bt_upload رو از ROUTER اصلی پاک کن (اگر جدا اضافه شده) ║
# ║  + MessageHandler برای filters.Document که این را صدا می‌زند  ║
# ║    را با هندلر اندیکاتور شخصی ادغام کن یا جدا نگه دار         ║
# ╚══════════════════════════════════════════════════════════════╝

BT_UPLOAD_REQUIRED_COLUMNS = ["نماد", "جهت", "تاریخ ورود", "ساعت ورود", "قیمت ورود", "حد ضرر"]
BT_UPLOAD_OPTIONAL_COLUMNS = ["تایم فریم"]

def bt_upload_sample_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📥 دریافت نمونه فایل Excel", callback_data="bt_upload_sample")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="fin_backtest")],
    ])

def bt_upload_timeframe_common_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("1 دقیقه", callback_data="btu_tf_1m"),
         InlineKeyboardButton("15 دقیقه", callback_data="btu_tf_15m")],
        [InlineKeyboardButton("1 ساعت", callback_data="btu_tf_1h")],
        [InlineKeyboardButton("📄 هر ردیف تایم‌فریم خودش را دارد (ستون فایل)", callback_data="btu_tf_percolumn")],
    ])

def build_bt_upload_sample_excel():
    """ساخت یک فایل نمونه Excel برای راهنمایی کاربر درباره ساختار درست فایل"""
    wb = Workbook()
    ws = wb.active
    ws.title = "لیست معاملات"
    headers = BT_UPLOAD_REQUIRED_COLUMNS + BT_UPLOAD_OPTIONAL_COLUMNS
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    ws.append(["EURUSD", "long", "2026-06-01", "14:30", "1.15842", "1.15640", "15m"])
    ws.append(["GOLD", "short", "2026-06-02", "09:00", "3350.5", "3360.0", "1h"])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    # شیت دوم: راهنمای کامل هر ستون (تا کاربر حتی بدون خواندن پیام تلگرام هم متوجه شود)
    guide = wb.create_sheet("راهنما")
    guide_rows = [
        ["ستون", "توضیح", "مثال معتبر"],
        ["نماد", "اسم ارز/جفت‌ارز دقیقاً مثل بازارهای مالی ربات", "EURUSD, GOLD, BTC, USD"],
        ["جهت", "long یا short (یا خرید/فروش به فارسی)", "long"],
        ["تاریخ ورود", "فرمت سال-ماه-روز (میلادی)", "2026-06-01"],
        ["ساعت ورود", "فرمت ۲۴ ساعته، به وقت ایران", "14:30"],
        ["قیمت ورود", "عدد اعشاری، بدون کاما یا واحد پول", "1.15842"],
        ["حد ضرر", "عدد اعشاری، بدون کاما یا واحد پول", "1.15640"],
        ["تایم فریم (اختیاری)", "1m یا 15m یا 1h -- اگر خالی بگذارید بعداً یک تایم‌فریم مشترک برای همه پرسیده می‌شود", "15m"],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 60
    guide.column_dimensions["C"].width = 20
    for row in guide.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def parse_bt_upload_file(file_bytes, filename):
    """
    پارس فایل CSV یا Excel به لیستی از دیکشنری‌های معامله.
    خروجی: (rows, error) -- rows لیست دیکشنری با کلیدهای استاندارد، یا
    (None, error_message) اگر ساختار فایل اشتباه بود.
    """
    column_map = {
        "نماد": "symbol", "جهت": "direction", "تاریخ ورود": "entry_date",
        "ساعت ورود": "entry_time", "قیمت ورود": "entry_price", "حد ضرر": "sl_price",
        "تایم فریم": "timeframe", "تایم‌فریم": "timeframe",
    }

    raw_rows = []
    try:
        if filename.lower().endswith(".csv"):
            text = file_bytes.decode("utf-8-sig")
            reader = csv.reader(text.splitlines())
            raw_rows = list(reader)
        else:
            from openpyxl import load_workbook
            import datetime as dt_module
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            # اگر کاربر شیت راهنما را هم نگه داشته، مطمئن شویم شیت داده‌ها
            # (نه شیت راهنما) خوانده می‌شود، نه صرفاً شیت فعال فایل
            ws = wb["لیست معاملات"] if "لیست معاملات" in wb.sheetnames else wb.active
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    formatted_row = []
                    for cell in row:
                        if cell is None:
                            formatted_row.append("")
                        elif isinstance(cell, dt_module.datetime):
                            # اکسل گاهی یک سلول «فقط ساعت» را به‌صورت datetime با
                            # تاریخ پایه ۱۸۹۹/۱۲/۳۰ یا ۱۹۰۰/۱/۱ ذخیره می‌کند (رفتار
                            # شناخته‌شده و رایج اکسل، نه خطای کاربر) -- این حالت را
                            # به‌عنوان ساعت (نه تاریخ کامل) تشخیص می‌دهیم
                            if cell.year in (1899, 1900):
                                formatted_row.append(cell.strftime("%H:%M"))
                            elif cell.hour == 0 and cell.minute == 0:
                                formatted_row.append(cell.strftime("%Y-%m-%d"))
                            else:
                                formatted_row.append(cell.strftime("%Y-%m-%d %H:%M"))
                        elif isinstance(cell, dt_module.date):
                            formatted_row.append(cell.strftime("%Y-%m-%d"))
                        elif isinstance(cell, dt_module.time):
                            formatted_row.append(cell.strftime("%H:%M"))
                        elif isinstance(cell, dt_module.timedelta):
                            # اگر در ستون قیمت (نه ساعت) یک بازه‌ی زمانی عجیب دیده شود،
                            # یعنی احتمالاً کاربر عدد را طوری تایپ کرده که اکسل آن را
                            # به اشتباه به‌عنوان زمان تفسیر کرده -- این یک مقدار خراب
                            # واقعی است و به AI ارجاع داده می‌شود (نه تبدیل نادرست)
                            formatted_row.append(f"[نامعتبر: {cell}]")
                        else:
                            formatted_row.append(str(cell))
                    raw_rows.append(formatted_row)
    except Exception as e:
        return None, f"❌ خطا در خواندن فایل: {e}"

    if len(raw_rows) < 2:
        return None, "❌ فایل خالی است یا فقط سطر عنوان دارد."

    header = [h.strip() for h in raw_rows[0]]
    missing = [col for col in BT_UPLOAD_REQUIRED_COLUMNS if col not in header]
    if missing:
        return None, f"❌ ستون‌های زیر در فایل پیدا نشد: {', '.join(missing)}\n\nاز دکمه «دریافت نمونه فایل» استفاده کنید."

    col_index = {}
    for i, h in enumerate(header):
        key = column_map.get(h)
        if key:
            col_index[key] = i

    parsed_rows = []
    problem_rows = []  # ردیف‌هایی که خطا دارند، برای بررسی بعدی با AI جمع می‌شوند
    for row_num, row in enumerate(raw_rows[1:], start=2):
        if all(not cell.strip() for cell in row):
            continue
        try:
            symbol = row[col_index["symbol"]].strip().upper()
            direction_raw = row[col_index["direction"]].strip().lower()
            direction = "long" if direction_raw in ("long", "خرید", "buy") else "short" if direction_raw in ("short", "فروش", "sell") else None
            entry_date = row[col_index["entry_date"]].strip()
            entry_time = row[col_index["entry_time"]].strip()
            entry_price_raw = row[col_index["entry_price"]].strip()
            sl_price_raw = row[col_index["sl_price"]].strip()
            timeframe = row[col_index["timeframe"]].strip() if "timeframe" in col_index and col_index["timeframe"] < len(row) else ""

            if not symbol or direction is None or not entry_date or not entry_time:
                raise ValueError("اطلاعات ناقص یا جهت نامعتبر (باید long/short باشد)")
            if "[نامعتبر:" in entry_price_raw or "[نامعتبر:" in sl_price_raw:
                raise ValueError("مقدار قیمت به‌جای عدد، به‌صورت زمان/تاریخ در اکسل ذخیره شده")
            entry_price = float(entry_price_raw)
            sl_price = float(sl_price_raw)
            datetime.strptime(entry_date, "%Y-%m-%d")
            datetime.strptime(entry_time, "%H:%M")

            parsed_rows.append({
                "symbol": symbol, "direction": direction, "entry_date": entry_date,
                "entry_time": entry_time, "entry_price": entry_price, "sl_price": sl_price,
                "timeframe": timeframe or None,
            })
        except (ValueError, IndexError) as e:
            problem_rows.append({"row_num": row_num, "raw_row": row, "error": str(e)})

    if not parsed_rows and not problem_rows:
        return None, "❌ هیچ ردیف معتبری در فایل پیدا نشد."

    return {"parsed_rows": parsed_rows, "problem_rows": problem_rows, "header": header, "col_index": col_index}, None


async def ai_fix_bt_upload_row(header, raw_row, error_message):
    """
    از هوش مصنوعی می‌خواهد یک ردیف مشکل‌دار از فایل معاملات را بررسی کند
    و بهترین حدس برای مقدار درست هر ستون بدهد. خروجی: dict با کلیدهای
    استاندارد (symbol/direction/entry_date/entry_time/entry_price/sl_price/
    timeframe) و یک فیلد اضافه fix_note (توضیح کوتاه اصلاحی که انجام شد)،
    یا None اگر AI هم نتوانست حدس معقولی بزند.
    """
    row_description = ", ".join(f"{h}: {v!r}" for h, v in zip(header, raw_row))
    prompt = (
        f"این یک ردیف از فایل معاملات فارکس/بازار مالی است که هنگام پردازش "
        f"خودکار خطا داده:\n{row_description}\n\n"
        f"خطای دقیق: {error_message}\n\n"
        f"این خطا معمولاً به‌خاطر این است که مایکروسافت اکسل هنگام ذخیره، "
        f"یک سلول ساعت یا قیمت را به‌اشتباه به‌صورت تاریخ/زمان تفسیر کرده. "
        f"بر اساس بقیه‌ی داده‌های همین ردیف (و منطق عادی معاملات فارکس)، "
        f"بهترین حدس خودت را برای مقدار درست هر ستون بزن.\n\n"
        f"دقیقاً و فقط به این فرمت جواب بده (اگر نمی‌توانی حدس معقولی بزنی، "
        f"فقط بنویس: غیرقابل‌حل):\n"
        f"نماد: <...>\n"
        f"جهت: <long یا short>\n"
        f"تاریخ ورود: <YYYY-MM-DD>\n"
        f"ساعت ورود: <HH:MM>\n"
        f"قیمت ورود: <عدد>\n"
        f"حد ضرر: <عدد>\n"
        f"توضیح اصلاح: <یک جمله کوتاه که چه چیزی را چطور حدس زدی>"
    )
    result = await get_ai_response([{"role": "user", "content": prompt}])
    if not result or "غیرقابل‌حل" in result.strip()[:20]:
        return None

    def extract(field, text):
        m = re.search(rf'{field}\s*:?\s*(.+)', text)
        return m.group(1).strip() if m else None

    try:
        symbol = extract("نماد", result)
        direction_raw = (extract("جهت", result) or "").lower()
        direction = "long" if "long" in direction_raw else "short" if "short" in direction_raw else None
        entry_date = extract("تاریخ ورود", result)
        entry_time = extract("ساعت ورود", result)
        entry_price = float(extract("قیمت ورود", result))
        sl_price = float(extract("حد ضرر", result))
        fix_note = extract("توضیح اصلاح", result) or "اصلاح شده توسط هوش مصنوعی"
        if not symbol or direction is None:
            return None
        datetime.strptime(entry_date, "%Y-%m-%d")
        datetime.strptime(entry_time, "%H:%M")
        return {
            "symbol": symbol.upper(), "direction": direction, "entry_date": entry_date,
            "entry_time": entry_time, "entry_price": entry_price, "sl_price": sl_price,
            "timeframe": None, "fix_note": fix_note,
        }
    except Exception:
        return None


def build_bt_upload_fixed_excel(good_rows, fixed_rows, unfixable_rows):
    """ساخت فایل اکسل نهایی: ردیف‌های سالم + ردیف‌های اصلاح‌شده (با ستون توضیح) + ردیف‌های حل‌نشده در انتها"""
    wb = Workbook()
    ws = wb.active
    ws.title = "لیست معاملات (اصلاح‌شده)"
    headers = BT_UPLOAD_REQUIRED_COLUMNS + BT_UPLOAD_OPTIONAL_COLUMNS + ["وضعیت"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    for r in good_rows:
        ws.append([r["symbol"], r["direction"], r["entry_date"], r["entry_time"],
                   r["entry_price"], r["sl_price"], r.get("timeframe") or "", "سالم"])

    for r in fixed_rows:
        row_idx = ws.max_row + 1
        ws.append([r["symbol"], r["direction"], r["entry_date"], r["entry_time"],
                   r["entry_price"], r["sl_price"], r.get("timeframe") or "",
                   f"✏️ اصلاح‌شده: {r['fix_note']}"])
        for cell in ws[row_idx]:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for problem in unfixable_rows:
        row_idx = ws.max_row + 1
        raw = problem["raw_row"]
        padded = raw + [""] * (len(BT_UPLOAD_REQUIRED_COLUMNS + BT_UPLOAD_OPTIONAL_COLUMNS) - len(raw))
        ws.append(padded[:len(BT_UPLOAD_REQUIRED_COLUMNS + BT_UPLOAD_OPTIONAL_COLUMNS)] + [f"❌ غیرقابل‌اصلاح: {problem['error']}"])
        for cell in ws[row_idx]:
            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

async def run_bt_upload_batch(user_id, update, rows, rules, strategy_name, common_timeframe):
    """اجرای بک‌تست برای همه ردیف‌های فایل با یک استراتژی مشترک"""
    batch_id = f"{user_id}_{int(time_module.time())}"
    results = []

    await update.message.reply_text(f"⏳ در حال اجرای بک‌تست روی {len(rows)} ردیف... (ممکن است کمی طول بکشد)")

    for row in rows:
        timeframe = row["timeframe"] or common_timeframe
        if not timeframe:
            results.append({**row, "error": "تایم‌فریم مشخص نشده"})
            continue
        try:
            candles, entry_dt_utc, error = await get_historical_candles(
                row["symbol"], row["entry_date"], row["entry_time"], timeframe
            )
            if error:
                results.append({**row, "error": error})
                continue
            result = run_custom_simulation(candles, row["entry_price"], row["sl_price"], row["direction"], rules)
            entry_dt_iran = datetime.strptime(f"{row['entry_date']} {row['entry_time']}", "%Y-%m-%d %H:%M")
            exit_dt_iran = to_iran_time(result["exit_time_utc"]) if result["exit_time_utc"] else None

            save_bt_result(
                user_id, batch_id, row["symbol"], row["direction"],
                entry_dt_iran.strftime("%Y-%m-%d %H:%M"), row["entry_price"], row["sl_price"], strategy_name,
                result["max_r"], result["result_r"],
                exit_dt_iran.strftime("%Y-%m-%d %H:%M") if exit_dt_iran else None,
                result["exit_price"], result["status"]
            )
            results.append({**row, "timeframe": timeframe, "result": result, "exit_dt_iran": exit_dt_iran})
        except Exception as e:
            results.append({**row, "error": f"خطای غیرمنتظره: {e}"})

    return results, batch_id

def build_bt_upload_result_excel(results, strategy_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Backtest"
    headers = ["نماد", "جهت", "تاریخ ورود", "ساعت ورود", "قیمت ورود", "حد ضرر",
               "تایم‌فریم", "بالاترین R", "نتیجه (R)", "وضعیت", "خطا"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        if "error" in r:
            ws.append([r["symbol"], r["direction"], r["entry_date"], r["entry_time"],
                       r["entry_price"], r["sl_price"], r.get("timeframe", ""), "-", "-", "خطا", r["error"]])
        else:
            result = r["result"]
            ws.append([
                r["symbol"], r["direction"], r["entry_date"], r["entry_time"],
                r["entry_price"], r["sl_price"], r["timeframe"],
                result["max_r"], result["result_r"],
                "بسته شده" if result["status"] == "closed" else "باز", ""
            ])
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ─── هندلر دکمه‌ها ────────────────────────────────────────────
async def handle_bt_upload_button(query, user_id):
    data = query.data

    if data == "btu_confirm_fixed":
        state, sdata = get_state(user_id)
        if state != "bt_upload_confirm_fixed":
            return
        import json
        payload = json.loads(sdata)
        await query.edit_message_text("✅ تأیید شد. در حال ادامه...")
        await continue_bt_upload_after_parsing(query, user_id, payload["rows"])
        return

    elif data == "btu_cancel_fixed":
        clear_state(user_id)
        await query.edit_message_text(
            "❌ لغو شد. می‌توانید فایل را دستی اصلاح کرده و دوباره از اول بفرستید.",
            reply_markup=backtest_menu()
        )
        return

    elif data == "bt_upload":
        await query.edit_message_text(
            "📁 آپلود فایل لیست معاملات\n\n"
            "برای جلوگیری از خطا، یک فایل نمونه‌ی آماده برایتان فرستادم. "
            "همان فایل را با معاملات خودتان پر کنید (بدون تغییر اسم ستون‌ها) و دوباره همین‌جا بفرستید.\n\n"
            f"ستون‌های اجباری: {', '.join(BT_UPLOAD_REQUIRED_COLUMNS)}\n"
            f"ستون اختیاری: تایم فریم (اگر خالی بگذارید، بعداً یک تایم‌فریم مشترک برای همه می‌پرسیم)\n\n"
            "جهت را می‌توانید long/short یا خرید/فروش بنویسید.",
            reply_markup=bt_upload_sample_menu()
        )
        sample = build_bt_upload_sample_excel()
        await query.message.reply_document(
            document=sample, filename="نمونه_لیست_معاملات.xlsx",
            caption="📥 این فایل را با معاملات خودتان پر کنید (ردیف‌های نمونه را پاک یا جایگزین کنید) و همین‌جا دوباره بفرستید."
        )
        set_state(user_id, "bt_upload_awaiting_file")

    elif data == "bt_upload_sample":
        sample = build_bt_upload_sample_excel()
        await query.message.reply_document(
            document=sample, filename="نمونه_لیست_معاملات.xlsx",
            caption="📥 این فایل نمونه را با معاملات خودتان پر کنید و دوباره ارسال کنید."
        )

    elif data.startswith("btu_tf_"):
        state, sdata = get_state(user_id)
        if state != "bt_upload_timeframe_choice":
            return
        import json
        payload = json.loads(sdata)
        common_tf = None if data == "btu_tf_percolumn" else data.replace("btu_tf_", "")
        payload["common_timeframe"] = common_tf

        strategies = get_strategies(user_id)
        set_state(user_id, "bt_upload_strategy_choice", json.dumps(payload))
        if strategies:
            keyboard = [[InlineKeyboardButton(f"📋 {name}", callback_data=f"btu_strat_{sid}")] for sid, name, _ in strategies[:10]]
            await query.edit_message_text("📋 کدام استراتژی مدیریت پوزیشن استفاده شود؟", reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            await query.edit_message_text(
                "❌ هنوز هیچ استراتژی مدیریت پوزیشنی نساخته‌اید.\n\n"
                "ابتدا از منوی «➕ بک‌تست جدید» یک استراتژی بسازید، سپس دوباره فایل را آپلود کنید.",
                reply_markup=backtest_menu()
            )
            clear_state(user_id)

    elif data.startswith("btu_strat_"):
        import json
        sid = int(data.replace("btu_strat_", ""))
        name, rules = get_strategy_by_id(sid)
        if not name:
            await query.answer("این استراتژی پیدا نشد.", show_alert=True)
            return
        state, sdata = get_state(user_id)
        payload = json.loads(sdata)
        rows = payload["rows"]
        common_timeframe = payload.get("common_timeframe")
        clear_state(user_id)

        await query.edit_message_text(f"⏳ در حال اجرای بک‌تست با استراتژی «{name}»...")
        results, batch_id = await run_bt_upload_batch(user_id, query, rows, rules, name, common_timeframe)

        success_results = [r for r in results if "error" not in r]
        error_results = [r for r in results if "error" in r]
        summary = f"📊 نتیجه بک‌تست گروهی از فایل\n\nاستراتژی: {name}\nتعداد کل ردیف‌ها: {len(rows)}\n"
        summary += f"✅ موفق: {len(success_results)} | ❌ خطا: {len(error_results)}"
        await query.message.reply_text(summary, reply_markup=backtest_menu())

        excel_buffer = build_bt_upload_result_excel(results, name)
        await query.message.reply_document(
            document=excel_buffer, filename=f"batch_backtest_{batch_id}.xlsx",
            caption="📎 فایل اکسل کامل نتایج (شامل ردیف‌های خطادار)"
        )

async def handle_bt_upload_document(update, context):
    """دریافت فایل CSV/Excel از کاربر وقتی در state انتظار فایل بک‌تست است"""
    user_id = update.effective_user.id
    state, data = get_state(user_id)
    if state != "bt_upload_awaiting_file":
        return False

    doc = update.message.document
    if not doc or not (doc.file_name.lower().endswith(".csv") or doc.file_name.lower().endswith(".xlsx")):
        await update.message.reply_text("❌ لطفاً فقط فایل CSV یا Excel (.xlsx) ارسال کنید.")
        return True

    file = await context.bot.get_file(doc.file_id)
    file_bytes = await file.download_as_bytearray()

    parse_result, error = parse_bt_upload_file(bytes(file_bytes), doc.file_name)
    if error:
        await update.message.reply_text(error, reply_markup=bt_upload_sample_menu())
        return True

    good_rows = parse_result["parsed_rows"]
    problem_rows = parse_result["problem_rows"]

    if not problem_rows:
        # هیچ ردیف مشکل‌داری نبود؛ مستقیم به مرحله بعد برو (رفتار قبلی، بدون تغییر)
        await continue_bt_upload_after_parsing(update, user_id, good_rows)
        return True

    # ردیف‌های مشکل‌دار وجود دارند؛ با کمک هوش مصنوعی سعی در اصلاح آن‌ها می‌شود
    await update.message.reply_text(
        f"⏳ {len(problem_rows)} ردیف مشکل ساختاری داشتند (معمولاً به‌خاطر تفسیر اشتباه "
        f"تاریخ/ساعت توسط اکسل). در حال بررسی و تلاش برای اصلاح خودکار با هوش مصنوعی..."
    )
    fixed_rows = []
    unfixable_rows = []
    header = parse_result["header"]
    for problem in problem_rows:
        fixed = await ai_fix_bt_upload_row(header, problem["raw_row"], problem["error"])
        if fixed:
            fixed_rows.append(fixed)
        else:
            unfixable_rows.append(problem)

    fixed_excel = build_bt_upload_fixed_excel(good_rows, fixed_rows, unfixable_rows)
    summary = (
        f"📋 نتیجه بررسی فایل:\n\n"
        f"✅ سالم: {len(good_rows)} ردیف\n"
        f"✏️ اصلاح‌شده توسط هوش مصنوعی: {len(fixed_rows)} ردیف\n"
        f"❌ غیرقابل‌اصلاح: {len(unfixable_rows)} ردیف\n\n"
        f"فایل اصلاح‌شده را بررسی کنید. ردیف‌های زرد رنگ اصلاح شده‌اند و "
        f"ردیف‌های قرمز رنگ (اگر بودند) از بک‌تست کنار گذاشته می‌شوند."
    )
    await update.message.reply_document(
        document=fixed_excel, filename="لیست_معاملات_اصلاح‌شده.xlsx", caption=summary
    )

    import json
    all_good_rows = good_rows + fixed_rows
    if not all_good_rows:
        await update.message.reply_text(
            "❌ متأسفانه هیچ ردیف قابل‌استفاده‌ای باقی نماند.",
            reply_markup=backtest_menu()
        )
        clear_state(user_id)
        return True

    set_state(user_id, "bt_upload_confirm_fixed", json.dumps({"rows": all_good_rows}))
    await update.message.reply_text(
        "آیا با این نسخه‌ی اصلاح‌شده ادامه دهیم؟",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ تأیید و ادامه بک‌تست", callback_data="btu_confirm_fixed")],
            [InlineKeyboardButton("❌ لغو، خودم دستی اصلاح می‌کنم", callback_data="btu_cancel_fixed")],
        ])
    )
    return True


async def continue_bt_upload_after_parsing(update, user_id, rows):
    """ادامه‌ی جریان عادی بعد از پارس موفق (چه بدون مشکل، چه بعد از تأیید نسخه‌ی اصلاح‌شده)"""
    import json
    has_all_timeframes = all(row.get("timeframe") for row in rows)
    payload = {"rows": rows}

    if has_all_timeframes:
        payload["common_timeframe"] = None
        strategies = get_strategies(user_id)
        set_state(user_id, "bt_upload_strategy_choice", json.dumps(payload))
        await update.message.reply_text(f"✅ {len(rows)} ردیف آماده بک‌تست است (هر ردیف تایم‌فریم خودش را دارد).")
        if strategies:
            keyboard = [[InlineKeyboardButton(f"📋 {name}", callback_data=f"btu_strat_{sid}")] for sid, name, _ in strategies[:10]]
            await update.message.reply_text("📋 کدام استراتژی مدیریت پوزیشن استفاده شود؟", reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            await update.message.reply_text(
                "❌ هنوز هیچ استراتژی مدیریت پوزیشنی نساخته‌اید.\n\n"
                "ابتدا از منوی «➕ بک‌تست جدید» یک استراتژی بسازید، سپس دوباره فایل را آپلود کنید.",
                reply_markup=backtest_menu()
            )
            clear_state(user_id)
    else:
        set_state(user_id, "bt_upload_timeframe_choice", json.dumps(payload))
        await update.message.reply_text(
            f"✅ {len(rows)} ردیف آماده است.\n\n"
            f"بعضی ردیف‌ها ستون تایم‌فریم ندارند. یک تایم‌فریم مشترک برای آن‌ها انتخاب کنید، "
            f"یا اگر همه‌ی ردیف‌ها را در فایل مشخص کرده بودید، گزینه‌ی مربوطه را بزنید:",
            reply_markup=bt_upload_timeframe_common_menu()
        )

# ═══════════════ پایان MODULE: BACKTEST FILE UPLOAD ══════════════




# ╔══════════════════════════════════════════════════════════════╗
# ║              MODULE: INDICATOR BACKTEST                      ║
# ║  بک‌تست استراتژی بر اساس اندیکاتور (MA Cross, RSI, MACD,     ║
# ║  Bollinger Bands) با اجرای ساده یا با مدیریت پوزیشن (R)       ║
# ║  برای حذف: این بلوک رو پاک کن                                ║
# ║  + خط ind_backtest رو از financial_menu پاک کن                ║
# ║  + خط ib_ رو از ROUTER اصلی پاک کن                            ║
# ║  + state های ib_ رو از message_handler پاک کن                 ║
# ╚══════════════════════════════════════════════════════════════╝

INDICATOR_DEFS = {
    "ma_cross": {
        "label": "📈 کراس میانگین متحرک (MA Cross)",
        "params": [
            {"key": "fast_period", "label": "دوره میانگین سریع", "default": 9},
            {"key": "slow_period", "label": "دوره میانگین کند", "default": 21},
        ],
        "explain": (
            "میانگین متحرک سریع (کوتاه‌مدت) و کند (بلندمدت) محاسبه می‌شود.\n"
            "📈 خرید: میانگین سریع از پایین به بالای میانگین کند عبور کند\n"
            "📉 فروش: میانگین سریع از بالا به پایین میانگین کند عبور کند"
        ),
    },
    "rsi": {
        "label": "📊 RSI (اشباع خرید/فروش)",
        "params": [
            {"key": "period", "label": "دوره RSI", "default": 14},
            {"key": "oversold", "label": "سطح اشباع فروش", "default": 30},
            {"key": "overbought", "label": "سطح اشباع خرید", "default": 70},
        ],
        "explain": (
            "شاخص قدرت نسبی بین ۰ تا ۱۰۰ محاسبه می‌شود.\n"
            "📈 خرید: RSI از زیر سطح اشباع فروش به بالای آن برگردد\n"
            "📉 فروش: RSI از بالای سطح اشباع خرید به زیر آن برگردد"
        ),
    },
    "macd": {
        "label": "📉 MACD",
        "params": [
            {"key": "fast_period", "label": "دوره EMA سریع", "default": 12},
            {"key": "slow_period", "label": "دوره EMA کند", "default": 26},
            {"key": "signal_period", "label": "دوره خط سیگنال", "default": 9},
        ],
        "explain": (
            "خط MACD = EMA سریع منهای EMA کند. خط سیگنال = EMA خط MACD.\n"
            "📈 خرید: خط MACD از پایین به بالای خط سیگنال عبور کند\n"
            "📉 فروش: خط MACD از بالا به پایین خط سیگنال عبور کند"
        ),
    },
    "bollinger": {
        "label": "📏 Bollinger Bands",
        "params": [
            {"key": "period", "label": "دوره میانگین", "default": 20},
            {"key": "std_mult", "label": "ضریب انحراف معیار", "default": 2},
        ],
        "explain": (
            "باند بالا و پایین بر اساس میانگین و انحراف معیار قیمت محاسبه می‌شود.\n"
            "📈 خرید: قیمت به باند پایین برسد یا پایین‌تر برود\n"
            "📉 فروش: قیمت به باند بالا برسد یا بالاتر برود"
        ),
    },
    "cvd": {
        "label": "📶 CVD (واگرایی حجم تجمعی)",
        "params": [
            {"key": "lookback", "label": "دوره مقایسه (تعداد کندل)", "default": 5},
        ],
        "explain": (
            "⚠️ توجه: در فارکس حجم واقعی خرید/فروش وجود ندارد (بازار متمرکز نیست)، "
            "پس این نسخه از CVD تقریبی است: بر اساس کندل‌های ۱ دقیقه‌ای، اگر کندل صعودی "
            "بسته شود حجمش «خرید» و اگر نزولی بسته شود «فروش» حساب می‌شود (فارغ از "
            "تایم‌فریم اصلی بک‌تست، همیشه از کندل ۱ دقیقه محاسبه می‌شود).\n"
            "📈 خرید: قیمت کف پایین‌تر می‌سازد ولی CVD کف بالاتر می‌سازد (واگرایی مثبت)\n"
            "📉 فروش: قیمت سقف بالاتر می‌سازد ولی CVD سقف پایین‌تر می‌سازد (واگرایی منفی)"
        ),
    },
}

# ─── محاسبات ریاضی اندیکاتورها (خود ربات محاسبه می‌کند، نه AI) ──
def calc_ema(prices, period):
    result = [None] * len(prices)
    if len(prices) < period:
        return result
    multiplier = 2 / (period + 1)
    result[period - 1] = sum(prices[:period]) / period
    for i in range(period, len(prices)):
        result[i] = (prices[i] - result[i - 1]) * multiplier + result[i - 1]
    return result

def calc_rsi(prices, period=14):
    if len(prices) < period + 1:
        return [None] * len(prices)
    result = [None] * len(prices)
    gains, losses = [], []
    for i in range(1, len(prices)):
        change = prices[i] - prices[i - 1]
        gains.append(max(change, 0))
        losses.append(max(-change, 0))
    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period
    result[period] = 100 if avg_loss == 0 else 100 - (100 / (1 + avg_gain / avg_loss))
    for i in range(period, len(gains)):
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period
        result[i + 1] = 100 if avg_loss == 0 else 100 - (100 / (1 + avg_gain / avg_loss))
    return result

def calc_macd(prices, fast=12, slow=26, signal=9):
    ema_fast = calc_ema(prices, fast)
    ema_slow = calc_ema(prices, slow)
    macd_line = [None] * len(prices)
    for i in range(len(prices)):
        if ema_fast[i] is not None and ema_slow[i] is not None:
            macd_line[i] = ema_fast[i] - ema_slow[i]
    valid = [v for v in macd_line if v is not None]
    signal_line = [None] * len(prices)
    if len(valid) >= signal:
        start_idx = next(i for i, v in enumerate(macd_line) if v is not None)
        sig_values = calc_ema(valid, signal)
        for i, v in enumerate(sig_values):
            signal_line[start_idx + i] = v
    return macd_line, signal_line

def calc_bollinger(prices, period=20, std_mult=2):
    import statistics
    middle = [None] * len(prices)
    upper = [None] * len(prices)
    lower = [None] * len(prices)
    for i in range(period - 1, len(prices)):
        window = prices[i - period + 1:i + 1]
        mean = sum(window) / period
        std = statistics.pstdev(window)
        middle[i] = mean
        upper[i] = mean + std_mult * std
        lower[i] = mean - std_mult * std
    return upper, middle, lower

def calc_cvd_from_1m(one_min_candles):
    """محاسبه CVD تجمعی از کندل‌های ۱ دقیقه (بر اساس جهت بسته شدن هر کندل)"""
    cvd_values = []
    cumulative = 0
    for c in one_min_candles:
        volume = c.get("tick_volume") or 0
        if c["close"] > c["open"]:
            cumulative += volume
        elif c["close"] < c["open"]:
            cumulative -= volume
        cvd_values.append(cumulative)
    return cvd_values

def aggregate_cvd_to_main_candles(one_min_candles, main_candles):
    """برای هر کندل اصلی، مقدار CVD تجمعی تا انتهای بازه آن کندل را برمی‌گرداند"""
    cvd_1m = calc_cvd_from_1m(one_min_candles)
    result = [None] * len(main_candles)
    j = 0
    last_val = None
    for i, main_c in enumerate(main_candles):
        main_time = main_c["time_utc"]
        while j < len(one_min_candles) and one_min_candles[j]["time_utc"] <= main_time:
            last_val = cvd_1m[j]
            j += 1
        result[i] = last_val
    return result

def generate_cvd_divergence_signals(closes, cvd_values, lookback=5):
    """واگرایی صعودی=خرید، واگرایی نزولی=فروش؛ بدون تکرار سیگنال یکسان پشت‌سرهم"""
    raw_signals = [None] * len(closes)
    for i in range(lookback * 2, len(closes)):
        if cvd_values[i] is None:
            continue
        window_prev = closes[i - lookback * 2:i - lookback]
        window_curr = closes[i - lookback:i + 1]
        cvd_prev = [v for v in cvd_values[i - lookback * 2:i - lookback] if v is not None]
        cvd_curr = [v for v in cvd_values[i - lookback:i + 1] if v is not None]
        if not window_prev or not window_curr or not cvd_prev or not cvd_curr:
            continue

        if min(window_curr) < min(window_prev) and min(cvd_curr) > min(cvd_prev):
            raw_signals[i] = "buy"
        if max(window_curr) > max(window_prev) and max(cvd_curr) < max(cvd_prev):
            raw_signals[i] = "sell"

    signals = [None] * len(closes)
    last_signal_type = None
    for i, s in enumerate(raw_signals):
        if s is not None and s != last_signal_type:
            signals[i] = s
            last_signal_type = s
    return signals

async def generate_signals(indicator_key, candles, params, symbol=None):
    closes = [c["close"] for c in candles]
    signals = [None] * len(candles)

    if indicator_key == "ma_cross":
        fast_ma = calc_ema(closes, params["fast_period"])
        slow_ma = calc_ema(closes, params["slow_period"])
        for i in range(1, len(closes)):
            if None in (fast_ma[i], slow_ma[i], fast_ma[i - 1], slow_ma[i - 1]):
                continue
            prev_diff = fast_ma[i - 1] - slow_ma[i - 1]
            curr_diff = fast_ma[i] - slow_ma[i]
            if prev_diff <= 0 < curr_diff:
                signals[i] = "buy"
            elif prev_diff >= 0 > curr_diff:
                signals[i] = "sell"

    elif indicator_key == "rsi":
        rsi = calc_rsi(closes, params["period"])
        oversold, overbought = params["oversold"], params["overbought"]
        for i in range(1, len(closes)):
            if rsi[i] is None or rsi[i - 1] is None:
                continue
            if rsi[i - 1] <= oversold < rsi[i]:
                signals[i] = "buy"
            elif rsi[i - 1] >= overbought > rsi[i]:
                signals[i] = "sell"

    elif indicator_key == "macd":
        macd_line, signal_line = calc_macd(closes, params["fast_period"], params["slow_period"], params["signal_period"])
        for i in range(1, len(closes)):
            if None in (macd_line[i], signal_line[i], macd_line[i - 1], signal_line[i - 1]):
                continue
            prev_diff = macd_line[i - 1] - signal_line[i - 1]
            curr_diff = macd_line[i] - signal_line[i]
            if prev_diff <= 0 < curr_diff:
                signals[i] = "buy"
            elif prev_diff >= 0 > curr_diff:
                signals[i] = "sell"

    elif indicator_key == "bollinger":
        upper, middle, lower = calc_bollinger(closes, params["period"], params["std_mult"])
        for i in range(len(closes)):
            if lower[i] is None or upper[i] is None:
                continue
            if closes[i] <= lower[i]:
                signals[i] = "buy"
            elif closes[i] >= upper[i]:
                signals[i] = "sell"

    elif indicator_key == "cvd":
        if not candles:
            return signals
        from_dt = candles[0]["time_utc"]
        to_dt = candles[-1]["time_utc"] + timedelta(minutes=1)
        one_min_candles, error = await get_historical_range_candles_raw(symbol, from_dt, to_dt, "1m")
        if error or not one_min_candles:
            return signals
        cvd_values = aggregate_cvd_to_main_candles(one_min_candles, candles)
        signals = generate_cvd_divergence_signals(closes, cvd_values, params["lookback"])

    return signals

# ─── موتور بک‌تست سیگنال‌محور ──────────────────────────────────
def run_signal_backtest_simple(candles, signals):
    trades = []
    position = None
    for i, candle in enumerate(candles):
        sig = signals[i]
        if sig == "buy" and position is None:
            position = {"entry_price": candle["close"], "entry_time": candle["time_utc"]}
        elif sig == "sell" and position is not None:
            exit_price = candle["close"]
            pnl_percent = (exit_price - position["entry_price"]) / position["entry_price"] * 100
            trades.append({
                "entry_time": position["entry_time"], "entry_price": position["entry_price"],
                "exit_time": candle["time_utc"], "exit_price": exit_price,
                "pnl_percent": round(pnl_percent, 2)
            })
            position = None
    return trades

def calc_sl_price_from_config(entry_price, sl_config, symbol):
    """
    محاسبه قیمت SL بر اساس واحد انتخابی کاربر:
    - percent: درصدی از قیمت ورود
    - dollar: مقدار دلاری ثابت (بر اساس ارزش هر واحد حرکت قیمت، مشابه ماژول میزان خرید)
    - pip: تعداد پیپ استاندارد (۰.۰۰۰۱ برای اکثر جفت‌ارزها، ۰.۰۱ برای جفت‌های JPY)
    """
    unit = sl_config["unit"]
    value = sl_config["value"]
    symbol_upper = symbol.upper()

    if unit == "percent":
        return entry_price * (1 - value / 100)

    elif unit == "pip":
        spec = FX_CONTRACT_SPECS.get(symbol_upper)
        pip_size = spec["pip_size"] if spec else 0.0001
        return entry_price - (value * pip_size)

    elif unit == "dollar":
        category = ps_classify_symbol(symbol_upper)
        if category == "forex_pair":
            spec = FX_CONTRACT_SPECS[symbol_upper]
            contract_size = spec["contract_size"]
            if spec["quote_is_usd"]:
                price_distance = value / contract_size
            else:
                price_distance = (value * entry_price) / contract_size
        elif category == "metal":
            contract_size = METAL_CONTRACT_SPECS[symbol_upper]["contract_size"]
            price_distance = value / contract_size
        else:
            # کریپتو، بازار ایران، یا فارکس ناشناخته: مستقیم بر اساس واحد قیمتی
            price_distance = value
        return entry_price - price_distance

    return entry_price * 0.98  # پیش‌فرض ایمن در صورت واحد نامشخص

def run_signal_backtest_with_management(candles, signals, sl_config, symbol, strategy_rules):
    trades = []
    for i, candle in enumerate(candles):
        if signals[i] != "buy":
            continue
        entry_price = candle["close"]
        sl_price = calc_sl_price_from_config(entry_price, sl_config, symbol)
        remaining = candles[i + 1:]
        if not remaining:
            continue
        result = run_custom_simulation(remaining, entry_price, sl_price, "long", strategy_rules)
        trades.append({
            "entry_time": candle["time_utc"], "entry_price": entry_price,
            "sl_price": sl_price, **result
        })
    return trades

# ─── دریافت داده تاریخی برای بازه کامل ─────────────────────────
async def get_historical_range_candles_raw(symbol, from_dt_utc, to_dt_utc, interval):
    """
    نسخه خام: from_dt_utc/to_dt_utc به‌صورت datetime مستقیم (نه رشته تاریخ).
    شامل فیلدهای open و tick_volume هم می‌شود (برای محاسبه CVD لازم است).
    """
    try:
        bars_per_day = {"1m": 1440, "15m": 96, "1h": 24}
        max_days_by_interval = {k: max(1, 2000 // v) for k, v in bars_per_day.items()}
        max_days = max_days_by_interval.get(interval, 20)
        span_days = (to_dt_utc - from_dt_utc).days
        if span_days > max_days:
            return None, f"⚠️ بازه انتخابی ({span_days} روز) برای این تایم‌فریم زیاد است.\nحداکثر حدود {max_days} روز پشتیبانی می‌شود."

        biquote_interval = BIQUOTE_INTERVAL_MAP.get(interval, "15m")
        symbol_map = {"GOLD": "XAUUSD", "SILVER": "XAGUSD"}
        biquote_symbol = symbol_map.get(symbol.upper(), symbol.upper())
        from_iso = from_dt_utc.strftime("%Y-%m-%dT%H:%M:%SZ")

        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.get(
                f"https://biquote.io/api/{biquote_symbol}/ohlc",
                params={"interval": biquote_interval, "limit": 2000, "from": from_iso}
            )
            if r.status_code == 404:
                return None, f"❌ نماد «{symbol}» در منبع داده پیدا نشد."
            data = r.json()
            bars = data.get("bars", [])

            candles = []
            to_ts = to_dt_utc.timestamp()
            for bar in bars:
                bar_time = datetime.strptime(bar["openTime"], "%Y-%m-%dT%H:%M:%SZ")
                if bar_time.timestamp() < to_ts and not bar.get("isOpen", False):
                    open_p, high, low, close = bar.get("open"), bar.get("high"), bar.get("low"), bar.get("close")
                    if None not in (open_p, high, low, close):
                        candles.append({
                            "time_utc": bar_time, "open": open_p, "high": high, "low": low, "close": close,
                            "tick_volume": bar.get("tickVolume", 0) or 0
                        })
            candles.sort(key=lambda c: c["time_utc"])
            if not candles:
                return None, "❌ داده‌ای برای این بازه زمانی یافت نشد."
            return candles, None
    except Exception as e:
        print(f"Historical range data error (raw): {e}")
        return None, "❌ خطا در دریافت داده‌های تاریخی."

async def get_historical_range_candles(symbol, from_date, to_date, interval):
    """نسخه سطح بالا: تاریخ‌ها به‌صورت رشته YYYY-MM-DD به وقت ایران دریافت می‌شوند"""
    try:
        from_dt_iran = datetime.strptime(from_date, "%Y-%m-%d")
        to_dt_iran = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
        from_dt_utc = from_dt_iran - IRAN_TZ_OFFSET
        to_dt_utc = to_dt_iran - IRAN_TZ_OFFSET
        return await get_historical_range_candles_raw(symbol, from_dt_utc, to_dt_utc, interval)
    except Exception as e:
        print(f"Historical range data error: {e}")
        return None, "❌ خطا در دریافت داده‌های تاریخی."

# ─── ساخت فایل اکسل نتایج ─────────────────────────────────────
def build_indicator_excel(trades, symbol, with_management):
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicator Backtest"

    if with_management:
        headers = ["تاریخ ورود", "قیمت ورود", "SL", "بالاترین R", "نتیجه (R)", "تاریخ خروج", "قیمت خروج", "وضعیت"]
    else:
        headers = ["تاریخ ورود", "قیمت ورود", "تاریخ خروج", "قیمت خروج", "سود/ضرر (%)"]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for t in trades:
        if with_management:
            ws.append([
                t["entry_time"].strftime("%Y-%m-%d %H:%M"), t["entry_price"], round(t["sl_price"], 5),
                t["max_r"], t["result_r"],
                t["exit_time_utc"].strftime("%Y-%m-%d %H:%M") if t.get("exit_time_utc") else "---",
                t["exit_price"] if t.get("exit_price") else "---",
                "بسته شده" if t["status"] == "closed" else "باز"
            ])
        else:
            ws.append([
                t["entry_time"].strftime("%Y-%m-%d %H:%M"), t["entry_price"],
                t["exit_time"].strftime("%Y-%m-%d %H:%M"), t["exit_price"], t["pnl_percent"]
            ])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ─── منوها ────────────────────────────────────────────────────
def indicator_list_menu():
    keyboard = []
    for key, info in INDICATOR_DEFS.items():
        keyboard.append([InlineKeyboardButton(info["label"], callback_data=f"ib_ind_{key}")])
    keyboard.append([InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")])
    return InlineKeyboardMarkup(keyboard)

def indicator_management_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔹 فقط سیگنال ساده", callback_data="ib_mgmt_simple"),
         InlineKeyboardButton("🔸 با مدیریت پوزیشن (R)", callback_data="ib_mgmt_full")],
    ])

def indicator_timeframe_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("1 دقیقه", callback_data="ib_tf_1m"),
         InlineKeyboardButton("15 دقیقه", callback_data="ib_tf_15m")],
        [InlineKeyboardButton("1 ساعت", callback_data="ib_tf_1h")],
    ])

def indicator_result_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 بک‌تست اندیکاتور جدید", callback_data="ind_backtest")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")],
    ])

def sl_unit_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 درصد از قیمت ورود", callback_data="ib_sl_unit_percent"),
         InlineKeyboardButton("💵 مقدار دلاری ثابت", callback_data="ib_sl_unit_dollar")],
        [InlineKeyboardButton("📏 تعداد پیپ (فقط فارکس)", callback_data="ib_sl_unit_pip")],
    ])

def ib_r_level_question_menu(r_mult):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ بله، جابجا کن", callback_data=f"ib_r_yes_{r_mult}"),
         InlineKeyboardButton("❌ نه، رد شو", callback_data=f"ib_r_no_{r_mult}")],
    ])

# ─── هندلر دکمه‌ها ────────────────────────────────────────────
async def handle_indicator_backtest(query, user_id):
    data = query.data

    if data == "ind_backtest":
        await query.edit_message_text(
            "📊 بک‌تست استراتژی اندیکاتوری\n\n"
            "یک اندیکاتور رایج انتخاب کنید تا روی داده‌ی تاریخی واقعی "
            "اجرا شود و سیگنال‌های خرید/فروش آن بررسی شود.",
            reply_markup=indicator_list_menu()
        )

    elif data.startswith("ib_ind_"):
        indicator_key = data.replace("ib_ind_", "")
        info = INDICATOR_DEFS[indicator_key]
        await query.edit_message_text(f"⏳ در حال دریافت توضیح از هوش مصنوعی درباره {info['label']}...")

        ai_prompt = (
            f"به فارسی و خیلی کوتاه (حداکثر ۳ خط) توضیح بده اندیکاتور "
            f"{info['label']} در تحلیل تکنیکال چه کاربردی دارد."
        )
        ai_text = await get_ai_response([{"role": "user", "content": ai_prompt}])
        explain_text = ai_text if ai_text else "توضیح هوش مصنوعی در دسترس نبود."

        params_text = "\n".join(f"• {p['label']}: پیش‌فرض {p['default']}" for p in info["params"])
        set_state(user_id, "ib_params_confirm", indicator_key)
        await query.message.reply_text(
            f"🤖 توضیح هوش مصنوعی:\n{explain_text}\n\n"
            f"📐 نحوه محاسبه سیگنال (دقیق):\n{info['explain']}\n\n"
            f"⚙️ پارامترهای پیش‌فرض:\n{params_text}\n\n"
            f"می‌خواهید پارامترها را تغییر دهید یا از پیش‌فرض استفاده شود؟",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⚡️ استفاده از پیش‌فرض", callback_data="ib_params_default")],
                [InlineKeyboardButton("✏️ تغییر پارامترها", callback_data="ib_params_custom")],
            ])
        )

    elif data == "ib_params_default":
        state, indicator_key = get_state(user_id)
        info = INDICATOR_DEFS[indicator_key]
        params = {p["key"]: p["default"] for p in info["params"]}
        import json
        set_state(user_id, "ib_management_choice", f"{indicator_key}|{json.dumps(params)}")
        await query.edit_message_text(
            "✅ پارامترهای پیش‌فرض انتخاب شد.\n\n"
            "می‌خواهید فقط سیگنال ساده بررسی شود یا با مدیریت پوزیشن (سیستم R)؟",
            reply_markup=indicator_management_menu()
        )

    elif data == "ib_params_custom":
        state, indicator_key = get_state(user_id)
        info = INDICATOR_DEFS[indicator_key]
        set_state(user_id, "ib_param_input_0", f"{indicator_key}|{{}}")
        first_param = info["params"][0]
        await query.edit_message_text(
            f"✏️ مقدار «{first_param['label']}» را وارد کنید (پیش‌فرض {first_param['default']}):"
        )

    elif data in ["ib_mgmt_simple", "ib_mgmt_full"]:
        state, sdata = get_state(user_id)
        indicator_key, params_str = sdata.split("|", 1)
        with_management = (data == "ib_mgmt_full")
        if with_management:
            strategies = get_strategies(user_id)
            if strategies:
                keyboard = []
                for sid, name, _ in strategies[:10]:
                    keyboard.append([InlineKeyboardButton(f"📋 {name}", callback_data=f"ib_strat_use_{sid}")])
                keyboard.append([InlineKeyboardButton("🆕 ساخت استراتژی جدید", callback_data="ib_strat_create")])
                set_state(user_id, "ib_strategy_choice", f"{indicator_key}|{params_str}")
                await query.edit_message_text(
                    "📋 کدام استراتژی مدیریت پوزیشن استفاده شود؟",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
            else:
                await start_ib_strategy_creation(query, user_id, indicator_key, params_str, is_message=False)
        else:
            set_state(user_id, "ib_symbol", f"{indicator_key}|{params_str}|none|none|none")
            await query.edit_message_text(
                "نماد ارز را وارد کنید:\nمثال: `EURUSD` `GOLD` `USDCAD`",
                parse_mode="Markdown"
            )

    elif data.startswith("ib_strat_use_"):
        sid = int(data.replace("ib_strat_use_", ""))
        name, rules = get_strategy_by_id(sid)
        state, sdata = get_state(user_id)
        indicator_key, params_str = sdata.split("|", 1)
        import json
        set_state(user_id, "ib_sl_unit_choice", f"{indicator_key}|{params_str}|{name}|{json.dumps(rules)}")
        await query.edit_message_text(
            f"✅ استراتژی «{name}» انتخاب شد.\n\n"
            f"📋 خلاصه:\n{describe_strategy(rules)}\n\n"
            f"حد ضرر خودکار هر معامله چطور تعیین شود؟",
            reply_markup=sl_unit_menu()
        )

    elif data == "ib_strat_create":
        state, sdata = get_state(user_id)
        indicator_key, params_str = sdata.split("|", 1)
        await start_ib_strategy_creation(query, user_id, indicator_key, params_str, is_message=False)

    elif data in ["ib_levels_custom", "ib_levels_default"]:
        state, sdata = get_state(user_id)
        indicator_key, params_str = sdata.split("|", 1)
        if data == "ib_levels_default":
            import json
            rules = {"__levels__": DEFAULT_R_LEVELS}
            first_level = DEFAULT_R_LEVELS[0]
            set_state(user_id, f"ib_strat_r_{first_level}", f"{indicator_key}|{params_str}|{json.dumps(rules)}")
            await query.edit_message_text(
                f"⚡️ سطوح پیش‌فرض انتخاب شد.\n\nدر سطح R{first_level}، حد ضرر را جابجا کنم؟",
                reply_markup=ib_r_level_question_menu(str(first_level))
            )
        else:
            set_state(user_id, "ib_levels_count", f"{indicator_key}|{params_str}")
            await query.edit_message_text("🎯 چند سطح ریوارد نیاز داری؟\n\nیک عدد وارد کنید (مثلاً 5):")

    elif data.startswith("ib_r_yes_") or data.startswith("ib_r_no_"):
        is_yes = data.startswith("ib_r_yes_")
        r_mult = data.replace("ib_r_yes_", "").replace("ib_r_no_", "")
        state, sdata = get_state(user_id)
        if is_yes:
            set_state(user_id, f"ib_strat_slval_{r_mult}", sdata)
            await query.edit_message_text(
                f"🎯 در سطح R{r_mult}، حد ضرر را روی چه ریواردی بگذارم؟\n\n"
                f"یک عدد بین `0` و `{r_mult}` وارد کنید.\n"
                f"برای ریسک‌فری دقیق بنویسید `0`",
                parse_mode="Markdown"
            )
        else:
            import json
            parts = sdata.split("|")
            indicator_key, params_str, rules_str = parts[0], parts[1], parts[2]
            rules = json.loads(rules_str)
            rules[r_mult] = None
            levels = get_strategy_levels(rules)
            next_idx = levels.index(float(r_mult)) + 1
            await proceed_to_next_ib_r_level(query, user_id, indicator_key, params_str, rules, next_idx)

    elif data in ["ib_sl_unit_percent", "ib_sl_unit_dollar", "ib_sl_unit_pip"]:
        state, sdata = get_state(user_id)
        unit_map = {"ib_sl_unit_percent": "percent", "ib_sl_unit_dollar": "dollar", "ib_sl_unit_pip": "pip"}
        sl_unit = unit_map[data]
        set_state(user_id, "ib_sl_value", f"{sdata}|{sl_unit}")
        unit_labels = {"percent": "درصد (مثلاً `2` یعنی ۲٪)", "dollar": "دلار (مثلاً `10`)", "pip": "پیپ (مثلاً `20`)"}
        await query.edit_message_text(
            f"مقدار حد ضرر را به {unit_labels[sl_unit]} وارد کنید:",
            parse_mode="Markdown"
        )

    elif data.startswith("ib_tf_"):
        timeframe = data.replace("ib_tf_", "")
        state, sdata = get_state(user_id)
        set_state(user_id, "ib_processing", f"{sdata}|{timeframe}")
        await query.edit_message_text("⏳ در حال دریافت داده‌های تاریخی و اجرای بک‌تست اندیکاتور...")
        await run_indicator_backtest_final(user_id, query)

async def start_ib_strategy_creation(query_or_update, user_id, indicator_key, params_str, is_message):
    text = (
        "🆕 هنوز استراتژی مدیریت پوزیشنی نساخته‌اید (یا می‌خواهید یکی جدید بسازید).\n\n"
        "می‌خواهید چند سطح ریوارد (R) خودتان تعیین کنید یا از سطوح "
        f"پیش‌فرض ({', '.join('R'+str(r) for r in DEFAULT_R_LEVELS)}) استفاده کنید؟"
    )
    markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("🎯 سطوح دلخواه خودم", callback_data="ib_levels_custom")],
        [InlineKeyboardButton("⚡️ همان ۸ سطح پیش‌فرض", callback_data="ib_levels_default")],
    ])
    set_state(user_id, "ib_strat_levels_choice", f"{indicator_key}|{params_str}")
    if is_message:
        await query_or_update.message.reply_text(text, reply_markup=markup)
    else:
        await query_or_update.edit_message_text(text, reply_markup=markup)

async def proceed_to_next_ib_r_level(query_or_update, user_id, indicator_key, params_str, rules, next_idx, is_message=False):
    import json
    levels = get_strategy_levels(rules)
    if next_idx < len(levels) - 1:
        next_r = levels[next_idx]
        set_state(user_id, f"ib_strat_r_{next_r}", f"{indicator_key}|{params_str}|{json.dumps(rules)}")
        text = f"در سطح R{next_r}، حد ضرر را جابجا کنم؟"
        markup = ib_r_level_question_menu(str(next_r))
        if is_message:
            await query_or_update.message.reply_text(text, reply_markup=markup)
        else:
            await query_or_update.edit_message_text(text, reply_markup=markup)
    else:
        set_state(user_id, "ib_strat_name", f"{indicator_key}|{params_str}|{json.dumps(rules)}")
        text = "✅ تمام سطوح تنظیم شد!\n\nاسمی برای این استراتژی وارد کنید (مثال: استراتژی اصلی):"
        if is_message:
            await query_or_update.message.reply_text(text)
        else:
            await query_or_update.edit_message_text(text)

async def handle_indicator_backtest_message(user_id, text, update):
    state, data = get_state(user_id)
    text = text.strip()
    import json

    # ─── دریافت پارامترهای سفارشی یکی‌یکی ───
    if state and state.startswith("ib_param_input_"):
        idx = int(state.replace("ib_param_input_", ""))
        indicator_key, params_json = data.split("|", 1)
        params = json.loads(params_json) if params_json != "{}" else {}
        info = INDICATOR_DEFS[indicator_key]
        param_def = info["params"][idx]
        try:
            value = float(text)
            if value == int(value):
                value = int(value)
        except:
            await update.message.reply_text(f"❌ لطفاً یک عدد معتبر برای «{param_def['label']}» وارد کنید:")
            return True

        params[param_def["key"]] = value
        next_idx = idx + 1
        if next_idx < len(info["params"]):
            next_param = info["params"][next_idx]
            set_state(user_id, f"ib_param_input_{next_idx}", f"{indicator_key}|{json.dumps(params)}")
            await update.message.reply_text(
                f"مقدار «{next_param['label']}» را وارد کنید (پیش‌فرض {next_param['default']}):"
            )
        else:
            set_state(user_id, "ib_management_choice", f"{indicator_key}|{json.dumps(params)}")
            params_text = "\n".join(f"• {info['params'][i]['label']}: {v}" for i, v in enumerate(params.values()))
            await update.message.reply_text(
                f"✅ پارامترها ثبت شد:\n{params_text}\n\n"
                f"می‌خواهید فقط سیگنال ساده بررسی شود یا با مدیریت پوزیشن (سیستم R)؟",
                reply_markup=indicator_management_menu()
            )
        return True

    # ─── تعداد سطوح دلخواه ───
    elif state == "ib_levels_count":
        try:
            count = int(text)
            if count < 2 or count > 15:
                await update.message.reply_text("❌ تعداد سطوح باید بین ۲ و ۱۵ باشد.\nدوباره وارد کنید:")
                return True
            indicator_key, params_str = data.split("|", 1)
            set_state(user_id, "ib_levels_input", f"{indicator_key}|{params_str}|{count}")
            await update.message.reply_text(
                f"🎯 حالا {count} سطح ریوارد را با کاما وارد کنید (به ترتیب صعودی):\n\n"
                f"مثال: `0.5, 1, 2, 3.5, 5`",
                parse_mode="Markdown"
            )
        except:
            await update.message.reply_text("❌ لطفاً یک عدد صحیح وارد کنید:")
        return True

    # ─── مقادیر سطوح دلخواه ───
    elif state == "ib_levels_input":
        indicator_key, params_str, count_str = data.split("|")
        count = int(count_str)
        try:
            parts = [p.strip() for p in text.split(",")]
            levels = [float(p) for p in parts]
        except:
            await update.message.reply_text("❌ فرمت اشتباه است.\nمثال: `0.5, 1, 2, 3.5, 5`", parse_mode="Markdown")
            return True
        if len(levels) != count:
            await update.message.reply_text(f"❌ شما گفتید {count} سطح، ولی {len(levels)} عدد وارد کردید.\nدوباره وارد کنید:")
            return True
        if any(l <= 0 for l in levels) or levels != sorted(levels) or len(set(levels)) != len(levels):
            await update.message.reply_text("❌ سطوح باید مثبت، صعودی و بدون تکرار باشند.\nدوباره وارد کنید:")
            return True

        rules = {"__levels__": levels}
        first_level = levels[0]
        set_state(user_id, f"ib_strat_r_{first_level}", f"{indicator_key}|{params_str}|{json.dumps(rules)}")
        await update.message.reply_text(
            f"✅ سطوح ثبت شد: {', '.join('R'+str(l) for l in levels)}\n\n"
            f"در سطح R{first_level}، حد ضرر را جابجا کنم؟",
            reply_markup=ib_r_level_question_menu(str(first_level))
        )
        return True

    # ─── مقدار SL نسبی برای یک سطح خاص ───
    elif state and state.startswith("ib_strat_slval_"):
        r_mult = state.replace("ib_strat_slval_", "")
        parts = data.split("|")
        indicator_key, params_str, rules_str = parts[0], parts[1], parts[2]
        rules = json.loads(rules_str)
        try:
            r_value = float(text)
        except:
            await update.message.reply_text(f"❌ لطفاً یک عدد معتبر بین `0` و `{r_mult}` وارد کنید:", parse_mode="Markdown")
            return True
        current_r = float(r_mult)
        if r_value < 0 or r_value >= current_r:
            await update.message.reply_text(
                f"❌ عدد باید بین `0` و `{r_mult}` باشد.\nدوباره وارد کنید:", parse_mode="Markdown"
            )
            return True
        rules[r_mult] = {"type": "r_trail", "value": r_value}
        levels = get_strategy_levels(rules)
        next_idx = levels.index(float(r_mult)) + 1
        await proceed_to_next_ib_r_level(update, user_id, indicator_key, params_str, rules, next_idx, is_message=True)
        return True

    # ─── ثبت نام استراتژی ───
    elif state == "ib_strat_name":
        parts = data.split("|")
        indicator_key, params_str, rules_str = parts[0], parts[1], parts[2]
        rules = json.loads(rules_str)
        strategy_name = text
        save_strategy(user_id, strategy_name, rules)
        set_state(user_id, "ib_sl_unit_choice", f"{indicator_key}|{params_str}|{strategy_name}|{json.dumps(rules)}")
        await update.message.reply_text(
            f"✅ استراتژی «{strategy_name}» ذخیره شد.\n\n"
            f"📋 خلاصه:\n{describe_strategy(rules)}\n\n"
            f"حد ضرر خودکار هر معامله چطور تعیین شود؟",
            reply_markup=sl_unit_menu()
        )
        return True

    # ─── درصد حد ضرر خودکار ───
    elif state == "ib_sl_value":
        try:
            sl_value = float(text)
            if sl_value <= 0:
                await update.message.reply_text("❌ مقدار باید بزرگ‌تر از صفر باشد:")
                return True
            parts = data.split("|", 4)
            indicator_key, params_str, strategy_name, rules_str, sl_unit = parts
            sl_config = json.dumps({"unit": sl_unit, "value": sl_value})
            set_state(user_id, "ib_symbol", f"{indicator_key}|{params_str}|{strategy_name}|{rules_str}|{sl_config}")
            await update.message.reply_text(
                "نماد ارز را وارد کنید:\nمثال: `EURUSD` `GOLD` `USDCAD`",
                parse_mode="Markdown"
            )
        except:
            await update.message.reply_text("❌ لطفاً یک عدد وارد کنید:")
        return True

    # ─── نماد ───
    elif state == "ib_symbol":
        symbol = text.upper()
        set_state(user_id, "ib_from_date", f"{data}|{symbol}")
        await update.message.reply_text(
            "تاریخ شروع بک‌تست را وارد کنید (فرمت: YYYY-MM-DD)\nمثال: `2026-06-01`",
            parse_mode="Markdown"
        )
        return True

    elif state == "ib_from_date":
        try:
            datetime.strptime(text, "%Y-%m-%d")
            set_state(user_id, "ib_to_date", f"{data}|{text}")
            await update.message.reply_text("تاریخ پایان بک‌تست را وارد کنید (فرمت: YYYY-MM-DD)")
        except:
            await update.message.reply_text("❌ فرمت تاریخ اشتباه است.\nمثال: `2026-06-01`", parse_mode="Markdown")
        return True

    elif state == "ib_to_date":
        try:
            datetime.strptime(text, "%Y-%m-%d")
            set_state(user_id, "ib_timeframe", f"{data}|{text}")
            await update.message.reply_text("⏱ تایم‌فریم را انتخاب کنید:", reply_markup=indicator_timeframe_menu())
        except:
            await update.message.reply_text("❌ فرمت تاریخ اشتباه است.\nمثال: `2026-06-20`", parse_mode="Markdown")
        return True

    return False

async def run_indicator_backtest_final(user_id, query):
    state, sdata = get_state(user_id)
    if state != "ib_processing" or not sdata:
        return

    import json
    parts = sdata.split("|")
    indicator_key = parts[0]
    params = json.loads(parts[1])
    with_management = parts[2] != "none"

    if with_management:
        strategy_name = parts[2]
        rules = json.loads(parts[3])
        sl_config = json.loads(parts[4])
        symbol = parts[5]
        from_date = parts[6]
        to_date = parts[7]
        timeframe = parts[8]
    else:
        symbol = parts[5]
        from_date = parts[6]
        to_date = parts[7]
        timeframe = parts[8]

    candles, error = await get_historical_range_candles(symbol, from_date, to_date, timeframe)
    if error:
        await query.message.reply_text(error, reply_markup=indicator_result_menu())
        clear_state(user_id)
        return

    is_custom = indicator_key.startswith("custom_")
    if is_custom:
        await query.message.reply_text("⏳ در حال اجرای کد اندیکاتور شخصی (ممکن است کمی طول بکشد)...")
        signals, exec_error = await run_custom_indicator_signals(params["code"], params["function_name"], candles)
        if exec_error:
            await query.message.reply_text(
                f"❌ خطا در اجرای اندیکاتور شخصی:\n{exec_error}",
                reply_markup=indicator_result_menu()
            )
            clear_state(user_id)
            return
        indicator_label = f"📶 {params.get('indicator_name', 'اندیکاتور شخصی')}"
    else:
        signals = await generate_signals(indicator_key, candles, params, symbol)
        indicator_label = INDICATOR_DEFS[indicator_key]["label"]

    if with_management:
        trades = run_signal_backtest_with_management(candles, signals, sl_config, symbol, rules)
    else:
        trades = run_signal_backtest_simple(candles, signals)

    if not trades:
        await query.message.reply_text(
            f"📊 نتیجه بک‌تست اندیکاتور — {symbol}\n\n"
            f"اندیکاتور: {indicator_label}\n"
            f"بازه: {from_date} تا {to_date}\n\n"
            f"❌ هیچ معامله‌ای در این بازه توسط این اندیکاتور سیگنال داده نشد.",
            reply_markup=indicator_result_menu()
        )
        clear_state(user_id)
        return

    if with_management:
        wins = [t for t in trades if t["result_r"] > 0]
        total_r = sum(t["result_r"] for t in trades)
        summary = (
            f"تعداد معاملات: {len(trades)}\n"
            f"سودده: {len(wins)} | ضررده: {len(trades)-len(wins)}\n"
            f"درصد برد: {len(wins)/len(trades)*100:.1f}%\n"
            f"مجموع نتیجه: {total_r:+.2f}R"
        )
        method_line = f"با مدیریت پوزیشن ({strategy_name})"
    else:
        wins = [t for t in trades if t["pnl_percent"] > 0]
        total_pct = sum(t["pnl_percent"] for t in trades)
        summary = (
            f"تعداد معاملات: {len(trades)}\n"
            f"سودده: {len(wins)} | ضررده: {len(trades)-len(wins)}\n"
            f"درصد برد: {len(wins)/len(trades)*100:.1f}%\n"
            f"مجموع سود/ضرر: {total_pct:+.2f}%"
        )
        method_line = "سیگنال ساده"

    text = (
        f"📊 نتیجه بک‌تست اندیکاتور — {symbol}\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
        f"اندیکاتور: {indicator_label}\n"
        f"بازه: {from_date} تا {to_date}\n"
        f"تایم‌فریم: {timeframe}\n"
        f"روش: {method_line}\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`\n"
        f"{summary}\n"
        f"`━━━━━━━━━━━━━━━━━━━━━━━━━━`"
    )
    await query.message.reply_text(text, parse_mode="Markdown", reply_markup=indicator_result_menu())

    excel_buffer = build_indicator_excel(trades, symbol, with_management)
    await query.message.reply_document(
        document=excel_buffer, filename=f"indicator_backtest_{symbol}.xlsx",
        caption="📎 فایل اکسل نتایج بک‌تست اندیکاتوری"
    )
    clear_state(user_id)

# ═══════════════ پایان MODULE: INDICATOR BACKTEST ═══════════════



# ╔══════════════════════════════════════════════════════════════╗
# ║              MODULE: CUSTOM INDICATOR                        ║
# ║  اندیکاتور شخصی -- کاربر کد پایتون می‌دهد یا با توضیح متنی    ║
# ║  از هوش مصنوعی می‌خواهد کد بسازد؛ کد قبل از اجرا از یک        ║
# ║  اسکنر امنیتی رد می‌شود و در یک PROCESS کاملاً ایزوله (با     ║
# ║  timeout واقعی و kill اجباری) اجرا می‌شود -- نه thread، چون   ║
# ║  thread واقعاً قابل kill کردن نیست و ریسک DoS دارد.           ║
# ║                                                                ║
# ║  ⚠️ توجه صادقانه: این یک محیط "ریسک‌کاهش‌یافته" است، نه یک    ║
# ║  sandbox با امنیت ۱۰۰٪ تضمین‌شده (که در پایتون خالص عملاً     ║
# ║  وجود ندارد بدون container/VM جدا). برای همین، کپی هر کد      ║
# ║  (از کاربر یا تولید AI) همیشه برای ادمین ارسال می‌شود.        ║
# ║                                                                ║
# ║  برای حذف: این بلوک رو پاک کن                                ║
# ║  + خط ci_ رو از ROUTER اصلی پاک کن                            ║
# ║  + state های ci_ رو از message_handler پاک کن                 ║
# ║  + MessageHandler filters.Document را پاک کن (اگر جای دیگر    ║
# ║    استفاده نمی‌شود)                                            ║
# ╚══════════════════════════════════════════════════════════════╝

CUSTOM_INDICATOR_ALLOWED_MODULES = {"math", "statistics", "datetime"}

CUSTOM_INDICATOR_FORBIDDEN_NAMES = {
    "os", "sys", "subprocess", "socket", "shutil", "importlib",
    "eval", "exec", "compile", "__import__", "open", "input",
    "globals", "locals", "vars", "dir", "getattr", "setattr", "delattr",
    "__builtins__", "__loader__", "__spec__", "pty", "ctypes",
    "multiprocessing", "threading", "asyncio", "requests", "httpx",
    "urllib", "ftplib", "telnetlib", "pickle", "marshal", "shelve",
    "file", "exit", "quit", "help", "breakpoint", "memoryview",
}

CUSTOM_INDICATOR_FORBIDDEN_ATTRS = {
    "__class__", "__bases__", "__subclasses__", "__globals__",
    "__code__", "__closure__", "__mro__", "__dict__", "__builtins__",
    "__getattribute__", "__reduce__", "__reduce_ex__",
}

CUSTOM_INDICATOR_SAFE_BUILTINS_NAMES = [
    "len", "range", "abs", "min", "max", "sum", "round", "sorted",
    "enumerate", "zip", "list", "dict", "tuple", "set", "int", "float",
    "str", "bool", "isinstance", "map", "filter", "reversed", "any", "all",
]

def scan_custom_code_safety(source_code):
    """
    اسکن AST کد پایتون قبل از اجرا (لایه اول دفاعی).
    خروجی: (is_safe: bool, reason: str یا None)
    """
    try:
        tree = ast.parse(source_code)
    except SyntaxError as e:
        return False, f"خطای نحوی در کد: {e}"

    for node in ast.walk(tree):
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            names = [alias.name for alias in node.names]
            for name in names:
                root_module = name.split(".")[0]
                if root_module not in CUSTOM_INDICATOR_ALLOWED_MODULES:
                    return False, f"ایمپورت غیرمجاز: {name} (فقط math, statistics, datetime مجاز است)"
        elif isinstance(node, ast.Name):
            if node.id in CUSTOM_INDICATOR_FORBIDDEN_NAMES:
                return False, f"استفاده از نام غیرمجاز: {node.id}"
        elif isinstance(node, ast.Attribute):
            if node.attr in CUSTOM_INDICATOR_FORBIDDEN_ATTRS:
                return False, f"دسترسی به attribute غیرمجاز: {node.attr}"
        elif isinstance(node, ast.Call):
            if isinstance(node.func, ast.Name) and node.func.id in ("eval", "exec", "compile", "__import__"):
                return False, f"فراخوانی تابع غیرمجاز: {node.func.id}"

    return True, None

def _custom_indicator_exec_worker(code, function_name, candles_data, result_queue):
    """
    این تابع داخل یک PROCESS کاملاً جدا اجرا می‌شود (نه در پروسه اصلی ربات).
    فقط به یک لیست candles (dict های ساده) و ماژول‌های ریاضی مجاز دسترسی دارد.
    """
    try:
        import builtins
        import math
        import statistics
        from datetime import datetime as dt_class

        safe_builtins = {}
        for name in CUSTOM_INDICATOR_SAFE_BUILTINS_NAMES:
            if hasattr(builtins, name):
                safe_builtins[name] = getattr(builtins, name)
        safe_builtins["True"] = True
        safe_builtins["False"] = False
        safe_builtins["None"] = None

        restricted_globals = {
            "__builtins__": safe_builtins,
            "math": math,
            "statistics": statistics,
            "datetime": dt_class,
        }
        exec(code, restricted_globals)
        func = restricted_globals.get(function_name)
        if not func:
            result_queue.put(("error", f"تابعی به اسم {function_name} در کد پیدا نشد"))
            return
        result = func(candles_data)
        if result not in ("buy", "sell", None):
            result_queue.put(("error", f"خروجی تابع باید 'buy'، 'sell' یا None باشد (دریافت‌شده: {result!r})"))
            return
        result_queue.put(("ok", result))
    except Exception as e:
        result_queue.put(("error", f"خطا در اجرای کد: {e}"))

async def run_custom_code_isolated(code, function_name, candles_data, timeout_seconds=5):
    """
    اجرای کد کاربر در یک PROCESS کاملاً ایزوله با timeout واقعی.
    اگر کد بیش از حد طول بکشد (مثل حلقه بی‌نهایت)، پروسه به‌صورت اجباری kill می‌شود
    -- این برخلاف thread، واقعاً کار می‌کند چون پروسه جدا از GIL اصلی است.
    """
    ctx = multiprocessing.get_context("spawn")
    result_queue = ctx.Queue()
    process = ctx.Process(target=_custom_indicator_exec_worker, args=(code, function_name, candles_data, result_queue))
    process.start()

    waited = 0.0
    step = 0.1
    while process.is_alive() and waited < timeout_seconds:
        await asyncio.sleep(step)
        waited += step

    if process.is_alive():
        process.terminate()
        await asyncio.sleep(0.3)
        if process.is_alive():
            process.kill()
        return None, "⏱ اجرای کد بیش از حد مجاز طول کشید و متوقف شد (احتمالاً حلقه بی‌پایان)."

    if not result_queue.empty():
        status, value = result_queue.get()
        if status == "ok":
            return value, None
        return None, value
    return None, "کد هیچ نتیجه‌ای برنگرداند."

async def run_custom_indicator_signals(code, function_name, candles):
    """
    برای هر نقطه از کندل‌ها (به‌تدریج، فقط با اطلاعات تا آن لحظه -- نه آینده)
    کد کاربر را اجرا می‌کند و سیگنال دریافت می‌کند.
    برای جلوگیری از اجرای بیش‌ازحد کند (هر کندل یک اجرای process جدا کند است)،
    یک سقف تعداد کل اجرا محدود می‌شود.
    """
    signals = [None] * len(candles)
    MAX_EXECUTIONS = 300
    step = max(1, len(candles) // MAX_EXECUTIONS)

    for i in range(1, len(candles), step):
        window = candles[:i + 1]
        candles_data = [
            {"open": c.get("open"), "high": c["high"], "low": c["low"], "close": c["close"],
             "time": c["time_utc"].strftime("%Y-%m-%d %H:%M")}
            for c in window
        ]
        result, error = await run_custom_code_isolated(code, function_name, candles_data, timeout_seconds=3)
        if error:
            return signals, error
        if result in ("buy", "sell"):
            signals[i] = result

    return signals, None

# ─── تشخیص نام تابع سیگنال با AI (وقتی کاربر کد دلخواه می‌فرستد) ──
async def ai_detect_signal_function(code):
    """
    از AI می‌خواهد یک تابع wrapper به نام generate_signal(candles) به انتهای
    کد کاربر اضافه کند که ورودی/خروجی استاندارد بدهد -- کد اصلی دست‌نخورده می‌ماند.
    """
    prompt = (
        "کد پایتون زیر را بررسی کن. باید یک تابع wrapper به نام "
        "generate_signal(candles) به انتهای این کد اضافه کنی که:\n"
        "- ورودی candles یک لیست دیکشنری با کلیدهای open/high/low/close/time است\n"
        "- خروجی باید دقیقاً یکی از سه مقدار باشد: 'buy' یا 'sell' یا None\n"
        "- باید از منطق موجود در کد اصلی کاربر (که پایین آمده) برای تولید سیگنال استفاده کند\n"
        "- کد اصلی کاربر را دست‌نخورده نگه دار، فقط تابع generate_signal را در انتهای آن اضافه کن\n"
        "- فقط و فقط کد پایتون کامل نهایی را برگردان، بدون هیچ توضیح یا متن اضافه، بدون ```\n\n"
        f"کد کاربر:\n{code}"
    )
    result = await get_ai_response([{"role": "user", "content": prompt}])
    if not result:
        return None
    cleaned = result.strip()
    if cleaned.startswith("```"):
        lines = cleaned.split("\n")
        cleaned = "\n".join(lines[1:-1]) if len(lines) > 2 else cleaned
    return cleaned

async def ai_generate_indicator_code(description):
    """تولید کد پایتون از توضیح متنی آزاد کاربر، طبق قوانین امنیتی مشخص"""
    prompt = (
        "یک اندیکاتور تحلیل تکنیکال به زبان پایتون بنویس که این توضیح را پیاده‌سازی کند:\n"
        f"«{description}»\n\n"
        "قوانین اجباری:\n"
        "- فقط یک تابع به نام generate_signal(candles) بنویس\n"
        "- candles یک لیست دیکشنری با کلیدهای open/high/low/close/time است\n"
        "- خروجی باید دقیقاً 'buy' یا 'sell' یا None باشد\n"
        "- فقط از ماژول‌های math, statistics, datetime می‌توانی import کنی (اگر لازم شد)\n"
        "- هیچ import دیگری، هیچ دسترسی به فایل/شبکه/سیستم نداشته باش\n"
        "- فقط و فقط کد پایتون کامل را برگردان، بدون هیچ توضیح یا متن اضافه، بدون ```\n"
    )
    result = await get_ai_response([{"role": "user", "content": prompt}])
    if not result:
        return None
    cleaned = result.strip()
    if cleaned.startswith("```"):
        lines = cleaned.split("\n")
        cleaned = "\n".join(lines[1:-1]) if len(lines) > 2 else cleaned
    return cleaned

# ─── دیتابیس اندیکاتورهای شخصی هر کاربر ─────────────────────────
def init_custom_indicator_db():
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS custom_indicators (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id    INTEGER,
        name       TEXT,
        code       TEXT,
        source     TEXT,
        created_at TEXT
    )''')
    conn.commit()
    conn.close()

def save_custom_indicator(user_id, name, code, source):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("INSERT INTO custom_indicators (user_id, name, code, source, created_at) VALUES (?, ?, ?, ?, ?)",
              (user_id, name, code, source, get_iran_now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()

def get_custom_indicators(user_id):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT id, name FROM custom_indicators WHERE user_id=? ORDER BY id DESC", (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_custom_indicator_by_id(cid):
    conn = sqlite3.connect("bot.db")
    c = conn.cursor()
    c.execute("SELECT name, code FROM custom_indicators WHERE id=?", (cid,))
    row = c.fetchone()
    conn.close()
    return row if row else (None, None)

# ─── منوها ────────────────────────────────────────────────────
def custom_indicator_entry_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📄 کد پایتون دارم", callback_data="ci_have_code"),
         InlineKeyboardButton("💬 کد ندارم، توضیح می‌دهم", callback_data="ci_no_code")],
        [InlineKeyboardButton("📋 اندیکاتورهای شخصی من", callback_data="ci_my_list")],
        [InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")],
    ])

def custom_indicator_list_menu(user_id):
    items = get_custom_indicators(user_id)
    keyboard = [[InlineKeyboardButton(f"📶 {name}", callback_data=f"ci_use_{cid}")] for cid, name in items]
    keyboard.append([InlineKeyboardButton("➕ ساخت اندیکاتور جدید", callback_data="ind_custom")])
    keyboard.append([InlineKeyboardButton("🔙 برگشت", callback_data="back_financial")])
    return InlineKeyboardMarkup(keyboard)

# ─── هندلر دکمه‌ها ────────────────────────────────────────────
async def handle_custom_indicator(query, user_id):
    data = query.data

    if data == "ind_custom":
        await query.edit_message_text(
            "📶 اندیکاتور شخصی\n\n"
            "می‌توانید کد پایتون خودتان را بدهید یا با توضیح متنی از هوش مصنوعی "
            "بخواهید کد بسازد.\n\n"
            "⚠️ نکته امنیتی: هر کد (از شما یا تولید AI) قبل از اجرا بررسی امنیتی "
            "می‌شود و کپی آن برای ادمین ارسال می‌شود.",
            reply_markup=custom_indicator_entry_menu()
        )

    elif data == "ci_have_code":
        set_state(user_id, "ci_awaiting_file")
        await query.edit_message_text(
            "📄 فایل کد پایتون خود را ارسال کنید (فرمت .py).\n\n"
            "کد باید یک تابع داشته باشد که بر اساس کندل‌های قیمتی سیگنال "
            "خرید/فروش تولید کند -- هوش مصنوعی خودش تابع مناسب را پیدا می‌کند.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="ind_custom")]])
        )

    elif data == "ci_no_code":
        set_state(user_id, "ci_awaiting_description")
        await query.edit_message_text(
            "💬 توضیح دهید چه اندیکاتور یا منطقی می‌خواهید.\n\n"
            "مثال: «وقتی قیمت ۳ کندل پشت‌سرهم صعودی باشد و حجم افزایش یابد، سیگنال خرید بده»",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="ind_custom")]])
        )

    elif data == "ci_my_list":
        items = get_custom_indicators(user_id)
        if not items:
            await query.answer("هنوز هیچ اندیکاتور شخصی نساخته‌اید.", show_alert=True)
            return
        await query.edit_message_text("📋 اندیکاتورهای شخصی شما:", reply_markup=custom_indicator_list_menu(user_id))

    elif data.startswith("ci_use_"):
        cid = int(data.replace("ci_use_", ""))
        name, code = get_custom_indicator_by_id(cid)
        if not code:
            await query.answer("این اندیکاتور پیدا نشد.", show_alert=True)
            return
        import json
        params = {"code": code, "function_name": "generate_signal", "indicator_name": name}
        set_state(user_id, "ib_management_choice", f"custom_{cid}|{json.dumps(params)}")
        await query.edit_message_text(
            f"✅ اندیکاتور «{name}» انتخاب شد.\n\n"
            f"می‌خواهید فقط سیگنال ساده بررسی شود یا با مدیریت پوزیشن (سیستم R)؟",
            reply_markup=indicator_management_menu()
        )

    elif data == "ci_confirm_save":
        state, sdata = get_state(user_id)
        if state != "ci_confirm":
            return
        import json
        payload = json.loads(sdata)
        set_state(user_id, "ci_naming", json.dumps(payload))
        await query.edit_message_text("✅ کد تأیید شد.\n\nاسمی برای این اندیکاتور شخصی انتخاب کنید:")

    elif data == "ci_reject":
        clear_state(user_id)
        await query.edit_message_text("❌ لغو شد.", reply_markup=custom_indicator_entry_menu())

async def handle_custom_indicator_message(user_id, text, update):
    import json
    state, data = get_state(user_id)

    if state == "ci_awaiting_description":
        description = text.strip()
        await update.message.reply_text("⏳ در حال تولید کد توسط هوش مصنوعی...")
        code = await ai_generate_indicator_code(description)
        if not code:
            await update.message.reply_text(
                "❌ هوش مصنوعی موقتاً در دسترس نیست. کمی بعد دوباره امتحان کنید.",
                reply_markup=custom_indicator_entry_menu()
            )
            clear_state(user_id)
            return True

        is_safe, reason = scan_custom_code_safety(code)
        if not is_safe:
            await update.message.reply_text(
                f"❌ کد تولیدشده توسط هوش مصنوعی از فیلتر امنیتی رد نشد:\n{reason}\n\n"
                f"لطفاً توضیح خود را واضح‌تر یا ساده‌تر بنویسید و دوباره امتحان کنید.",
                reply_markup=custom_indicator_entry_menu()
            )
            clear_state(user_id)
            return True

        try:
            code_bytes = code.encode("utf-8")
            for aid in get_all_admin_ids():
                await update.get_bot().send_document(
                    aid,
                    document=io.BytesIO(code_bytes),
                    filename=f"ai_generated_indicator_{user_id}.py",
                    caption=f"🤖 کد اندیکاتور تولیدشده توسط AI\n👤 کاربر: {user_id}\n📝 توضیح: {description}"
                )
        except Exception as e:
            print(f"خطا در ارسال کد به ادمین: {e}")

        payload = {"code": code, "function_name": "generate_signal", "source": "ai_generated"}
        set_state(user_id, "ci_confirm", json.dumps(payload))
        await update.message.reply_text(
            f"✅ کد تولید و از فیلتر امنیتی عبور کرد.\n\n"
            f"```python\n{code[:1500]}\n```\n\n"
            f"می‌خواهید این اندیکاتور ذخیره و استفاده شود؟",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ تأیید و ذخیره", callback_data="ci_confirm_save")],
                [InlineKeyboardButton("❌ لغو", callback_data="ci_reject")],
            ])
        )
        return True

    elif state == "ci_naming":
        payload = json.loads(data)
        name = text.strip()
        save_custom_indicator(user_id, name, payload["code"], payload.get("source", "unknown"))
        await update.message.reply_text(
            f"✅ اندیکاتور «{name}» ذخیره شد و به لیست اندیکاتورهای شخصی شما اضافه شد.",
            reply_markup=custom_indicator_entry_menu()
        )
        clear_state(user_id)
        return True

    return False

async def handle_custom_indicator_document(update, context):
    """دریافت فایل .py از کاربر وقتی در state انتظار فایل است"""
    user_id = update.effective_user.id
    state, data = get_state(user_id)
    if state != "ci_awaiting_file":
        return False

    doc = update.message.document
    if not doc or not doc.file_name.endswith(".py"):
        await update.message.reply_text("❌ لطفاً یک فایل با پسوند .py ارسال کنید.")
        return True

    file = await context.bot.get_file(doc.file_id)
    file_bytes = await file.download_as_bytearray()
    try:
        code = file_bytes.decode("utf-8")
    except Exception:
        await update.message.reply_text("❌ فایل قابل خوانده شدن نیست (باید متن ساده UTF-8 باشد).")
        return True

    try:
        for aid in get_all_admin_ids():
            await context.bot.send_document(
                aid,
                document=io.BytesIO(bytes(file_bytes)),
                filename=f"user_indicator_{user_id}_{doc.file_name}",
                caption=f"📄 کد اندیکاتور شخصی ارسال‌شده توسط کاربر\n👤 کاربر: {user_id}"
            )
    except Exception as e:
        print(f"خطا در ارسال فایل کاربر به ادمین: {e}")

    is_safe, reason = scan_custom_code_safety(code)
    if not is_safe:
        await update.message.reply_text(
            f"❌ کد شما از فیلتر امنیتی رد نشد:\n{reason}\n\n"
            f"لطفاً کد را اصلاح کنید (فقط math, statistics, datetime قابل import هستند؛ "
            f"دسترسی به فایل، شبکه، یا سیستم مجاز نیست).",
            reply_markup=custom_indicator_entry_menu()
        )
        clear_state(user_id)
        return True

    await update.message.reply_text("⏳ در حال بررسی کد توسط هوش مصنوعی برای تشخیص تابع سیگنال...")
    adapted_code = await ai_detect_signal_function(code)
    if not adapted_code:
        await update.message.reply_text(
            "❌ هوش مصنوعی موقتاً در دسترس نیست. کمی بعد دوباره امتحان کنید.",
            reply_markup=custom_indicator_entry_menu()
        )
        clear_state(user_id)
        return True

    is_safe2, reason2 = scan_custom_code_safety(adapted_code)
    if not is_safe2:
        await update.message.reply_text(
            f"❌ کد نهایی (بعد از افزودن adapter) از فیلتر امنیتی رد نشد:\n{reason2}\n\n"
            f"این یعنی کد اصلی شما شامل بخش‌های غیرمجاز است.",
            reply_markup=custom_indicator_entry_menu()
        )
        clear_state(user_id)
        return True

    import json
    payload = {"code": adapted_code, "function_name": "generate_signal", "source": "user_uploaded"}
    set_state(user_id, "ci_confirm", json.dumps(payload))
    await update.message.reply_text(
        f"✅ کد شما بررسی و آماده شد (تابع generate_signal به آن اضافه شد).\n\n"
        f"می‌خواهید این اندیکاتور ذخیره و استفاده شود؟",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ تأیید و ذخیره", callback_data="ci_confirm_save")],
            [InlineKeyboardButton("❌ لغو", callback_data="ci_reject")],
        ])
    )
    return True

# ═══════════════ پایان MODULE: CUSTOM INDICATOR ═════════════════







# ╔══════════════════════════════════════════════════════════════╗
# ║                      MAIN MENU                               ║
# ║                                                              ║
# ║  برای اضافه کردن دکمه: یه خط InlineKeyboardButton اضافه کن ║
# ║  برای حذف دکمه: اون خط رو پاک کن                           ║
# ╚══════════════════════════════════════════════════════════════╝

DEFAULT_HELP_TEXT = (
    "❓ راهنمای ربات\n\n"
    "💰 بازارهای مالی: قیمت لحظه‌ای طلا/ارز/کریپتو، واچ‌لیست، هشدار قیمت، "
    "تحلیل هوش مصنوعی، بک‌تست استراتژی، ربات مدیریت سرمایه (نسخه‌ی تست و لایسنس)\n\n"
    "🎮 سرگرمی: پیگیری سلامت و ورزش خانواده، جستجوی موزیک\n\n"
    "👶 فرزندان: قصه‌گویی هوشمند، ثبت و یادآوری تکالیف\n\n"
    "⏰ یادآور و تسک: یادآور‌های یک‌باره یا تکرارشونده\n\n"
    "🛠 ابزارهای کاربردی: دستیار هوش مصنوعی، فیلترشکن\n\n"
    "⚙️ تنظیمات: پروفایل، تکمیل اطلاعات، پیام به ادمین\n\n"
    "سوالی داشتید، از ⚙️ تنظیمات → 📩 پیام به ادمین استفاده کنید."
)

def main_menu(user_id=None):
    rows = [
        [InlineKeyboardButton("💰 بازارهای مالی", callback_data="menu_financial"),      # ← حذف = پاک کن این خط
         InlineKeyboardButton("🎮 سرگرمی", callback_data="menu_fun")],                  # ← حذف = پاک کن این خط
        [InlineKeyboardButton("👶 فرزندان", callback_data="menu_kids"),                 # ← حذف = پاک کن این خط
         InlineKeyboardButton("⏰ یادآور و تسک", callback_data="menu_reminder")],       # ← حذف = پاک کن این خط
        [InlineKeyboardButton("🛠 ابزارهای کاربردی", callback_data="menu_tools"),       # ← حذف ماژول ابزارها = پاک کن این خط
         InlineKeyboardButton("⚙️ تنظیمات", callback_data="menu_settings")],            # ← حذف = پاک کن این خط
        [InlineKeyboardButton("❓ راهنما", callback_data="menu_help")],                 # ← حذف راهنما = پاک کن این خط
    ]
    if is_admin(user_id):
        rows.append([InlineKeyboardButton("🛠 پنل مدیریت", callback_data="admin_panel")])  # ← حذف ماژول پنل ادمین = پاک کن این خط
    return InlineKeyboardMarkup(rows)


# ╔══════════════════════════════════════════════════════════════╗
# ║                       ROUTER                                 ║
# ║                                                              ║
# ║  هر دکمه به ماژول مربوطه هدایت میشه                        ║
# ║  برای حذف ماژول: خط مربوطه رو پاک کن                       ║
# ╚══════════════════════════════════════════════════════════════╝

REQUIRED_REFERRALS = 3

def membership_choice_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👤 تأیید توسط ادمین", callback_data="join_admin")],
        [InlineKeyboardButton("🎁 دعوت ۳ نفر (رایگان و فوری)", callback_data="join_referral")],
        [InlineKeyboardButton("💳 خرید اشتراک (به‌زودی)", callback_data="join_paid")],
    ])

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    was_new_user = get_user(user.id) is None
    prior_state, _ = get_state(user.id)  # قبل از پاک شدن می‌خوانیمش تا بفهمیم اولین /start بعد از تأیید است یا نه
    save_user(user.id, user.username or "", user.first_name)
    ensure_referral_code(user.id)
    clear_state(user.id)

    # ─── ادمین‌ها (مالک اصلی یا اضافه‌شده از پنل) هیچ‌وقت نیازی به تأیید ندارند ───
    just_auto_approved_admin = False
    if is_admin(user.id):
        db_user_check = get_user(user.id)
        if db_user_check and db_user_check[3] != 'active':
            approve_user(user.id)
            just_auto_approved_admin = True

    # ─── اطلاع‌رسانی فقط برای ورود اول (عضویت جدید) به ادمین ───
    if not is_admin(user.id) and was_new_user:
        await notify_admins(
            context.bot,
            f"🆕 کاربر جدید\n👤 {user.first_name}\n🆔 {user.id}\n📛 @{user.username or 'ندارد'}"
        )

    # ─── پردازش لینک دعوت (/start R123456) ───
    if context.args:
        code = context.args[0].strip()
        referrer_id = get_user_by_referral_code(code)
        db_user_check = get_user(user.id)
        # فقط اگر این کاربر تا الان هیچ‌وقت وضعیت فعال/در انتظار تأیید نداشته
        # (یعنی همین الان برای اولین‌بار در حال ثبت‌نام است) دعوت ثبت می‌شود
        if referrer_id and db_user_check and db_user_check[3] == 'pending' and not db_user_check[8]:
            conn = sqlite3.connect("bot.db")
            c = conn.cursor()
            c.execute("UPDATE users SET referred_by=? WHERE user_id=?", (referrer_id, user.id))
            conn.commit()
            conn.close()
            recorded = record_referral(referrer_id, user.id)
            if recorded:
                new_count = get_referred_count(referrer_id)
                try:
                    if new_count >= REQUIRED_REFERRALS:
                        referrer = get_user(referrer_id)
                        if referrer and referrer[3] == 'pending':
                            approve_user_by_referral(referrer_id)
                            await context.bot.send_message(
                                referrer_id,
                                f"🎉 تبریک! {REQUIRED_REFERRALS} نفر با لینک شما عضو شدند.\n"
                                f"✅ حساب شما فعال شد.\n\n/start بزنید تا وارد شوید."
                            )
                    else:
                        await context.bot.send_message(
                            referrer_id,
                            f"👥 یک نفر جدید با لینک دعوت شما عضو شد.\n"
                            f"تعداد دعوت‌های موفق: {new_count} از {REQUIRED_REFERRALS}"
                        )
                except Exception as e:
                    print(f"خطای اطلاع‌رسانی به معرف: {e}")

    db_user = get_user(user.id)

    if db_user[3] == 'active':
        expires = datetime.strptime(db_user[5], "%Y-%m-%d %H:%M")
        if get_iran_now() >= expires:
            reject_user(user.id)
            await update.message.reply_text(f"⏳ {user.first_name} عزیز،\nاشتراک شما منقضی شده.")
            await notify_admins(context.bot, f"🔄 منقضی:\n👤 {user.first_name}\n🆔 {user.id}")
            return
        if prior_state == "first_active_start_pending" or just_auto_approved_admin:
            await update.message.reply_text(
                f"✅ {user.first_name} عزیز، خوش آمدید! 🎉\n\n"
                f"این چند بخش اصلی ربات هستند:\n\n"
                f"💰 بازارهای مالی — قیمت لحظه‌ای، هشدار قیمت، تحلیل هوش مصنوعی، ربات مدیریت سرمایه\n"
                f"🎮 سرگرمی — سلامت خانواده، موزیک\n"
                f"👶 فرزندان — قصه‌گویی هوشمند، یادآور تکالیف\n"
                f"⏰ یادآور و تسک\n"
                f"🛠 ابزارهای کاربردی — دستیار هوش مصنوعی، فیلترشکن\n\n"
                f"هر وقت خواستید توضیح کامل‌تر ببینید، دکمه‌ی ❓ راهنما را بزنید.\n\n"
                f"🏠 منوی اصلی:",
                reply_markup=main_menu(user.id)
            )
        else:
            await update.message.reply_text(f"✅ {user.first_name} عزیز، خوش آمدید!\n\n🏠 منوی اصلی:",
                reply_markup=main_menu(user.id))
        return

    if db_user[3] == 'rejected':
        await update.message.reply_text(f"❌ {user.first_name} عزیز،\nدسترسی شما تأیید نشده.")
        return

    # کاربر pending است و هنوز روش عضویت را انتخاب نکرده
    await update.message.reply_text(
        f"👋 سلام {user.first_name} عزیز!\n\n"
        f"برای استفاده از ربات، یکی از روش‌های زیر را انتخاب کنید:",
        reply_markup=membership_choice_menu()
    )

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = query.from_user.id
    db_user = get_user(user_id)

    # ─── ادمین ───
    if data.startswith("approve_") or data.startswith("reject_"):
        if not is_admin(user_id):
            await query.answer("⛔ این دکمه فقط برای ادمین است.", show_alert=True)
            return
        action, uid = data.split("_")
        uid = int(uid)
        target = get_user(uid)
        name = target[2] if target else str(uid)
        if action == "approve":
            approve_user(uid)
            target = get_user(uid)
            await query.edit_message_text(f"✅ {name} تأیید شد.\n📅 انقضا: {format_date_fa(target[5])}")
            await context.bot.send_message(
                uid,
                f"🎉 {name} عزیز!\n✅ تأیید شدید.\n📅 اشتراک تا {format_date_fa(target[5])}\n\n"
                f"💡 پیشنهاد می‌کنیم از بخش ⚙️ تنظیمات، اطلاعات پروفایل خود (ایمیل، تلفن، تاریخ تولد، "
                f"شهر، کشور و شغل) را تکمیل کنید تا در آینده بتوانید از امکانات بیشتر ربات استفاده کنید.\n\n"
                f"/start بزنید."
            )
        else:
            reject_user(uid)
            await query.edit_message_text(f"❌ {name} رد شد.")
            await context.bot.send_message(uid, f"❌ {name} عزیز،\nمتأسفانه تأیید نشدید.")
        return

    # ─── انتخاب روش عضویت ───
    if data == "join_admin":
        if db_user and db_user[3] == 'pending':
            await query.edit_message_text("✅ درخواست شما ثبت شد.\n⏳ منتظر تأیید ادمین باشید.")
            await notify_admins(
                context.bot,
                f"🔔 کاربر جدید (درخواست تأیید ادمین):\n👤 {query.from_user.first_name}\n"
                f"🆔 {user_id}\n📛 @{query.from_user.username or 'ندارد'}",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("✅ تأیید", callback_data=f"approve_{user_id}"),
                    InlineKeyboardButton("❌ رد", callback_data=f"reject_{user_id}")
                ]])
            )
        return

    if data == "join_referral":
        if db_user and db_user[3] == 'pending':
            code = ensure_referral_code(user_id)
            bot_username = (await context.bot.get_me()).username
            link = f"https://t.me/{bot_username}?start={code}"
            current_count = get_referred_count(user_id)
            await query.edit_message_text(
                f"🎁 لینک دعوت اختصاصی شما:\n{link}\n\n"
                f"این لینک را برای {REQUIRED_REFERRALS} نفر بفرستید. وقتی هرکدام با همین "
                f"لینک وارد ربات شوند و /start بزنند، به‌صورت خودکار به‌عنوان دعوت شما "
                f"ثبت می‌شود. توجه: کسی که قبلاً (حتی قبل از این لینک) عضو ربات بوده "
                f"باشد، دوباره به‌عنوان دعوت جدید محسوب نمی‌شود.\n\n"
                f"تعداد دعوت‌های موفق شما تا الان: {current_count} از {REQUIRED_REFERRALS}"
            )
        return

    if data == "join_paid":
        await query.answer("💳 پرداخت اشتراک به‌زودی فعال می‌شود.", show_alert=True)
        return

    if not db_user or db_user[3] != 'active':
        await query.answer("❌ دسترسی شما فعال نیست.", show_alert=True)
        return

    # ─── منوی اصلی ───
    if data == "back_main":
        clear_state(user_id)
        await query.edit_message_text("🏠 منوی اصلی:", reply_markup=main_menu(user_id))
        return

    # ─── ROUTER: هر prefix به ماژول مربوطه ───
    if data.startswith("ps_"):
        await handle_position_size(query, user_id)  # ← حذف ماژول میزان خرید = پاک کن این خط

    elif data == "bt_upload" or data.startswith("btu_"):
        await handle_bt_upload_button(query, user_id)  # ← حذف ماژول آپلود فایل بک‌تست = پاک کن این خط

    elif data.startswith("bt_") or data == "fin_backtest":
        handled = await handle_backtest_buttons_extra(query, user_id, data)  # ← حذف ماژول بک‌تست = پاک کن این خط
        if not handled:
            await handle_backtest(query, user_id)  # ← حذف ماژول بک‌تست = پاک کن این خط

    elif data.startswith("ib_") or data == "ind_backtest":
        await handle_indicator_backtest(query, user_id)  # ← حذف ماژول بک‌تست اندیکاتوری = پاک کن این خط

    elif data.startswith("ci_") or data == "ind_custom":
        await handle_custom_indicator(query, user_id)  # ← حذف ماژول اندیکاتور شخصی = پاک کن این خط

    elif data.startswith("admin_send_music_"):
        if not is_admin(user_id):
            await query.answer("⛔ این دکمه فقط برای ادمین است.", show_alert=True)
            return
        request_id = int(data.replace("admin_send_music_", ""))
        request_row = get_music_request(request_id)
        if not request_row:
            await query.answer("این درخواست پیدا نشد.", show_alert=True)
            return
        requester_user_id, request_query_text, fulfilled = request_row
        set_state(user_id, "admin_awaiting_music_file", str(request_id))
        await query.edit_message_text(
            f"📤 فایل صوتی «{request_query_text}» را همین‌جا بفرستید تا خودکار برای کاربر ارسال شود."
        )

    elif data.startswith("admin_send_capital_"):  # ← حذف ماژول ربات مدیریت سرمایه = پاک کن این خط
        if not is_admin(user_id):
            await query.answer("⛔ این دکمه فقط برای ادمین است.", show_alert=True)
            return
        request_id = int(data.replace("admin_send_capital_", ""))
        request_row = get_capital_request(request_id)
        if not request_row:
            await query.answer("این درخواست پیدا نشد.", show_alert=True)
            return
        set_state(user_id, "admin_awaiting_capital_content", str(request_id))
        await query.edit_message_text(
            "📤 متن، عکس، فایل، صدا یا ویدیوی موردنظر را همین‌جا بفرستید.\n"
            "می‌توانید چند پیام پشت‌سرهم بفرستید؛ وقتی کارتان تمام شد، دکمه‌ی «✅ پایان ارسال» را بزنید.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ پایان ارسال", callback_data=f"admin_finish_capital_{request_id}")]
            ])
        )

    elif data.startswith("admin_finish_capital_"):  # ← حذف ماژول ربات مدیریت سرمایه = پاک کن این خط
        if not is_admin(user_id):
            await query.answer("⛔ این دکمه فقط برای ادمین است.", show_alert=True)
            return
        request_id = int(data.replace("admin_finish_capital_", ""))
        clear_state(user_id)
        await query.edit_message_text(f"✅ ارسال برای این کاربر تمام شد (درخواست #{request_id}).")

    elif data.startswith("msgc_start_"):  # ← حذف فیچر پیام به ادمین = پاک کن این ۵ بلوک elif
        if not is_admin(user_id):
            await query.answer("⛔ این دکمه فقط برای ادمین است.", show_alert=True)
            return
        chat_id = int(data.replace("msgc_start_", ""))
        chat_row = get_msg_chat(chat_id)
        if not chat_row:
            await query.answer("این پیام پیدا نشد.", show_alert=True)
            return
        target_user_id, status, _existing_admin_id = chat_row
        if status == "closed":
            await query.answer("این گفتگو قبلاً بسته شده.", show_alert=True)
            return
        update_msg_chat_status(chat_id, "active")
        assign_msg_chat_admin(chat_id, user_id)
        set_state(user_id, "admin_msg_chat_active", f"{chat_id}:{target_user_id}")
        set_state(target_user_id, "user_msg_chat_active", f"{chat_id}:{user_id}")
        try:
            await context.bot.send_message(
                target_user_id,
                "✅ ادمین به پیام شما پاسخ داد؛ از همین‌جا می‌توانید ادامه بدهید (متن، عکس، فایل، صدا یا ویدیو)."
            )
        except Exception as e:
            print(f"خطای اطلاع‌رسانی شروع پاسخ به کاربر: {e}")
        await query.edit_message_text(
            "✅ گفتگو باز شد. از همین‌جا هرچه بفرستید مستقیم برای کاربر می‌رود.",
            reply_markup=msg_chat_action_menu(chat_id)
        )

    elif data.startswith("msgc_transfer_pick_"):
        if not is_primary_admin(user_id):
            await query.answer("⛔ فقط مالک اصلی می‌تواند گفتگو را منتقل کند.", show_alert=True)
            return
        chat_id = int(data.replace("msgc_transfer_pick_", ""))
        chat_row = get_msg_chat(chat_id)
        if not chat_row or chat_row[1] == "closed":
            await query.answer("این گفتگو دیگر معتبر نیست.", show_alert=True)
            return
        extra_admins = get_all_extra_admins()
        if not extra_admins:
            await query.answer("هنوز هیچ ادمین اضافه‌ای تعریف نکرده‌اید (پنل مدیریت → مدیریت ادمین‌ها).", show_alert=True)
            return
        rows = [[InlineKeyboardButton(f"👤 {aid}", callback_data=f"msgc_transfer_to_{chat_id}_{aid}")]
                for aid, _added_at in extra_admins]
        rows.append([InlineKeyboardButton("🔙 انصراف", callback_data=f"msgc_cancel_transfer_{chat_id}")])
        await query.edit_message_text("↪️ این پیام به کدام ادمین منتقل شود؟", reply_markup=InlineKeyboardMarkup(rows))

    elif data.startswith("msgc_transfer_to_"):
        if not is_primary_admin(user_id):
            await query.answer("⛔ فقط مالک اصلی می‌تواند گفتگو را منتقل کند.", show_alert=True)
            return
        remainder = data.replace("msgc_transfer_to_", "")
        chat_id_str, target_admin_id_str = remainder.split("_")
        chat_id, target_admin_id = int(chat_id_str), int(target_admin_id_str)
        chat_row = get_msg_chat(chat_id)
        if not chat_row or chat_row[1] == "closed":
            await query.answer("این گفتگو دیگر معتبر نیست.", show_alert=True)
            return
        target_user_id = chat_row[0]
        original_text = get_msg_chat_original_text(chat_id)
        assign_msg_chat_admin(chat_id, target_admin_id)
        try:
            requester_target = await context.bot.get_chat(target_user_id)
            await context.bot.send_message(
                target_admin_id,
                msg_request_text(requester_target.first_name or "کاربر", requester_target.username, target_user_id, original_text),
                reply_markup=msg_request_assigned_menu(chat_id)
            )
        except Exception as e:
            print(f"خطای ارسال پیام منتقل‌شده به ادمین {target_admin_id}: {e}")
        await query.edit_message_text(f"✅ پیام به ادمین {target_admin_id} منتقل شد.")

    elif data.startswith("msgc_cancel_transfer_"):
        chat_id = int(data.replace("msgc_cancel_transfer_", ""))
        chat_row = get_msg_chat(chat_id)
        if not chat_row:
            await query.answer("این پیام پیدا نشد.", show_alert=True)
            return
        target_user_id, status, _admin_id = chat_row
        original_text = get_msg_chat_original_text(chat_id)
        try:
            requester_target = await context.bot.get_chat(target_user_id)
            req_name, req_username = requester_target.first_name or "کاربر", requester_target.username
        except Exception:
            req_name, req_username = "کاربر", None
        await query.edit_message_text(
            msg_request_text(req_name, req_username, target_user_id, original_text),
            reply_markup=msg_request_primary_menu(chat_id)
        )

    elif data.startswith("msgc_end_"):
        if not is_admin(user_id):
            await query.answer("⛔ این دکمه فقط برای ادمین است.", show_alert=True)
            return
        chat_id = int(data.replace("msgc_end_", ""))
        chat_row = get_msg_chat(chat_id)
        if not chat_row:
            await query.answer("این گفتگو پیدا نشد.", show_alert=True)
            return
        target_user_id, status, chat_admin_id = chat_row
        update_msg_chat_status(chat_id, "closed")
        clear_state(chat_admin_id or user_id)
        clear_state(target_user_id)
        await query.edit_message_text("✅ گفتگو بسته شد.")
        try:
            await context.bot.send_message(
                target_user_id,
                "🔒 گفتگو با ادمین بسته شد. در صورت نیاز دوباره از ⚙️ تنظیمات → 📩 پیام به ادمین اقدام کنید."
            )
        except Exception as e:
            print(f"خطای اطلاع‌رسانی پایان گفتگو به کاربر: {e}")

    elif data.startswith("lic_start_"):  # ← حذف فیچر چت لایسنس = پاک کن این بلوک
        if not is_admin(user_id):
            await query.answer("⛔ این دکمه فقط برای ادمین است.", show_alert=True)
            return
        chat_id = int(data.replace("lic_start_", ""))
        chat_row = get_license_chat(chat_id)
        if not chat_row:
            await query.answer("این درخواست پیدا نشد.", show_alert=True)
            return
        target_user_id, status, _existing_admin_id = chat_row
        if status == "closed":
            await query.answer("این مکالمه قبلاً بسته شده.", show_alert=True)
            return
        update_license_chat_status(chat_id, "active")
        assign_license_chat_admin(chat_id, user_id)
        set_state(user_id, "admin_chat_active", f"{chat_id}:{target_user_id}")
        set_state(target_user_id, "user_chat_active", f"{chat_id}:{user_id}")
        try:
            await context.bot.send_message(
                target_user_id,
                "✅ ادمین مکالمه را شروع کرد؛ از همین‌جا می‌توانید متن، عکس، فایل، صدا "
                "یا ویدیو مستقیم بفرستید."
            )
        except Exception as e:
            print(f"خطای اطلاع‌رسانی شروع مکالمه به کاربر: {e}")
        await query.edit_message_text(
            "✅ مکالمه شروع شد. از همین‌جا هرچه بفرستید (حتی فوروارد از کانال دیگر) "
            "مستقیم برای کاربر می‌رود؛ یا از دکمه‌های زیر برای ارسال سریع فایل/متن آماده استفاده کنید.",
            reply_markup=license_chat_action_menu(chat_id)
        )

    elif data.startswith("lic_transfer_pick_"):  # ← حذف انتقال چت لایسنس = پاک کن این ۲ بلوک elif
        if not is_primary_admin(user_id):
            await query.answer("⛔ فقط مالک اصلی می‌تواند مکالمه را منتقل کند.", show_alert=True)
            return
        chat_id = int(data.replace("lic_transfer_pick_", ""))
        chat_row = get_license_chat(chat_id)
        if not chat_row or chat_row[1] == "closed":
            await query.answer("این مکالمه دیگر معتبر نیست.", show_alert=True)
            return
        extra_admins = get_all_extra_admins()
        if not extra_admins:
            await query.answer("هنوز هیچ ادمین اضافه‌ای تعریف نکرده‌اید (پنل مدیریت → مدیریت ادمین‌ها).", show_alert=True)
            return
        rows = [[InlineKeyboardButton(f"👤 {aid}", callback_data=f"lic_transfer_to_{chat_id}_{aid}")]
                for aid, _added_at in extra_admins]
        rows.append([InlineKeyboardButton("🔙 انصراف", callback_data=f"lic_cancel_transfer_{chat_id}")])
        await query.edit_message_text("↪️ این مکالمه به کدام ادمین منتقل شود؟", reply_markup=InlineKeyboardMarkup(rows))

    elif data.startswith("lic_transfer_to_"):
        if not is_primary_admin(user_id):
            await query.answer("⛔ فقط مالک اصلی می‌تواند مکالمه را منتقل کند.", show_alert=True)
            return
        remainder = data.replace("lic_transfer_to_", "")
        chat_id_str, target_admin_id_str = remainder.split("_")
        chat_id, target_admin_id = int(chat_id_str), int(target_admin_id_str)
        chat_row = get_license_chat(chat_id)
        if not chat_row or chat_row[1] == "closed":
            await query.answer("این مکالمه دیگر معتبر نیست.", show_alert=True)
            return
        target_user_id = chat_row[0]
        requester_target = await context.bot.get_chat(target_user_id)
        assign_license_chat_admin(chat_id, target_admin_id)
        try:
            await context.bot.send_message(
                target_admin_id,
                license_request_text(requester_target.first_name or "کاربر", requester_target.username, target_user_id),
                reply_markup=license_request_assigned_menu(chat_id)
            )
        except Exception as e:
            print(f"خطای ارسال درخواست منتقل‌شده به ادمین {target_admin_id}: {e}")
        await query.edit_message_text(f"✅ مکالمه به ادمین {target_admin_id} منتقل شد.")

    elif data.startswith("lic_cancel_transfer_"):
        chat_id = int(data.replace("lic_cancel_transfer_", ""))
        chat_row = get_license_chat(chat_id)
        if not chat_row:
            await query.answer("این مکالمه پیدا نشد.", show_alert=True)
            return
        target_user_id, status, _admin_id = chat_row
        try:
            requester_target = await context.bot.get_chat(target_user_id)
            req_name, req_username = requester_target.first_name or "کاربر", requester_target.username
        except Exception:
            req_name, req_username = "کاربر", None
        await query.edit_message_text(
            license_request_text(req_name, req_username, target_user_id),
            reply_markup=license_request_primary_menu(chat_id)
        )

    elif data.startswith("lic_end_"):  # ← حذف فیچر چت لایسنس = پاک کن این بلوک
        if not is_admin(user_id):
            await query.answer("⛔ این دکمه فقط برای ادمین است.", show_alert=True)
            return
        chat_id = int(data.replace("lic_end_", ""))
        chat_row = get_license_chat(chat_id)
        if not chat_row:
            await query.answer("این مکالمه پیدا نشد.", show_alert=True)
            return
        target_user_id, status, chat_admin_id = chat_row
        update_license_chat_status(chat_id, "closed")
        clear_state(chat_admin_id or user_id)
        clear_state(target_user_id)
        await query.edit_message_text("✅ مکالمه بسته شد.")
        try:
            await context.bot.send_message(
                target_user_id,
                "🔒 مکالمه با ادمین بسته شد. در صورت نیاز دوباره از منوی «ربات مدیریت سرمایه» اقدام کنید."
            )
        except Exception as e:
            print(f"خطای اطلاع‌رسانی پایان مکالمه به کاربر: {e}")

    elif data.startswith("lic_send_paid_"):  # ← حذف کتابخانه/چت لایسنس = پاک کن این بلوک
        if not is_admin(user_id):
            await query.answer("⛔ این دکمه فقط برای ادمین است.", show_alert=True)
            return
        chat_id = int(data.replace("lic_send_paid_", ""))
        chat_row = get_license_chat(chat_id)
        if not chat_row or chat_row[1] == "closed":
            await query.answer("این مکالمه دیگر فعال نیست.", show_alert=True)
            return
        target_user_id = chat_row[0]
        paid_setting = get_bot_setting("paid_file")
        if not paid_setting:
            await query.answer("هنوز فایل نسخه‌ی پولی از پنل مدیریت تنظیم نشده.", show_alert=True)
            return
        import json
        info = json.loads(paid_setting)
        sent = await send_content_by_type(
            context, target_user_id, info["file_type"], info.get("file_id"), info.get("caption"),
            default_caption="💳 نسخه‌ی پولی ربات مدیریت سرمایه"
        )
        if sent:
            await query.answer("✅ ارسال شد.")
            await query.message.reply_text("✅ فایل نسخه‌ی پولی برای کاربر ارسال شد.", reply_markup=license_chat_action_menu(chat_id))
        else:
            await query.answer("❌ خطا در ارسال.", show_alert=True)

    elif data.startswith("lic_pick_file_"):
        chat_id = int(data.replace("lic_pick_file_", ""))
        files = get_admin_files()
        if not files:
            await query.answer("هنوز فایلی در کتابخانه ثبت نشده.", show_alert=True)
            return
        rows = [[InlineKeyboardButton(f"📄 {label}", callback_data=f"lic_sendfile_{chat_id}_{fid}")]
                for fid, label, *_rest in files]
        rows.append([InlineKeyboardButton("🔙 برگشت", callback_data=f"lic_back_{chat_id}")])
        await query.edit_message_text("📚 کدام فایل ارسال شود؟", reply_markup=InlineKeyboardMarkup(rows))

    elif data.startswith("lic_sendfile_"):
        remainder = data.replace("lic_sendfile_", "")
        chat_id_str, file_db_id_str = remainder.split("_")
        chat_id, file_db_id = int(chat_id_str), int(file_db_id_str)
        chat_row = get_license_chat(chat_id)
        if not chat_row or chat_row[1] == "closed":
            await query.answer("این مکالمه دیگر فعال نیست.", show_alert=True)
            return
        target_user_id = chat_row[0]
        row = get_admin_file(file_db_id)
        if not row:
            await query.answer("این فایل پیدا نشد.", show_alert=True)
            return
        label, file_type, file_id, caption = row
        sent = await send_content_by_type(context, target_user_id, file_type, file_id, caption, default_caption=label)
        if sent:
            await query.edit_message_text(f"✅ «{label}» برای کاربر ارسال شد.", reply_markup=license_chat_action_menu(chat_id))
        else:
            await query.answer("❌ خطا در ارسال.", show_alert=True)

    elif data.startswith("lic_pick_text_"):
        chat_id = int(data.replace("lic_pick_text_", ""))
        texts = get_admin_texts()
        if not texts:
            await query.answer("هنوز متنی در کتابخانه ثبت نشده.", show_alert=True)
            return
        rows = [[InlineKeyboardButton(f"💬 {label}", callback_data=f"lic_sendtext_{chat_id}_{tid}")]
                for tid, label, _content in texts]
        rows.append([InlineKeyboardButton("🔙 برگشت", callback_data=f"lic_back_{chat_id}")])
        await query.edit_message_text("💬 کدام متن ارسال شود؟", reply_markup=InlineKeyboardMarkup(rows))

    elif data.startswith("lic_sendtext_"):
        remainder = data.replace("lic_sendtext_", "")
        chat_id_str, text_id_str = remainder.split("_")
        chat_id, text_id = int(chat_id_str), int(text_id_str)
        chat_row = get_license_chat(chat_id)
        if not chat_row or chat_row[1] == "closed":
            await query.answer("این مکالمه دیگر فعال نیست.", show_alert=True)
            return
        target_user_id = chat_row[0]
        row = get_admin_text(text_id)
        if not row:
            await query.answer("این متن پیدا نشد.", show_alert=True)
            return
        label, content = row
        try:
            await context.bot.send_message(target_user_id, content)
            await query.edit_message_text(f"✅ متن «{label}» برای کاربر ارسال شد.", reply_markup=license_chat_action_menu(chat_id))
        except Exception as e:
            await query.answer(f"❌ خطا در ارسال: {e}", show_alert=True)

    elif data.startswith("lic_back_"):
        chat_id = int(data.replace("lic_back_", ""))
        await query.edit_message_text(
            "💬 در حال مکالمه. از دکمه‌ها استفاده کنید یا مستقیم پیام بفرستید.",
            reply_markup=license_chat_action_menu(chat_id)
        )

    elif data == "menu_ai_market" or data.startswith("aim_"):
        await handle_market_ai(query, user_id, context)  # ← حذف ماژول تحلیل بازار = پاک کن این خط

    elif data == "fin_news":
        await handle_forex_news(query, user_id)  # ← حذف ماژول اخبار = پاک کن این خط

    elif data.startswith("menu_financial") or data.startswith("fin_") or data.startswith("wl_") or data.startswith("alarm_") or data.startswith("back_financial") or data.startswith("cap_"):
        await handle_financial(query, user_id, context)  # ← حذف ماژول = پاک کن این خط

    elif data.startswith("menu_reminder") or data.startswith("rem_"):
        await handle_reminder(query, user_id)  # ← حذف ماژول = پاک کن این خط

    elif data.startswith("menu_ai") or data.startswith("ai_"):
        await handle_ai(query, user_id)  # ← حذف ماژول = پاک کن این خط

    elif data.startswith("menu_kids") or data.startswith("kids_"):
        await handle_kids(query, user_id)  # ← حذف ماژول = پاک کن این خط

    elif (data.startswith("menu_fun") or data == "back_fun" or
          data.startswith("menu_health") or data.startswith("health_") or
          data.startswith("menu_music") or data.startswith("music_")):
        await handle_fun(query, user_id)  # ← حذف ماژول سرگرمی = پاک کن این خط

    elif data == "menu_tools":
        await handle_tools(query, user_id)  # ← حذف ماژول ابزارهای کاربردی = پاک کن این خط

    elif data.startswith("menu_settings") or data.startswith("set_") or data == "menu_social_links":
        await handle_settings(query, user_id, context)  # ← حذف ماژول = پاک کن این خط

    elif data == "menu_help":  # ← حذف راهنما = پاک کن این بلوک
        help_text = get_bot_setting("help_text") or DEFAULT_HELP_TEXT
        await query.edit_message_text(
            help_text[:4000],
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 برگشت", callback_data="back_main")]])
        )

    elif data.startswith("menu_antifilter") or data.startswith("af_"):
        await handle_antifilter(query, user_id)  # ← حذف ماژول فیلترشکن = پاک کن این خط

    elif data == "admin_panel" or data.startswith("adm_"):
        await handle_admin_panel(query, user_id, context)  # ← حذف ماژول پنل ادمین = پاک کن این خط

    else:
        await query.answer("🔧 به زودی اضافه می‌شود!", show_alert=True)
 
async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text_raw = update.message.text.strip() if update.message.text else ""

    db_user = get_user(user_id)
    if not db_user or db_user[3] != 'active':
        return

    text = text_raw
    state, _ = get_state(user_id)

    if state in ("admin_awaiting_capital_content", "adm_awaiting_trial_file_content",
                  "adm_awaiting_paid_file_content", "adm_awaiting_new_file_content", "adm_awaiting_edit_file_content",
                  "admin_chat_active", "user_chat_active",
                  "admin_msg_chat_active", "user_msg_chat_active"):  # ← حذف ماژول ربات مدیریت سرمایه/پیام به ادمین/چت لایسنس/کتابخانه = پاک کن این خط
        handled = await handle_admin_generic_content(update, context)
        if handled:
            return

    if state in ["wl_add", "alarm_symbol", "alarm_price"]:
        await handle_financial_message(user_id, text, update)  # ← حذف ماژول = پاک کن این خط

    elif state and (state.startswith("music_") or state.startswith("health_")):
        await handle_fun_message(user_id, text, update)  # ← حذف ماژول سرگرمی = پاک کن این خط

    elif state and state.startswith("adm_"):
        await handle_admin_panel_message(user_id, text, update)  # ← حذف ماژول پنل ادمین = پاک کن این خط

    elif state and state.startswith("set_"):
        await handle_settings_message(user_id, text, update, context)  # ← حذف ماژول تنظیمات = پاک کن این خط

    elif state and state.startswith("aim_"):
        await handle_market_ai_message(user_id, text, update)  # ← حذف ماژول تحلیل بازار = پاک کن این خط

    elif state == "ai_chatting":
        await handle_ai_message(user_id, text, update)  # ← حذف ماژول دستیار AI = پاک کن این خط

    elif state and state.startswith("kids_"):
        await handle_kids_message(user_id, text, update)  # ← حذف ماژول فرزندان = پاک کن این خط

    elif state and state.startswith("ps_"):
        await handle_position_size_message(user_id, text, update)  # ← حذف ماژول میزان خرید = پاک کن این خط

    elif state and state.startswith("bt_upload_"):
        if state == "bt_upload_awaiting_file":
            await update.message.reply_text("📁 لطفاً فایل CSV یا Excel را به‌عنوان فایل (نه متن) ارسال کنید.")
        # سایر state های bt_upload_ (مثل انتخاب تایم‌فریم/استراتژی) فقط با دکمه پیش می‌روند

    elif state and state.startswith("bt_"):
        await handle_backtest_message(user_id, text, update)  # ← حذف ماژول بک‌تست = پاک کن این خط

    elif state and state.startswith("ib_"):
        await handle_indicator_backtest_message(user_id, text, update)  # ← حذف ماژول بک‌تست اندیکاتوری = پاک کن این خط

    elif state and state.startswith("ci_"):
        await handle_custom_indicator_message(user_id, text, update)  # ← حذف ماژول اندیکاتور شخصی = پاک کن این خط

    elif state and state.startswith("rem_"):
        await handle_reminder_message(user_id, text, update)  # ← حذف ماژول یادآور = پاک کن این خط

    else:
        # یا هیچ state فعالی نبود، یا state موجود بود ولی منتظر دکمه بود نه متن
        # (مثلاً انتخاب جهت آلارم) و کاربر به‌جای زدن دکمه متن فرستاد. قبلاً در
        # هر دو حالت هیچ پاسخی داده نمی‌شد (سکوت کامل)؛ کاربر فکر می‌کرد ربات خراب شده
        await update.message.reply_text(
            "🤔 متوجه این پیام نشدم.\n\n"
            "از دکمه‌های منو استفاده کنید، یا دستور /start را بزنید تا به منوی اصلی برگردید.\n"
            "برای راهنما هم می‌توانید دستور /help را بزنید."
        )

# ╔══════════════════════════════════════════════════════════════╗
# ║                       RUN BOT                                ║
# ╚══════════════════════════════════════════════════════════════╝

async def post_init(app):
    asyncio.create_task(alarm_loop(app))
    asyncio.create_task(reminder_loop(app))  # ← حذف ماژول یادآور = پاک کن این خط
    asyncio.create_task(license_escalation_loop(app))  # ← حذف انتقال خودکار چت لایسنس = پاک کن این خط
    asyncio.create_task(msg_chat_escalation_loop(app))  # ← حذف انتقال خودکار پیام به ادمین = پاک کن این خط
    print("⏰ سیستم آلارم فعال شد")

async def license_escalation_loop(app):
    """
    هر ۶۰ ثانیه چک می‌کند که آیا درخواست لایسنسی وجود دارد که هنوز
    'pending' مانده، دست مالک اصلی است (منتقل نشده)، و بیش از ۱۵ دقیقه
    از ساختش گذشته. اگر بله و حداقل یک ادمین اضافه وجود داشته باشد،
    به‌صورت round-robin (به‌ترتیب و مساوی) به یکی از آن‌ها واگذار می‌شود.
    """
    while True:
        try:
            overdue = get_overdue_pending_license_chats(minutes=15)
            for chat_id, target_user_id in overdue:
                next_admin_id = get_next_round_robin_admin()
                if next_admin_id is None:
                    continue  # هیچ ادمین اضافه‌ای نیست -- کاری نمی‌شود کرد، دست مالک اصلی می‌ماند
                assign_license_chat_admin(chat_id, next_admin_id)
                try:
                    requester_target = await app.bot.get_chat(target_user_id)
                    req_name, req_username = requester_target.first_name or "کاربر", requester_target.username
                except Exception:
                    req_name, req_username = "کاربر", None
                try:
                    await app.bot.send_message(
                        next_admin_id,
                        f"⏰ این درخواست بیش از ۱۵ دقیقه بی‌پاسخ مانده و خودکار به شما واگذار شد:\n\n"
                        + license_request_text(req_name, req_username, target_user_id),
                        reply_markup=license_request_assigned_menu(chat_id)
                    )
                except Exception as e:
                    print(f"خطای اطلاع‌رسانی واگذاری خودکار به ادمین {next_admin_id}: {e}")
                for primary_id in PRIMARY_ADMIN_IDS:
                    try:
                        await app.bot.send_message(
                            primary_id,
                            f"ℹ️ درخواست لایسنس کاربر {target_user_id} بعد از ۱۵ دقیقه بی‌پاسخی، "
                            f"خودکار به ادمین {next_admin_id} واگذار شد."
                        )
                    except Exception as e:
                        print(f"خطای اطلاع کوتاه واگذاری خودکار به مالک اصلی {primary_id}: {e}")
        except Exception as e:
            print(f"خطای license_escalation_loop: {e}")
        await asyncio.sleep(60)

async def msg_chat_escalation_loop(app):
    """
    مشابه license_escalation_loop ولی برای گفتگوهای «📩 پیام به ادمین» --
    هر ۶۰ ثانیه چک می‌کند و پیام‌های بیش‌از-۱۵-دقیقه‌ی بی‌پاسخ را به‌صورت
    round-robin مستقل (صف جدا از چت لایسنس) واگذار می‌کند. همچنین در هر
    دور، لاگ گفتگوهای قدیمی‌تر از ۱۰ روز را هم پاک می‌کند.
    """
    while True:
        try:
            purge_old_msg_chat_logs(days=10)
            overdue = get_overdue_pending_msg_chats(minutes=15)
            for chat_id, target_user_id in overdue:
                next_admin_id = get_next_round_robin_admin(setting_key="msg_round_robin_index")
                if next_admin_id is None:
                    continue
                assign_msg_chat_admin(chat_id, next_admin_id)
                original_text = get_msg_chat_original_text(chat_id)
                try:
                    requester_target = await app.bot.get_chat(target_user_id)
                    req_name, req_username = requester_target.first_name or "کاربر", requester_target.username
                except Exception:
                    req_name, req_username = "کاربر", None
                try:
                    await app.bot.send_message(
                        next_admin_id,
                        f"⏰ این پیام بیش از ۱۵ دقیقه بی‌پاسخ مانده و خودکار به شما واگذار شد:\n\n"
                        + msg_request_text(req_name, req_username, target_user_id, original_text),
                        reply_markup=msg_request_assigned_menu(chat_id)
                    )
                except Exception as e:
                    print(f"خطای اطلاع‌رسانی واگذاری خودکار پیام به ادمین {next_admin_id}: {e}")
                for primary_id in PRIMARY_ADMIN_IDS:
                    try:
                        await app.bot.send_message(
                            primary_id,
                            f"ℹ️ پیام کاربر {target_user_id} بعد از ۱۵ دقیقه بی‌پاسخی، "
                            f"خودکار به ادمین {next_admin_id} واگذار شد."
                        )
                    except Exception as e:
                        print(f"خطای اطلاع کوتاه واگذاری خودکار پیام به مالک اصلی {primary_id}: {e}")
        except Exception as e:
            print(f"خطای msg_chat_escalation_loop: {e}")
        await asyncio.sleep(60)

# ─── Web server ساده برای Railway (بدون این Railway بات رو خاموش می‌کنه) ───
class HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"Bot is running!")
    def log_message(self, *args):
        pass

def run_web():
    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), HealthHandler)
    server.serve_forever()

Thread(target=run_web, daemon=True).start()

init_db()
init_backtest_db()  # ← حذف ماژول بک‌تست = پاک کن این خط
init_fun_db()  # ← حذف ماژول سرگرمی = پاک کن این خط
init_ai_db()  # ← حذف ماژول دستیار AI = پاک کن این خط
async def handle_document_dispatcher(update, context):
    """
    دیسپچر مرکزی برای همه‌ی فایل‌های ورودی کاربر (چون تلگرام فقط اولین
    MessageHandler با فیلتر Document.ALL را اجرا می‌کند، همه‌ی منطق
    دریافت فایل باید از یک نقطه‌ی واحد عبور کند). بر اساس state فعلی
    کاربر، به تابع مناسب هدایت می‌شود.
    """
    handled = await handle_admin_generic_content(update, context)  # ← حذف ماژول ربات مدیریت سرمایه/پیام به ادمین = پاک کن این خط
    if handled:
        return
    handled = await handle_bt_upload_document(update, context)  # ← حذف ماژول آپلود فایل بک‌تست = پاک کن این خط
    if handled:
        return
    await handle_custom_indicator_document(update, context)  # ← حذف ماژول اندیکاتور شخصی = پاک کن این خط

async def handle_admin_capital_content(update, context):
    """
    دریافت هر نوع محتوا (متن، عکس، فایل، صدا، ویدیو) از ادمین -- وقتی روی
    دکمه‌ی «آماده‌ام ارسال کنم» برای یک درخواست ربات مدیریت سرمایه زده و
    در state انتظار آن است -- و ارسال خودکار همان محتوا برای کاربری که
    درخواست داده بود. طبق همان الگوی درخواست موزیک، این پیام از طرف خود
    ربات برای کاربر ارسال می‌شود (نه فوروارد مستقیم، که هویت ادمین را
    برای کاربر آشکار می‌کرد). ادمین می‌تواند چند پیام پشت‌سرهم بفرستد؛
    state فقط با زدن دکمه‌ی «پایان ارسال» بسته می‌شود. خروجی: True اگر
    پردازش شد، False اگر این ماژول منتظر محتوا نبود (تا دیسپچر مربوطه
    به مسیر بعدی برود).
    """
    admin_user_id = update.effective_user.id
    if not is_admin(admin_user_id):
        return False
    state, data = get_state(admin_user_id)
    if state != "admin_awaiting_capital_content":
        return False

    request_id = int(data)
    request_row = get_capital_request(request_id)
    if not request_row:
        await update.message.reply_text("❌ این درخواست دیگر معتبر نیست.")
        clear_state(admin_user_id)
        return True

    requester_user_id, fulfilled = request_row
    msg = update.message
    default_caption = "📊 اطلاعات مربوط به ربات مدیریت سرمایه که درخواست کرده بودید:"
    finish_button = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ پایان ارسال", callback_data=f"admin_finish_capital_{request_id}")]
    ])

    prefix = f"{default_caption}\n\n"  # relay_free_message خودش برای متن، رسانه‌ی بدون کپشن، و
                                        # رسانه‌ی دارای کپشن به‌درستی از این استفاده می‌کند
    sent = await relay_free_message(msg, context, requester_user_id, prefix=prefix)

    if not sent:
        await update.message.reply_text(
            "❌ نوع پیام پشتیبانی نمی‌شود؛ لطفاً متن، عکس، فایل، صدا یا ویدیو بفرستید.",
            reply_markup=finish_button
        )
        return True

    mark_capital_request_fulfilled(request_id)
    await update.message.reply_text(
        "✅ برای کاربر ارسال شد. می‌توانید پیام بعدی را بفرستید یا پایان بدهید.",
        reply_markup=finish_button
    )
    return True

def summarize_message_for_log(msg):
    """خلاصه‌ی کوتاه یک پیام تلگرام برای ثبت در لاگ گفتگو (برای اکسل بعدی) -- برای متن خود متن، برای رسانه فقط نوعش"""
    if msg.text:
        return msg.text[:500]
    if msg.document:
        base = f"[فایل: {msg.document.file_name or 'بدون‌نام'}]"
        return f"{base} -- {msg.caption}" if msg.caption else base
    if msg.photo:
        return f"[عکس] -- {msg.caption}" if msg.caption else "[عکس]"
    if msg.audio:
        return f"[فایل صوتی] -- {msg.caption}" if msg.caption else "[فایل صوتی]"
    if msg.video:
        return f"[ویدیو] -- {msg.caption}" if msg.caption else "[ویدیو]"
    if msg.voice:
        return "[پیام صوتی]"
    return "[نوع پیام ناشناخته]"

async def handle_msg_chat_content(update, context):
    """
    رله‌ی دوطرفه‌ی محتوا (متن/عکس/فایل/صدا/ویدیو) بین ادمین و کاربری که
    از طریق «📩 پیام به ادمین» گفتگو باز کرده -- دقیقاً مثل چت لایسنس،
    با این تفاوت که هر پیام (هر دو جهت) در msg_chat_logs هم ثبت می‌شود
    (برای گزارش اکسل ۱۰-روزه‌ی ادمین). خروجی: True اگر پردازش شد،
    False اگر این ماژول منتظر محتوا نبود.
    """
    sender_id = update.effective_user.id
    state, data = get_state(sender_id)
    msg = update.message

    if state == "admin_msg_chat_active" and is_admin(sender_id):
        try:
            chat_id_str, target_user_id_str = data.split(":")
            chat_id = int(chat_id_str)
            target_user_id = int(target_user_id_str)
        except Exception:
            await update.message.reply_text("❌ خطای داخلی؛ لطفاً دوباره از پیام کاربر شروع کنید.")
            clear_state(sender_id)
            return True

        action_menu = msg_chat_action_menu(chat_id)
        sent = await relay_free_message(msg, context, target_user_id)
        if sent:
            log_msg_chat_message(chat_id, target_user_id, sender_id, "admin_to_user", summarize_message_for_log(msg))
            await update.message.reply_text("✅ برای کاربر ارسال شد.", reply_markup=action_menu)
        else:
            await update.message.reply_text(
                "❌ نوع پیام پشتیبانی نمی‌شود؛ لطفاً متن، عکس، فایل، صدا یا ویدیو بفرستید.",
                reply_markup=action_menu
            )
        return True

    if state == "user_msg_chat_active":
        try:
            chat_id_str, owning_admin_id_str = data.split(":")
            chat_id = int(chat_id_str)
            owning_admin_id = int(owning_admin_id_str)
        except Exception:
            await update.message.reply_text("❌ خطای داخلی؛ لطفاً دوباره از ⚙️ تنظیمات → 📩 پیام به ادمین شروع کنید.")
            clear_state(sender_id)
            return True
        sent = await relay_free_message(msg, context, owning_admin_id, prefix=f"💬 پیام از کاربر (آیدی {sender_id}):\n")
        if sent:
            log_msg_chat_message(chat_id, sender_id, owning_admin_id, "user_to_admin", summarize_message_for_log(msg))
        else:
            await update.message.reply_text("❌ نوع پیام پشتیبانی نمی‌شود؛ لطفاً متن، عکس، فایل، صدا یا ویدیو بفرستید.")
        return True

    return False

async def handle_admin_keyed_file_upload(update, context):
    """
    آپلود محتوای ادمین برای یکی از دو «اسلات ثابت» در bot_settings --
    فایل نسخه‌ی تست (trial_file) و فایل نسخه‌ی پولی (paid_file). هر دو
    منطق کاملاً یکسانی دارند (یک کلید ثابت، همیشه جایگزین می‌شود، نه
    لیستی از چند مورد -- برخلاف کتابخانه‌ی فایل‌های آماده).
    خروجی: True اگر پردازش شد، False در غیر این صورت.
    """
    admin_user_id = update.effective_user.id
    if not is_admin(admin_user_id):
        return False
    state, _ = get_state(admin_user_id)
    state_map = {
        "adm_awaiting_trial_file_content": ("trial_file", "نسخه‌ی تست", admin_panel_menu(admin_user_id)),
        "adm_awaiting_paid_file_content": ("paid_file", "نسخه‌ی پولی", admin_library_menu()),
    }
    if state not in state_map:
        return False
    setting_key, label, back_menu = state_map[state]

    msg = update.message
    file_type, file_id = detect_message_content(msg)
    if file_type is None:
        await update.message.reply_text("❌ نوع پیام پشتیبانی نمی‌شود؛ لطفاً متن، عکس، فایل، صدا یا ویدیو بفرستید.")
        return True

    import json
    caption = msg.caption if file_type != "text" else msg.text
    set_bot_setting(setting_key, json.dumps({"file_type": file_type, "file_id": file_id, "caption": caption}))
    clear_state(admin_user_id)
    await update.message.reply_text(f"✅ فایل {label} ذخیره شد.", reply_markup=back_menu)
    return True

async def handle_admin_library_file_upload(update, context):
    """
    آپلود یک فایل جدید در کتابخانه‌ی فایل‌های آماده‌ی ادمین (حداکثر ۱۰
    عدد). مرحله‌ی دوم از یک فرایند دومرحله‌ای: مرحله‌ی اول (گرفتن اسم)
    در handle_admin_panel_message مدیریت می‌شود؛ این تابع فقط مرحله‌ی
    دوم (گرفتن خود محتوا) را انجام می‌دهد. خروجی: True اگر پردازش شد.
    """
    admin_user_id = update.effective_user.id
    if not is_admin(admin_user_id):
        return False
    state, label = get_state(admin_user_id)
    if state != "adm_awaiting_new_file_content":
        return False

    if count_admin_files() >= 10:
        await update.message.reply_text(
            "⛔ در حال حاضر ۱۰ فایل ثبت شده (حداکثر مجاز). اول یکی را از لیست حذف کنید.",
            reply_markup=admin_library_menu()
        )
        clear_state(admin_user_id)
        return True

    msg = update.message
    file_type, file_id = detect_message_content(msg)
    if file_type is None:
        await update.message.reply_text("❌ نوع پیام پشتیبانی نمی‌شود؛ لطفاً متن، عکس، فایل، صدا یا ویدیو بفرستید.")
        return True

    caption = msg.caption if file_type != "text" else msg.text
    add_admin_file(label, file_type, file_id, caption)
    clear_state(admin_user_id)
    await update.message.reply_text(f"✅ فایل «{label}» به کتابخانه اضافه شد.", reply_markup=admin_library_menu())
    return True

async def handle_admin_library_file_edit(update, context):
    """
    مرحله‌ی دوم ویرایش یک فایل موجود در کتابخانه (جایگزینی کامل بدون
    حذف و افزودن دوباره). مرحله‌ی اول (گرفتن اسم جدید) در
    handle_admin_panel_message مدیریت می‌شود؛ این تابع فقط مرحله‌ی دوم
    (گرفتن محتوای جدید) را انجام می‌دهد. خروجی: True اگر پردازش شد.
    """
    admin_user_id = update.effective_user.id
    if not is_admin(admin_user_id):
        return False
    state, data = get_state(admin_user_id)
    if state != "adm_awaiting_edit_file_content":
        return False

    try:
        file_db_id_str, new_label = data.split(":", 1)
        file_db_id = int(file_db_id_str)
    except Exception:
        await update.message.reply_text("❌ خطای داخلی؛ لطفاً دوباره از لیست فایل‌ها شروع کنید.", reply_markup=admin_library_menu())
        clear_state(admin_user_id)
        return True

    msg = update.message
    file_type, file_id = detect_message_content(msg)
    if file_type is None:
        await update.message.reply_text("❌ نوع پیام پشتیبانی نمی‌شود؛ لطفاً متن، عکس، فایل، صدا یا ویدیو بفرستید.")
        return True

    caption = msg.caption if file_type != "text" else msg.text
    update_admin_file(file_db_id, new_label, file_type, file_id, caption)
    clear_state(admin_user_id)
    await update.message.reply_text(f"✅ فایل «{new_label}» به‌روزرسانی شد.", reply_markup=admin_library_menu())
    return True

async def handle_admin_broadcast_content_media(update, context):
    """
    دریافت محتوای غیرمتنی (عکس/فایل/صدا/ویدیو -- حتی فوروارد از یک
    کانال دیگر) برای ارسال پیام همگانی از پنل مدیریت. پیام‌های متنی
    ساده از این تابع عبور می‌کنند (return False) تا توسط منطق قدیمی
    و ساده‌تر در handle_admin_panel_message پردازش شوند؛ این تابع فقط
    مسئول انواع رسانه‌ای است که آن تابع (چون فقط پیام‌های متنی را
    می‌بیند) اصلاً نمی‌تواند دریافت کند.
    خروجی: True اگر پردازش شد، False در غیر این صورت.
    """
    admin_user_id = update.effective_user.id
    if not is_admin(admin_user_id):
        return False
    state, target_data = get_state(admin_user_id)
    if state != "adm_awaiting_broadcast_content":
        return False

    msg = update.message
    if msg.text and not (msg.document or msg.photo or msg.audio or msg.video or msg.voice):
        return False  # متن خالص -- بگذار handle_admin_panel_message پردازشش کند

    file_type, file_id = detect_message_content(msg)
    if file_type is None:
        await update.message.reply_text("❌ نوع پیام پشتیبانی نمی‌شود؛ لطفاً متن، عکس، فایل، صدا یا ویدیو بفرستید.")
        return True

    target_ids = parse_broadcast_target_ids(target_data)
    if not target_ids:
        await update.message.reply_text("❌ هیچ مقصدی برای ارسال پیدا نشد.", reply_markup=admin_panel_menu(admin_user_id))
        clear_state(admin_user_id)
        return True

    caption = msg.caption
    import json
    set_state(admin_user_id, "adm_broadcast_ready",
              json.dumps({"target_ids": target_ids, "file_type": file_type, "file_id": file_id, "caption": caption}))
    await send_content_by_type(context, admin_user_id, file_type, file_id, caption)
    await update.message.reply_text(
        f"👆 پیش‌نمایش بالا. این محتوا به {len(target_ids)} کاربر ارسال خواهد شد. تأیید می‌کنید؟",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ بله، ارسال کن", callback_data="adm_broadcast_confirm_send")],
            [InlineKeyboardButton("❌ لغو", callback_data="admin_panel")],
        ])
    )
    return True

async def handle_license_chat_content(update, context):
    """
    رله‌ی دوطرفه‌ی محتوا (متن/عکس/فایل/صدا/ویدیو، حتی فوروارد از یک
    کانال دیگر) بین ادمین و کاربری که برای لایسنس پولی تأیید شده.
    هم سمت ادمین (state=admin_chat_active) و هم سمت کاربر عادی
    (state=user_chat_active) از همین یک تابع عبور می‌کنند. خروجی:
    True اگر پردازش شد، False اگر این ماژول منتظر محتوا نبود.
    """
    sender_id = update.effective_user.id
    state, data = get_state(sender_id)
    msg = update.message

    if state == "admin_chat_active" and is_admin(sender_id):
        try:
            chat_id_str, target_user_id_str = data.split(":")
            chat_id = int(chat_id_str)
            target_user_id = int(target_user_id_str)
        except Exception:
            await update.message.reply_text("❌ خطای داخلی؛ لطفاً دوباره از پیام درخواست شروع کنید.")
            clear_state(sender_id)
            return True

        action_menu = license_chat_action_menu(chat_id)
        sent = await relay_free_message(msg, context, target_user_id)
        if sent:
            await update.message.reply_text("✅ برای کاربر ارسال شد.", reply_markup=action_menu)
        else:
            await update.message.reply_text(
                "❌ نوع پیام پشتیبانی نمی‌شود؛ لطفاً متن، عکس، فایل، صدا یا ویدیو بفرستید.",
                reply_markup=action_menu
            )
        return True

    if state == "user_chat_active":
        # data به‌فرمت "chat_id:admin_id" است تا پیام کاربر دقیقاً به همان
        # ادمینی که مکالمه را شروع کرده برسد، نه لزوماً مالک اصلی
        try:
            _chat_id_str, owning_admin_id_str = data.split(":")
            owning_admin_id = int(owning_admin_id_str)
        except Exception:
            owning_admin_id = ADMIN_ID  # حالت قدیمی/خرابی داده -- به مالک اصلی برمی‌گردیم
        sent = await relay_free_message(msg, context, owning_admin_id, prefix=f"💬 پیام از کاربر (آیدی {sender_id}):\n")
        if not sent:
            await update.message.reply_text("❌ نوع پیام پشتیبانی نمی‌شود؛ لطفاً متن، عکس، فایل، صدا یا ویدیو بفرستید.")
        return True

    return False

async def handle_admin_generic_content(update, context):
    """
    دیسپچر ترکیبی برای همه‌ی فیچرهایی که نیاز به رد و بدل کردن محتوای
    آزاد (متن/عکس/فایل/صدا/ویدیو) بین ادمین و یک کاربر خاص دارند --
    ربات مدیریت سرمایه، پاسخ به پیام کاربر، تنظیم فایل نسخه‌ی تست/پولی،
    کتابخانه‌ی فایل‌های آماده، ارسال پیام همگانی، و چت زنده‌ی لایسنس.
    هر تابع خودش بر اساس state فعلی فرستنده تشخیص می‌دهد که آیا باید
    پردازش کند یا نه.
    """
    if await handle_admin_capital_content(update, context):
        return True
    if await handle_msg_chat_content(update, context):
        return True
    if await handle_admin_keyed_file_upload(update, context):
        return True
    if await handle_admin_library_file_upload(update, context):
        return True
    if await handle_admin_library_file_edit(update, context):
        return True
    if await handle_admin_broadcast_content_media(update, context):
        return True
    if await handle_license_chat_content(update, context):
        return True
    return False

async def handle_admin_music_audio(update, context):
    """
    دریافت فایل صوتی از ادمین (وقتی روی دکمه‌ی «آماده‌ام فایل را بفرستم»
    زده و در state انتظار فایل موزیک است) و ارسال خودکار همان فایل برای
    کاربری که این آهنگ را جستجو کرده بود -- طبق درخواست صریح شما، این
    پیام از طرف خود ربات برای کاربر ارسال می‌شود (نه فوروارد مستقیم از
    ادمین، که هویت ادمین را برای کاربر آشکار می‌کرد).
    """
    handled = await handle_admin_generic_content(update, context)  # ← حذف ماژول ربات مدیریت سرمایه/پیام به ادمین = پاک کن این خط
    if handled:
        return
    admin_user_id = update.effective_user.id
    if not is_admin(admin_user_id):
        return
    state, data = get_state(admin_user_id)
    if state != "admin_awaiting_music_file":
        return
    if not update.message.audio:
        await update.message.reply_text("❌ لطفاً فایل را به‌صورت Audio (نه Document) بفرستید.")
        return

    request_id = int(data)
    request_row = get_music_request(request_id)
    if not request_row:
        await update.message.reply_text("❌ این درخواست دیگر معتبر نیست.")
        clear_state(admin_user_id)
        return

    requester_user_id, query_text, fulfilled = request_row
    try:
        await context.bot.send_audio(
            chat_id=requester_user_id,
            audio=update.message.audio.file_id,
            caption=f"🎵 آهنگ «{query_text}» که جستجو کرده بودید، پیدا شد!"
        )
        mark_music_request_fulfilled(request_id)
        await update.message.reply_text(f"✅ آهنگ با موفقیت برای کاربر ارسال شد.")
    except Exception as e:
        await update.message.reply_text(f"❌ خطا در ارسال آهنگ به کاربر: {e}")
    clear_state(admin_user_id)

init_reminder_db()  # ← حذف ماژول یادآور = پاک کن این خط
init_custom_indicator_db()  # ← حذف ماژول اندیکاتور شخصی = پاک کن این خط
app = Application.builder().token(TOKEN).post_init(post_init).build()
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    help_text = get_bot_setting("help_text") or DEFAULT_HELP_TEXT
    await update.message.reply_text(help_text[:4000])

app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("help", help_command))  # ← حذف راهنما = پاک کن این خط
app.add_handler(CallbackQueryHandler(button_handler))
app.add_handler(MessageHandler(filters.Document.ALL, handle_document_dispatcher))
app.add_handler(MessageHandler(filters.AUDIO, handle_admin_music_audio))
app.add_handler(MessageHandler(filters.PHOTO | filters.VIDEO | filters.VOICE, handle_admin_generic_content))  # ← حذف ماژول ربات مدیریت سرمایه/پیام به ادمین = پاک کن این خط
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))
app.add_handler(MessageHandler(filters.ChatType.CHANNEL, handle_channel_post))  # ← حذف ماژول سرگرمی = پاک کن این خط
print("Bot is running...")
app.run_polling()